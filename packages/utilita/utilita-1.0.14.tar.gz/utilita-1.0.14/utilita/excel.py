import openpyxl as xl
import pandas as pd
import numpy as np
from deprecated import deprecated
import os
from typing import TypedDict
from io import BytesIO
from tempfile import NamedTemporaryFile
import re

class ExcelTableRows(TypedDict):
    headers: list
    rows: list

def workbook_to_bytes(wb: xl.Workbook, lock_worksheets: bool=False) -> BytesIO:
    """Takes an openpyxl workbook, and saves it as a BytesIO object.

    Params:
        wb (openpyxl.Workbook): Workbook object to save

        lock_worksheets (bool): Lock worksheets with no password.
    """
    if lock_worksheets:
        for x in wb.worksheets:
            x.protection.enable() # always do this before exporting because they automatically become unprotected somehow

    # If on windows, add delete=False to NamedTemporaryFile()
    ntf_delete = False if os.name == 'nt' else True
    with NamedTemporaryFile(delete=ntf_delete) as tmp:
        wb.save(tmp.name)
        excel_bytes = BytesIO(tmp.read())
        excel_bytes.seek(0)
        excel_bytes = excel_bytes.read()
    return excel_bytes

def df_to_worksheet(df: pd.DataFrame, cell: xl.cell.cell.Cell) -> None:
    """Takes a DataFrame and writes it to a worksheet.
    
    Params:
        df (DataFrame): DataFrame to insert to worksheet

        cell (openpyxl.cell.cell.Cell): Cell to begin inserting data at.
    """
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell.offset(r+1, c).value = df.iloc[r, c]

@deprecated(version='0.3.0', reason='renamed to df_to_worksheet')
def df_to_sh(df: pd.DataFrame, cell: xl.cell.cell.Cell):
    df_to_worksheet(df, cell)

def df_to_excel_table_resize(df: pd.DataFrame, wb: xl.Workbook, sheet_name: str, table_name: str, convert_np_bools: str=True) -> None:
    """Takes a pandas dataframe, and writes it into an excel table.
    
    It will also resize the table rowcount to match. It will not resize the columns. Pandas will also not
    respect the excel column arrangement, it will insert the data in the dataframe's column order.

    Params:
        df (DataFrame): Dataframe to insert

        wb (openpyxl.Workbook): Opened workbook object

        sheet_name (str): Worksheet name in workbook

        table_name (str): Table name in Worksheet

        convert_np_bools (bool): Booleans in dataframes get converted to an int. This is to force them as python bools.
            WARNING: This will overwrite the original dataframe

    Returns:
        None. TODO: Return workbook object?

    Exceptions:
        ValueError: If a dataframe's column count is different from the excel table's column count.
    
    """
    data_db = wb[sheet_name]
    assert table_name in data_db.tables, f"Table {table_name} does not exist in the worksheet {sheet_name}"
    ref_start, ref_end = data_db.tables[table_name].ref.split(':')

    start_col, start_row = split_string(ref_start)
    end_col, end_row = split_string(ref_end)
    table_length = (excel_letter_to_numeric(end_col))-(excel_letter_to_numeric(start_col)-1)

    if table_length != len(df.columns):
        raise ValueError(f"Dataframe width does not match table: df has {len(df.columns)} columns and {table_name} has {table_length} columns.")
    
    if convert_np_bools:
        for col in df.columns:
            if df[col].dtype == np.bool_:
                df[col] = df[col].astype(object)

    if len(df) > 0:
        # Blank out range (if data already exists)
        # TODO: Add check for keeping headers?
        for row in data_db[f"{start_col}{start_row+1}:{end_col}{end_row}"]:
            for cell in row:
                cell.value = None

        df_to_worksheet(df, data_db[f"{start_col}{start_row}"])
        data_db.tables[table_name].ref = f"{start_col}{start_row}:{end_col}{start_row+len(df)}"

def excel_table_to_rows(wb: xl.Workbook, sheet_name: str, table_name: str) -> ExcelTableRows:
    """Takes an openpyxl workbook, sheet_name, and table_name, and returns a dict of headers and rows.

    Args:
        wb (workbook): openpyxl workbook

        sheet_name (str): worksheet name

        table_name (str): table name

    Returns:
        dict of {'headers': <list of headers>, 'rows': <list of rows>}
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Workbook does not have the sheet name {sheet_name}. worksheet names are case sensitive.')

    if sheet_name in wb.sheetnames and table_name not in wb[sheet_name].tables:
        raise ValueError(f'Worksheet does not have the table name {table_name}. table names are case sensitive.')
    
    sheet = wb[sheet_name]
    tbl_ref = sheet.tables[table_name].ref
    start_ref, end_ref = map(excel_cell_to_coordinates, tbl_ref.split(':'))

    headers = None
    data = []

    for row in sheet.iter_rows(min_row=start_ref[1], max_row=end_ref[1], min_col=start_ref[0], max_col=end_ref[0]):
        if not isinstance(headers, list):
            headers = [cell.value for cell in row]
        else:
            data.append((cell.value for cell in row))
    
    return {
        'headers': headers,
        'rows': data
    }


def excel_table_to_df(wb: xl.Workbook, sheet_name: str, table_name: str) -> pd.DataFrame:
    """Takes an openpyxl workbook, sheet_name, and table_name, and returns a dataframe

    Args:
        wb (workbook): openpyxl workbook

        sheet_name (str): worksheet name

        table_name (str): table name

    Returns:
    """
    row_data = excel_table_to_rows(wb=wb, sheet_name=sheet_name, table_name=table_name)

    return pd.DataFrame(data=row_data['rows'], columns=row_data['headers'])


def get_named_range_cell(wb: xl.Workbook, rng_name: str) -> xl.cell.cell.Cell:
    '''return cell object from range name'''
    r = list(wb.defined_names[rng_name].destinations)[0]
    return wb[r[0]][r[1]]


def excel_letter_to_numeric(letters) -> int:
    """Takes a full excel letter column refernece and convert it into a number.
    
    Params:
        letters (str): Excel column in letters eg: a=1, aa=26, aaa=703
        
    Returns:
        Converted int of the individual letters.
    """
    result = 0
    for letter in letters:
        result = result * 26 + __letter_to_number(letter)
    return result

def numeric_to_excel_letter(number: int) -> str:
    """Takes an column number, and returns a letter that matches excel, where 1=A.
    
    Params:
        number (int): Column Number
        
    Returns:
        Letters that can be used in excel where 1=A, 27=AA etc...
    """
    if isinstance(number, int) and number < 1:
        raise ValueError('Number must be positive')
    
    result = ''
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        result = __number_to_letter(remainder + 1) + result
    return result

def excel_cell_to_coordinates(cell: str) -> tuple:
    """Takes an excel reference eg: A1, and returns a tuple with numeric coordinates eg: (1, 1).
    
    Params:
        cell (str): Cell String

    Returns:
        tuple of cell coordinates.
    """
    if re.match(r'^([A-Za-z])+(\d+)$', cell) is not None:
        match = re.split('(\d+)', cell)
        return (excel_letter_to_numeric(match[0]), int(match[1]))
    
    else:
        raise ValueError('Cell reference must be in excel format eg: A1, ZZ21')

def split_string(string: str) -> list:
    """Converts excel cell to list of col, row eg: B2 -> ['B', 2]"""
    col, row = re.findall(r'\D+|\d+', string)
    return [col, int(row)]

def __letter_to_number(letter: str) -> int:
    """Takes an excel single cell letter, and converts it into a number.
    
    Params:
        letter (str): single letter A-Z

    Returns:
        letter position, where A=1 and Z=26.
    """
    return ord(letter.upper()) - ord('A') + 1

def __number_to_letter(number: int) -> str:
    """Takes a numeric column and converts it into an letter.
    
    Params:
        number (str): 1-26 to convert to a letter.

    Returns:
        Converted letter from number, where 1=A, 26=Z.

    Exceptions:
        ValueError: If the input number is not between 1 and 26.
    """
    if isinstance(number, int) and number >= 1 and number <= 26:
        return chr(number + ord('A') - 1)
    else:
        raise ValueError('number must be 1-26')
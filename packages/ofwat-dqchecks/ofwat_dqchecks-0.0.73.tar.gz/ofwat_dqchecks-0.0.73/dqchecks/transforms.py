"""
Functions to transform data from Excel.
Main function being:

process_fout_sheets
"""
import datetime
import re
import logging
from typing import Optional
from dataclasses import dataclass
from collections import namedtuple
from openpyxl.workbook.workbook import Workbook
import pandas as pd
from dqchecks.exceptions import (
    EmptyRowsPatternCheckError,
    ColumnHeaderValidationError,)

logging.basicConfig(level=logging.INFO)

# Define namedtuple for context
ProcessingContext = namedtuple(
    'ProcessingContext', ['org_cd', 'submission_period_cd', 'process_cd',
                          'template_version', 'last_modified']
)

@dataclass
class FoutProcessConfig:
    """
    Configuration options for processing fOut sheets in an Excel workbook.

    Attributes:
        observation_patterns (list[str]): List of regex patterns used to identify
            observation period columns in the data sheets.
        fout_patterns (list[str]): List of regex patterns to match sheet names
            that should be processed.
        column_rename_map (Optional[dict[str, str]]): Optional mapping dictionary to rename
            columns from their source names to standardized output names.
            If None, a default mapping will be used.
        run_validations (bool): Flag to determine whether to run validation checks on
            sheets (e.g., empty row checks, header validations). Defaults to True.
        skip_rows (int): Number of rows to skip from the sheet when loading. Defaults to 2.
        reshape (bool): Whether to reshape the data using melt (long format). If False,
            data remains in wide format. Defaults to True.
    """
    observation_patterns: list[str]
    fout_patterns: list[str]
    column_rename_map: Optional[dict[str, str]] = None
    run_validations: bool = True
    skip_rows: int = 2
    reshape: bool = True

def is_valid_regex(pattern: str) -> bool:
    """
    Check if a given string is a valid regex
    """
    try:
        re.compile(pattern)  # Try to compile the regex pattern
        return True  # If no exception, it's a valid regex
    except re.error:  # If an exception is raised, it's not a valid regex
        return False

def validate_workbook(wb: Workbook):
    """
    Validates if the provided workbook is an instance of openpyxl Workbook.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("The 'wb' argument must be a valid openpyxl workbook object.")

def validate_context(context: ProcessingContext):
    """
    Validates the fields in the context object.
    """
    if not isinstance(context.org_cd, str) or not context.org_cd:
        raise ValueError("The 'org_cd' argument must be a non-empty string.")

    if not isinstance(context.submission_period_cd, str) or not context.submission_period_cd:
        raise ValueError("The 'submission_period_cd' argument must be a non-empty string.")

    if not isinstance(context.process_cd, str) or not context.process_cd:
        raise ValueError("The 'process_cd' argument must be a non-empty string.")

    if not isinstance(context.template_version, str) or not context.template_version:
        raise ValueError("The 'template_version' argument must be a non-empty string.")

    if not isinstance(context.last_modified, datetime.datetime):
        raise ValueError("The 'last_modified' argument must be a datetime object.")

def validate_observation_patterns(observation_patterns: list[str]):
    """
    Validates the observation patterns argument.
    """
    if not isinstance(observation_patterns, list)\
            or not all(isinstance(i, str) for i in observation_patterns)\
            or not all(is_valid_regex(i) for i in observation_patterns):
        raise ValueError("The 'observation_patterns' argument needs to be a list of regex strings.")

def extract_fout_sheets(wb: Workbook, fout_patterns: list[str]):
    """
    Extracts sheets from the workbook whose names match any of the given regex patterns.
    
    Args:
        wb (Workbook): The Excel workbook object.
        fout_patterns (list[str]): A list of regex patterns to match sheet names.
    
    Returns:
        List[str]: A list of matching sheet names.

    Raises:
        ValueError: If no matching sheets are found.
    """
    regexes = [re.compile(p) for p in fout_patterns]

    matching_sheets = [
        sheet for sheet in wb.sheetnames
        if any(regex.match(sheet) for regex in regexes)
    ]

    if not matching_sheets:
        raise ValueError(
            f"No sheets matching patterns {fout_patterns} found. Available sheets: {wb.sheetnames}"
        )

    return matching_sheets


def read_sheets_data(wb: Workbook, fout_sheets: list, skip_rows: int = 2):
    """
    Reads data from the sheets into pandas DataFrames.
    """
    df_list = []
    for sheetname in fout_sheets:
        data = wb[sheetname].iter_rows(min_row=skip_rows, values_only=True)
        try:
            headers = next(data)
        except StopIteration as exc:
            raise ValueError(f"Sheet '{sheetname}' is empty or has no data.") from exc

        df = pd.DataFrame(data, columns=headers)
        df["Sheet_Cd"] = sheetname  # Add a column for the sheet name
        df_list.append(df)
    return df_list

def clean_data(df_list: list):
    """
    Drops rows with NaN values and checks if any dataframe is empty.
    """
    df_list = [
        df.dropna(how='all', subset=df.columns.difference(['Sheet_Cd']))
        for df in df_list
    ]
    if any(i.empty for i in df_list):
        raise ValueError("No valid data found after removing rows with NaN values.")
    return df_list

def process_observation_columns(df: pd.DataFrame, observation_patterns: list[str]):
    """
    Identifies and returns the observation period columns based on the provided patterns.
    """
    observation_period_columns = []
    for observation_pattern in observation_patterns:
        observation_period_columns += list(df.filter(regex=observation_pattern).columns.tolist())
    return set(observation_period_columns)

def process_df(
    df: pd.DataFrame,
    context: ProcessingContext,
    observation_patterns: list[str],
    column_rename_map: dict[str, str],
) -> pd.DataFrame:
    """
    Processes a single dataframe by melting it and adding context columns.

    Args:
        df (pd.DataFrame): The input dataframe to process.
        context (ProcessingContext): The processing context with metadata.
        observation_patterns (list[str]): Regex patterns to identify observation columns.
        column_rename_map (dict[str, str]): Mapping from original to final column names.

    Returns:
        pd.DataFrame: The processed and reshaped dataframe.
    """
    observation_period_columns = process_observation_columns(df, observation_patterns)
    if not observation_period_columns:
        raise ValueError("No observation period columns found in the data.")

    # Get the ID columns (all columns except observation period columns)
    id_columns = set(df.columns.tolist()) - observation_period_columns

    # Pivot the DataFrame to melt observation period columns into rows
    pivoted_df = df.melt(
        id_vars=list(id_columns),
        var_name="Observation_Period_Cd",
        value_name="Measure_Value"
    )

    # Add static context columns to the pivoted DataFrame
    pivoted_df["Organisation_Cd"] = context.org_cd
    pivoted_df["Submission_Period_Cd"] = context.submission_period_cd
    pivoted_df["Process_Cd"] = context.process_cd
    pivoted_df["Template_Version"] = context.template_version
    pivoted_df["Submission_Date"] = context.last_modified  # Use the last modified date
    if "Cell_Cd" not in pivoted_df.columns:
        pivoted_df["Cell_Cd"] = "--placeholder--"
    if "Section_Cd" not in pivoted_df.columns:
        pivoted_df["Section_Cd"] = "--placeholder--"

    # Convert all columns to strings for consistency
    pivoted_df = pivoted_df.astype(str)

    # Rename the columns according to the provided mapping
    pivoted_df = pivoted_df.rename(columns=column_rename_map)

    # Reorder the columns to match the desired output format
    ordered_columns = list(column_rename_map.values())
    pivoted_df = pivoted_df[ordered_columns]

    return pivoted_df


def check_empty_rows(wb: Workbook, sheet_names: list[str]):
    # pylint: disable=C0301
    """
    Validates that specified sheets in a workbook contain only empty cells in specific rows.

    This function performs two checks on each provided worksheet:
      1. Verifies that all cells in row 3 (under the header) are empty.
      2. Verifies that all cells in row 1 (the top row), excluding the third column, are empty.

    If any sheet fails either check, a custom EmptyRowsPatternCheckError is raised, indicating which sheets failed.

    Parameters:
        wb (Workbook): An openpyxl Workbook instance containing the sheets to check.
        sheet_names (list[str]): A list of worksheet names to validate.

    Returns:
        bool: True if all sheets pass the checks.

    Raises:
        TypeError: If 'wb' is not a Workbook or 'sheet_names' is not a list of strings.
        ValueError: If 'sheet_names' is empty or contains names not found in the workbook.
        EmptyRowsPatternCheckError: If any sheet contains non-empty values in the checked rows.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")

    under_header_bad_sheet_names = []
    top_row_bad_sheet_names = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]

        # Check under header row (row 3)
        under_header_row = sheet.iter_rows(min_row=3, values_only=True)
        under_header_row_vals = list(next(under_header_row, []))
        if set(under_header_row_vals) not in [{None}, {"", None}, {""}]:
            under_header_bad_sheet_names.append(sheet_name)

        # Check top row (row 1), with 3rd and 2nd element removed
        top_row = sheet.iter_rows(min_row=1, values_only=True)
        top_row_vals = list(next(top_row, []))
        if len(top_row_vals) > 2:
            del top_row_vals[2] # Remove C1
            del top_row_vals[1] # Remove B1
        if set(top_row_vals) not in [{None}, {"", None}, {""}]:
            top_row_bad_sheet_names.append(sheet_name)

    if under_header_bad_sheet_names or top_row_bad_sheet_names:
        raise EmptyRowsPatternCheckError(under_header_bad_sheet_names, top_row_bad_sheet_names)

    return True  # Validation passed

def check_column_headers(wb: Workbook, sheet_names: list[str]):
    """
    Validates that each sheet has the required columns in the correct order starting from row 2.

    Args:
        wb (Workbook): The openpyxl workbook object.
        sheet_names (list[str]): List of sheet names to check.

    Raises:
        TypeError: If wb is not a Workbook, or sheet_names is not a list of strings.
        ValueError: If sheet_names is empty or contains names not in the workbook.
        ColumnHeaderValidationError: If any sheet has missing or misordered expected columns.

    Returns:
        True: If all sheets pass the header validation.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")
    if not all(name in wb.sheetnames for name in sheet_names):
        raise ValueError("One or more sheet names are not present in the workbook.")

    expected_columns = ["Acronym", "Reference", "Item description", "Unit", "Model"]
    bad_sheets = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        header_rows = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
        header = next(header_rows, ())

        # Keep only the expected columns in the order they appear in the sheet
        filtered_header = [col for col in header if col in expected_columns]

        if filtered_header != expected_columns:
            bad_sheets.append(sheet_name)

    if bad_sheets:
        raise ColumnHeaderValidationError(bad_sheets, expected_columns)

    return True

def get_qd_column_rename_map() -> dict[str, str]:
    """
    Returns a dictionary mapping column names to themselves for use in 
    renaming or standardizing columns in a DataFrame related to quarterly data.

    This mapping ensures consistent column naming conventions across processing steps.

    Returns:
        dict[str, str]: A dictionary where keys and values are column names.
    """
    return {
        "Organisation_Cd": "Organisation_Cd",
        "Submission_Period_Cd": "Submission_Period_Cd",
        "Observation_Period_Cd": "Observation_Period_Cd",
        "Process_Cd": "Process_Cd",
        "Template_Version": "Template_Version",
        "Sheet_Cd": "Sheet_Cd",
        "Measure_Cd": "Measure_Cd",
        "Measure_Value": "Measure_Value",
        "Measure_Desc": "Measure_Desc",
        "Measure_Unit": "Measure_Unit",
        "Model_Cd": "Model_Cd",  # missing?
        "Submission_Date": "Submission_Date",
        "Section_Cd": "Section_Cd",  # missing?
        "Cell_Cd": "Cell_Cd",
    }

def finalize_dataframe(
    df: pd.DataFrame,
    context: ProcessingContext,
    column_rename_map: dict[str, str]
) -> pd.DataFrame:
    """
    Adds context columns, renames columns based on the mapping, and reorders columns.

    Args:
        df (pd.DataFrame): The DataFrame to finalize.
        context (ProcessingContext): Context metadata to embed into the DataFrame.
        column_rename_map (dict[str, str]): Mapping from original to standardized column names.

    Returns:
        pd.DataFrame: Finalized and standardized DataFrame.
    """
    df["Organisation_Cd"] = context.org_cd
    df["Submission_Period_Cd"] = context.submission_period_cd
    df["Process_Cd"] = context.process_cd
    df["Template_Version"] = context.template_version
    df["Submission_Date"] = context.last_modified
    if "Cell_Cd" not in df.columns:
        df["Cell_Cd"] = "--placeholder--"
    if "Section_Cd" not in df.columns:
        df["Section_Cd"] = "--placeholder--"

    df = df.astype(str)
    df = df.rename(columns=column_rename_map)

    # Keep only columns in the final output, in the specified order
    ordered_columns = [col for col in column_rename_map.values() if col in df.columns]
    return df[ordered_columns]

def get_default_column_rename_map() -> dict[str, str]:
    """
    Returns the default mapping dictionary for renaming dataframe columns.

    This mapping translates original column names from the input data
    into the standardized output column names used in the processed DataFrame.

    Returns:
        dict[str, str]: A dictionary where keys are original column names,
                        and values are the corresponding standardized column names.
    """
    return {
        'Organisation_Cd': 'Organisation_Cd',
        'Submission_Period_Cd': 'Submission_Period_Cd',
        'Observation_Period_Cd': 'Observation_Period_Cd',
        'Process_Cd': 'Process_Cd',
        'Template_Version': 'Template_Version',
        'Sheet_Cd': 'Sheet_Cd',
        'Reference': 'Measure_Cd',
        'Measure_Value': 'Measure_Value',
        'Item description': 'Measure_Desc',
        'Unit': 'Measure_Unit',
        'Model': 'Model_Cd',
        'Submission_Date': 'Submission_Date',
        "Section_Cd": "Section_Cd",
        "Cell_Cd": "Cell_Cd",
    }

def process_fout_sheets(
        wb: Workbook,
        context: ProcessingContext,
        config: FoutProcessConfig,
    ) -> pd.DataFrame:
    """
    Processes all sheets in the given Excel workbook matching the specified patterns,
    transforming and normalizing their data into a consolidated DataFrame.

    Args:
        wb (Workbook): The openpyxl Workbook object.
        context (ProcessingContext): Processing context metadata.
        config (FoutProcessConfig): Configuration options including patterns, column mapping,
            and reshape flag.

    Returns:
        pd.DataFrame: The consolidated processed DataFrame.
    """
    # Validate inputs
    validate_workbook(wb)
    validate_context(context)
    validate_observation_patterns(config.observation_patterns)

    if not wb.data_only:
        logging.warning("Reading in non data_only mode. Some data may not be accessible.")

    logging.info("Using observation patterns: %s", config.observation_patterns)

    # Extract matching sheets
    fout_sheets = extract_fout_sheets(wb, config.fout_patterns)

    if config.run_validations:
        assert check_empty_rows(wb, fout_sheets)
        assert check_column_headers(wb, fout_sheets)

    # Read and clean data
    df_list = read_sheets_data(wb, fout_sheets, skip_rows=config.skip_rows)
    df_list = clean_data(df_list)

    column_rename_map = config.column_rename_map or get_default_column_rename_map()

    processed_dfs = []

    for df in df_list:
        if config.reshape:
            df = process_df(df, context, config.observation_patterns, column_rename_map)

        processed_df = finalize_dataframe(df, context, column_rename_map)
        processed_dfs.append(processed_df)

    final_df = pd.concat(processed_dfs, ignore_index=True)
    return final_df

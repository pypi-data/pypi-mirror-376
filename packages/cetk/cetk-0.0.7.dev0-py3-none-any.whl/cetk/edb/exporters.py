import ast
import os

# need to import fiona before geopandas due to gpd bug causing circular imports.
import fiona  # noqa
import geopandas as gpd
import numpy as np
import pandas as pd
import pyogrio
import rasterio as rio
from django.db import connection
from openpyxl import Workbook
from shapely import wkt

from cetk.edb.const import DEFAULT_EMISSION_UNIT, WGS84_SRID
from cetk.edb.importers.source_import import (
    OPTIONAL_COLUMNS_POINT,
    REQUIRED_COLUMNS_AREA,
    REQUIRED_COLUMNS_POINT,
)
from cetk.edb.models import (
    Activity,
    ActivityCode,
    AreaSource,
    AreaSourceSubstance,
    CodeSet,
    ColdstartTimevar,
    CongestionProfile,
    EmissionFactor,
    Facility,
    Fleet,
    FleetMemberFuel,
    FlowTimevar,
    GridSource,
    GridSourceSubstance,
    PointSource,
    PointSourceSubstance,
    RoadAttribute,
    RoadClass,
    RoadSource,
    Substance,
    TrafficSituation,
    VehicleEF,
    VehicleFuel,
    VehicleFuelComb,
)
from cetk.edb.models.timevar_models import Timevar
from cetk.edb.units import activity_rate_unit_from_si, emis_conversion_factor_from_si
from cetk.tools.utils import get_db

REQUIRED_COLUMNS_GRID = {
    "name": np.str_,
    "rastername": np.str_,
    "timevar": np.str_,
    "path": np.str_,
}

days_header = [
    "ID",
    "typeday",
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "sunday",
]
months_header = [
    " ",
    "month",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]
time_intervals = [
    "00-01",
    "01-02",
    "02-03",
    "03-04",
    "04-05",
    "05-06",
    "06-07",
    "07-08",
    "08-09",
    "09-10",
    "10-11",
    "11-12",
    "12-13",
    "13-14",
    "14-15",
    "15-16",
    "16-17",
    "17-18",
    "18-19",
    "19-20",
    "20-21",
    "21-22",
    "22-23",
    "23-24",
]


def export_sources(export_filepath, srid=WGS84_SRID, unit=DEFAULT_EMISSION_UNIT):
    # Create a new Excel workbook and remove standard first Sheet
    workbook = Workbook()
    del workbook["Sheet"]
    if PointSource.objects.count() > 0:
        point_columns = REQUIRED_COLUMNS_POINT | OPTIONAL_COLUMNS_POINT
        worksheet = workbook.create_sheet(title="PointSource")
        create_source_sheet(
            worksheet, PointSource, point_columns, PointSourceSubstance, unit
        )
        workbook.save(export_filepath)

    if AreaSource.objects.count() > 0:
        worksheet = workbook.create_sheet(title="AreaSource")
        create_source_sheet(
            worksheet, AreaSource, REQUIRED_COLUMNS_AREA, AreaSourceSubstance, unit
        )
        workbook.save(export_filepath)

    if GridSource.objects.count() > 0:
        worksheet = workbook.create_sheet(title="GridSource")
        create_source_sheet(
            worksheet, GridSource, REQUIRED_COLUMNS_GRID, GridSourceSubstance, unit
        )
        workbook.save(export_filepath)

    if EmissionFactor.objects.count() > 0:
        worksheet = workbook.create_sheet(title="EmissionFactor")
        header = [
            "activity_name",
            "substance",
            "factor",
            "emissionfactor_unit",
            "activity_unit",
        ]
        worksheet.append(header)
        for emfac in EmissionFactor.objects.all():
            # all factors stored in SI, original factor unit at import not stored in db.
            factor_unit = "kg/" + emfac.activity.unit.split("/")[0]
            row_data = [
                emfac.activity.name,
                emfac.substance.slug,
                emfac.factor,
                factor_unit,
                emfac.activity.unit,
            ]
            worksheet.append(row_data)
        workbook.save(export_filepath)

    if CodeSet.objects.count() > 0:
        worksheet = workbook.create_sheet(title="CodeSet")
        header = ["name", "slug", "description"]
        worksheet.append(header)
        for cs in CodeSet.objects.all():
            worksheet.append([cs.name, cs.slug, cs.description])
        workbook.save(export_filepath)

    if ActivityCode.objects.count() > 0:
        worksheet = workbook.create_sheet(title="ActivityCode")
        header = ["codeset_slug", "activitycode", "label", "vertical_distribution_slug"]
        worksheet.append(header)
        for ac in ActivityCode.objects.all():
            if ac.vertical_dist is not None:
                worksheet.append(
                    [ac.code_set.slug, ac.code, ac.label, ac.vertical_dist.slug]
                )
            else:
                worksheet.append([ac.code_set.slug, ac.code, ac.label, ""])
        workbook.save(export_filepath)

    if Timevar.objects.count() > 0:
        worksheet = workbook.create_sheet(title="Timevar")
        create_timevar_sheet(worksheet, Timevar)
        workbook.save(export_filepath)

    if FlowTimevar.objects.count() > 0:
        worksheet = workbook.create_sheet(title="FlowTimevar")
        create_timevar_sheet(worksheet, FlowTimevar)
        workbook.save(export_filepath)

    if ColdstartTimevar.objects.count() > 0:
        worksheet = workbook.create_sheet(title="ColdstartTimevar")
        create_timevar_sheet(worksheet, ColdstartTimevar)
        workbook.save(export_filepath)

    if CongestionProfile.objects.count() > 0:
        worksheet = workbook.create_sheet(title="CongestionProfile")
        create_timevar_sheet(worksheet, CongestionProfile)
        workbook.save(export_filepath)

    if RoadSource.objects.count() > 0:
        create_roadsource_sheet(workbook)
        workbook.save(export_filepath)

    # RoadAttributes could be defined before importing any roads
    if RoadAttribute.objects.count() > 0:
        worksheet = workbook.create_sheet(title="RoadAttribute")
        header = ["name", "slug"]
        worksheet.append(header)
        for ra in RoadAttribute.objects.all():
            worksheet.append([ra.name, ra.slug])
        workbook.save(export_filepath)

    if VehicleFuelComb.objects.count() > 0:
        create_vehiclefuel_sheet(workbook)
        workbook.save(export_filepath)

    if Fleet.objects.count() > 0:
        create_fleet_sheet(workbook)
        workbook.save(export_filepath)

    if TrafficSituation.objects.count() > 0:
        create_traffic_sheet(workbook)
        workbook.save(export_filepath)

    if VehicleEF.objects.count() > 0:
        create_vehicle_ef_sheet(workbook)
        workbook.save(export_filepath)


def create_vehicle_ef_sheet(workbook):
    worksheet = workbook.create_sheet(title="VehicleEmissionFactor")
    header = [
        "traffic_situation",
        "vehicle",
        "fuel",
        "substance",
        "freeflow",
        "heavy",
        "saturated",
        "stopngo",
        "coldstart",
        "unit",
    ]
    worksheet.append(header)
    for vef in VehicleEF.objects.all():
        # TODO could convert to more human readable values with for example unit mg/km.
        row = [
            vef.traffic_situation.ts_id,
            vef.vehicle.name,
            vef.fuel.name,
            vef.substance.slug,
            vef.freeflow,
            vef.heavy,
            vef.saturated,
            vef.stopngo,
            vef.coldstart,
            "kg/m",
        ]
        worksheet.append(row)


def create_traffic_sheet(workbook):
    worksheet = workbook.create_sheet(title="TrafficSituation")
    header = ["traffic_situation"]
    attributes = []
    for attribute in RoadAttribute.objects.all():
        header.append("attr:" + attribute.slug)
        attributes.append(attribute.slug)
    attributes = np.array(attributes)
    worksheet.append(header)
    for i, rc in enumerate(RoadClass.objects.all()):
        row = [""] * len(header)
        row[0] = rc.traffic_situation.ts_id
        for av in rc.attribute_values.all():
            ind = np.where(attributes == av.attribute.slug)[0][0]
            row[ind + 1] = av.value
        worksheet.append(row)


def create_fleet_sheet(workbook):
    worksheet = workbook.create_sheet(title="Fleet")
    header = [
        "name",
        "default_heavy_vehicle_share",
        "vehicle",
        "vehicle_fraction",
        "coldstart_fraction",
        "flow_timevar",
        "coldstart_timevar",
    ]
    fuel_ids = []
    for fuel in VehicleFuel.objects.all():
        fuel_ids.append(fuel.id)
        header.append(f"fuel:{fuel.name}")
    worksheet.append(header)
    for fleet in Fleet.objects.all():
        for fleetmember in fleet.vehicles.all():
            veh = fleetmember.vehicle
            row = [
                fleet.name,
                fleet.default_heavy_vehicle_share,
                veh.name,
                fleetmember.fraction,
                fleetmember.coldstart_fraction,
            ]
            if fleetmember.timevar is None:
                row.append("")
            else:
                row.append(fleetmember.timevar.name)
            if fleetmember.coldstart_timevar is None:
                row.append("")
            else:
                row.append(fleetmember.coldstart_timevar.name)
            for fuel_id in fuel_ids:
                try:
                    row.append(fleetmember.fuels.get(fuel_id=fuel_id).fraction)
                except FleetMemberFuel.DoesNotExist:
                    row.append("")
            worksheet.append(row)


def create_vehiclefuel_sheet(workbook):
    worksheet = workbook.create_sheet(title="VehicleFuel")
    header = ["name", "isheavy", "info", "fuel"]
    codeset_slugs = [code.slug for code in CodeSet.objects.all()]
    codeset_ids = [CodeSet.objects.get(slug=slug).id for slug in codeset_slugs]
    codeset_columns = [f"activitycode_{slug}" for slug in codeset_slugs]
    header = header + codeset_columns
    worksheet.append(header)
    for vf in VehicleFuelComb.objects.all():
        row = [vf.vehicle.name, str(vf.vehicle.isheavy), vf.vehicle.info, vf.fuel.name]
        for i in codeset_ids:
            try:
                row.append(
                    ActivityCode.objects.get(id=getattr(vf, f"activitycode{i}_id")).code
                )
            except ActivityCode.DoesNotExist:
                row.append("")
        worksheet.append(row)


def create_roadsource_sheet(workbook):
    worksheet = workbook.create_sheet(title="RoadSource")
    header = [
        "filepath",
        "name",
        "nlanes",
        "width",
        "median_strip_width",
        "aadt",
        "heavy_vehicle_share",
        "fleet",
        "congestion_profile",
        "speed",
        "slope",
    ]
    # add attributes, always at least 1 to define which emission factors to use
    road_attributes = ""
    for attribute in RoadAttribute.objects.all():
        header.append("attr:" + attribute.slug)
        road_attributes += f", rc.{attribute.slug} AS {attribute.slug} "
    worksheet.append(header)
    # slight adaptions to header to form actual content
    db_name = os.path.basename(get_db()).split(".")[0]
    roadpath = os.path.join(os.path.dirname(get_db()), db_name + "-roadsource.gpkg")
    header[0] = roadpath
    header = [attr.replace("attr:", "") for attr in header]
    worksheet.append(header)

    sql = (
        "SELECT rs.name, rs.nolanes, rs.width, rs.median_strip_width, rs.aadt, "
        "rs.heavy_vehicle_share, f.name AS fleet, cp.name AS congestion_profile, "
        "rs.speed, rs.slope, rs.roadclass_id, ST_AsText(geom) AS geometry "
        "FROM edb_roadsource AS rs "
        "LEFT JOIN edb_fleet AS F ON rs.fleet_id = f.id "
        "LEFT JOIN edb_congestionprofile AS cp ON rs.congestion_profile_id = cp.id "
    )
    cur = connection.cursor()
    cur.execute(sql)

    df = pd.DataFrame(cur.fetchall(), columns=[col[0] for col in cur.description])
    roadclass_mapping = {}
    for rc in RoadClass.objects.all():
        roadclass_mapping[rc.id] = rc.attributes
    for col in roadclass_mapping[1].keys():
        df[col] = df["roadclass_id"].map(
            lambda x: roadclass_mapping[x][col] if x in roadclass_mapping else None
        )
    df = df.drop(columns=["roadclass_id"])
    df["geometry"] = df["geometry"].apply(wkt.loads)
    gdf = gpd.GeoDataFrame(df, geometry="geometry")
    crs = "EPSG:4326"
    gdf.crs = crs
    # does not work from QGIS gdf.to_file(roadpath, driver="GPKG")
    pyogrio.write_dataframe(gdf, roadpath, layer="roads")


def create_timevar_sheet(worksheet, tvar_type):
    for tvar in tvar_type.objects.all():
        worksheet.append(days_header)
        if tvar_type is CongestionProfile:
            typeday_list = ast.literal_eval(tvar.traffic_condition)
        else:
            typeday_list = ast.literal_eval(tvar.typeday)
        for i in range(len(typeday_list)):
            if i == 0:
                row_data = [tvar.name] + [time_intervals[i]] + typeday_list[i]
            else:
                row_data = [""] + [time_intervals[i]] + typeday_list[i]
            worksheet.append(row_data)
        if tvar_type is not CongestionProfile:
            month_list = ast.literal_eval(tvar.month)
            worksheet.append(months_header)
            worksheet.append(["", ""] + month_list)


def create_source_sheet(
    worksheet, model_type, REQUIRED_COLUMNS, SourceSubstanceModel, unit
):
    """create source sheet for Point, Area or GridSource"""
    emis_conversion_factor = emis_conversion_factor_from_si(unit)
    header = list(REQUIRED_COLUMNS.keys())
    codeset_slugs = [code.slug for code in CodeSet.objects.all()]
    codeset_ids = [CodeSet.objects.get(slug=slug).id for slug in codeset_slugs]
    codeset_columns = [f"activitycode_{slug}" for slug in codeset_slugs]
    substance_slugs = list(
        set([ss.substance.slug for ss in SourceSubstanceModel.objects.all()])
    )
    substance_columns = [f"subst:{subst}" for subst in substance_slugs]
    header = header + codeset_columns + substance_columns + ["emission_unit"]
    activities = Activity.objects.all()
    if len(activities) > 0:
        activity_names = [activity.name for activity in activities]
        activity_columns = [f"act:{name}" for name in activity_names]
        header = header + activity_columns

    worksheet.append(header)
    for source in model_type.objects.all():
        if source.timevar_id is not None:
            timevar_name = Timevar.objects.get(id=source.timevar_id).name
        else:
            timevar_name = ""

        activitycodes = {}
        for i in codeset_ids:
            activitycode = getattr(source, f"activitycode{i}")
            activitycodes[i] = activitycode.code if activitycode is not None else ""
        source_substances = [ss.substance.slug for ss in source.substances.all()]
        if model_type == GridSource:
            if len(source_substances) > 0:
                rasternames = set(
                    [substance.raster for substance in source.substances.all()]
                )
                if len(rasternames) == 1:
                    rastername = source.substances.first().raster
                else:
                    try:
                        rastername = (
                            os.path.commonprefix(list(rasternames))
                            + "{subst}"
                            + common_suffix(list(rasternames))
                        )
                    except Exception:
                        raise ValueError(
                            "Could not find a generic raster name "
                            + f"for GridSource {source.name}"
                        )
            elif source.activities.count() > 0:
                rastername = source.activities.first().raster
                rasternames = [rastername]
            else:
                rastername = ""
            if rastername != "":
                rasterpath = os.path.join(
                    os.path.dirname(get_db()), rastername + ".tif"
                )
                for name in rasternames:
                    # storing each raster, not generic name
                    tif_path = os.path.join(os.path.dirname(get_db()), name + ".tif")
                    with rio.open(f"GPKG:{get_db()}:raster_{name}") as src:
                        meta = src.meta
                        meta.update(driver="GTiff")
                        with rio.open(tif_path, "w", **meta) as dst:
                            # always only 1 band
                            dst.write(src.read(1), 1)
            else:
                rasterpath = ""
            row_data = [source.name, rastername, timevar_name, rasterpath]
        else:
            row_data = [
                str(source.facility),
                Facility.objects.get(id=source.facility_id).name,
                source.name,
                source.geom.coords[1] if model_type == PointSource else source.geom.wkt,
            ]
            if model_type == PointSource:
                row_data.append(source.geom.coords[0])
            row_data.append(timevar_name)

            if hasattr(source, "chimney_height"):
                row_data.extend(
                    [
                        source.chimney_height,
                        source.chimney_outer_diameter,
                        source.chimney_inner_diameter,
                        source.chimney_gas_speed,
                        source.chimney_gas_temperature,
                        source.house_width,
                        source.house_height,
                    ]
                )

        for i in codeset_ids:
            row_data.append(activitycodes[i])

        emis_row = [
            source.substances.get(substance=Substance.objects.get(slug=slug).id).value
            if slug in source_substances
            else 0
            for slug in substance_slugs
        ]
        emis_row = [emis * emis_conversion_factor for emis in emis_row]
        row_data = row_data + emis_row + [unit]

        if len(activities) > 0:
            source_activities = [
                Activity.objects.get(id=sa.activity_id).name
                for sa in source.activities.all()
            ]
            source_activity_rates = dict()
            for act in source.activities.all():
                activity_unit = Activity.objects.get(id=act.activity_id).unit
                activity_rate = activity_rate_unit_from_si(act.rate, activity_unit)
                source_activity_rates[Activity.objects.get(id=act.activity_id).name] = (
                    activity_rate
                )
            act_row = [
                source_activity_rates[name] if name in source_activities else 0
                for name in activity_names
            ]
            row_data = row_data + act_row
        worksheet.append(row_data)


def common_suffix(strings):
    reversed_strings = [s[::-1] for s in strings]
    common_suffix_reversed = os.path.commonprefix(reversed_strings)
    return common_suffix_reversed[::-1]

"""Data importers for the edb application."""

import numpy as np
import pandas as pd
from django.contrib.gis.geos import GEOSGeometry, Point
from openpyxl import load_workbook

from cetk import logging
from cetk.edb.cache import cache_queryset
from cetk.edb.const import SHEET_NAMES, WGS84_SRID
from cetk.edb.models import (
    Activity,
    AreaSource,
    AreaSourceActivity,
    AreaSourceSubstance,
    CodeSet,
    Facility,
    PointSource,
    PointSourceActivity,
    PointSourceSubstance,
    Substance,
)
from cetk.edb.models.timevar_models import Timevar
from cetk.edb.units import activity_rate_unit_to_si, emission_unit_to_si

from .activity_import import import_emissionfactorsheet
from .codeset_import import import_activitycodesheet, import_codesetsheet
from .timevar_import import import_timevarsheet
from .utils import cache_codeset, import_error, worksheet_to_dataframe

# from cetk.edb.models.common_models import Settings
# import sys
# from datetime import datetime
# uncomment for automatic debugging when receiving error
# def info(type, value, tb):
#     # if hasattr(sys, "ps1") or not sys.stderr.isatty():
#     #     # we are in interactive mode or we don't have a tty-like
#     #     # device, so we call the default hook
#     #     sys.__excepthook__(type, value, tb)
#     # else:
#     import pdb
#     import traceback

#     # we are NOT in interactive mode, print the exception...
#     traceback.print_exception(type, value, tb)
#     print
#     # ...then start the debugger in post-mortem mode.
#     # pdb.pm() # deprecated
#     pdb.post_mortem(tb)  # more "modern"


# sys.excepthook = info

# FIXME: get from settings instead, but then tests won't work
MAX_ERROR_MESSAGES = 10


# column facility and name are used as index and is therefore not included here
REQUIRED_COLUMNS_AREA = {
    "facility_id": np.str_,
    "facility_name": np.str_,
    "source_name": np.str_,
    "geometry": np.str_,
    "timevar": np.str_,
}

REQUIRED_COLUMNS_POINT = {
    "facility_id": np.str_,
    "facility_name": np.str_,
    "source_name": np.str_,
    "lat": float,
    "lon": float,
}

OPTIONAL_COLUMNS_POINT = {
    "timevar": np.str_,
    "chimney_height": float,
    "outer_diameter": float,
    "inner_diameter": float,
    "gas_speed": float,
    "gas_temperature[K]": float,
    "house_width": float,
    "house_height": float,
}


log = logging.getLogger(__name__)


def cache_sources(queryset):
    """Return dict of model instances with (facility__official_id, name): instance"""
    sources = {}
    for source in queryset:
        if source.facility is not None:
            sources[source.facility.official_id, source.name] = source
        else:
            sources[None, source.name] = source
    return sources


def import_sources(
    filepath,
    validation=False,
    encoding=None,
    srid=None,
    sourcetype="point",
):
    """Import point- or area-sources from xlsx or csv-file.
    PointSource- and AreaSourceSubstances only, not Activities.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is same srid as domain
    """
    return_message = []
    extension = filepath.suffix
    if extension == ".csv":
        # read csv-file
        if sourcetype == "point":
            with open(filepath, encoding=encoding or "utf-8") as csvfile:
                log.debug("reading point-sources from csv-file")
                df = pd.read_csv(
                    csvfile,
                    sep=";",
                    skip_blank_lines=True,
                    comment="#",
                    dtype=REQUIRED_COLUMNS_POINT,
                )
        else:
            with open(filepath, encoding=encoding or "utf-8") as csvfile:
                log.debug("reading area-sources from csv-file")
                df = pd.read_csv(
                    csvfile,
                    sep=";",
                    skip_blank_lines=True,
                    comment="#",
                    dtype=REQUIRED_COLUMNS_AREA,
                )
    elif extension == ".xlsx":
        # read spreadsheet
        try:
            workbook = load_workbook(filename=filepath, data_only=True, read_only=True)
        except Exception as exc:
            return_message.append(import_error(str(exc), validation))
        worksheet = workbook.worksheets[0]
        if len(workbook.worksheets) > 1:
            if sourcetype == "point":
                log.debug(
                    "Multiple sheets in spreadsheet, importing sheet 'PointSource'."
                )
                data = workbook["PointSource"].values
            elif sourcetype == "area":
                log.debug(
                    "Multiple sheets in spreadsheet, importing sheet 'AreaSource'."
                )
                data = workbook["AreaSource"].values
        else:
            data = worksheet.values
        df = worksheet_to_dataframe(data)
        workbook.close()
    else:
        return_message.append(
            import_error(
                "Only xlsx and csv files are supported for import",
                validation=validation,
            )
        )

    df = set_datatypes(df, sourcetype)

    return create_or_update_sources(
        df,
        validation=validation,
        srid=None,
        sourcetype=sourcetype,
    )


def set_datatypes(df, sourcetype):
    if sourcetype == "point":
        df = df.astype(dtype=REQUIRED_COLUMNS_POINT)
        optional_columns = {
            k: OPTIONAL_COLUMNS_POINT[k]
            for k in OPTIONAL_COLUMNS_POINT
            if k in df.columns
        }
        df = df.astype(dtype=optional_columns)
    else:
        df = df.astype(dtype=REQUIRED_COLUMNS_AREA)
    # below is necessary not to create facilities with name 'None'
    df = df.replace(to_replace="None", value=None)
    df = df.replace(to_replace="nan", value=None)
    return df


# @profile
def create_or_update_sources(
    df,
    validation=False,
    srid=None,
    sourcetype="point",
    cache=True,
):
    return_message = []
    # user defined SRID for import or WGS84 if nothing specified
    # as long as we do not have functions in Eclair to edit the "settings_SRID"
    # it does not make sense to use that SRID as default for import.
    srid = srid or WGS84_SRID
    # cache related models
    substances = cache_queryset(Substance.objects.all(), "slug")
    timevars = cache_queryset(Timevar.objects.all(), "name")
    facilities = cache_queryset(Facility.objects.all(), "official_id")

    if cache:
        if sourcetype == "point":
            sources = cache_sources(
                PointSource.objects.select_related("facility").all()
            )  # .prefetch_related("substances")
        elif sourcetype == "area":
            sources = cache_sources(
                AreaSource.objects.select_related("facility").all()
            )  # .prefetch_related("substances")
        else:
            return_message.append(
                import_error(
                    "this sourcetype is not implemented",
                    validation=validation,
                )
            )

    code_sets = [
        cache_codeset(CodeSet.objects.filter(id=i).first()) for i in range(1, 4)
    ]
    code_set_slugs = {}
    for i in range(1, 4):
        try:
            code_set_slug = CodeSet.objects.get(id=i).slug
        except CodeSet.DoesNotExist:
            code_set_slug = None
        code_set_slugs[i] = code_set_slug

    if sourcetype == "point":
        for col in REQUIRED_COLUMNS_POINT.keys():
            if col not in df.columns:
                return_message.append(
                    import_error(
                        f"Missing required column '{col}'", validation=validation
                    )
                )
    else:
        for col in REQUIRED_COLUMNS_AREA.keys():
            if col not in df.columns:
                return_message.append(
                    import_error(
                        f"Missing required column '{col}'", validation=validation
                    )
                )

    # set dataframe index
    try:
        df.set_index(
            ["facility_id", "source_name"], verify_integrity=True, inplace=True
        )
    except ValueError as err:
        return_message.append(
            import_error(
                f"Non-unique combination of facility_id and source_name: {err}",
                validation=validation,
            )
        )
    update_facilities = []
    create_facilities = {}
    drop_substances = []
    create_substances = []
    update_sources = []
    create_sources = {}
    activitycode_columns = [key for key in df.columns if key.startswith("activitycode")]
    row_nr = 2
    for row_key, row in df.iterrows():
        # If we have many errors, no point in continuing
        if len(return_message) > MAX_ERROR_MESSAGES:
            return_dict = {
                "facility": {
                    "updated": 0,
                    "created": 0,
                },
                "pointsource": {
                    "updated": 0,
                    "created": 0,
                },
            }
            return return_dict, return_message

        row_dict = row.to_dict()

        # initialize activitycodes
        source_data = {
            "activitycode1": None,
            "activitycode2": None,
            "activitycode3": None,
        }

        if sourcetype == "point":
            # get pointsource coordinates
            try:
                if pd.isna(row_dict["lat"]) or pd.isna(row_dict["lon"]):
                    return_message.append(
                        import_error(
                            f"missing coordinates for source '{row_key}'",
                            validation=validation,
                        )
                    )
                    continue
                x = float(row_dict["lon"])
                y = float(row_dict["lat"])
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid {sourcetype} coordinates on row {row_nr}",
                        validation=validation,
                    )
                )
            # create geometry
            source_data["geom"] = Point(x, y, srid=srid).transform(4326, clone=True)
            # get chimney properties
            for attr, key in {
                "chimney_height": "chimney_height",
                "chimney_inner_diameter": "inner_diameter",
                "chimney_outer_diameter": "outer_diameter",
                "chimney_gas_speed": "gas_speed",
                "chimney_gas_temperature": "gas_temperature[K]",
            }.items():
                if pd.isna(row_dict[key]):
                    return_message.append(
                        import_error(
                            "Missing value in PointSource sheet "
                            f"for {key} on row {row_nr}",
                            validation=validation,
                        )
                    )
                    continue
                else:
                    source_data[attr] = row_dict[key]

            # get downdraft parameters
            try:
                if not pd.isna(row_dict["house_width"]):
                    source_data["house_width"] = row_dict["house_width"]
            except KeyError:
                if row_nr == 2:
                    log.debug("house_width is skipped from import.")
            try:
                if not pd.isna(row_dict["house_height"]):
                    source_data["house_height"] = row_dict["house_height"]
            except KeyError:
                if row_nr == 2:
                    log.debug("house_heigth is skipped from import.")

        elif sourcetype == "area":
            try:
                if pd.isna(row_dict["geometry"]):
                    return_message.append(
                        import_error(
                            f"missing area polygon for source '{row_key}'",
                            validation=validation,
                        )
                    )
                    continue
                wkt_polygon = row_dict["geometry"]
                # TODO add check that valid WKT polygon
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid polygon geometry in AreaSource sheet on row {row_nr}",
                        validation=validation,
                    )
                )
                continue
            source_data["geom"] = GEOSGeometry(f"SRID={4326};" + wkt_polygon)
        else:
            return_message.append(
                import_error(
                    "this sourcetype is not implemented", validation=validation
                )
            )

        # get activitycodes
        for code_ind, code_set in enumerate(code_sets, 1):
            try:
                code_set_slug = code_set_slugs[code_ind]
                code_attribute = f"activitycode_{code_set_slug}"
                if code_attribute in row_dict:
                    code = row_dict[code_attribute]
                    if len(code_set) == 0:
                        if code is not None and code is not np.nan:
                            return_message.append(
                                import_error(
                                    f"Unknown activitycode_{code_set_slug} '{code}'"
                                    f" for {sourcetype} source on row {row_nr}",
                                    validation=validation,
                                )
                            )
                    if not pd.isna(code):
                        try:
                            # note this can be problematic with codes 01 etc as SNAP
                            # TODO activitycodes should be string directly on import!
                            activity_code = code_set[str(code)]
                            codeset_id = activity_code.code_set_id
                            source_data[f"activitycode{codeset_id}"] = activity_code
                        except KeyError:
                            return_message.append(
                                import_error(
                                    f"Unknown activitycode_{code_set_slug} '{code}'"
                                    f" for {sourcetype} source on row {row_nr}",
                                    validation=validation,
                                )
                            )
            except AttributeError:
                # no such codeset exists
                if len(activitycode_columns) > CodeSet.objects.count():
                    # need to check if activitycode is specified for unimported codeset
                    codeset_slug = [
                        column.split("_", 1)[-1] for column in activitycode_columns
                    ]
                    for index, column in enumerate(activitycode_columns):
                        if not pd.isna(row_dict[column]):
                            try:
                                CodeSet.objects.get(slug=codeset_slug[index])
                            except CodeSet.DoesNotExist:
                                return_message.append(
                                    import_error(
                                        "Specified activitycode "
                                        f"{row_dict[column]} for "
                                        f" unknown codeset {codeset_slug[index]}"
                                        f" for {sourcetype} source on row {row_nr}",
                                        validation=validation,
                                    )
                                )
                pass

        # get columns with tag values for the current row
        tag_keys = [key for key in row_dict.keys() if key.startswith("tag:")]
        # set tags dict for source
        source_data["tags"] = {
            key[4:]: row_dict[key] for key in tag_keys if pd.notna(row_dict[key])
        }

        # get timevar name and corresponding timevar
        timevar_name = row_dict["timevar"]
        if pd.notna(timevar_name):
            try:
                source_data["timevar"] = timevars[timevar_name]
            except KeyError:
                return_message.append(
                    import_error(
                        f"Timevar '{timevar_name}' "
                        f"on row {row_nr} for {sourcetype} source does not exist",
                        validation=validation,
                    )
                )
        # get all column-names starting with "subst" whith value for the current row
        subst_keys = [
            key
            for key in row_dict.keys()
            if key.startswith("subst:") and pd.notna(row_dict[key])
        ]
        # create list of data dict for each substance emission
        emissions = {}
        for subst_key in subst_keys:
            subst = subst_key[6:]
            # dict with substance emission properties (value and substance)
            emis = {}
            emissions[subst] = emis

            # get substance
            try:
                emis["substance"] = substances[subst]
            except KeyError:
                return_message.append(
                    import_error(f"Undefined substance {subst}", validation)
                )

            try:
                if "emission_unit" in row_dict and not pd.isnull(
                    row_dict["emission_unit"]
                ):
                    emis["value"] = emission_unit_to_si(
                        float(row_dict[subst_key]), row_dict["emission_unit"]
                    )
                else:
                    return_message.append(
                        import_error(
                            f"No unit specified for {sourcetype}-source"
                            f" emissions on row {row_nr}",
                            validation=validation,
                        )
                    )
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid {sourcetype}-source emission value "
                        f"{row_dict[subst_key]} on row {row_nr}",
                        validation=validation,
                    )
                )
            except KeyError as err:
                return_message.append(
                    import_error(
                        f"Missing data {err} on row {row_nr} for {sourcetype} sources.",
                        validation=validation,
                    )
                )

        official_facility_id, source_name = row_key
        if pd.isna(official_facility_id):
            official_facility_id = None

        if pd.isna(source_name):
            return_message.append(
                import_error(
                    f"No name specified for {sourcetype} source on row {row_nr}",
                    validation=validation,
                )
            )
        if pd.isna(row_dict["facility_name"]):
            facility_name = None
        else:
            facility_name = row_dict["facility_name"]

        try:
            facility = facilities[official_facility_id]
            update_facilities.append(facility)
        except KeyError:
            if official_facility_id is not None:
                if official_facility_id in create_facilities:
                    facility = create_facilities[official_facility_id]
                else:
                    if facility_name is None:
                        return_message.append(
                            import_error(
                                f"No name specified for facility on row {row_nr}",
                                validation=validation,
                            )
                        )
                        facility_name = "unspecified"

                    facility = Facility(
                        name=facility_name,
                        official_id=official_facility_id,
                    )
                    create_facilities[official_facility_id] = facility
            else:
                facility = None

        source_data["facility"] = facility
        source_key = (official_facility_id, source_name)
        try:
            if cache:
                source = sources[source_key]
            else:
                if official_facility_id is None:
                    facility_id = None
                else:
                    # this could through a keyerror for the wrong reason,
                    # because facility is None, not because pointsource exists already
                    facility_id = facilities[str(official_facility_id)].id
                if sourcetype == "point":
                    source = PointSource.objects.get(
                        name=str(source_name), facility_id=facility_id
                    )
                elif sourcetype == "area":
                    source = AreaSource.objects.get(
                        name=str(source_name), facility_id=facility_id
                    )
            for key, val in source_data.items():
                setattr(source, key, val)
            update_sources.append(source)
            drop_substances += list(source.substances.all())
            if sourcetype == "point":
                create_substances += [
                    PointSourceSubstance(source=source, **emis)
                    for emis in emissions.values()
                ]
            elif sourcetype == "area":
                create_substances += [
                    AreaSourceSubstance(source=source, **emis)
                    for emis in emissions.values()
                ]
        except (PointSource.DoesNotExist, AreaSource.DoesNotExist, KeyError):
            if sourcetype == "point":
                source = PointSource(name=source_name, **source_data)
                if source_key not in create_sources:
                    create_sources[source_key] = source
                    create_substances += [
                        PointSourceSubstance(source=source, **emis)
                        for emis in emissions.values()
                    ]
                else:
                    return_message.append(
                        import_error(
                            f"multiple rows for the same point-source '{source_name}'",
                            validation=validation,
                        )
                    )
            else:
                source = AreaSource(name=source_name, **source_data)
                if source_key not in create_sources:
                    create_sources[source_key] = source
                    create_substances += [
                        AreaSourceSubstance(source=source, **emis)
                        for emis in emissions.values()
                    ]
                else:
                    return_message.append(
                        import_error(
                            f"multiple rows for the same area-source '{source_name}'",
                            validation=validation,
                        )
                    )
        row_nr += 1

    existing_facility_names = set([f.name for f in facilities.values()])
    duplicate_facility_names = []
    for official_id, f in create_facilities.items():
        if f.name in existing_facility_names:
            duplicate_facility_names.append(f.name)
    if len(duplicate_facility_names) > 0:
        return_message.append(
            import_error(
                "The following facility names are already "
                "used in inventory but for facilities with "
                f"different official_id: {duplicate_facility_names}",
                validation=validation,
            )
        )
    duplicate_facility_names = {}
    for f in create_facilities.values():
        if f.name in duplicate_facility_names:
            duplicate_facility_names[f.name] += 1
        else:
            duplicate_facility_names[f.name] = 1
    duplicate_facility_names = [
        name for name, nr in duplicate_facility_names.items() if nr > 1
    ]
    if len(duplicate_facility_names) > 0:
        return_message.append(
            import_error(
                "The same facility name is used on multiple rows but "
                f"with different facility_id: {duplicate_facility_names}",
                validation=validation,
            )
        )

    Facility.objects.bulk_create(create_facilities.values())
    Facility.objects.bulk_update(update_facilities, ["name"])

    facilities = cache_queryset(Facility.objects.all(), "id")
    # ensure PointSource.facility_id is not None if facility exists.
    for source in create_sources.values():
        if source.facility is not None:
            # find the facility_id corresponding to official id, or set None
            source.facility_id = next(
                (
                    key
                    for key, value in facilities.items()
                    if str(value) == str(source.facility)
                ),
                None,
            )
            if source.facility_id is None:
                raise ImportError(
                    f"Could not link pointsource {source.name} to "
                    + f"facility {source.facility.name}"
                )
    if sourcetype == "point":
        PointSource.objects.bulk_create(create_sources.values())
        PointSource.objects.bulk_update(
            update_sources,
            [
                "name",
                "geom",
                "tags",
                "chimney_gas_speed",
                "chimney_gas_temperature",
                "chimney_height",
                "chimney_inner_diameter",
                "chimney_outer_diameter",
                "house_height",
                "house_width",
                "activitycode1",
                "activitycode2",
                "activitycode3",
            ],
        )

        # drop existing substance emissions of point-sources that will be updated
        PointSourceSubstance.objects.filter(
            pk__in=[inst.id for inst in drop_substances]
        ).delete()

        # ensure PointSourceSubstance.source_id is not None
        for emis in create_substances:
            emis.source_id = PointSource.objects.get(
                name=emis.source, facility_id=emis.source.facility_id
            ).id
        PointSourceSubstance.objects.bulk_create(create_substances)
        return_dict = {
            "facility": {
                "updated": len(update_facilities),
                "created": len(create_facilities),
            },
            "pointsource": {
                "updated": len(update_sources),
                "created": len(create_sources),
            },
        }
    if sourcetype == "area":
        AreaSource.objects.bulk_create(create_sources.values())
        AreaSource.objects.bulk_update(
            update_sources,
            [
                "name",
                "geom",
                "tags",
                "activitycode1",
                "activitycode2",
                "activitycode3",
            ],
        )

        # drop existing substance emissions of point-sources that will be updated
        AreaSourceSubstance.objects.filter(
            pk__in=[inst.id for inst in drop_substances]
        ).delete()

        # ensure PointSourceSubstance.source_id is not None
        for emis in create_substances:
            emis.source_id = AreaSource.objects.get(
                name=emis.source, facility_id=emis.source.facility_id
            ).id
        AreaSourceSubstance.objects.bulk_create(create_substances)
        return_dict = {
            "facility": {
                "updated": len(update_facilities),
                "created": len(create_facilities),
            },
            "areasource": {
                "updated": len(update_sources),
                "created": len(create_sources),
            },
        }
    # print(datetime.now().strftime("%H:%M:%S") + "finish point import")
    return return_dict, return_message


# @profile
def import_sourceactivities(
    filepath,
    encoding=None,
    srid=None,
    import_sheets=SHEET_NAMES,
    validation=False,
):
    """Import sheets Timevar, Codeset, ActivityCode, EmissionFactor,
    PointSource and AreaSource from xlsx. If GridSource in import_sheets,
    it will be ignored.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is wgs84
        import_sheets: sheets to be imported, default is all.
    """
    return_message = []
    try:
        workbook = load_workbook(filename=filepath, data_only=True, read_only=True)
    except Exception as exc:
        return_message.append(import_error(str(exc), validation))

    return_dict = {}
    sheet_names = [sheet.title for sheet in workbook.worksheets]
    if ("Timevar" in sheet_names) and ("Timevar" in import_sheets):
        log.debug("validating/importing timevars")
        updates, msgs = import_timevarsheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("CodeSet" in sheet_names) and ("CodeSet" in import_sheets):
        log.debug("validating/importing code-sets")
        updates, msgs = import_codesetsheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("ActivityCode" in sheet_names) and ("ActivityCode" in import_sheets):
        log.debug("validating/importing activity-codes")
        updates, msgs = import_activitycodesheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("EmissionFactor" in sheet_names) and ("EmissionFactor" in import_sheets):
        log.debug("validating/importing emission-factors")
        updates, msgs = import_emissionfactorsheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("PointSource" in sheet_names) and ("PointSource" in import_sheets):
        log.debug("validating/importing sheet PointSource")
        data = workbook["PointSource"].values
        df_pointsource = worksheet_to_dataframe(data)
        df_pointsource = set_datatypes(df_pointsource, "point")
        # import pointsources and pointsourcesubstances
        caching_sources = len(df_pointsource) > PointSource.objects.count()
        ps, msgs = create_or_update_sources(
            df_pointsource,
            srid=srid,
            validation=validation,
            sourcetype="point",
            cache=caching_sources,
        )
        return_dict.update(ps)
        # If the pointsource data is not OK, no point in continuing
        return_message += msgs
        if validation and len(msgs) > 0:
            return return_dict, return_message

        # import pointsourceactivities
        activities = cache_queryset(Activity.objects.all(), "name")
        facilities = cache_queryset(Facility.objects.all(), "official_id")
        if caching_sources:
            log.debug("caching sources to speed up updates")
            pointsourceactivities = cache_queryset(
                PointSourceActivity.objects.select_related("activity", "source").all(),
                ["activity", "source"],
            )
            pointsources = cache_sources(
                PointSource.objects.select_related("facility").all()
            )
        log.debug("Reading sources")
        create_pointsourceactivities = []
        update_pointsourceactivities = []
        # NB: does not work if column header starts with space, but same for subst:
        activity_keys = [k for k in df_pointsource.columns if k.startswith("act:")]
        row_nr = 1
        for row_key, row in df_pointsource.iterrows():
            # original unit stored in activity.unit, but
            # pointsourceactivity.rate stored as activity / s.
            if pd.isna(row.name[0]):
                facility_id = None
            else:
                facility_id = str(row.name[0])
            if caching_sources:
                # row index set as ["facility_id", "source_name"]
                # in create_or_update_source
                pointsource = pointsources[facility_id, str(row.name[1])]
            else:
                facility = facilities[facility_id] if facility_id is not None else None
                pointsource = PointSource.objects.get(
                    name=str(row.name[1]), facility=facility
                )
            log.debug(f"pointsource {row_nr}: {pointsource.name}")
            for activity_key in activity_keys:
                if not pd.isnull(row[activity_key]):
                    rate = row[activity_key]
                    activity_name = activity_key[4:]
                    try:
                        activity = activities[activity_name]
                    except KeyError:
                        return_message += import_error(
                            f"unknown activity '{activity_name}'"
                            + f" for pointsource '{row['source_name']}'",
                            validation=validation,
                        )
                    rate = activity_rate_unit_to_si(rate, activity.unit)
                    try:
                        if caching_sources:
                            psa = pointsourceactivities[activity, pointsource]
                        else:
                            psa = PointSourceActivity.objects.get(
                                activity_id=activity.id, source_id=pointsource.id
                            )
                        setattr(psa, "rate", rate)
                        update_pointsourceactivities.append(psa)
                    except (PointSourceActivity.DoesNotExist, KeyError):
                        psa = PointSourceActivity(
                            activity=activity, source=pointsource, rate=rate
                        )
                        create_pointsourceactivities.append(psa)
            row_nr += 1
        log.debug("Creating point-sources")
        PointSourceActivity.objects.bulk_create(create_pointsourceactivities)
        PointSourceActivity.objects.bulk_update(
            update_pointsourceactivities, ["activity", "source", "rate"]
        )
        return_dict.update(
            {
                "pointsourceactivity": {
                    "updated": len(update_pointsourceactivities),
                    "created": len(create_pointsourceactivities),
                }
            }
        )
    if ("AreaSource" in sheet_names) and ("AreaSource" in import_sheets):
        data = workbook["AreaSource"].values
        df_areasource = worksheet_to_dataframe(data)
        df_areasource = set_datatypes(df_areasource, "area")
        ps, msgs = create_or_update_sources(
            df_areasource,
            srid=srid,
            validation=validation,
            sourcetype="area",
        )
        return_dict.update(ps)
        return_message += msgs
        # If the areasource data is not OK, no point in continuing
        if validation and len(msgs) > 0:
            return return_dict, return_message

        # for now always caching areasources, change if case with many areasources
        # becomes relevant.
        areasourceactivities = cache_queryset(
            AreaSourceActivity.objects.select_related("activity", "source").all(),
            ["activity", "source"],
        )
        activities = cache_queryset(Activity.objects.all(), "name")
        areasources = cache_sources(
            AreaSource.objects.select_related("facility")
            .prefetch_related("substances")
            .all()
        )
        create_areasourceactivities = []
        update_areasourceactivities = []
        for row_key, row in df_areasource.iterrows():
            # NB: does not work if column header starts with space, but same for subst:
            activity_keys = [k for k in row.keys() if k.startswith("act:")]
            for activity_key in activity_keys:
                if not pd.isna(row[activity_key]):
                    rate = float(row[activity_key])
                    try:
                        activity = activities[activity_key[4:]]
                    except KeyError:
                        return_message += import_error(
                            f"unknown activity '{activity_name}'"
                            f" for areasource '{row['source_name']}'",
                            validation=validation,
                        )

                    rate = activity_rate_unit_to_si(rate, activity.unit)
                    # original unit stored in activity.unit, but
                    # areasourceactivity.rate stored as activity / s.
                    # facility_id and source_name set as df index
                    if pd.isna(row.name[0]):
                        facility_id = None
                    else:
                        facility_id = str(row.name[0])
                    areasource = areasources[facility_id, str(row.name[1])]
                    try:
                        psa = areasourceactivities[activity, areasource]
                        setattr(psa, "rate", rate)
                        update_areasourceactivities.append(psa)
                    except KeyError:
                        psa = AreaSourceActivity(
                            activity=activity, source=areasource, rate=rate
                        )
                        create_areasourceactivities.append(psa)

        AreaSourceActivity.objects.bulk_create(create_areasourceactivities)
        AreaSourceActivity.objects.bulk_update(
            update_areasourceactivities, ["activity", "source", "rate"]
        )
        return_dict.update(
            {
                "areasourceactivity": {
                    "updated": len(update_areasourceactivities),
                    "created": len(create_areasourceactivities),
                }
            }
        )

    workbook.close()
    # print(datetime.now().strftime("%H:%M:%S") + "finish import")
    return return_dict, return_message

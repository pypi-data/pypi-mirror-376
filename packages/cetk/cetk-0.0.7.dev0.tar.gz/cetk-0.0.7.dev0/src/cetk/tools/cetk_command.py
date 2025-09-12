"""Command line interface for managing a Clair emission inventory offline."""

import argparse
import datetime
import os
import sys
from math import ceil
from pathlib import Path

from django.core import serializers
from openpyxl import load_workbook
from pyproj import CRS
from pyproj.exceptions import CRSError

import cetk
from cetk import logging
from cetk.db import run_migrate
from cetk.edb.utils import delete_sources
from cetk.tools.utils import (
    CalledProcessError,
    SubprocessError,
    add_standard_command_options,
    check_and_get_path,
    create_from_template,
    get_db,
    get_template_db,
)

log = logging.getLogger("cetk")

settings = cetk.configure()

from cetk.edb.const import DEFAULT_SRID  # noqa
from cetk.edb.const import DEFAULT_EMISSION_UNIT, SHEET_NAMES, WGS84_SRID  # noqa
from cetk.edb.exporters import export_sources  # noqa
from cetk.edb.importers import import_activitycodesheet  # noqa
from cetk.edb.importers import (  # noqa
    import_gridsources,
    import_sourceactivities,
    import_traffic,
)
from cetk.edb.importers.utils import ValidationError  # noqa
from cetk.edb.models import Settings, Substance  # noqa
from cetk.edb.rasterize.rasterizer import EmissionRasterizer, Output  # noqa
from cetk.emissions.calc import aggregate_emissions  # noqa
from cetk.emissions.calc import get_used_substances  # noqa
from cetk.emissions.views import create_emission_table  # noqa

SOURCETYPES = ("point", "area", "grid", "road")

sheet_choices = ["All"]
sheet_choices.extend(SHEET_NAMES)


def adjust_extent(extent, srid, cellsize):
    """adjust extent to include an integer nr of cells."""
    x1, y1, x2, y2 = (
        extent or Settings.get_current().extent.transform(srid, clone=True).extent
    )
    nx = ceil((x2 - x1) / cellsize)
    ny = ceil((y2 - y1) / cellsize)
    x2 = x1 + nx * cellsize
    y2 = y1 + ny * cellsize
    extent = (x1, y1, x2, y2)
    # Settings.extent is a Polygon, Settings.extent.extent a tuple (x1, y1, x2, y2)
    if extent is None:
        log.error("could not rasterize emissions, default extent not set")
        sys.exit(1)
    return extent, ny, nx


class Editor(object):
    def __init__(self):
        self.db_path = settings.DATABASES["default"]["NAME"]

    def migrate(self, template=False, db_path=None):
        if db_path is None:
            if template:
                db_path = get_template_db()
            else:
                db_path = get_db()
        log.debug(f"Running migrations for database {db_path}")
        try:
            std_out, std_err = run_migrate(db_path=db_path)
        except (CalledProcessError, SubprocessError) as err:
            log.error(f"Error while migrating {db_path}: {err}")
        else:
            log.debug(f"Successfully migrated database {db_path}")

    def info(self, **kwargs):
        settings = Settings.get_current()
        print(serializers.serialize("json", [settings]))

    def update_settings(self, srid=None, codeset1=None, codeset2=None, codeset3=None):
        from cetk.edb.models import CodeSet, Settings

        settings = Settings.get_current()
        if srid == WGS84_SRID:
            log.error(
                "SRID in settings must be a "
                "projected coordinate system in meters, not lat/lon.\n"
            )
            sys.exit(1)

        elif srid is not None:
            try:
                crs = CRS.from_epsg(srid)  # noqa
            except CRSError:
                log.error(f"SRID={srid} does not define a valid CRS.\n")
                sys.exit(1)
            settings.srid = srid
            log.info(f"Updated srid to {srid}.\n")
        try:
            if codeset1 is not None:
                slug = codeset1
                settings.codeset1 = CodeSet.objects.get(slug=slug)
                log.info(f"Updated code-set1 to {slug}.\n")
            if codeset2 is not None:
                slug = codeset2
                settings.codeset2 = CodeSet.objects.get(slug=slug)
                log.info(f"Updated code-set2 to {slug}.\n")
            if codeset3 is not None:
                slug = codeset3
                settings.codeset3 = CodeSet.objects.get(slug=slug)
                log.info(f"Updated code-set3 to {slug}.\n")
        except CodeSet.DoesNotExist:
            log.error(f"No code-set defined with slug: '{slug}'")
            sys.exit(1)
        settings.save()

    def import_workbook(self, filename, sheets=SHEET_NAMES, dry_run=False):
        return_msg = []
        db_updates = {}
        workbook = load_workbook(filename=filename, data_only=True, read_only=True)
        import_sheets = [
            s
            for s in SHEET_NAMES
            if s in frozenset(sheets) and s in frozenset(workbook.sheetnames)
        ]

        try:
            missing_sheets = set(import_sheets).difference(workbook.sheetnames)
            if len(missing_sheets) > 0:
                log.info(f"Workbook has no sheets named: {missing_sheets}")

            # PointSource and AreaSource
            sourceact = [
                "CodeSet",
                "ActivityCode",
                "EmissionFactor",
                "Timevar",
                "PointSource",
                "AreaSource",
            ]

            if any(name in sourceact for name in import_sheets):
                updates, msgs = import_sourceactivities(
                    filename,
                    import_sheets=import_sheets,
                    validation=dry_run,
                )
                db_updates.update(updates)
                return_msg += msgs

            traffic = [
                "RoadSource",
                "VehicleFuel",
                "Fleet",
                "CongestionProfile",
                "FlowTimevar",
                "ColdstartTimevar",
                "RoadAttribute",
                "TrafficSituation",
                "VehicleEmissionFactor",
            ]
            if any(name in traffic for name in import_sheets):
                updates, msgs = import_traffic(
                    filename, sheets=import_sheets, validation=dry_run
                )
                return_msg += msgs
                db_updates.update(updates)
            if "GridSource" in import_sheets:
                updates, msgs = import_gridsources(filename)
                db_updates.update(updates)
                return_msg += msgs
            # skip empty messages
            # return_msg = [entry.strip() for entry in return_msg if entry]
            if len(return_msg) > 0:
                raise ValidationError()
        except ValidationError as err:
            return_msg.append(str(err))
            if not dry_run:
                run_type = "import"
            else:
                run_type = "validation"
            if len(return_msg) > settings.MAX_ERROR_MESSAGES:
                log.error(
                    f">10 errors during {run_type}, first 10:{os.linesep}"
                    f"{os.linesep}{return_msg[:10]}"
                )
            else:
                log.error(
                    f"Errors during {run_type}:{os.linesep}"
                    f"{os.linesep.join(return_msg)}"
                )
        else:
            log.info(f"getting here {datetime.datetime.now()}")
            if not dry_run:
                log.info(
                    f"{datetime.datetime.now()} successfully imported {db_updates}"
                )
            else:
                log.info(
                    f"{datetime.datetime.now()} successfully validated {db_updates}"
                )
        workbook.close()
        return db_updates, return_msg

    def update_emission_tables(
        self,
        sourcetypes=None,
        unit=DEFAULT_EMISSION_UNIT,
        substances=None,
    ):
        sourcetypes = sourcetypes or ("point", "area", "road")
        substances = substances or get_used_substances()
        if len(substances) == 0:
            log.error("No emission factors or direct emissions found in database")
            sys.exit(1)

        for sourcetype in sourcetypes:
            create_emission_table(sourcetype, substances=substances, unit=unit)

    def aggregate_emissions(
        self,
        filename,
        sourcetypes=None,
        unit=DEFAULT_EMISSION_UNIT,
        codeset=None,
        substances=None,
    ):
        df = aggregate_emissions(
            sourcetypes=sourcetypes, unit=unit, codeset=codeset, substances=substances
        )
        try:
            df.to_excel(filename)
        except Exception as err:
            log.error(
                f"could not write aggregated emission to file {filename}: {str(err)}"
            )
            sys.exit(1)

    def rasterize_emissions(
        self,
        outputpath,
        cellsize,
        sourcetypes=None,
        point_ids=None,
        area_ids=None,
        grid_ids=None,
        road_ids=None,
        unit=DEFAULT_EMISSION_UNIT,
        codeset=None,
        substances=None,
        begin=None,
        end=None,
        extent=None,
        srid=None,
        timezone=None,
    ):
        substances = substances or get_used_substances()
        timezone = timezone or datetime.timezone.utc
        srid = srid or DEFAULT_SRID
        extent, ny, nx = adjust_extent(extent, srid, cellsize)

        if codeset:
            from cetk.edb.models.source_models import CodeSet

            codeset_index = Settings.get_current().get_codeset_index(codeset)
            code_labels = dict(
                CodeSet.objects.filter(id=codeset_index)
                .first()
                .codes.values_list("code", "label")
            )
            for code, label in code_labels.items():
                basename = f"eclair_{codeset}{code}_"
                log.debug(f"rasterizing for code: {code} : {label}")
                # try:
                output = Output(
                    extent=extent,
                    timezone=timezone,
                    path=outputpath,
                    srid=srid,
                    basename=basename,
                )
                kwargs = {f"ac{codeset_index}": code}
                rasterizer = EmissionRasterizer(output, nx=nx, ny=ny)
                rasterizer.process(
                    substances,
                    begin=begin,
                    end=end,
                    unit=unit,
                    sourcetypes=sourcetypes,
                    point_ids=point_ids,
                    area_ids=area_ids,
                    grid_ids=grid_ids,
                    road_ids=road_ids,
                    **kwargs,
                )
        else:
            try:
                output = Output(
                    extent=extent, timezone=timezone, path=outputpath, srid=srid
                )
                rasterizer = EmissionRasterizer(output, nx=nx, ny=ny)
                rasterizer.process(
                    substances,
                    begin=begin,
                    end=end,
                    unit=unit,
                    sourcetypes=sourcetypes,
                    point_ids=point_ids,
                    area_ids=area_ids,
                    grid_ids=grid_ids,
                    road_ids=road_ids,
                )
            except Exception as err:
                log.error(f"could not rasterize emissions: {str(err)}")
                sys.exit(1)

    def export_data(self, filename):
        export_sources(filename)
        return True


def main():
    db_path = get_db() or "unspecified"
    parser = argparse.ArgumentParser(
        description="Manage Clair offline emission inventories",
        usage=f"""cetk <command> [<args>]

        Main commands are:
        create   create an sqlite inventory
        migrate  migrate an sqlite inventory
        info     print settings
        import   import data
        delete   delete data
        export   export data
        calc     calculate emissions
        settings change database settings

        Current database is {db_path} (set by $CETK_DATABASE_PATH)
        """,
    )
    add_standard_command_options(parser)
    parser.add_argument(
        "command",
        help="Subcommand to run",
        choices=(
            "migrate",
            "create",
            "info",
            "import",
            "delete",
            "export",
            "calc",
            "settings",
        ),
    )
    verbosity = [arg for arg in sys.argv if arg == "-v"]
    sys_args = [arg for arg in sys.argv if arg != "-v"]
    sub_args = sys_args[2:]
    main_args = parser.parse_args(args=sys_args[1:2] + verbosity)
    editor = Editor()
    if main_args.command == "create":
        sub_parser = argparse.ArgumentParser(
            description="Create database from template.",
            usage="cetk create <filename>",
        )
        sub_parser.add_argument(
            "filename",
            help="Path of new database",
        )
        args = sub_parser.parse_args(sub_args)
        create_from_template(args.filename)
        log.debug(
            f"Created new database '{args.filename}' "
            f"from template '{get_template_db()}'"
        )
        sys.exit(0)

    if (
        len(sys.argv) < 2
        and sys.argv[2] not in ("-h", "--help", "--version")
        and db_path == "unspecified"
    ):
        sys.stderr.write("No database specified, set by $CETK_DATABASE_PATH\n")
        sys.exit(1)

    if main_args.command == "info":
        sub_parser = argparse.ArgumentParser(
            description="Print or update settings.",
            usage="usage: cetk info ",
        )
        args = sub_parser.parse_args(sub_args)
        editor.info()

    if main_args.command == "migrate":
        sub_parser = argparse.ArgumentParser(
            description=f"Migrate database {db_path}.",
            usage="usage: cetk migrate",
        )
        sub_parser.add_argument(
            "--template",
            action="store_true",
            help="Migrate the template database",
        )
        sub_parser.add_argument(
            "--dbpath",
            help="Specify database path manually",
        )
        args = sub_parser.parse_args(sub_args)
        editor.migrate(template=args.template, db_path=args.dbpath)
    elif main_args.command == "import":
        sub_parser = argparse.ArgumentParser(
            description="Import data from an xlsx-file",
            usage="cetk import <filename> [options]",
        )
        sub_parser.add_argument(
            "filename", help="Path to xslx-file", type=check_and_get_path
        )
        sub_parser.add_argument(
            "--sheets",
            nargs="+",
            default=SHEET_NAMES,
            help=f"List of sheets to import, valid names {SHEET_NAMES}",
        )
        sub_parser.add_argument(
            "--dryrun",
            action="store_true",
            help="Do dry run to validate import file without actually importing data",
        )
        args = sub_parser.parse_args(sub_args)
        if not Path(db_path).exists():
            sys.stderr.write(
                f"Database {db_path} does not exist, first run "
                "'cetk create' or 'cetk migrate'\n"
            )
            sys.exit(1)

        editor.import_workbook(args.filename, sheets=args.sheets, dry_run=args.dryrun)
        sys.exit(0)
    elif main_args.command == "delete":
        sub_parser = argparse.ArgumentParser(
            description="Delete sources",
            usage="cetk delete [options]",
        )
        sub_parser.add_argument(
            "--sourcetype", help="Only one sourcetype out of list", choices=SOURCETYPES
        )
        sub_parser.add_argument("--id", nargs="*", help="Id for sources to be deleted")
        args = sub_parser.parse_args(sub_args)
        from cetk.edb.models import AreaSource, GridSource, PointSource, RoadSource

        models = {
            "point": PointSource,
            "area": AreaSource,
            "road": RoadSource,
            "grid": GridSource,
        }
        if args.sourcetype in models.keys():
            try:
                delete_sources(models[args.sourcetype], args.id)
                sys.stdout.write(
                    f"Successfully deleted {args.sourcetype}sources {args.id}\n"
                )
                sys.exit(0)
            except ValueError as e:
                sys.stderr.write(f"Sources could not be deleted: \n {e} ")
                sys.exit(1)

    elif main_args.command == "calc":
        sub_parser = argparse.ArgumentParser(
            description="Calculate emissions",
            usage="cetk calc [options]",
        )
        sub_parser.add_argument(
            "--unit",
            default=DEFAULT_EMISSION_UNIT,
            help="Unit of emissions, default=%(default)s",
        )
        sub_parser.add_argument(
            "--sourcetypes", nargs="*", help="Only sourcetypes", choices=SOURCETYPES
        )
        sub_parser.add_argument(
            "--substances",
            nargs="*",
            help="Only substances (default is all with emissions)",
            choices=Substance.objects.values_list("slug", flat=True),
            metavar=("NOx", "PM10"),
        )
        sub_parser.add_argument(
            "--codeset",
            help="Aggregate or rasterize emissions by codeset",
            metavar="SLUG",
        )
        calc_grp = sub_parser.add_mutually_exclusive_group()
        calc_grp.add_argument(
            "--update", help="Create/update emission tables", action="store_true"
        )
        calc_grp.add_argument(
            "--aggregate", help="Aggregate emissions", metavar="FILENAME"
        )
        calc_grp.add_argument(
            "--rasterize", help="Rasterize emissions", metavar="OUTPUTPATH"
        )
        rasterize_grp = sub_parser.add_argument_group(
            "rasterize emissions", description="Settings to rasterize emissions"
        )
        rasterize_grp.add_argument(
            "--cellsize", help="Cellsize (meter) in output raster", type=float
        )
        rasterize_grp.add_argument(
            "--extent",
            help="Extent of output raster. Settings.extent is taken otherwise",
            nargs=4,
            type=float,
            metavar=("x1", "y1", "x2", "y2"),
        )
        rasterize_grp.add_argument(
            "--srid",
            help="EPSG of output raster. 4-5 digits integer",
            metavar="EPSG",
        )
        rasterize_grp.add_argument(
            "--begin",
            help="when hourly rasters are desired, specify begin date."
            + " Time 00:00 assumed",
            metavar="YYMMDDHH",
        )
        rasterize_grp.add_argument(
            "--end",
            help="when hourly rasters are desired, specify end date"
            + " Time 00:00 assumed",
            metavar="YYMMDDHH",
        )
        rasterize_grp.add_argument(
            "--point-ids",
            nargs="+",
            help="List of pointsource id's to include in rasterizer. "
            + "All sources will be rasterized for other sourcetypes if not limited "
            + "using argument sourcetypes or filtering specified for ids for "
            + " other sourcetypes too.",
        )
        rasterize_grp.add_argument(
            "--area-ids",
            nargs="+",
            help="List of areasource id's to include in rasterizer",
        )
        rasterize_grp.add_argument(
            "--road-ids",
            nargs="+",
            help="List of roadsource id's to include in rasterizer",
        )
        rasterize_grp.add_argument(
            "--grid-ids",
            nargs="+",
            help="List of gridsource id's to include in rasterizer",
        )
        # TODO add argument begin/end for rasterize
        # TODO add argument to aggregate emissions within polygon

        args = sub_parser.parse_args(sub_args)
        try:
            if args.substances is not None:
                substances = []
                for s in args.substances:
                    substances.append(Substance.objects.get(slug=s))
            else:
                substances = None
        except Substance.DoesNotExist:
            sys.stderr.write(f"Substance {s} does not exist.\n")
        if not Path(db_path).exists():
            sys.stderr.write("Database does not exist.\n")
            sys.exit(1)
        if args.update:
            editor.update_emission_tables(
                sourcetypes=args.sourcetypes, unit=args.unit, substances=args.substances
            )
            sys.stdout.write("Successfully updated tables\n")
            sys.exit(0)
        if args.aggregate is not None:
            editor.aggregate_emissions(
                args.aggregate,
                substances=substances,
                sourcetypes=args.sourcetypes,
                unit=args.unit,
                codeset=args.codeset,
            )
            sys.stdout.write("Successfully aggregated emissions\n")
            sys.exit(0)
        if args.rasterize is not None:
            if args.begin is not None:
                args.begin = datetime.datetime.strptime(args.begin, "%y%m%d%H").replace(
                    tzinfo=datetime.timezone.utc
                )
                if args.end is not None:
                    args.end = datetime.datetime.strptime(args.end, "%y%m%d%H").replace(
                        tzinfo=datetime.timezone.utc
                    )
                else:
                    sys.stderr.write(
                        "If begin is specified," + " end has to be specified too.\n"
                    )
                    sys.exit(1)
            editor.rasterize_emissions(
                args.rasterize,
                args.cellsize,
                substances=substances,
                sourcetypes=args.sourcetypes,
                point_ids=args.point_ids,
                area_ids=args.area_ids,
                grid_ids=args.grid_ids,
                road_ids=args.road_ids,
                unit=args.unit,
                extent=args.extent,
                srid=args.srid,
                begin=args.begin,
                end=args.end,
                codeset=args.codeset,
            )  # TODO add arguments for codeset, substances, begin/end, timezone!
            sys.stdout.write("Successfully rasterized emissions\n")
            sys.exit(0)

    elif main_args.command == "export":
        sub_parser = argparse.ArgumentParser(
            description="Export data to xlsx-file",
            usage="cetk export <filename> [options]",
        )
        sub_parser.add_argument("filename", help="Path to xslx-file")
        if not Path(db_path).exists():
            sys.stderr.write(
                "Database " + db_path + " does not exist, first run "
                "'cetk create' or 'cetk migrate'\n"
            )
            sys.exit(1)

        args = sub_parser.parse_args(sub_args)
        try:
            # Check if the file can be created at the given path
            os.access(args.filename, os.W_OK | os.X_OK)
        except Exception as e:
            sys.stderr.write(f"File {args.filename} cannot be created: \n {e} ")
            sys.exit(1)
        status = editor.export_data(args.filename)
        if status:
            sys.stdout.write(f"Exported data from {db_path} to {args.filename}.\n")
            sys.exit(0)
        else:
            sys.stderr.write("Did not export data, something went wrong.")
            sys.exit(1)

    elif main_args.command == "settings":
        sub_parser = argparse.ArgumentParser(
            description="Update settings",
            usage="cetk settings [--srid EPSG] [--codeset1 SLUG] [--codeset2 SLUG] [--codeset3 SLUG]",
        )
        sub_parser.add_argument("--srid", type=int, help="Update srid in settings")
        sub_parser.add_argument("--codeset1", help="Update code-set1")
        sub_parser.add_argument("--codeset2", help="Update code-set2")
        sub_parser.add_argument("--codeset3", help="Update code-set3")
        args = sub_parser.parse_args(sub_args)
        editor.update_settings(
            srid=args.srid,
            codeset1=args.codeset1,
            codeset2=args.codeset2,
            codeset3=args.codeset3,
        )

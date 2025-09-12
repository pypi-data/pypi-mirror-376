"""Tests for emission model exporters."""

import glob
import os
from importlib import resources
from tempfile import gettempdir

import pytest
from openpyxl import load_workbook

from cetk.db import run_migrate
from cetk.edb.const import SHEET_NAMES
from cetk.edb.importers.utils import worksheet_to_dataframe
from cetk.tools.utils import run_export, run_import


@pytest.fixture
def all_source_xlsx(tmpdir, settings):
    return resources.files("edb.data") / "all_sources.xlsx"


@pytest.fixture
def tmp_db(tmpdir):
    db_path = tmpdir / "test.eclair.gpkg"
    os.environ["CETK_DATABASE_PATH"] = str(db_path)
    return db_path


def test_export_sources(tmp_db, all_source_xlsx):
    run_migrate(db_path=tmp_db)
    backup_path, proc = run_import(all_source_xlsx, db_path=tmp_db)
    proc.wait()

    outpath = os.path.join(gettempdir(), "export.xlsx")
    proc = run_export(outpath, db_path=tmp_db)
    proc.wait()

    workbook = load_workbook(filename=outpath, data_only=True, read_only=True)
    assert len(workbook.sheetnames) == len(SHEET_NAMES)
    assert worksheet_to_dataframe(workbook["PointSource"].values).shape[0] == 4
    assert worksheet_to_dataframe(workbook["AreaSource"].values).shape[0] == 4
    assert worksheet_to_dataframe(workbook["GridSource"].values).shape[0] == 4
    assert worksheet_to_dataframe(workbook["TrafficSituation"].values).shape[0] == 18

    # check workbook.sheetnames should have all, but now only pointsources?
    workbook.close()

    # best test for correct format is to check if output can be imported again
    backup_path, proc = run_import(outpath, db_path=tmp_db)
    proc.wait()

    stderr_files = glob.glob(os.path.join(gettempdir(), "cetk_import_*_stderr.log"))
    stderr_files.sort(key=lambda f: int(f.split("_")[-2]))
    if stderr_files:
        latest_stderr_file = stderr_files[-1]
        with open(latest_stderr_file, "r") as f:
            stderr_content = f.read()

    assert "successfully" in stderr_content

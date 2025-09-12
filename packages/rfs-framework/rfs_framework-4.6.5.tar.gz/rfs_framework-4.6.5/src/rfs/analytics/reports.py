"""
RFS Advanced Analytics - Report Generation System (RFS v4.1)

리포트 생성 및 내보내기 시스템
"""

import asyncio
import base64
import io
import json
import tempfile
import uuid
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional, Union

from ..core.result import Failure, Result, Success
from .charts import Chart
from .dashboard import Dashboard
from .data_source import DataQuery, DataSource


class ReportFormat(Enum):
    PDF = "pdf"
    HTML = "html"
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"


class ReportSchedule(Enum):
    ONCE = "once"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"


@dataclass
class ReportSection:
    """리포트 섹션"""

    section_id: str
    title: str
    content_type: str
    content: Any
    metadata: Dict[str, Any] = field(default_factory=dict)
    width: Optional[str] = None
    height: Optional[str] = None
    padding: str = "10px"
    margin: str = "5px"
    condition: Optional[str] = None
    visible: bool = True


@dataclass
class ReportTemplate:
    """리포트 템플릿"""

    template_id: str
    name: str
    description: str
    sections: List[ReportSection]
    variables: Dict[str, Any] = field(default_factory=dict)
    styles: Dict[str, Any] = field(default_factory=dict)
    header_template: Optional[str] = None
    footer_template: Optional[str] = None
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)
    version: str = "1.0"


@dataclass
class ReportConfig:
    """리포트 생성 설정"""

    title: str
    subtitle: str = ""
    author: str = ""
    organization: str = ""
    page_size: str = "A4"
    orientation: str = "portrait"
    margin_top: str = "2cm"
    margin_bottom: str = "2cm"
    margin_left: str = "2cm"
    margin_right: str = "2cm"
    font_family: str = "Arial, sans-serif"
    font_size: str = "12pt"
    line_height: str = "1.5"
    include_header: bool = True
    include_footer: bool = True
    include_page_numbers: bool = True
    include_date: bool = True
    include_toc: bool = False
    include_summary: bool = False
    compress: bool = False


class Report(ABC):
    """리포트 추상 클래스"""

    def __init__(
        self, report_id: str, template: ReportTemplate, config: ReportConfig
    ) -> None:
        self.report_id = report_id
        self.template = template
        self.config = config
        self.sections: List[ReportSection] = []
        self.variables: Dict[str, Any] = {}
        self.generated_at: Optional[datetime] = None

    @abstractmethod
    async def generate(self) -> Result[bytes, str]:
        """리포트 생성

        Returns:
            Result[bytes, str]: 생성된 리포트 바이트 또는 오류
        """
        raise NotImplementedError("Subclasses must implement generate method")

    @abstractmethod
    def get_mime_type(self) -> str:
        """MIME 타입 반환

        Returns:
            str: MIME 타입 문자열
        """
        raise NotImplementedError("Subclasses must implement get_mime_type method")

    def add_section(self, section: ReportSection) -> Result[bool, str]:
        """섹션 추가"""
        try:
            self.sections = self.sections + [section]
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to add section: {str(e)}")

    def remove_section(self, section_id: str) -> Result[bool, str]:
        """섹션 제거"""
        try:
            self.sections = [s for s in self.sections if s.section_id != section_id]
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to remove section: {str(e)}")

    def set_variable(self, key: str, value: Any) -> Result[bool, str]:
        """변수 설정"""
        try:
            self.variables = {**self.variables, key: value}
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to set variable: {str(e)}")

    def get_variable(self, key: str, default: Any = None) -> Any:
        """변수 조회"""
        return self.variables.get(key, default)

    async def _process_sections(self) -> Result[List[ReportSection], str]:
        """섹션 처리 (조건부 표시, 변수 치환 등)"""
        try:
            processed_sections = []
            for section in self.sections:
                if section.condition:
                    if not self._evaluate_condition(section.condition):
                        continue
                if not section.visible:
                    continue
                processed_section = await self._substitute_variables(section)
                processed_sections = processed_sections + [processed_section]
            return Success(processed_sections)
        except Exception as e:
            return Failure(f"Section processing failed: {str(e)}")

    def _evaluate_condition(self, condition: str) -> bool:
        """조건 평가 (간단한 구현)"""
        try:
            for key, value in self.variables.items():
                condition = condition.replace(f"${{{key}}}", str(value))
            allowed_names = {"True": True, "False": False, "None": None}
            code = compile(condition, "<string>", "eval")
            return eval(code, {"__builtins__": {}}, allowed_names)
        except Exception:
            return True

    async def _substitute_variables(self, section: ReportSection) -> ReportSection:
        """변수 치환"""
        title = section.title
        for key, value in self.variables.items():
            title = title.replace(f"${{{key}}}", str(value))
        return ReportSection(
            section_id=section.section_id,
            title=title,
            content_type=section.content_type,
            content=section.content,
            metadata=section.metadata,
            width=section.width,
            height=section.height,
            padding=section.padding,
            margin=section.margin,
            condition=section.condition,
            visible=section.visible,
        )


class PDFReport(Report):
    """PDF 리포트"""

    def get_mime_type(self) -> str:
        return "application/pdf"

    async def generate(self) -> Result[bytes, str]:
        """PDF 리포트 생성"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )

            buffer = io.BytesIO()
            page_size = A4 if self.config.page_size == "A4" else letter
            doc = SimpleDocTemplate(
                buffer,
                pagesize=page_size,
                topMargin=72,
                bottomMargin=72,
                leftMargin=72,
                rightMargin=72,
            )
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Title"],
                fontSize=24,
                spaceAfter=30,
                alignment=1,
            )
            story = []
            if self.config.title:
                story = story + [Paragraph(self.config.title, title_style)]
                story = story + [Spacer(1, 20)]
            if self.config.subtitle:
                story = story + [
                    Paragraph(self.config.subtitle, styles.get("Heading2"))
                ]
                story = story + [Spacer(1, 20)]
            if self.config.author or self.config.organization:
                meta_text = []
                if self.config.author:
                    meta_text = meta_text + [f"Author: {self.config.author}"]
                if self.config.organization:
                    meta_text = meta_text + [
                        f"Organization: {self.config.organization}"
                    ]
                if self.config.include_date:
                    meta_text = meta_text + [
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    ]
                story = story + [
                    Paragraph("<br/>".join(meta_text), styles.get("Normal"))
                ]
                story = story + [Spacer(1, 30)]
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            sections = sections_result.unwrap()
            for section in sections:
                if section.title:
                    story = story + [Paragraph(section.title, styles.get("Heading3"))]
                    story = story + [Spacer(1, 12)]
                match section.content_type:
                    case "text":
                        story = story + [
                            Paragraph(str(section.content), styles.get("Normal"))
                        ]
                    case "table":
                        if (
                            hasattr(section.content, "__class__")
                            and section.content.__class__.__name__ == "list"
                        ) and section.content:
                            if (
                                hasattr(section.content[0], "__class__")
                                and section.content[0].__class__.__name__ == "dict"
                            ):
                                headers = list(section.content[0].keys())
                                table_data = [headers]
                                for row in section.content:
                                    table_data = table_data + [
                                        [str(row.get(h, "")) for h in headers]
                                    ]
                            else:
                                table_data = section.content
                            table = Table(table_data)
                            table.setStyle(
                                TableStyle(
                                    [
                                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                        (
                                            "TEXTCOLOR",
                                            (0, 0),
                                            (-1, 0),
                                            colors.whitesmoke,
                                        ),
                                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                        ("FONTSIZE", (0, 0), (-1, 0), 14),
                                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                                    ]
                                )
                            )
                            story = story + [table]
                    case "chart":
                        if (
                            hasattr(section.content, "__class__")
                            and section.content.__class__.__name__ == "dict"
                        ) and "image_base64" in section.content:
                            image_data = base64.b64decode(
                                section.content["image_base64"]
                            )
                            img_buffer = io.BytesIO(image_data)
                            from reportlab.lib.utils import ImageReader
                            from reportlab.platypus import Image

                            img = Image(ImageReader(img_buffer), width=400, height=300)
                            story = story + [img]
                story = story + [Spacer(1, 20)]
            doc.build(story)
            self.generated_at = datetime.now()
            return Success(buffer.getvalue())
        except ImportError:
            return Failure("reportlab library required for PDF generation")
        except Exception as e:
            return Failure(f"PDF generation failed: {str(e)}")


class HTMLReport(Report):
    """HTML 리포트"""

    def get_mime_type(self) -> str:
        return "text/html"

    async def generate(self) -> Result[bytes, str]:
        """HTML 리포트 생성"""
        try:
            html_parts = []
            html_parts = html_parts + [self._generate_html_header()]
            html_parts = html_parts + [self._generate_title_section()]
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            sections = sections_result.unwrap()
            html_parts = html_parts + ['<div class="report-body">']
            for section in sections:
                html_parts = html_parts + [await self._generate_section_html(section)]
            html_parts = html_parts + ["</div>"]
            html_parts = html_parts + [self._generate_html_footer()]
            html_content = "\n".join(html_parts)
            self.generated_at = datetime.now()
            return Success(html_content.encode("utf-8"))
        except Exception as e:
            return Failure(f"HTML generation failed: {str(e)}")

    def _generate_html_header(self) -> str:
        """HTML 헤더 생성"""
        return f'\n<!DOCTYPE html>\n<html lang="en">\n<head>\n    <meta charset="UTF-8">\n    <meta name="viewport" content="width=device-width, initial-scale=1.0">\n    <title>{self.config.title}</title>\n    <style>\n        body {{\n            font-family: {self.config.font_family};\n            font-size: {self.config.font_size};\n            line-height: {self.config.line_height};\n            margin: 0;\n            padding: 20px;\n            background-color: #f9f9f9;\n        }}\n        .report-container {{\n            max-width: 1200px;\n            margin: 0 auto;\n            background: white;\n            padding: 40px;\n            box-shadow: 0 0 10px rgba(0,0,0,0.1);\n        }}\n        .report-title {{\n            text-align: center;\n            color: #333;\n            border-bottom: 3px solid #007bff;\n            padding-bottom: 20px;\n            margin-bottom: 30px;\n        }}\n        .report-subtitle {{\n            text-align: center;\n            color: #666;\n            margin-bottom: 20px;\n        }}\n        .report-meta {{\n            text-align: center;\n            color: #888;\n            font-size: 0.9em;\n            margin-bottom: 40px;\n        }}\n        .section {{\n            margin-bottom: 30px;\n        }}\n        .section-title {{\n            color: #007bff;\n            border-left: 4px solid #007bff;\n            padding-left: 15px;\n            margin-bottom: 15px;\n        }}\n        .table-container {{\n            overflow-x: auto;\n        }}\n        table {{\n            width: 100%;\n            border-collapse: collapse;\n            margin-bottom: 20px;\n        }}\n        th, td {{\n            padding: 12px;\n            text-align: left;\n            border-bottom: 1px solid #ddd;\n        }}\n        th {{\n            background-color: #007bff;\n            color: white;\n            font-weight: bold;\n        }}\n        tr:nth-child(even) {{\n            background-color: #f2f2f2;\n        }}\n        .chart-container {{\n            text-align: center;\n            margin: 20px 0;\n        }}\n    </style>\n</head>\n<body>\n<div class="report-container">\n'

    def _generate_title_section(self) -> str:
        """제목 섹션 생성"""
        parts = []
        if self.config.title:
            parts = parts + [f'<h1 class="report-title">{self.config.title}</h1>']
        if self.config.subtitle:
            parts = parts + [f'<h2 class="report-subtitle">{self.config.subtitle}</h2>']
        meta_parts = []
        if self.config.author:
            meta_parts = meta_parts + [f"Author: {self.config.author}"]
        if self.config.organization:
            meta_parts = meta_parts + [f"Organization: {self.config.organization}"]
        if self.config.include_date:
            meta_parts = meta_parts + [
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ]
        if meta_parts:
            parts = parts + [f'<div class="report-meta">{"| ".join(meta_parts)}</div>']
        return "\n".join(parts)

    async def _generate_section_html(self, section: ReportSection) -> str:
        """섹션 HTML 생성"""
        parts = ['<div class="section">']
        if section.title:
            parts = parts + [f'<h3 class="section-title">{section.title}</h3>']
        match section.content_type:
            case "text":
                parts = parts + [f"<p>{section.content}</p>"]
            case "table":
                if (
                    hasattr(section.content, "__class__")
                    and section.content.__class__.__name__ == "list"
                ) and section.content:
                    parts = parts + ['<div class="table-container">']
                    parts = parts + ["<table>"]
                    if (
                        hasattr(section.content[0], "__class__")
                        and section.content[0].__class__.__name__ == "dict"
                    ):
                        headers = list(section.content[0].keys())
                        parts = parts + ["<thead><tr>"]
                        for header in headers:
                            parts = parts + [f"<th>{header}</th>"]
                        parts = parts + ["</tr></thead>"]
                        parts = parts + ["<tbody>"]
                        for row in section.content:
                            parts = parts + ["<tr>"]
                            for header in headers:
                                parts = parts + [f"<td>{row.get(header, '')}</td>"]
                            parts = parts + ["</tr>"]
                        parts = parts + ["</tbody>"]
                    parts = parts + ["</table>"]
                    parts = parts + ["</div>"]
            case "chart":
                if (
                    hasattr(section.content, "__class__")
                    and section.content.__class__.__name__ == "dict"
                ):
                    if "html" in section.content:
                        parts = parts + ['<div class="chart-container">']
                        parts = parts + [section.content["html"]]
                        parts = parts + ["</div>"]
                    elif "image_base64" in section.content:
                        parts = parts + ['<div class="chart-container">']
                        parts = parts + [
                            f'<img src="data:image/png;base64,{section.content["image_base64"]}" alt="Chart" style="max-width: 100%; height: auto;">'
                        ]
                        parts = parts + ["</div>"]
        parts = parts + ["</div>"]
        return "\n".join(parts)

    def _generate_html_footer(self) -> str:
        """HTML 푸터 생성"""
        return "\n</div>\n</body>\n</html>\n"


class ExcelReport(Report):
    """Excel 리포트"""

    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    async def generate(self) -> Result[bytes, str]:
        """Excel 리포트 생성"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            current_row = 1
            if self.config.title:
                ws.cell(row=current_row, column=1, value=self.config.title)
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal="center")
                current_row = current_row + 2
            if self.config.author:
                ws.cell(
                    row=current_row, column=1, value=f"Author: {self.config.author}"
                )
                current_row = current_row + 1
            if self.config.include_date:
                ws.cell(
                    row=current_row,
                    column=1,
                    value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                )
                current_row = current_row + 2
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            sections = sections_result.unwrap()
            for section in sections:
                if section.title:
                    ws.cell(row=current_row, column=1, value=section.title)
                    section_title_cell = ws.cell(row=current_row, column=1)
                    section_title_cell.font = Font(size=14, bold=True)
                    current_row = current_row + 1
                if section.content_type == "text":
                    ws.cell(row=current_row, column=1, value=str(section.content))
                    current_row = current_row + 1
                elif section.content_type == "table":
                    if (
                        hasattr(section.content, "__class__")
                        and section.content.__class__.__name__ == "list"
                    ) and section.content:
                        if (
                            hasattr(section.content[0], "__class__")
                            and section.content[0].__class__.__name__ == "dict"
                        ):
                            headers = list(section.content[0].keys())
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(
                                    row=current_row, column=col, value=header
                                )
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(
                                    start_color="366092",
                                    end_color="366092",
                                    fill_type="solid",
                                )
                            current_row = current_row + 1
                            for row_data in section.content:
                                for col, header in enumerate(headers, 1):
                                    ws.cell(
                                        row=current_row,
                                        column=col,
                                        value=str(row_data.get(header, "")),
                                    )
                                current_row = current_row + 1
                current_row = current_row + 1
            buffer = io.BytesIO()
            wb.save(buffer)
            self.generated_at = datetime.now()
            return Success(buffer.getvalue())
        except ImportError:
            return Failure("openpyxl library required for Excel generation")
        except Exception as e:
            return Failure(f"Excel generation failed: {str(e)}")


class ReportBuilder:
    """리포트 빌더"""

    def __init__(self) -> None:
        self._template: Optional[ReportTemplate] = None
        self._config: Optional[ReportConfig] = None
        self._format: ReportFormat = ReportFormat.HTML
        self._data_sources: Dict[str, DataSource] = {}

    def template(self, template: ReportTemplate) -> "ReportBuilder":
        """템플릿 설정"""
        self._template = template
        return self

    def config(self, config: ReportConfig) -> "ReportBuilder":
        """설정 지정"""
        self._config = config
        return self

    def format(self, format: ReportFormat) -> "ReportBuilder":
        """포맷 설정"""
        self._format = format
        return self

    def data_source(self, name: str, source: DataSource) -> "ReportBuilder":
        """데이터 소스 추가"""
        self._data_sources = {**self._data_sources, name: source}
        return self

    async def build(self) -> Result[Report, str]:
        """리포트 생성"""
        if not self._template:
            return Failure("Template not specified")
        if not self._config:
            return Failure("Config not specified")
        report_id = str(uuid.uuid4())
        try:
            match self._format:
                case ReportFormat.PDF:
                    report = PDFReport(report_id, self._template, self._config)
                case ReportFormat.HTML:
                    report = HTMLReport(report_id, self._template, self._config)
                case ReportFormat.EXCEL:
                    report = ExcelReport(report_id, self._template, self._config)
                case _:
                    return Failure(f"Unsupported format: {self._format}")
            for template_section in self._template.sections:
                report.add_section(template_section)
            for key, value in self._template.variables.items():
                report.set_variable(key, value)
            return Success(report)
        except Exception as e:
            return Failure(f"Report build failed: {str(e)}")


async def generate_report(
    template: ReportTemplate,
    config: ReportConfig,
    format: ReportFormat = ReportFormat.HTML,
    variables: Optional[Dict[str, Any]] = None,
) -> Result[bytes, str]:
    """리포트 생성 헬퍼 함수"""
    builder = ReportBuilder()
    builder.template(template).config(config).format(format)
    report_result = await builder.build()
    if report_result.is_failure():
        return report_result
    report = report_result.unwrap()
    if variables:
        for key, value in variables.items():
            report.set_variable(key, value)
    return await report.generate()


async def schedule_report(
    template: ReportTemplate,
    config: ReportConfig,
    schedule: ReportSchedule,
    format: ReportFormat = ReportFormat.HTML,
    output_path: Optional[str] = None,
) -> Result[str, str]:
    """리포트 스케줄링 (간단한 구현)"""
    try:
        schedule_id = str(uuid.uuid4())
        schedule_info = {
            "schedule_id": schedule_id,
            "template": template,
            "config": config,
            "schedule": schedule,
            "format": format,
            "output_path": output_path,
            "created_at": datetime.now().isoformat(),
            "next_run": _calculate_next_run(schedule).isoformat(),
        }
        schedule_path = (
            Path(tempfile.gettempdir()) / f"rfs_report_schedule_{schedule_id}.json"
        )
        with open(schedule_path, "w") as f:
            serializable_info = {
                "schedule_id": schedule_id,
                "template_id": template.template_id,
                "schedule": schedule.value,
                "format": format.value,
                "output_path": output_path,
                "created_at": schedule_info["created_at"],
                "next_run": schedule_info["next_run"],
            }
            json.dump(serializable_info, f, indent=2)
        return Success(schedule_id)
    except Exception as e:
        return Failure(f"Schedule creation failed: {str(e)}")


def _calculate_next_run(schedule: ReportSchedule) -> datetime:
    """다음 실행 시간 계산"""
    now = datetime.now()
    match schedule:
        case ReportSchedule.ONCE:
            return now
        case ReportSchedule.DAILY:
            return now + timedelta(days=1)
        case ReportSchedule.WEEKLY:
            return now + timedelta(weeks=1)
        case ReportSchedule.MONTHLY:
            if now.month == 12:
                return now.replace(year=now.year + 1, month=1)
            else:
                return now.replace(month=now.month + 1)
        case ReportSchedule.QUARTERLY:
            return now + timedelta(days=90)
        case ReportSchedule.YEARLY:
            return now.replace(year=now.year + 1)
    return now


def create_report_template(
    template_id: str, name: str, description: str = ""
) -> ReportTemplate:
    """리포트 템플릿 생성 헬퍼 함수"""
    return ReportTemplate(
        template_id=template_id, name=name, description=description, sections=[]
    )


def create_text_section(section_id: str, title: str, content: str) -> ReportSection:
    """텍스트 섹션 생성"""
    return ReportSection(
        section_id=section_id, title=title, content_type="text", content=content
    )


def create_table_section(
    section_id: str, title: str, data: List[Dict[str, Any]]
) -> ReportSection:
    """테이블 섹션 생성"""
    return ReportSection(
        section_id=section_id, title=title, content_type="table", content=data
    )


def create_chart_section(
    section_id: str, title: str, chart_data: Dict[str, Any]
) -> ReportSection:
    """차트 섹션 생성"""
    return ReportSection(
        section_id=section_id, title=title, content_type="chart", content=chart_data
    )

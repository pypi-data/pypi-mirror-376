from typing import List

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.colors import BLACK, WHITE
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.numbers import NumberFormat
from openpyxl.worksheet.worksheet import Worksheet

__border_side = Side(border_style=BORDER_THIN, color=BLACK)

thin_border = Border(
        left=__border_side,
        right=__border_side,
        top=__border_side,
        bottom=__border_side
)

blue_bg = PatternFill("solid", fgColor="203763")
yellow_bg = PatternFill("solid", fgColor="FDF2CC")

blue_fill = PatternFill(
        start_color="4bacc6",
        end_color="4bacc6",
        fill_type="solid"
)

green_fill = PatternFill(
        start_color="92d050",
        end_color="92d050",
        fill_type="solid"
)
grey_fill = PatternFill(
        start_color="d3d3d3",
        end_color="d3d3d3",
        fill_type="solid"
)
yellow_fill = PatternFill(
        start_color="ffd966",
        end_color="ffd966",
        fill_type="solid"
)
orange_fill = PatternFill(
        start_color="FFB61E",
        end_color="FFB61E",
        fill_type="solid"
)
black_fill = PatternFill(
        start_color="000000",
        end_color="000000",
        fill_type="solid"
)

red_highlight = PatternFill(bgColor="FFC7CE")
green_highlight = PatternFill(bgColor="C6EFCD")

red_font = Font(name="Calibri", size=12, bold=True, color="ff0000")

date_number_format = NumberFormat(10, formatCode="m/d/yyyy")
date_style = NamedStyle("date_style", number_format=date_number_format.formatCode, alignment=Alignment())


def style_table_title(cell: Cell) -> None:
    cell.fill = blue_bg
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=14, bold=True, color=WHITE)


def style_table_headers(worksheet: Worksheet, row_num: int) -> None:
    for cell in worksheet[row_num]:
        cell.fill = blue_bg
        cell.font = Font(color=WHITE)


def style_table_desc(cell: Cell) -> None:
    cell.fill = yellow_bg
    cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_adjust_width_of_columns(worksheet: Worksheet):
    min_col_width = 10
    max_col_width = 50

    def _as_text(value) -> str:
        if value is None:
            return ""
        return str(value)

    for column_cells in worksheet.columns:
        max_cell_length = max(len(_as_text(cell.value).strip()) for cell in column_cells)
        length = max(max_cell_length, min_col_width)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length * 1.2, max_col_width)


def apply_wrap_text(cell: Cell) -> None:
    alignment = cell.alignment.copy()
    alignment.wrap_text = True
    cell.alignment = alignment


def apply_wrap_text_to_col(ws: Worksheet, col_letter: str) -> None:
    row = ws[col_letter]

    for cell in row:
        apply_wrap_text(cell)


def apply_width_to_columns(ws: Worksheet, col_letters: List[str], width: float) -> None:
    for letter in col_letters:
        ws.column_dimensions[letter].width = width


def sum_cols(sheet: Worksheet, sum_cell: str, range_to_sum: str) -> None:
    sheet[sum_cell] = f"=SUM({range_to_sum})"
    sheet[sum_cell].font = Font(bold=True)


def highlight_cells_red_if_trending_down(sheet: Worksheet, rule_range: str) -> None:
    rule = CellIsRule(operator="greaterThan", formula=[0], fill=red_highlight)
    sheet.conditional_formatting.add(
            rule_range, rule
    )


def highlight_cells_green_if_trending_up(sheet: Worksheet, rule_range: str) -> None:
    rule = CellIsRule(operator="lessThan", formula=[0], fill=green_highlight)
    sheet.conditional_formatting.add(
            rule_range, rule
    )


def highlight_cells_red_based_off_text(sheet: Worksheet, rule_range: str, rule_text: str) -> None:
    dxf = DifferentialStyle(fill=red_highlight)
    rule = Rule(type="containsText", operator="containsText", text=rule_text, dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("FALSE",A1)))']
    sheet.conditional_formatting.add(rule_range, rule)


def remove_worksheet(wb: Workbook, sheet_name: str) -> None:
    if sheet_name in wb.get_sheet_names():
        wb.remove_sheet(wb[sheet_name])


def create_col_names_dict(worksheet: Worksheet) -> dict:
    """
    Returns a dictionary that maps column names to excel column letters
    """
    tmp = [(c[0].value, c[0].column_letter) for c in worksheet.iter_cols(1, worksheet.max_column)]
    return dict(tmp)


def style_worksheet_headers(ws: Worksheet, data_table=False):
    for cell in ws[1]:
        if data_table:
            cell.font = Font(name="Calibri", size=12, bold=True, color="ffffff")
            cell.fill = PatternFill(start_color="4472c4", end_color="4472c4", fill_type="solid")
        else:
            cell.font = Font(name="Calibri", size=12, bold=True)

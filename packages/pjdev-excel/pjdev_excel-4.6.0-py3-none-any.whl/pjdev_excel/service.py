from typing import List, TypeVar

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from pydantic import BaseModel

from pjdev_excel.models import RowDefinition
from pjdev_excel.utilities import auto_adjust_width_of_columns

T = TypeVar("T", bound=BaseModel)


def create_workbook(
    sheet_name: str, row_definitions: List[RowDefinition[T]], data: List[T]
) -> Workbook:
    wb = Workbook()
    wb = add_worksheet_to_workbook(wb, sheet_name, row_definitions, data)
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    return wb


def add_worksheet_to_workbook(
    wb: Workbook,
    sheet_name: str,
    row_definitions: List[RowDefinition[T]],
    data: List[T],
) -> Workbook:
    sheet = wb.create_sheet(sheet_name)
    sheet.title = sheet_name
    sheet.append([d.column_name for d in row_definitions])
    for d in data:
        sheet.append([r.value_getter(d) for r in row_definitions])

    auto_adjust_width_of_columns(sheet)
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(row_definitions))}{len(data) + 1}"
    )

    return wb

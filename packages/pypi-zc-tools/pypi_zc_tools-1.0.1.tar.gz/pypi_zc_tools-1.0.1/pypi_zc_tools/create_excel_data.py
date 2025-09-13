# -*- coding：utf-8 -*-
# 项目名称：pypi_zc_tools
# 编辑文件名：create_excel_data.py
# 系统日期：2025/9/13
# 编辑用户：ZC


import click
from faker import Faker
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# 创建实例
fake = Faker(locale='zh_CN')


def optimized_column_width(worksheet):
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)

        # 获取该列所有单元格的内容长度
        lengths = [len(str(cell.value)) if cell.value else 0 for cell in column]

        if lengths:  # 确保列不为空
            max_length = max(lengths)
            # 设置列宽，限制最大宽度
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 30)


@click.command()
@click.option('--output',default='data.xlsx',help='输出文件名')
@click.option('--count',type=click.IntRange(1,100),default=20,help='生成数量(行数)')
def generate_data(output, count):
    # 创建空列表来存储所有行数据
    all_data = []
    for num in range(count):
        row_data = [
            num + 1,
            '苏州市',
            fake.random_element(elements=('姑苏区', '相城区', '吴中区', '吴江区')),
            fake.building_number(),
            fake.ssn(),
            fake.name(),
            fake.password(),
            fake.phone_number(),
            fake.address(),
            fake.date_time_this_year(),
            fake.random_int(min=1, max=10),
            fake.company(),
            fake.job(),
            fake.email()
        ]
        all_data.append(row_data)

    headers = ['序号', '市', '区', '编号', '身份证号', '姓名', '密码', '电话', '住址',
           '时间', '批次','公司','工作','邮箱']


    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in all_data:
        ws.append(row)

    optimized_column_width(ws)
    wb.save(output)
    click.echo(f'数据已保存到{output}')


if __name__ == '__main__':
    generate_data()

# 运行命令：python create_excel_data.py --output=data.xlsx --count=20
import contextlib
import dataclasses
from typing import Any, NamedTuple

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

__version__ = "1.2.4"


class CellCord(NamedTuple):
    column: int
    row: int

    def offset(self, column: int = 0, row: int = 0):
        return CellCord(column=self.column + column, row=self.row + row)

    @classmethod
    def from_string(cls, cell: str):
        c, r = coordinate_from_string(cell)
        return cls(column=column_index_from_string(c) - 1, row=r - 1)


def get_coord(coord: CellCord) -> str:
    return f"{get_column_letter(coord.column + 1)}{coord.row + 1}"


def get_area(coord: CellCord, width: int, height: int) -> str:
    return f"{get_coord(coord)}:{get_coord(coord.offset(column=width - 1, row=height - 1))}"


type Fieldnames = list[str | tuple[str, int]] | list[str] | list[tuple[str, int]]


@dataclasses.dataclass
class _DictWriter:
    fieldnames: Fieldnames
    cell: CellCord
    ws: Worksheet
    auto_header: bool
    _height: int = 0

    def __post_init__(self):
        if self.auto_header:
            self.writeheader()

    @property
    def current_first_cell(self):
        return self.cell.offset(row=self._height)

    @property
    def area(self):
        return get_area(self.cell, len(self.fieldnames), self._height)

    def _write_cell(self, coord: CellCord, value: Any):
        self.ws.cell(row=coord.row + 1, column=coord.column + 1, value=value)

    def writeheader(self):
        cell = self.current_first_cell
        for i, fieldname in enumerate(self.fieldnames):
            if isinstance(fieldname, tuple):
                fieldname, width = fieldname
                self.ws.column_dimensions[get_column_letter(cell.column + i + 1)].width = width
            self._write_cell(cell.offset(column=i), fieldname)
        self._height += 1

    def writerow(self, row: dict[str, Any]):
        cell = self.current_first_cell
        for i, fieldname in enumerate(self.fieldnames):
            if isinstance(fieldname, tuple):
                fieldname, _ = fieldname
            try:
                value = row[fieldname]
                self._write_cell(cell.offset(column=i), value)
            except KeyError:
                pass
        self._height += 1


@contextlib.contextmanager
def DictWriter(
    ws: Worksheet,
    cell: str | CellCord,
    fieldnames: Fieldnames,
    auto_header: bool = False,
    displayName: str = "Table1",
    style: str = "TableStyleMedium9",
):
    if isinstance(cell, str):
        cell = CellCord.from_string(cell)

    writer = _DictWriter(fieldnames, cell, ws, auto_header)

    yield writer

    tab = Table(displayName=displayName, ref=f"{writer.area}")
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)

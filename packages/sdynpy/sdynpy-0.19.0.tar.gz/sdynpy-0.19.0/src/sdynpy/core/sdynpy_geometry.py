"""
Objects and procedures to handle operations on test or model geometries

This module defines a Geometry object as well as all of the subcomponents of
a geometry object: nodes, elements, tracelines and coordinate system.  Geometry
plotting is also handled in this module.

Copyright 2022 National Technology & Engineering Solutions of Sandia,
LLC (NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import copy
import numpy as np
import pyvista as pv
import pyvistaqt as pvqt
import pyqtgraph as pg
import vtk
from qtpy import uic, QtCore
from qtpy.QtGui import QKeySequence
try:
    from qtpy.QtGui import QAction
except ImportError:
    from qtpy.QtWidgets import QAction
from qtpy.QtWidgets import QApplication, QMainWindow, QSizePolicy, QFileDialog, QMenu
from .sdynpy_array import SdynpyArray
from .sdynpy_colors import colormap, coord_colormap
from .sdynpy_coordinate import CoordinateArray, coordinate_array, from_nodelist
from ..signal_processing.sdynpy_rotation import R, lstsq_rigid_transform
from ..signal_processing.sdynpy_camera import point_on_pixel
from ..core.sdynpy_matrix import matrix
import time
import os
from PIL import Image
from scipy.spatial import Delaunay
import pandas as pd
import openpyxl
import warnings

try:
    repr(pv.GPUInfo())
    IGNORE_PLOTS = False
    # Adding this to help out non-interactive consoles
    if QApplication.instance() is None:
        app = QApplication([''])

except RuntimeError:
    IGNORE_PLOTS = True
    print('No GPU Found, Geometry plotting will not work!')

# Enumerations

_cs_type_strings = {0: 'Cartesian',
                    1: 'Polar',
                    2: 'Spherical'}

_element_types = {
    11: 'Rod',
    21: 'Linear beam',
    22: 'Tapered beam',
    23: 'Curved beam',
    24: 'Parabolic beam',
    31: 'Straight pipe',
    32: 'Curved pipe',
    41: 'Plane Stress Linear Triangle',
    42: 'Plane Stress Parabolic Triangle',
    43: 'Plane Stress Cubic Triangle',
    44: 'Plane Stress Linear Quadrilateral',
    45: 'Plane Stress Parabolic Quadrilateral',
    46: 'Plane Strain Cubic Quadrilateral',
    51: 'Plane Strain Linear Triangle',
    52: 'Plane Strain Parabolic Triangle',
    53: 'Plane Strain Cubic Triangle',
    54: 'Plane Strain Linear Quadrilateral',
    55: 'Plane Strain Parabolic Quadrilateral',
    56: 'Plane Strain Cubic Quadrilateral',
    61: 'Plate Linear Triangle',
    62: 'Plate Parabolic Triangle',
    63: 'Plate Cubic Triangle',
    64: 'Plate Linear Quadrilateral',
    65: 'Plate Parabolic Quadrilateral',
    66: 'Plate Cubic Quadrilateral',
    71: 'Membrane Linear Quadrilateral',
    72: 'Membrane Parabolic Triangle',
    73: 'Membrane Cubic Triangle',
    74: 'Membrane Linear Triangle',
    75: 'Membrane Parabolic Quadrilateral',
    76: 'Membrane Cubic Quadrilateral',
    81: 'Axisymetric Solid Linear Triangle',
    82: 'Axisymetric Solid Parabolic Triangle',
    84: 'Axisymetric Solid Linear Quadrilateral',
    85: 'Axisymetric Solid Parabolic Quadrilateral',
    91: 'Thin Shell Linear Triangle',
    92: 'Thin Shell Parabolic Triangle',
    93: 'Thin Shell Cubic Triangle',
    94: 'Thin Shell Linear Quadrilateral',
    95: 'Thin Shell Parabolic Quadrilateral',
    96: 'Thin Shell Cubic Quadrilateral',
    101: 'Thick Shell Linear Wedge',
    102: 'Thick Shell Parabolic Wedge',
    103: 'Thick Shell Cubic Wedge',
    104: 'Thick Shell Linear Brick',
    105: 'Thick Shell Parabolic Brick',
    106: 'Thick Shell Cubic Brick',
    111: 'Solid Linear Tetrahedron',
    112: 'Solid Linear Wedge',
    113: 'Solid Parabolic Wedge',
    114: 'Solid Cubic Wedge',
    115: 'Solid Linear Brick',
    116: 'Solid Parabolic Brick',
    117: 'Solid Cubic Brick',
    118: 'Solid Parabolic Tetrahedron',
    121: 'Rigid Bar',
    122: 'Rigid Element',
    136: 'Node To Node Translational Spring',
    137: 'Node To Node Rotational Spring',
    138: 'Node To Ground Translational Spring',
    139: 'Node To Ground Rotational Spring',
    141: 'Node To Node Damper',
    142: 'Node To Gound Damper',
    151: 'Node To Node Gap',
    152: 'Node To Ground Gap',
    161: 'Lumped Mass',
    171: 'Axisymetric Linear Shell',
    172: 'Axisymetric Parabolic Shell',
    181: 'Constraint',
    191: 'Plastic Cold Runner',
    192: 'Plastic Hot Runner',
    193: 'Plastic Water Line',
    194: 'Plastic Fountain',
    195: 'Plastic Baffle',
    196: 'Plastic Rod Heater',
    201: 'Linear node-to-node interface',
    202: 'Linear edge-to-edge interface',
    203: 'Parabolic edge-to-edge interface',
    204: 'Linear face-to-face interface',
    208: 'Parabolic face-to-face interface',
    212: 'Linear axisymmetric interface',
    213: 'Parabolic axisymmetric interface',
    221: 'Linear rigid surface',
    222: 'Parabolic rigid surface',
    231: 'Axisymetric linear rigid surface',
    232: 'Axisymentric parabolic rigid surface'}

_exodus_elem_type_map = {'hex8': 115,
                         'tet4': 111,
                         'quad4': 64,
                         'tri3': 61,
                         'shell4': 64,
                         'shell3': 61,
                         'bar2': 21,
                         'truss2': 21,
                         'bar': 21,
                         'shell8': 65,
                         'hex20': 116,
                         'tetra10': 118,
                         'hex': 115,
                         'beam': 21,
                         'rod':21, # 21 : 'Linear beam',
                         'truss':21, # 21 : 'Linear beam',
                         'tri':41, # 41 : 'Plane Stress Linear Triangle',
                         'triangle':41, # 41 : 'Plane Stress Linear Triangle',
                         'quad':44, # 44 : 'Plane Stress Linear Quadrilateral',
                         'quadrilateral':44, # 44 : 'Plane Stress Linear Quadrilateral',
                         'hexahedral':115,
                         'tet':111,
                         'tetrahedral':111,
                         # Add new elements here
                         }

_beam_elem_types = [11, 21, 22, 23, 24, 31, 32, 121]
_face_element_types = [v for v in range(41, 107)]
_solid_element_types = [v for v in range(111, 119)]
_vtk_element_map = {
    111: vtk.VTK_TETRA,
    112: vtk.VTK_WEDGE,
    113: vtk.VTK_QUADRATIC_WEDGE,
    115: vtk.VTK_HEXAHEDRON,
    116: vtk.VTK_QUADRATIC_HEXAHEDRON,
    118: vtk.VTK_QUADRATIC_TETRA,
}

_vtk_connectivity_reorder = {
    vtk.VTK_QUADRATIC_HEXAHEDRON: [0, 1, 2, 3, 4, 5, 6,
                                   7, 8, 9, 10, 11, 16, 17, 18, 19, 12, 13, 14, 15]
}

MAX_NUMBER_REPR = 100


class GeometryPlotter(pvqt.BackgroundPlotter):
    """Class used to plot geometry

    This class is essentially identical to PyVista's BackgroundPlotter;
    however a small amount of additional functionality is added."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for action in self.main_menu.actions():
            if action.menu() and action.text().lower() == 'view':
                view_menu = action.menu()
                break
        for action in view_menu.actions():
            if action.menu() and action.text().lower() == 'orientation marker':
                orien_menu = action.menu()
                break
        orien_menu.addSeparator()
        orien_menu.addAction('Show Origin', self.show_origin_marker)
        orien_menu.addAction('Hide Origin', self.hide_origin_marker)
        self._origin_marker = None

    def show_origin_marker(self, arrow_scale=0.05):
        if self._origin_marker is None:
            bbox_diagonal = self.length
            arrow_factor = bbox_diagonal * arrow_scale
            self._origin_marker = pv.tools.create_axes_marker(label_size=np.array([0.25, 0.1])*arrow_scale)
            self._origin_marker.SetTotalLength(arrow_factor, arrow_factor, arrow_factor)
        self.renderer.AddActor(self._origin_marker)
        self.renderer.Modified()

    def hide_origin_marker(self):
        if self._origin_marker is not None:
            self.renderer.remove_actor(self._origin_marker)
            self.renderer.Modified()

    def save_rotation_animation(self, filename=None, frames=20, frame_rate=20,
                                individual_images=False):
        """
        Saves an animation of the rotating geometry to a file

        Parameters
        ----------
        filename : str, optional
            Path to the file being saved. The default is None, which returns
            the images as an array rather than saves them to a file
        frames : int, optional
            Number of frames in the animation. The default is 20.
        frame_rate : float, optional
            Number of frames per second if the animation is saved to a GIF.
            The default is 20.
        individual_images : bool, optional
            Boolean to specify whether the images are saved as individual PNG
            files or a single GIF. The default is False.

        Returns
        -------
        imgs : np.ndarray or None
            Returns array of images if filename is None, otherwise returns None

        """
        imgs = []
        phases = np.linspace(0, 2 * np.pi, frames, endpoint=False)
        focus = np.array(self.camera.focal_point)
        distance = np.array(self.camera.distance)
        initial_view_from = np.array(self.camera.position) - focus
        view_up = self.camera.up
        for phase in phases:
            # Compute a rotation matrix
            rotmat = R(view_up, phase)
            view_from = (rotmat @ initial_view_from[:, np.newaxis]).squeeze()
            self.camera.position = np.array(focus) + distance * \
                np.array(view_from) / np.linalg.norm(view_from)
            self.camera.focal_point = focus
            self.camera.up = view_up
            self.render()
            # Force GUI Update to ensure the OS doesn't delay events
            QApplication.processEvents()
            imgs.append(self.screenshot())
        if filename is None:
            return imgs
        else:
            if individual_images:
                for i, img in enumerate(imgs):
                    fname = os.path.splitext(filename)
                    fname = fname[0] + '-{:d}'.format(i) + fname[1]
                    Image.fromarray(img).save(fname)
            else:
                imgs = [Image.fromarray(img).convert(
                    'P', palette=Image.ADAPTIVE, colors=256) for img in imgs]
                imgs[0].save(fp=filename, format='GIF', append_images=imgs[1:],
                             save_all=True, duration=int(1 / frame_rate * 1000), loop=0)


class TransientPlotter(GeometryPlotter):
    """Class used to plot transient deformations"""

    def __init__(self, geometry, displacement_data, displacement_scale=1.0,
                 frames_per_second=20,
                 undeformed_opacity=0.0, deformed_opacity=1.0, plot_kwargs={},
                 transformation_shapes=None, num_curves=50,
                 show: bool = True,
                 app=None,
                 window_size=None,
                 off_screen=None,
                 allow_quit_keypress=True,
                 toolbar=True,
                 menu_bar=True,
                 editor=False,
                 update_app_icon=None,
                 **kwargs):
        """
        Create a TransientPlotter object to plot displacements over time

        Parameters
        ----------
        geometry : Geometry
            Geometry on which the displacements will be plotted
        displacement_data : TimeHistoryArray
            Transient displacement data that will be applied
        displacement_scale : float, optional
            Scale factor applied to displacements. The default is 1.0.
        frames_per_second : float, optional
            Number of time steps to plot per second while the displacement is
            animating.  Default is 20.
        undeformed_opacity : float, optional
            Opacity of the undeformed geometry. The default is 0.0, or
            completely transparent.
        deformed_opacity : float, optional
            Opacity of the deformed geometry. The default is 1.0, or completely
            opaque.
        plot_kwargs : dict, optional
            Keyword arguments passed to the Geometry.plot function
        transformation_shapes : ShapeArray
            Shape matrix that will be used to expand the data.  Must be the
            same size as the `displacement_data`
        num_curves : int, optional
            Maximum number of curves to plot on the time selector.  Default is
            50
        show : bool, optional
            Show the plotting window.  If ``False``, show this window by
            running ``show()``. The default is True.
        app : QApplication, optional
            Creates a `QApplication` if left as `None`.  The default is None.
        window_size : list of int, optional
            Window size in pixels.  Defaults to ``[1024, 768]``
        off_screen : TYPE, optional
            Renders off screen when True.  Useful for automated
            screenshots or debug testing. The default is None.
        allow_quit_keypress : bool, optional
            Allow user to exit by pressing ``"q"``. The default is True.
        toolbar : bool, optional
            If True, display the default camera toolbar. Defaults to True.
        menu_bar : bool, optional
            If True, display the default main menu. Defaults to True.
        editor : TYPE, optional
            If True, display the VTK object editor. Defaults to False.
        update_app_icon : bool, optional
            If True, update_app_icon will be called automatically to update the
            Qt app icon based on the current rendering output. If None, the
            logo of PyVista will be used. If False, no icon will be set.
            Defaults to None. The default is None.
        title : str, optional
            Title of plotting window.
        multi_samples : int, optional
            The number of multi-samples used to mitigate aliasing. 4 is a
            good default but 8 will have better results with a potential
            impact on performance.
        line_smoothing : bool, optional
            If True, enable line smothing
        point_smoothing : bool, optional
            If True, enable point smothing
        polygon_smoothing : bool, optional
            If True, enable polygon smothing
        auto_update : float, bool, optional
            Automatic update rate in seconds.  Useful for automatically
            updating the render window when actors are change without
            being automatically ``Modified``.  If set to ``True``, update
            rate will be 1 second.

        Returns
        -------
        None.

        """
        self.shape_select_toolbar: pvqt.plotting.QToolBar = None
        self.animation_control_toolbar: pvqt.plotting.QToolBar = None
        self.loop_action = None
        self.plot_abscissa_action = None
        self.timer: pvqt.plotting.QTimer = None
        self.frames_per_second = frames_per_second
        self.animation_speed = 1.0
        self.last_time = time.time()

        super(TransientPlotter, self).__init__(show, app, None if window_size is None else tuple(window_size),
                                               off_screen,
                                               allow_quit_keypress, toolbar,
                                               menu_bar, editor, update_app_icon,
                                               **kwargs)

        self.text = self.add_text('', font_size=10)
        self.loop_action.setChecked(False)
        self.plot_abscissa_action.setChecked(True)

        # Get the layout so we can add a new widget to it
        layout = self.frame.layout()
        self.time_selector_plot = pg.PlotWidget()
        layout.addWidget(self.time_selector_plot)
        sp = self.time_selector_plot.sizePolicy()
        sp.setVerticalPolicy(QSizePolicy.Minimum)
        self.time_selector_plot.setSizePolicy(sp)
        self.time_selector_plot.sizeHint = lambda: QtCore.QSize(1006, 80)
        # self.time_selector_plot.getPlotItem().hideAxis('left')
        # self.time_selector_plot.getPlotItem().axes['left']['item'].setStyle(showValues=False)

        # Now we need to plot the data in the plot
        if num_curves < displacement_data.size:
            indices = np.linspace(0, displacement_data.size - 1, num_curves).astype(int)
        else:
            indices = np.arange(displacement_data.size)
        for data in displacement_data.flatten()[indices]:
            self.time_selector_plot.plot(data.abscissa, data.ordinate)
        # Plot the envelope
        max_envelope = displacement_data.flatten().ordinate.max(axis=0)
        min_envelope = displacement_data.flatten().ordinate.min(axis=0)
        self.time_selector_plot.plot(displacement_data.flatten(
        )[0].abscissa, max_envelope, pen=pg.mkPen(color='k', width=2))
        self.time_selector_plot.plot(displacement_data.flatten(
        )[0].abscissa, min_envelope, pen=pg.mkPen(color='k', width=2))

        # Add an infinite line for a cursor
        self.time_selector = pg.InfiniteLine(
            0, movable=True, bounds=[displacement_data.abscissa.min(),
                                     displacement_data.abscissa.max()])
        self.time_selector_plot.addItem(self.time_selector)
        self.time_selector.sigPositionChanged.connect(self.modify_abscissa)

        # Set initial values
        self.displacement_scale = displacement_scale
        self.abscissa_index = 0
        self.displacement_data = displacement_data.flatten()
        self.transformation_shapes = None if transformation_shapes is None else transformation_shapes.flatten()
        self.plot_geometry = geometry
        self.num_abscissa = self.displacement_data.num_elements
        self.abscissa = self.displacement_data[0].abscissa
        self.abscissa_sig_figs = int(np.ceil(-np.log10(np.min(np.diff(self.abscissa))))) + 1
        self.abscissa_format_spec = 'Time: {{:0.{:}f}}'.format(
            self.abscissa_sig_figs if self.abscissa_sig_figs > 0 else 0)

        # Plot the geometry
        plotter, self.face_mesh_undeformed, self.point_mesh_undeformed, self.solid_mesh_undeformed = geometry.plot(plotter=self,
                                                                                                                   opacity=undeformed_opacity,
                                                                                                                   **plot_kwargs)

        plotter, self.face_mesh_deformed, self.point_mesh_deformed, self.solid_mesh_deformed = geometry.plot(plotter=self,
                                                                                                             opacity=deformed_opacity,
                                                                                                             **plot_kwargs)

        # Compute the node indices up front to make things faster when switching steps
        if self.transformation_shapes is None:
            coordinates = np.unique(self.displacement_data.coordinate)
        else:
            coordinates = np.unique(self.transformation_shapes.coordinate)
        coordinates = coordinates[np.in1d(coordinates.node, self.plot_geometry.node.id)
                                  & (np.abs(coordinates.direction) > 0)
                                  & (np.abs(coordinates.direction) < 4)]
        global_deflections = self.plot_geometry.global_deflection(coordinates).T
        if self.transformation_shapes is None:
            shape_array = self.displacement_data[coordinates[:, np.newaxis]].ordinate
        else:
            shape_array = self.transformation_shapes[coordinates].T
        # Compute global displacements for each node
        self.node_deflections = np.zeros((3 * self.plot_geometry.node.size, shape_array.shape[-1]))
        for node_index, node in self.plot_geometry.node.ndenumerate():
            coord_indices = coordinates.node == node.id
            self.node_deflections[node_index[0] * 3:(node_index[0] + 1) *
                                  3] = global_deflections[:, coord_indices] @ shape_array[coord_indices]
        if self.transformation_shapes is None:
            self.weighting = None
        else:
            self.weighting = self.displacement_data.ordinate
        self.update_displacement()

    def modify_abscissa(self):
        abscissa = self.time_selector.value()
        self.abscissa_index = np.argmin(np.abs(self.abscissa - abscissa))
        self.update_displacement(update_selector=False)

    def update_displacement(self, update_selector=True):
        index = int(round(self.abscissa_index))
        abscissa = self.abscissa[index]
        if self.plot_abscissa_action.isChecked():
            self.text.SetText(2, self.abscissa_format_spec.format(abscissa))
        else:
            self.text.SetText(2, '')
        if self.weighting is None:
            displacement = self.node_deflections[:, index] * self.displacement_scale
        else:
            displacement = self.node_deflections @ self.weighting[:,
                                                                  np.newaxis, index] * self.displacement_scale
        for mesh_deformed, mesh_undeformed in (
            (self.face_mesh_deformed, self.face_mesh_undeformed),
            (self.point_mesh_deformed, self.point_mesh_undeformed),
            (self.solid_mesh_deformed, self.solid_mesh_undeformed)
        ):
            if mesh_deformed is not None:
                mesh_deformed.points = (mesh_undeformed.points +
                                        displacement.reshape(-1, 3) * self.displacement_scale
                                        )
        # Update the time selector
        if update_selector:
            self.time_selector.blockSignals(True)
            self.time_selector.setValue(abscissa)
            self.time_selector.blockSignals(False)
        self.render()

    def add_toolbars(self) -> None:
        """
        Adds toolbars to the BackgroundPlotter

        """
        super().add_toolbars()
        self.shape_select_toolbar = self.app_window.addToolBar(
            "Shape Selector"
        )

        self._add_action(
            self.shape_select_toolbar, '|<', self.go_to_first_step
        )
        self._add_action(
            self.shape_select_toolbar, '<', self.go_to_previous_step
        )
        self._add_action(
            self.shape_select_toolbar, '>', self.go_to_next_step
        )
        self._add_action(
            self.shape_select_toolbar, '>|', self.go_to_last_step
        )

        self.animation_control_toolbar = self.app_window.addToolBar(
            "Animation Control"
        )

        self._add_action(
            self.animation_control_toolbar, '<  Play', self.play_animation_reverse
        )
        self._add_action(
            self.animation_control_toolbar, 'Stop', self.stop_animation
        )
        self._add_action(
            self.animation_control_toolbar, 'Play >', self.play_animation
        )

    def add_menu_bar(self) -> None:
        """
        Adds the menu bar to the BackgroundPlotter

        """
        super().add_menu_bar()

        shape_menu = self.main_menu.addMenu("Shape")
        scaling_menu = shape_menu.addMenu('Scaling')
        scaling_menu.addAction('1/4x', self.select_scaling_0p25)
        scaling_menu.addAction('1/2x', self.select_scaling_0p5)
        self.scale_0p8_action = QAction('0.8x')
        self.scale_0p8_action.triggered.connect(self.select_scaling_0p8)
        self.scale_0p8_action.setShortcut(QKeySequence(','))
        scaling_menu.addAction(self.scale_0p8_action)
        self.scale_reset_action = QAction('Reset')
        self.scale_reset_action.triggered.connect(self.select_scaling_1)
        self.scale_reset_action.setShortcut(QKeySequence('/'))
        scaling_menu.addAction(self.scale_reset_action)
        self.scale_1p25_action = QAction('1.25x')
        self.scale_1p25_action.triggered.connect(self.select_scaling_1p25)
        self.scale_1p25_action.setShortcut(QKeySequence('.'))
        scaling_menu.addAction(self.scale_1p25_action)
        scaling_menu.addAction('2x', self.select_scaling_2p0)
        scaling_menu.addAction('4x', self.select_scaling_4p0)
        speed_menu = shape_menu.addMenu('Animation Speed')
        self.speed_0p8_action = QAction('0.8x')
        self.speed_0p8_action.triggered.connect(self.select_speed_0p8)
        self.speed_0p8_action.setShortcut(QKeySequence(';'))
        speed_menu.addAction(self.speed_0p8_action)
        speed_menu.addAction('Reset', self.select_speed_1)
        self.speed_1p25_action = QAction('1.25x')
        self.speed_1p25_action.triggered.connect(self.select_speed_1p25)
        self.speed_1p25_action.setShortcut(QKeySequence("'"))
        speed_menu.addAction(self.speed_1p25_action)
        shape_menu.addSeparator()
        self.plot_abscissa_action = pvqt.plotting.QAction('Show Abscissa', checkable=True)
        shape_menu.addAction(self.plot_abscissa_action)
        self.loop_action = pvqt.plotting.QAction('Loop Animation', checkable=True)
        shape_menu.addAction(self.loop_action)

    def go_to_first_step(self):
        self.stop_animation()
        self.abscissa_index = 0
        self.update_displacement()

    def go_to_previous_step(self):
        self.stop_animation()
        self.abscissa_index -= 1
        if self.abscissa_index < 0:
            if self.loop_action.isChecked():
                self.abscissa_index = self.num_abscissa - 1
            else:
                self.abscissa_index = 0
        self.update_displacement()

    def go_to_next_step(self):
        self.stop_animation()
        self.abscissa_index += 1
        if self.abscissa_index > self.num_abscissa - 1:
            if self.loop_action.isChecked():
                self.abscissa_index = 0
            else:
                self.abscissa_index = self.num_abscissa - 1
        self.update_displacement()

    def go_to_last_step(self):
        self.stop_animation()
        self.abscissa_index = self.num_abscissa - 1
        self.update_displacement()

    def select_scaling_0p25(self):
        """
        Adjusts the current shape scaling by 0.25x
        """
        self.displacement_scale *= 0.25
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_0p5(self):
        """
        Adjusts the current shape scaling by 0.5x
        """
        self.displacement_scale *= 0.5
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_0p8(self):
        """
        Adjusts the current shape scaling by 0.8x
        """
        self.displacement_scale *= 0.8
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_1(self):
        """
        Resets shape scaling to 1
        """
        self.displacement_scale = 1.0
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_1p25(self):
        """
        Adjusts the current shape scaling by 1.25x
        """
        self.displacement_scale *= 1.25
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_2p0(self):
        """
        Adjusts the current shape scaling by 2x
        """
        self.displacement_scale *= 2.0
        self.last_time = time.time()
        self.update_displacement()

    def select_scaling_4p0(self):
        """
        Adjusts the current shape scaling by 4x
        """
        self.displacement_scale *= 4.0
        self.last_time = time.time()
        self.update_displacement()

    def select_speed_1(self):
        """
        Resets animation speed to default quantities
        """
        self.animation_speed = 1.0

    def select_speed_0p8(self):
        """
        Adjusts the current animation speed by 0.8
        """
        self.animation_speed *= 0.8

    def select_speed_1p25(self):
        """
        Adjusts the current animation speed by 1.25
        """
        self.animation_speed *= 1.25

    def play_animation(self):
        self.timer = pvqt.plotting.QTimer()
        self.timer.timeout.connect(self.update_abscissa)
        self.timer.start(10)
        self.last_time = time.time()

    def play_animation_reverse(self):
        self.timer = pvqt.plotting.QTimer()
        self.timer.timeout.connect(self.update_abscissa_reverse)
        self.timer.start(10)
        self.last_time = time.time()

    def stop_animation(self):
        if self.timer is not None:
            self.timer.stop()
            self.timer = None

    def update_abscissa(self):
        now = time.time()
        time_elapsed = now - self.last_time
        index_change = self.frames_per_second * self.animation_speed * time_elapsed
        self.last_time = now
        self.abscissa_index += index_change
        if self.abscissa_index > self.num_abscissa - 1:
            if self.loop_action.isChecked():
                self.abscissa_index = 0
            else:
                self.stop_animation()
                self.abscissa_index = self.num_abscissa - 1
        self.update_displacement()

    def update_abscissa_reverse(self):
        now = time.time()
        time_elapsed = now - self.last_time
        index_change = self.frames_per_second * self.animation_speed * time_elapsed
        self.last_time = now
        self.abscissa_index -= index_change
        if self.abscissa_index < 0:
            if self.loop_action.isChecked():
                self.abscissa_index = self.num_abscissa - 1
            else:
                self.stop_animation()
                self.abscissa_index = 0
        self.update_displacement()

    def save_animation(self, filename=None, frame_rate=20,
                       individual_images=False, start_time=None, stop_time=None,
                       step=None):
        """
        Saves the current shape animation to a file

        Parameters
        ----------
        filename : str, optional
            Path to the file being saved. The default is None, which returns
            the images as an array rather than saves them to a file
        frame_rate : float, optional
            Number of frames per second if the animation is saved to a GIF.
            The default is 20.
        individual_images : bool, optional
            Boolean to specify whether the images are saved as individual PNG
            files or a single GIF. The default is False.
        start_time : float, optional
            Time value at which the animation will start.  Default is first
            time step
        stop_time : float, optional
            Time value at which the animation will end.  Default is last time
            step
        step : int, optional
            Only render every ``step`` time steps.  Default is to render all
            timesteps
        Returns
        -------
        imgs : np.ndarray or None
            Returns array of images if filename is None, otherwise returns None

        """
        self.stop_animation()
        imgs = []
        indices = np.arange(self.abscissa.size)
        if start_time is not None:
            indices = indices[self.abscissa[indices] >= start_time]
        if stop_time is not None:
            indices = indices[self.abscissa[indices] <= stop_time]
        if step is not None:
            indices = indices[::step]
        for index in indices:
            self.abscissa_index = index
            self.update_displacement()
            # Force GUI Update to ensure the OS doesn't delay events
            QApplication.processEvents()
            imgs.append(self.screenshot())
        if filename is None:
            return imgs
        else:
            if individual_images:
                for i, img in enumerate(imgs):
                    fname = os.path.splitext(filename)
                    fname = fname[0] + '-{:d}'.format(i) + fname[1]
                    Image.fromarray(img).save(fname)
            else:
                imgs = [Image.fromarray(img).convert(
                    'P', palette=Image.ADAPTIVE, colors=256) for img in imgs]
                imgs[0].save(fp=filename, format='GIF', append_images=imgs[1:],
                             save_all=True, duration=int(1 / frame_rate * 1000), loop=0)

    def _close(self):
        self.stop_animation()
        super(TransientPlotter, self)._close()


class ShapePlotter(GeometryPlotter):
    """Class used to plot animated shapes"""

    def __init__(self, geometry, shapes, plot_kwargs,
                 background_plotter_kwargs={'auto_update': 0.01},
                 undeformed_opacity=0.25, starting_scale=1.0,
                 deformed_opacity=1.0, shape_name='Mode',
                 show_damping=True):
        """
        Create a Shape Plotter object to plot shapes

        Parameters
        ----------
        geometry : Geometry
            Geometry on which the shapes will be plotted
        shapes : ShapeArray
            Shapes to plot on the geometry
        plot_kwargs : dict
            Keyword arguments passed to the Geometry.plot function
        background_plotter_kwargs : dict, optional
            Keyword arguments passed to the BackgroundPlotter constructor.
            The default is {'auto_update':0.01}.
        undeformed_opacity : float, optional
            Opacity of the undeformed geometry.  Set to zero for no undeformed
            geometry. The default is 0.25.
        starting_scale : float, optional
            Starting scale of the shapes on the plot. The default is 1.0.
        shape_name : str, optional
            Name that will be used in the plotter to describe the shape. The
            default is "Mode"
        show_damping : bool, optional
            Boolean that specifies whether the damping should be displayed in
            the comment.
        """
        self.shape_select_toolbar: pvqt.plotting.QToolBar = None
        self.animation_control_toolbar: pvqt.plotting.QToolBar = None
        self.signed_amplitude_action = None
        self.complex_action = None
        self.real_action = None
        self.imag_action = None
        self.loop_action = None
        self.plot_comments_action = None
        self.save_animation_action = None
        self.display_undeformed = None
        self.geometry = geometry
        self.shapes = shapes.ravel()
        self.node_displacements = None
        self.current_shape = 0
        self.displacement_scale = 1.0
        self.displacement_scale_modification = starting_scale
        self.animation_speed = 1.0
        self.last_time = time.time()
        self.phase = 0.0
        self.timer: pvqt.plotting.QTimer = None
        self.shape_name = shape_name
        self.show_damping = show_damping
        self.shape_comment_table = None

        super().__init__(**background_plotter_kwargs)

        self.text = self.add_text('', font_size=10)

        self.complex_action.setChecked(True)
        self.loop_action.setChecked(True)
        self.plot_comments_action.setChecked(True)

        plotter, self.face_mesh_undeformed, self.point_mesh_undeformed, self.solid_mesh_undeformed = geometry.plot(plotter=self,
                                                                                                                   opacity=undeformed_opacity,
                                                                                                                   **plot_kwargs)

        plotter, self.face_mesh_deformed, self.point_mesh_deformed, self.solid_mesh_deformed = geometry.plot(plotter=self,
                                                                                                             opacity=deformed_opacity,
                                                                                                             **plot_kwargs)

        # Compute the node indices up front to make things faster when switching shapes
        coordinates = self.shapes[self.current_shape].coordinate
        nodes = coordinates.node
        # Only do coordinates that are in the geometry node array
        self.indices = np.in1d(coordinates.node, self.geometry.node.id)
        coordinates = coordinates[self.indices]
        nodes = nodes[self.indices]
        direction = coordinates.direction
        self.coordinate_node_index_map = {}
        for coordinate_index, node_id in np.ndenumerate(nodes):
            if direction[coordinate_index] == 0 or abs(direction[coordinate_index]) > 3:
                continue
            if node_id not in self.coordinate_node_index_map:
                self.coordinate_node_index_map[node_id] = []
            self.coordinate_node_index_map[node_id].append(coordinate_index[0])
        local_deformations = coordinates.local_direction()
        ordered_nodes = self.geometry.node(nodes)
        coordinate_systems = self.geometry.coordinate_system(ordered_nodes.disp_cs)
        points = global_coord(self.geometry.coordinate_system(
            ordered_nodes.def_cs), ordered_nodes.coordinate)
        self.global_deflections = global_deflection(coordinate_systems, local_deformations, points)

        self.compute_displacements()
        self.show_comment()
        self.render()
        self.update_shape()
        self.update_shape_mode(0)
        QApplication.processEvents()

    def compute_displacements(self, compute_scale=True) -> np.ndarray:
        """
        Computes displacements to apply to the geometry

        Parameters
        ----------
        compute_scale : bool, optional
            Renormalize the displacement scaling. The default is True.

        """
        shape_displacements = self.shapes[self.current_shape].shape_matrix[..., self.indices]
        node_displacements = np.zeros(self.geometry.node.shape + (3,),
                                      dtype=shape_displacements.dtype)
        for node_index, node in self.geometry.node.ndenumerate():
            try:
                node_indices = self.coordinate_node_index_map[node.id]
            except KeyError:
                print('Node {:} not found in shape array'.format(node.id))
                node_displacements[node_index] = 0
                continue
            node_deflection_scales = shape_displacements[node_indices]
            node_deflection_directions = self.global_deflections[node_indices]
            node_displacements[node_index] += np.sum(node_deflection_scales[:, np.newaxis] *
                                                     node_deflection_directions, axis=0)
        self.node_displacements = node_displacements.reshape(-1, 3)
        # Compute maximum displacement
        if compute_scale:
            max_displacement = np.max(np.linalg.norm(self.node_displacements, axis=-1))
            global_coords = self.geometry.global_node_coordinate()
            bbox_diagonal = np.linalg.norm(
                np.max(global_coords, axis=0) - np.min(global_coords, axis=0))
            self.displacement_scale = 0.05 * bbox_diagonal / \
                max_displacement  # Set to 5% of bounding box dimension

    def update_shape_mode(self, phase=None):
        """
        Updates the mode that is being plotted

        Parameters
        ----------
        phase : float, optional
            Sets the current phase of the shape. The default is None, which
            computes the new phase based on the time elapsed.

        """
        if phase is None:
            now = time.time()
            time_elapsed = now - self.last_time
            phase_change = 2 * np.pi / self.animation_speed * time_elapsed
            self.last_time = now
            self.phase += phase_change
        else:
            self.phase = phase
        if self.phase > 2 * np.pi:
            self.phase -= 2 * np.pi
            if not self.loop_action.isChecked():
                self.stop_animation()
        # print('Updating Phase: {:}'.format(self.phase))
        if self.complex_action.isChecked():
            # print('Complex')
            deformation = np.real(np.exp(self.phase * 1j) * self.node_displacements)
        elif self.real_action.isChecked():
            # print('Real')
            deformation = np.cos(self.phase) * np.real(self.node_displacements)
        elif self.imag_action.isChecked():
            # print('Imag')
            deformation = np.cos(self.phase) * np.imag(self.node_displacements)
        elif self.signed_amplitude_action.isChecked():
            # print('Signed Amplitude')
            imag = np.imag(self.node_displacements).flatten()[:, np.newaxis]
            real = np.real(self.node_displacements).flatten()[:, np.newaxis]
            best_fit = np.linalg.lstsq(real, imag)[0][0, 0]
            best_fit_phase = np.arctan(best_fit)
            rectified_displacements = self.node_displacements / np.exp(1j * best_fit_phase)
            deformation = np.cos(self.phase) * np.real(rectified_displacements)
        for mesh_deformed, mesh_undeformed in (
                (self.face_mesh_deformed, self.face_mesh_undeformed),
                (self.point_mesh_deformed, self.point_mesh_undeformed),
                (self.solid_mesh_deformed, self.solid_mesh_undeformed)
        ):
            if mesh_deformed is not None:
                mesh_deformed.points = (mesh_undeformed.points +
                                        deformation * self.displacement_scale * self.displacement_scale_modification
                                        )
        self.render()

    def update_shape(self):
        """
        Updates the shape that is being plotted
        """
        self.update_shape_mode()

    def save_animation_from_action(self):
        filename, file_filter = QFileDialog.getSaveFileName(
            self, 'Select File to Save Animation', filter='GIF (*.gif)')
        if filename == '':
            return
        self.save_animation(filename)

    def save_animation(self, filename=None, frames=20, frame_rate=20,
                       individual_images=False):
        """
        Saves the current shape animation to a file

        Parameters
        ----------
        filename : str, optional
            Path to the file being saved. The default is None, which returns
            the images as an array rather than saves them to a file
        frames : int, optional
            Number of frames in the animation. The default is 20.
        frame_rate : float, optional
            Number of frames per second if the animation is saved to a GIF.
            The default is 20.
        individual_images : bool, optional
            Boolean to specify whether the images are saved as individual PNG
            files or a single GIF. The default is False.

        Returns
        -------
        imgs : np.ndarray or None
            Returns array of images if filename is None, otherwise returns None

        """
        self.stop_animation()
        imgs = []
        phases = np.linspace(0, 2 * np.pi, frames, endpoint=False)
        for phase in phases:
            self.update_shape_mode(phase)
            # Force GUI Update to ensure the OS doesn't delay events
            QApplication.processEvents()
            imgs.append(self.screenshot())
        if filename is None:
            return imgs
        else:
            if individual_images:
                for i, img in enumerate(imgs):
                    fname = os.path.splitext(filename)
                    fname = fname[0] + '-{:d}'.format(i) + fname[1]
                    Image.fromarray(img).save(fname)
            else:
                imgs = [Image.fromarray(img).convert(
                    'P', palette=Image.ADAPTIVE, colors=256) for img in imgs]
                imgs[0].save(fp=filename, format='GIF', append_images=imgs[1:],
                             save_all=True, duration=int(1 / frame_rate * 1000), loop=0)

    def save_animation_all_shapes(self, filename_base='Shape_{:}.gif',
                                  frames=20, frame_rate=20,
                                  individual_images=False):
        """
        Helper function to easily save animations for all shapes

        Parameters
        ----------
        filename_base : str, optional
            Filename that will be passed to the format function to produce the
            actual file name for each shape. The default is 'Shape_{:}.gif'.
        frames : int, optional
            Number of frames in the animation. The default is 20.
        frame_rate : float, optional
            Number of frames per second if the animation is saved to a GIF.
            The default is 20.
        individual_images : bool, optional
            Boolean to specify whether the images are saved as individual PNG
            files or a single GIF. The default is False.

        """
        for shape in range(self.shapes.size):
            self.current_shape = shape
            self.compute_displacements()
            self.show_comment()
            self.save_animation(filename_base.format(shape + 1),
                                frames, frame_rate, individual_images)

    def add_menu_bar(self) -> None:
        """
        Adds the menu bar to the BackgroundPlotter

        """
        super().add_menu_bar()

        # Get the file menu
        file_menu = [child for child in self.main_menu.children()
                     if child.__class__.__name__ == 'QMenu'
                     and child.title() == 'File'][0]
        self.save_animation_action = pvqt.plotting.QAction('Save Animation')
        self.save_animation_action.triggered.connect(self.save_animation_from_action)
        file_menu.insertAction(file_menu.actions()[1], self.save_animation_action)

        shape_menu = self.main_menu.addMenu("Shape")
        complex_display_menu = shape_menu.addMenu("Complex Display")
        self.signed_amplitude_action = pvqt.plotting.QAction('Signed Amplitude', checkable=True)
        self.signed_amplitude_action.triggered.connect(self.select_complex)
        complex_display_menu.addAction(self.signed_amplitude_action)
        self.complex_action = pvqt.plotting.QAction('Complex', checkable=True)
        self.complex_action.triggered.connect(self.select_complex)
        complex_display_menu.addAction(self.complex_action)
        self.real_action = pvqt.plotting.QAction('Real', checkable=True)
        self.real_action.triggered.connect(self.select_complex)
        complex_display_menu.addAction(self.real_action)
        self.imag_action = pvqt.plotting.QAction('Imaginary', checkable=True)
        self.imag_action.triggered.connect(self.select_complex)
        complex_display_menu.addAction(self.imag_action)
        scaling_menu = shape_menu.addMenu('Scaling')
        scaling_menu.addAction('1/4x', self.select_scaling_0p25)
        scaling_menu.addAction('1/2x', self.select_scaling_0p5)
        self.scale_0p8_action = QAction('0.8x')
        self.scale_0p8_action.triggered.connect(self.select_scaling_0p8)
        self.scale_0p8_action.setShortcut(QKeySequence(','))
        scaling_menu.addAction(self.scale_0p8_action)
        self.scale_reset_action = QAction('Reset')
        self.scale_reset_action.triggered.connect(self.select_scaling_1)
        self.scale_reset_action.setShortcut(QKeySequence('/'))
        scaling_menu.addAction(self.scale_reset_action)
        self.scale_1p25_action = QAction('1.25x')
        self.scale_1p25_action.triggered.connect(self.select_scaling_1p25)
        self.scale_1p25_action.setShortcut(QKeySequence('.'))
        scaling_menu.addAction(self.scale_1p25_action)
        scaling_menu.addAction('2x', self.select_scaling_2p0)
        scaling_menu.addAction('4x', self.select_scaling_4p0)
        speed_menu = shape_menu.addMenu('Animation Speed')
        self.speed_0p8_action = QAction('0.8x')
        self.speed_0p8_action.triggered.connect(self.select_speed_0p8)
        self.speed_0p8_action.setShortcut(QKeySequence(';'))
        speed_menu.addAction(self.speed_0p8_action)
        speed_menu.addAction('Reset', self.select_speed_1)
        self.speed_1p25_action = QAction('1.25x')
        self.speed_1p25_action.triggered.connect(self.select_speed_1p25)
        self.speed_1p25_action.setShortcut(QKeySequence("'"))
        speed_menu.addAction(self.speed_1p25_action)
        shape_menu.addSeparator()
        self.plot_comments_action = pvqt.plotting.QAction('Show Comments', checkable=True)
        self.plot_comments_action.triggered.connect(self.show_comment)
        shape_menu.addAction(self.plot_comments_action)
        self.loop_action = pvqt.plotting.QAction('Loop Animation', checkable=True)
        self.loop_action.triggered.connect(self.select_loop)
        shape_menu.addAction(self.loop_action)

        display_menu = self.main_menu.addMenu('Display')
        self.display_undeformed = pvqt.plotting.QAction('Show Undeformed', checkable=True)
        self.display_undeformed.triggered.connect(self.toggle_undeformed)
        display_menu.addAction(self.display_undeformed)

    def add_toolbars(self) -> None:
        """
        Adds toolbars to the BackgroundPlotter

        """
        super().add_toolbars()
        self.shape_select_toolbar = self.app_window.addToolBar(
            "Shape Selector"
        )

        self._add_action(
            self.shape_select_toolbar, '<<', self.prev_shape
        )
        self._add_action(
            self.shape_select_toolbar, '?', self.select_shape
        )
        self._add_action(
            self.shape_select_toolbar, '>>', self.next_shape
        )

        self.animation_control_toolbar = self.app_window.addToolBar(
            "Animation Control"
        )

        self._add_action(
            self.animation_control_toolbar, 'Play', self.play_animation
        )
        self._add_action(
            self.animation_control_toolbar, 'Stop', self.stop_animation
        )

    def play_animation(self):
        """Starts the animation playing"""
        self.timer = pvqt.plotting.QTimer()
        self.timer.timeout.connect(self.update_shape)
        self.timer.start(10)
        self.phase = 0.0
        self.last_time = time.time()

    def stop_animation(self):
        """Stops the animation from playing"""
        if self.timer is not None:
            self.timer.stop()
            self.timer = None

    def toggle_undeformed(self) -> None:
        pass

    def select_scaling_0p25(self):
        """
        Adjusts the current shape scaling by 0.25x
        """
        self.displacement_scale_modification *= 0.25
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_0p5(self):
        """
        Adjusts the current shape scaling by 0.5x
        """
        self.displacement_scale_modification *= 0.5
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_0p8(self):
        """
        Adjusts the current shape scaling by 0.8x
        """
        self.displacement_scale_modification *= 0.8
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_1(self):
        """
        Resets shape scaling to 1
        """
        self.displacement_scale_modification = 1.0
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_1p25(self):
        """
        Adjusts the current shape scaling by 1.25x
        """
        self.displacement_scale_modification *= 1.25
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_2p0(self):
        """
        Adjusts the current shape scaling by 2x
        """
        self.displacement_scale_modification *= 2.0
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_scaling_4p0(self):
        """
        Adjusts the current shape scaling by 4x
        """
        self.displacement_scale_modification *= 4.0
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_speed_1(self):
        """
        Resets animation speed to default quantities
        """
        self.animation_speed = 1.0

    def select_speed_0p8(self):
        """
        Adjusts the current animation speed by 0.8
        """
        self.animation_speed /= 0.8

    def select_speed_1p25(self):
        """
        Adjusts the current animation speed by 1.25
        """
        self.animation_speed /= 1.25

    def select_complex(self) -> None:
        """
        Adjusts how complex shapes are plotted
        """
        complex_actions = [
            self.signed_amplitude_action,
            self.complex_action,
            self.real_action,
            self.imag_action
        ]
        sender = self.sender()
        sender_index = complex_actions.index(sender)
        complex_actions.pop(sender_index)
        if not sender.isChecked():
            sender.setChecked(True)
        else:
            for action in complex_actions:
                action.setChecked(False)
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()

    def select_loop(self) -> None:
        pass

    def next_shape(self) -> None:
        """
        Increases the index of the shape being plotted
        """
        self.current_shape += 1
        if self.current_shape >= self.shapes.size:
            self.current_shape = 0
        self.compute_displacements()
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()
        self.show_comment()

    def prev_shape(self) -> None:
        """
        Decreases the index of the shape being plotted
        """
        self.current_shape -= 1
        if self.current_shape < 0:
            self.current_shape = self.shapes.size - 1
        self.compute_displacements()
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()
        self.show_comment()

    def reset_shape(self) -> None:
        """
        Resets a shape if the value of the current shape is changed
        """
        self.compute_displacements()
        self.phase = 0.0
        self.last_time = time.time()
        self.update_shape()
        self.show_comment()

    def select_shape(self) -> None:
        from .sdynpy_shape import ShapeCommentTable
        self.shape_comment_table = ShapeCommentTable(self.shapes, self)

    def show_comment(self):
        """
        Shows or hides the shape comment
        """
        if self.plot_comments_action.isChecked():
            self.text.SetText(2, '\n'.join([
                '{:} {:}'.format(self.shape_name, self.current_shape + 1),
                '  Frequency: {:0.2f}'.format(self.shapes[self.current_shape].frequency)] +
                (['  Damping: {:.2f}%'.format(self.shapes[self.current_shape].damping * 100)] if self.show_damping else []) +
                ['  ' + self.shapes[self.current_shape].comment1,
                 '  ' + self.shapes[self.current_shape].comment2,
                 '  ' + self.shapes[self.current_shape].comment3,
                 '  ' + self.shapes[self.current_shape].comment4,
                 '  ' + self.shapes[self.current_shape].comment5]))
        else:
            self.text.SetText(2, '')

    def _close(self):
        self.stop_animation()
        super(ShapePlotter, self)._close()


class DeflectionShapePlotter(ShapePlotter):
    """Class used to plot animated deflection shapes from spectra"""

    def __init__(self, geometry, deflection_shape_data, plot_kwargs,
                 background_plotter_kwargs={'auto_update': 0.01},
                 undeformed_opacity=0.25, starting_scale=1.0,
                 deformed_opacity=1.0, num_curves=50):
        """
        Create a DeflectionShapePlotter object to plot shapes

        Parameters
        ----------
        geometry : Geometry
            Geometry on which the shapes will be plotted
        deflection_shape_data : NDDataArray
            Data array containing the deflection shapes to plot
        plot_kwargs : dict
            Keyword arguments passed to the Geometry.plot function
        background_plotter_kwargs : dict, optional
            Keyword arguments passed to the BackgroundPlotter constructor.
            The default is {'auto_update':0.01}.
        undeformed_opacity : float, optional
            Opacity of the undeformed geometry.  Set to zero for no undeformed
            geometry. The default is 0.25.
        starting_scale : float, optional
            Starting scale of the shapes on the plot. The default is 1.0.
        num_curves : int, optional
            Maximum number of curves to plot on the frequency selector.
            Default is 50
        """
        # Convert function data into shapes
        from .sdynpy_shape import shape_array
        flattened_data = deflection_shape_data.flatten()
        response_coordinates = flattened_data.response_coordinate
        if not np.unique(response_coordinates).size == response_coordinates.size:
            raise ValueError('all response coordinates for operating data must be unique')
        shapes = shape_array(response_coordinates, flattened_data.ordinate.T,
                             flattened_data[0].abscissa)
        super().__init__(geometry, shapes, plot_kwargs, background_plotter_kwargs,
                         undeformed_opacity, starting_scale, deformed_opacity, shape_name='Shape',
                         show_damping=False)

        # Get the layout so we can add a new widget to it
        layout = self.frame.layout()
        self.frequency_selector_plot = pg.PlotWidget()
        layout.addWidget(self.frequency_selector_plot)
        sp = self.frequency_selector_plot.sizePolicy()
        sp.setVerticalPolicy(QSizePolicy.Minimum)
        self.frequency_selector_plot.setSizePolicy(sp)
        self.frequency_selector_plot.sizeHint = lambda: QtCore.QSize(1006, 120)

        # Now we need to plot the data in the plot
        if num_curves < flattened_data.size:
            indices = np.linspace(0, flattened_data.size - 1, num_curves).astype(int)
        else:
            indices = np.arange(flattened_data.size)
        for data in flattened_data[indices]:
            self.frequency_selector_plot.plot(data.abscissa, np.abs(data.ordinate))
        # Plot the envelope
        max_envelope = np.linalg.svd(
            flattened_data.ordinate.T[..., np.newaxis], False, False, False).squeeze()
        self.frequency_selector_plot.plot(
            flattened_data[0].abscissa, max_envelope, pen=pg.mkPen(color='k', width=2))

        # Add an infinite line for a cursor
        self.frequency_selector = pg.InfiniteLine(
            0, movable=True, bounds=[flattened_data.abscissa.min(),
                                     flattened_data.abscissa.max()])
        self.frequency_selector_plot.addItem(self.frequency_selector)
        self.frequency_selector.sigPositionChanged.connect(self.modify_abscissa)

        self.frequency_selector_plot.setLogMode(False, True)
        self.frequency_selector_plot.setRange(
            yRange=(np.log10(max_envelope.min()), np.log10(max_envelope.max())))

    def modify_abscissa(self):
        frequency = self.frequency_selector.value()
        self.current_shape = np.argmin(np.abs(self.shapes.frequency - frequency))
        self.compute_displacements()
        self.update_shape()
        self.show_comment()

    def save_multiline_animation(self, filename=None, frame_rate=20,
                                 phase_change_per_frame=0.087,
                                 frequency_change_per_frame=0.5,
                                 start_phase=0.0,
                                 start_frequency=None, end_frequency=None,
                                 individual_images=False):
        self.stop_animation()
        imgs = []
        if start_frequency is None:
            start_frequency = self.shapes.frequency[0]
        if end_frequency is None:
            end_frequency = self.shapes.frequency[-1]
        frequency_frames = np.arange(start_frequency, end_frequency, frequency_change_per_frame)
        phase_frames = start_phase + (np.arange(frequency_frames.size) * phase_change_per_frame)
        for frequency, phase in zip(frequency_frames, phase_frames):
            self.current_shape = np.argmin(np.abs(self.shapes.frequency - frequency))
            self.compute_displacements()
            self.update_shape_mode(phase)
            self.show_comment()
            # Force GUI Update to ensure the OS doesn't delay events
            QApplication.processEvents()
            imgs.append(self.screenshot())
        if filename is None:
            return imgs
        else:
            if individual_images:
                for i, img in enumerate(imgs):
                    fname = os.path.splitext(filename)
                    fname = fname[0] + '-{:d}'.format(i) + fname[1]
                    Image.fromarray(img).save(fname)
            else:
                imgs = [Image.fromarray(img).convert(
                    'P', palette=Image.ADAPTIVE, colors=256) for img in imgs]
                imgs[0].save(fp=filename, format='GIF', append_images=imgs[1:],
                             save_all=True, duration=int(1 / frame_rate * 1000), loop=0)


def split_list(seq, value):
    """
    Splits a list by a value that exists in the list

    Parameters
    ----------
    seq : iterable
        The sequence to split up
    value :
        Value within the list to use to split up the list

    Yields
    ------
    group : list
        Sublist of seq separated into groups.

    """
    group = []
    for val in seq:
        if val != value:
            group.append(val)
        elif group:
            yield group
            group = []
    yield group


def global_coord(coordinate_system, local_coord):
    '''Compute global coordinates from local coordinates

    Compute the global coordinates of a node from its local coordinates in the
    corresponding coordinate system.

    Parameters
    ----------

    coordinate_system : CoordinateSystemArray
        An array of coordinate systems of shape [...]
    local_coord : ndarray
        An array of coordinates of shape [...,3]

    Returns
    ------
    global_coord : ndarray
        An array of coordinates of shape [...,3]
    '''
    # Make sure we don't change local_coord
    local_coord = np.array(copy.deepcopy(local_coord))

    # First convert to cartesian if necessary
    cylindrical_css = coordinate_system.cs_type == 1
    cyl_coord = local_coord[cylindrical_css]
    cyl_coord[..., 0], cyl_coord[..., 1] = (cyl_coord[..., 0] * np.cos(cyl_coord[..., 1] * np.pi / 180),
                                            cyl_coord[..., 0] * np.sin(cyl_coord[..., 1] * np.pi / 180))
    local_coord[cylindrical_css] = cyl_coord

    spherical_css = coordinate_system.cs_type == 2
    sph_coord = local_coord[spherical_css]
    sph_coord[..., 0], sph_coord[..., 1], sph_coord[..., 2] = (
        sph_coord[..., 0] * np.sin(sph_coord[..., 1] * np.pi / 180) *
        np.cos(sph_coord[..., 2] * np.pi / 180),
        sph_coord[..., 0] * np.sin(sph_coord[..., 1] * np.pi / 180) *
        np.sin(sph_coord[..., 2] * np.pi / 180),
        sph_coord[..., 0] * np.cos(sph_coord[..., 1] * np.pi / 180))
    local_coord[spherical_css] = sph_coord
    # Convert to global cartesian from local cartesian
    return np.einsum('...ij,...i->...j', coordinate_system.matrix[..., :3, :], local_coord) + coordinate_system.matrix[..., 3, :]


def local_coord(coordinate_system, global_coord):
    """Compute local coordinates from global coordinates

    Compute the local coordinates of a node in the
    corresponding coordinate system from its global coordinates.

    Parameters
    ----------

    coordinate_system : CoordinateSystemArray
        An array of coordinate systems of shape [...]
    global_coord : ndarray
        An array of coordinates of shape [...,3]

    Returns
    ------
    local_coord : ndarray
        An array of coordinates of shape [...,3]
    """
    # Make sure we don't change global_coord
    global_coord = np.array(copy.deepcopy(global_coord))
    # Make sure we're working with a numpy array
    # Convert to the local cartesian
    local_coordinate = np.einsum(
        '...ji,...i->...j', coordinate_system.matrix[..., :3, :], global_coord - coordinate_system.matrix[..., 3, :])

    # Convert to cylindrical or spherical if necessary.
    cylindrical_css = coordinate_system.cs_type == 1
    cyl_coord = local_coordinate[cylindrical_css]
    cyl_coord[..., 0], cyl_coord[..., 1] = (np.sqrt(cyl_coord[..., 0]**2 + cyl_coord[..., 1]**2),
                                            np.arctan2(cyl_coord[..., 1], cyl_coord[..., 0]) * 180 / np.pi)
    local_coordinate[cylindrical_css] = cyl_coord

    spherical_css = coordinate_system.cs_type == 2
    sph_coord = local_coordinate[spherical_css]
    sph_coord[..., 0], sph_coord[..., 1] = (
        np.sqrt(sph_coord[..., 0]**2 + sph_coord[..., 1]**2 + sph_coord[..., 2]**2),
        np.arctan2(sph_coord[..., 1], sph_coord[..., 0]) * 180 / np.pi)
    # At this point local_coordinate[2,:] is still z and local_coordinate[0,:] is
    # now r instead of x
    sph_coord[..., 2] = np.arccos(sph_coord[..., 2] / sph_coord[..., 0]) * 180 / np.pi
    # Flip the 1 and 2 rows to be consistent with IMAT
    sph_coord = sph_coord[..., [0, 2, 1]]
    local_coordinate[spherical_css] = sph_coord
    return local_coordinate


def global_deflection(coordinate_system, local_deflection, global_point=None):
    """
    Compute deflections in the global coordinate system

    Parameters
    ----------
    coordinate_system : CoordinateSystemArray
        An array of coordinate systems of shape [...]
    local_deflection : ndarray
        An array of deflections of shape [...,3]
    global_point : np.ndarray, optional
        An array of coordinates in the global coordinate system of shape
        [...,3]. Must be specified if there are spherical or cylindrical
        coordinate systems.  The default is None.

    Raises
    ------
    ValueError
        If cylindrical or spherical coordinate systems exist and global_point
        is not specified

    Returns
    -------
    global_deflection : ndarray
        An array of coordinates of shape [...,3] specifying the deflection
        directions

    """
    # Make sure we don't change local_deflection or point
    local_deflection = np.array(copy.deepcopy(local_deflection))
    if global_point is not None:
        global_point = np.array(copy.deepcopy(global_point))
    # Convert cylindrical and spherical to a local cartesian AT THE POINT
    # OF INTEREST (i.e. rotated in the coordinate system)
    local_rotation = np.zeros(coordinate_system.shape + (3, 3))
    for key, cs in coordinate_system.ndenumerate():
        if cs.cs_type == 0:
            this_local_rotation = np.eye(3)
        elif cs.cs_type == 1:
            if global_point is None:
                raise ValueError('point must be specified for Cylindrical coordinate systems')
            local_point = local_coord(cs, global_point[key])
            this_local_rotation = R(2, local_point[1], degrees=True)
        elif cs.cs_type == 2:
            if global_point is None:
                raise ValueError('point must be specified for Spherical coordinate systems')
            local_point = local_coord(cs, global_point[key])
            this_local_rotation = R(
                2, local_point[2], degrees=True) @ R(1, local_point[1], degrees=True)
        else:
            raise ValueError('invalid coordinate system type {:}'.format(cs.cs_type))
        local_rotation[key] = this_local_rotation
    return np.einsum('...ij,...j->...i', np.einsum('...ij,...ik->...jk', coordinate_system.matrix[..., :3, :], local_rotation), local_deflection)


class CoordinateSystemArray(SdynpyArray):
    """Coordinate System Information specifying position and directions.

    Use the coordinate_system_array helper function to create the array.
        """
    data_dtype = [('id', 'uint64'), ('name', '<U40'), ('color', 'uint16'),
                  ('cs_type', 'uint16'), ('matrix', 'float64', (4, 3))]

    def __new__(subtype, shape, buffer=None, offset=0,
                strides=None, order=None):
        # Create the ndarray instance of our type, given the usual
        # ndarray input arguments.  This will call the standard
        # ndarray constructor, but return an object of our type.
        # It also triggers a call to InfoArray.__array_finalize__
        obj = super(CoordinateSystemArray, subtype).__new__(subtype, shape,
                                                            CoordinateSystemArray.data_dtype, buffer, offset, strides, order)
        # Finally, we must return the newly created object:
        return obj

    def __repr__(self):
        string_out = '{:>8s}, {:>6s}, {:>20s}, {:>5s}, {:>10s}\n'.format(
            'Index', 'ID', 'Name', 'Color', 'Type')
        if self.size == 0:
            string_out += '----------- Empty -------------\n'
        for i, (key, val) in enumerate(self.ndenumerate()):
            if i >= MAX_NUMBER_REPR:
                string_out += '  .\n  .\n  .\n'
                break
            string_out += '{:>8}, {:>6}, {:>20}, {:>5}, {:>10}\n'.format(
                str(key), val.id, val.name, val.color, _cs_type_strings[val.cs_type[()]])
        return string_out

    def __call__(self, index_by_id):
        """
        Select coordinate system by id rather than by index

        Parameters
        ----------
        index_by_id : int or np.ndarray
            ID number(s) to use to select coordinate systems.

        Raises
        ------
        ValueError
            If specified ID(s) not found in array.

        Returns
        -------
        output : CoordinateSystemArray
            Subset of CoordinateSystemArray with the specified IDs.

        """
        index_dict = {node.id[()]: index for index, node in self.ndenumerate()}
        ids = np.array(index_by_id)
        output = np.empty(ids.shape, dtype=self.dtype).view(self.__class__)
        for key, val in np.ndenumerate(ids):
            try:
                index = index_dict[val]
            except KeyError:
                raise ValueError('ID {:} not found in array'.format(val))
            output[key] = self[index]
        return output

    @staticmethod
    def from_unv(unv_data_dict, combine=True):
        """
        Load CoordinateSystemArrays from universal file data from read_unv

        Parameters
        ----------
        unv_data_dict : dict
            Dictionary containing data from read_unv
        combine : bool, optional
            If True, return as a single CoordinateSytemArray

        Returns
        -------
        output_arrays : CoordinateSystemArray
            Coordinate Systems read from unv

        """
        try:
            datasets = unv_data_dict[2420]
        except KeyError:
            if combine:
                return CoordinateSystemArray(0)
            else:
                return []
        output_arrays = []
        for dataset in datasets:
            output_arrays.append(coordinate_system_array(dataset.cs_labels,
                                                         dataset.cs_names,
                                                         dataset.cs_colors,
                                                         dataset.cs_types,
                                                         dataset.cs_matrices))
        if combine:
            output_arrays = np.concatenate(output_arrays)
        return output_arrays


def coordinate_system_array(id=1, name='', color=1, cs_type=0, matrix=np.concatenate((np.eye(3), np.zeros((1, 3))), axis=0),
                            structured_array=None):
    """
    Creates an array that specifies coordinate systems in the geometry

    Creates an array of coordinate systems that specify directions of sensors
    in a test or analysis.  Coordinate system arrays can be created using a
    numpy structured array or individual arrays for each attribute.
    Multidimensional arrays can be used.

    Parameters
    ----------
    id : ndarray
        Integer array corresponding to the id of the coordinate systems. Input
        will be cast to an integer (i.e. 2.0 -> 2, 1.9 -> 1)
    name : ndarray
        name of the coordinate systems
    color : ndarray
        color of the coordinate systems as integers
    cs_type : ndarray
        Coordinate system types (0 = Cartesian, 1 = Polar, 2 = Spherical)
    matrix : ndarray
        Coordinate system transformation matrix with shape [... x 4 x 3]
    structured_array : ndarray (structured)
        Alternatively to the individual attributes, a single numpy structured
        array can be passed, which should have the same name as the inputs to
        the function listed above.

    Returns
    -------
    coordinate_system_array : CoordinateSystemArray
    """
    if structured_array is not None:
        try:
            id = structured_array['id']
            name = structured_array['name']
            color = structured_array['color']
            cs_type = structured_array['cs_type']
            matrix = structured_array['matrix']
        except (ValueError, TypeError):
            raise ValueError(
                'structured_array must be numpy.ndarray with dtype names "id", "name", "color", "cs_type" and "matrix"')
    else:
        id = np.array(id)
        name = np.array(name)
        color = np.array(color)
        cs_type = np.array(cs_type)
        matrix = np.array(matrix)

    # Don't check shapes because we want them to be broadcastable, but id have to
    # be unique so we will use that for the shape
    coord_sys_array = CoordinateSystemArray(id.shape)
    coord_sys_array.id = id
    coord_sys_array.name = name
    coord_sys_array.color = color
    coord_sys_array.cs_type = cs_type
    coord_sys_array.matrix = matrix

    return coord_sys_array


class ElementArray(SdynpyArray):
    """Element information array

    Use the element_array helper function to create the array.
        """

    data_dtype = [('id', 'uint64'), ('type', 'uint8'), ('color', 'uint16'),
                  ('connectivity', 'object')]

    def __new__(subtype, shape, buffer=None, offset=0,
                strides=None, order=None):
        # Create the ndarray instance of our type, given the usual
        # ndarray input arguments.  This will call the standard
        # ndarray constructor, but return an object of our type.
        # It also triggers a call to InfoArray.__array_finalize__

        obj = super(ElementArray, subtype).__new__(subtype, shape,
                                                   ElementArray.data_dtype, buffer, offset, strides, order)
        # Finally, we must return the newly created object:
        return obj

    def __repr__(self):
        string_out = '{:>8s}, {:>6s}, {:>4s}, {:>5s}, {:>7s}\n'.format(
            'Index', 'ID', 'Type', 'Color', '# Nodes')
        if self.size == 0:
            string_out += '----------- Empty -------------\n'
        for i, (key, val) in enumerate(self.ndenumerate()):
            if i >= MAX_NUMBER_REPR:
                string_out += '  .\n  .\n  .\n'
                break
            string_out += '{:>8}, {:>6}, {:>4}, {:>5}, {:>7}\n'.format(
                str(key), val.id, val.type, val.color, len(val.connectivity))
        return string_out

    def __call__(self, index_by_id):
        """
        Select elements by id rather than by index

        Parameters
        ----------
        index_by_id : int or np.ndarray
            ID number(s) to use to select elements.

        Raises
        ------
        ValueError
            If specified ID(s) not found in array.

        Returns
        -------
        output : ElementArray
            Subset of ElementArray with the specified IDs.

        """
        index_dict = {node.id[()]: index for index, node in self.ndenumerate()}
        ids = np.array(index_by_id)
        output = np.empty(ids.shape, dtype=self.dtype).view(self.__class__)
        for key, val in np.ndenumerate(ids):
            try:
                index = index_dict[val]
            except KeyError:
                raise ValueError('ID {:} not found in array'.format(val))
            output[key] = self[index]
        return output

    def reduce(self, node_list):
        """
        Keep only elements that have all nodes contained in node_list

        Parameters
        ----------
        node_list : iterable
            Iterable containing nodes to keep.

        Returns
        -------
        ElementArray
            ElementArray containing only elements with all nodes in node_list.

        """
        output_list = []
        for key, element in self.ndenumerate():
            if np.all(np.in1d(element.connectivity, node_list)):
                output_list.append(element)
        return np.array(output_list, self.dtype).view(ElementArray)


    def remove(self, element_list):
        """
        Removes elements with id numbers in traceline_list
    
        Parameters
        ----------
        traceline_list : interable
            Iterable containing tracelines to discard.
    
        Returns
        -------
        TracelineArray
            TracelineArray containing tracelines with id numbers not it
            traceline_list.
    
        """
        output_list = []
        for key,element in self.ndenumerate():
            if element.id not in element_list:
                output_list.append(element)
        return np.array(output_list, self.dtype).view(ElementArray)

    @staticmethod
    def from_unv(unv_data_dict, combine=True):
        """
        Load ElementArrays from universal file data from read_unv

        Parameters
        ----------
        unv_data_dict : dict
            Dictionary containing data from read_unv
        combine : bool, optional
            If True, return as a single ElementArray

        Returns
        -------
        output_arrays : ElementArray
            Elements read from unv

        """
        try:
            datasets = unv_data_dict[2412]
        except KeyError:
            if combine:
                return ElementArray(0)
            else:
                return []
        output_arrays = []
        for dataset in datasets:
            output_arrays.append(element_array(dataset.element_labels,
                                               dataset.fe_descriptor_ids,
                                               dataset.colors,
                                               dataset.connectivities))
        if combine:
            output_arrays = np.concatenate(output_arrays)
        return output_arrays


def element_array(id=1, type=None, color=1,
                  connectivity=None, structured_array=None):
    """
    Creates an array that specifies elements in the geometry

    Creates an array of elements that specify connectivity of sensors
    in a test or analysis.  Element arrays can be created using a
    numpy structured array or individual arrays for each attribute.
    Multidimensional arrays can be used.

    Parameters
    ----------
    id : ndarray
        Integer array corresponding to the id of the node. Input
        will be cast to an integer (i.e. 2.0 -> 2, 1.9 -> 1)
    type : ndarray
        Element types.  See notes for type mapping
    color : ndarray
        color of the elements as integers
    connectivity : ndarray
        An object array of iterables defining the connectivity for each element.
    structured_array : ndarray (structured)
        Alternatively to the individual attributes, a single numpy structured
        array can be passed, which should have the same name as the inputs to
        the function listed above.


    Returns
    -------
    element_array : ElementArray


    Notes
    -----
    Here is a list of element types

       11 : 'Rod',
       21 : 'Linear beam',
       22 : 'Tapered beam',
       23 : 'Curved beam',
       24 : 'Parabolic beam',
       31 : 'Straight pipe',
       32 : 'Curved pipe',
       41 : 'Plane Stress Linear Triangle',
       42 : 'Plane Stress Parabolic Triangle',
       43 : 'Plane Stress Cubic Triangle',
       44 : 'Plane Stress Linear Quadrilateral',
       45 : 'Plane Stress Parabolic Quadrilateral',
       46 : 'Plane Strain Cubic Quadrilateral',
       51 : 'Plane Strain Linear Triangle',
       52 : 'Plane Strain Parabolic Triangle',
       53 : 'Plane Strain Cubic Triangle',
       54 : 'Plane Strain Linear Quadrilateral',
       55 : 'Plane Strain Parabolic Quadrilateral',
       56 : 'Plane Strain Cubic Quadrilateral',
       61 : 'Plate Linear Triangle',
       62 : 'Plate Parabolic Triangle',
       63 : 'Plate Cubic Triangle',
       64 : 'Plate Linear Quadrilateral',
       65 : 'Plate Parabolic Quadrilateral',
       66 : 'Plate Cubic Quadrilateral',
       71 : 'Membrane Linear Quadrilateral',
       72 : 'Membrane Parabolic Triangle',
       73 : 'Membrane Cubic Triangle',
       74 : 'Membrane Linear Triangle',
       75 : 'Membrane Parabolic Quadrilateral',
       76 : 'Membrane Cubic Quadrilateral',
       81 : 'Axisymetric Solid Linear Triangle',
       82 : 'Axisymetric Solid Parabolic Triangle',
       84 : 'Axisymetric Solid Linear Quadrilateral',
       85 : 'Axisymetric Solid Parabolic Quadrilateral',
       91 : 'Thin Shell Linear Triangle',
       92 : 'Thin Shell Parabolic Triangle',
       93 : 'Thin Shell Cubic Triangle',
       94 : 'Thin Shell Linear Quadrilateral',
       95 : 'Thin Shell Parabolic Quadrilateral',
       96 : 'Thin Shell Cubic Quadrilateral',
       101: 'Thick Shell Linear Wedge',
       102: 'Thick Shell Parabolic Wedge',
       103: 'Thick Shell Cubic Wedge',
       104: 'Thick Shell Linear Brick',
       105: 'Thick Shell Parabolic Brick',
       106: 'Thick Shell Cubic Brick',
       111: 'Solid Linear Tetrahedron',
       112: 'Solid Linear Wedge',
       113: 'Solid Parabolic Wedge',
       114: 'Solid Cubic Wedge',
       115: 'Solid Linear Brick',
       116: 'Solid Parabolic Brick',
       117: 'Solid Cubic Brick',
       118: 'Solid Parabolic Tetrahedron',
       121: 'Rigid Bar',
       122: 'Rigid Element',
       136: 'Node To Node Translational Spring',
       137: 'Node To Node Rotational Spring',
       138: 'Node To Ground Translational Spring',
       139: 'Node To Ground Rotational Spring',
       141: 'Node To Node Damper',
       142: 'Node To Gound Damper',
       151: 'Node To Node Gap',
       152: 'Node To Ground Gap',
       161: 'Lumped Mass',
       171: 'Axisymetric Linear Shell',
       172: 'Axisymetric Parabolic Shell',
       181: 'Constraint',
       191: 'Plastic Cold Runner',
       192: 'Plastic Hot Runner',
       193: 'Plastic Water Line',
       194: 'Plastic Fountain',
       195: 'Plastic Baffle',
       196: 'Plastic Rod Heater',
       201: 'Linear node-to-node interface',
       202: 'Linear edge-to-edge interface',
       203: 'Parabolic edge-to-edge interface',
       204: 'Linear face-to-face interface',
       208: 'Parabolic face-to-face interface',
       212: 'Linear axisymmetric interface',
       213: 'Parabolic axisymmetric interface',
       221: 'Linear rigid surface',
       222: 'Parabolic rigid surface',
       231: 'Axisymetric linear rigid surface',
       232: 'Axisymentric parabolic rigid surface'
    """
    if structured_array is not None:
        try:
            id = structured_array['id']
            type = structured_array['type']
            color = structured_array['color']
            connectivity = structured_array['connectivity']
        except (ValueError, TypeError):
            raise ValueError(
                'structured_array must be numpy.ndarray with dtype names "id", "type", "color", and "connectivity"')
    else:
        id = np.array(id)
        type = np.array(type)
        color = np.array(color)
        # Need to make connectivity an object array so don't do this!
        # connectivity = np.array(connectivity,dtype=object)

    # Don't check shapes because we want them to be broadcastable, but id have to
    # be unique so we will use that for the shape
    earray = ElementArray(id.shape)
    earray.id = id
    earray.type = type
    earray.color = color
    if earray.ndim == 0:
        # This is gross, but it works!
        np.ndarray.__getitem__(earray, 'connectivity')[()] = np.array(connectivity, dtype='uint32')
    else:
        connectivity = np.array(connectivity, dtype=object)
        for key, val in np.ndenumerate(id):
            earray.connectivity[key] = np.array(connectivity[key], dtype='uint32')
    return earray


class NodeArray(SdynpyArray):
    """Node information array

    Use the node_array helper function to create the array.
        """

    data_dtype = [('id', 'uint64'), ('coordinate', 'float64', (3,)), ('color', 'uint16'),
                  ('def_cs', 'uint64'), ('disp_cs', 'uint64')]

    def __new__(subtype, shape, buffer=None, offset=0,
                strides=None, order=None):
        # Create the ndarray instance of our type, given the usual
        # ndarray input arguments.  This will call the standard
        # ndarray constructor, but return an object of our type.
        # It also triggers a call to InfoArray.__array_finalize__

        obj = super(NodeArray, subtype).__new__(subtype, shape,
                                                NodeArray.data_dtype, buffer, offset, strides, order)
        # Finally, we must return the newly created object:
        return obj

    def __repr__(self):
        string_out = '{:>8s}, {:>6s}, {:>8s}, {:>8s}, {:>8s}, {:>5s}, {:>5s}\n'.format(
            'Index', 'ID', 'X', 'Y', 'Z', 'DefCS', 'DisCS')
        if self.size == 0:
            string_out += '----------- Empty -------------\n'
        for i, (key, val) in enumerate(self.ndenumerate()):
            if i >= MAX_NUMBER_REPR:
                string_out += '  .\n  .\n  .\n'
                break
            string_out += '{:>8}, {:>6}, {:>8.3f}, {:>8.3f}, {:>8.3f}, {:>5}, {:>5}\n'.format(
                str(key), val.id, *val.coordinate, val.def_cs, val.disp_cs)
        return string_out

    def __call__(self, index_by_id):
        """
        Select nodes by id rather than by index

        Parameters
        ----------
        index_by_id : int or np.ndarray
            ID number(s) to use to select nodes.

        Raises
        ------
        ValueError
            If specified ID(s) not found in array.

        Returns
        -------
        output : NodeArray
            Subset of NodeArray with the specified IDs.

        """
        index_dict = {node.id[()]: index for index, node in self.ndenumerate()}
        ids = np.array(index_by_id)
        output = np.empty(ids.shape, dtype=self.dtype).view(self.__class__)
        for key, val in np.ndenumerate(ids):
            try:
                index = index_dict[val]
            except KeyError:
                raise ValueError('ID {:} not found in array'.format(val))
            output[key] = self[index]
        return output

    def reduce(self, node_list):
        """
        Keep only nodes that are contained in node_list

        Parameters
        ----------
        node_list : iterable
            Iterable containing nodes to keep.

        Returns
        -------
        NodeArray
            NodeArray containing only nodes in node_list.

        """
        return self(node_list)

    @staticmethod
    def from_unv(unv_data_dict, combine=True):
        """
        Load NodeArrays from universal file data from read_unv

        Parameters
        ----------
        unv_data_dict : dict
            Dictionary containing data from read_unv
        combine : bool, optional
            If True, return as a single NodeArray

        Returns
        -------
        output_arrays : NodeArray
            Nodes read from unv

        """
        try:
            datasets = unv_data_dict[2411]
        except KeyError:
            if combine:
                return NodeArray(0)
            else:
                return []
        output_arrays = []
        for dataset in datasets:
            output_arrays.append(node_array(dataset.node_labels,
                                            dataset.coordinates,
                                            dataset.colors,
                                            dataset.export_coordinate_systems,
                                            dataset.displacement_coordinate_systems))
        if combine:
            output_arrays = np.concatenate(output_arrays)
        return output_arrays

    def by_position(self, position_array):
        """
        Select node by closest position

        Parameters
        ----------
        position_array : np.ndarray
            A (...,3) shape array containing positions of nodes to keep

        Returns
        -------
        NodeArray
            NodArray containing nodes that were closest to the positions in
            position_array.

        """
        node_differences = np.linalg.norm(
            self.coordinate - position_array[(Ellipsis,) + (np.newaxis,) * self.ndim + (slice(None),)], axis=-1)
        node_differences = node_differences.reshape(*position_array.shape[:-1], -1)
        node_indices = np.argmin(
            node_differences,
            axis=-1)
        return self.flatten()[node_indices]

    @staticmethod
    def project_to_minimum_plane(coordinates, return_3D=True):
        """
        Projects coordinates to a single best-fit plane

        Parameters
        ----------
        coordinates : np.ndarray
            A (...,3) coordinate array.
        return_3D : bool, optional
            If True, return the 3D coordinates of the projected points.
            Otherwise return the projected 2D coordinate. The default is True.

        Returns
        -------
        np.ndarray
            Points projected to a best-fit plane.

        """
        coords = coordinates.T.copy()
        mean = np.mean(coords, axis=-1, keepdims=True)
        coords -= mean
        U, S, V = np.linalg.svd(coords)
        coords = U.T @ coords
        coords[2, :] = 0
        if return_3D:
            return (U @ coords + mean).T
        else:
            return coords[:2].T

    def global_coordinate(self, cs_array):
        """
        Get the global coordinate of each node

        Parameters
        ----------
        cs_array : CoordinateSystemArray
            CoordinateSystemArray consisting of the local coordinate systems for
            each node

        Returns
        -------
        points : np.ndarray
            Coordinates of the nodes in the global coordinate system.

        """
        points = global_coord(cs_array(self.def_cs), self.coordinate)
        return points

    def triangulate(self, cs_array, projection_function=None,
                    return_element_array=True, element_color=1,
                    condition_threshold=None):
        """
        Creates elements for a node set using Delaunay Triangulation

        Parameters
        ----------
        cs_array : CoordinateSystemArray
            CoordinateSystemArray containing coordinate systems for each node.
        projection_function : function, optional
            Function to use to project 3D coordinates to 2D coordinates for
            triangulation. The default is None.
        return_element_array : bool, optional
            Returns an ElementArray if True, otherwise it simply returns the
            triangle simplices. The default is True.
        element_color : np.ndarray or int, optional
            Integers representing colors applied to the elements.
            The default is 1.
        condition_threshold : float, optional
            Condition number threshold used to remove triangles that are poorly
            shaped. The default is None.

        Returns
        -------
        ElementArray or np.ndarray
            ElementArray containing elements or np.ndarray containing triangle
            simplices.

        """
        if projection_function is None:
            def projection_function(coords): return self.project_to_minimum_plane(coords, False)
        # Compute global positions of the nodes
        global_positions = self.global_coordinate(cs_array)
        # Now compute the projection function
        projected_positions = projection_function(global_positions)
        # Now triangulate
        tri = Delaunay(projected_positions)
        simplices = tri.simplices
        if condition_threshold is not None:
            # Now remove small triangles via condition number
            tri_points = projected_positions[tri.simplices, :]
            tri_points -= np.mean(tri_points, axis=1, keepdims=True)
            simplices = simplices[np.linalg.cond(tri_points) < condition_threshold]
        if not return_element_array:
            return simplices
        else:
            # Create an element array
            connectivity = self.id[simplices]
            return element_array(id=np.arange(connectivity.shape[0]) + 1,
                                 type=61, color=element_color,
                                 connectivity=connectivity)

    def by_grid(self, grid_spacing, x_min=None, y_min=None, z_min=None,
                x_max=None, y_max=None, z_max=None):
        """
        Selects nodes in a grid

        Parameters
        ----------
        grid_spacing : float
            Approximate grid spacing between selected nodes
        x_min : float, optional
            Minimum X-value of the grid.
            The default is the minimum node x-coordinate.
        y_min : float, optional
            Minimun Y-value of the grid.
            The default is the minimum node y-coordinate.
        z_min : float, optional
            Minimum Z-value of the grid.
            The default is the minimum node z-coordinate.
        x_max : float, optional
            Maximum X-value of the grid.
            The default is the maximum node x-coordinate.
        y_max : float, optional
            Maximum Y-value of the grid.
            The default is the maximum node y-coordinate.
        z_max : float, optional
            Maximum Z-value of the grid.
            The default is the maximum node z-coordinate.

        Returns
        -------
        NodeArray
            NodeArray containing the closest nodes to the grid points
        """
        if x_min is None:
            x_min = self.coordinate[:, 0].min()
        if y_min is None:
            y_min = self.coordinate[:, 1].min()
        if z_min is None:
            z_min = self.coordinate[:, 2].min()
        if x_max is None:
            x_max = self.coordinate[:, 0].max()
        if y_max is None:
            y_max = self.coordinate[:, 1].max()
        if z_max is None:
            z_max = self.coordinate[:, 2].max()
        min_coords = np.array((x_min, y_min, z_min))
        max_coords = np.array((x_max, y_max, z_max))
        range_coords = max_coords - min_coords
        num_grids = np.ceil(range_coords / grid_spacing).astype(int)
        # Now we will create the grid in each dimension using linspace
        grid_arrays = [np.linspace(min_coord, max_coord, num_coord)
                       for min_coord, max_coord, num_coord
                       in zip(min_coords, max_coords, num_grids)]
        # We will then use meshgrid to assemble the array grid.  We will flatten the
        # point dimension and transpose so the array has shape (n_points x 3)
        grid_points = np.array(np.meshgrid(*grid_arrays, indexing='ij')).reshape(3, -1).T
        # Now we will select nodes that are closest to the points in the grid
        candidate_nodes = self.by_position(grid_points)
        # Reduce to only nodes that are within one grid spacing of their requested point
        # this reduces the nodes that are selected by points far from the model
        candidate_nodes = candidate_nodes[
            np.linalg.norm(candidate_nodes.coordinate - grid_points, axis=-1) < grid_spacing]
        # Remove any duplicates
        candidate_node_ids = np.unique(candidate_nodes.id)
        return candidate_nodes(candidate_node_ids)


def node_array(id=1, coordinate=np.array((0, 0, 0)), color=1, def_cs=1, disp_cs=1,
               structured_array=None):
    """
    Creates an array that specifies nodes in the geometry

    Creates an array of nodes that specify positions of sensors
    in a test or analysis.  Node arrays can be created using a
    numpy structured array or individual arrays for each attribute.
    Multidimensional arrays can be used.

    Parameters
    ----------
    id : ndarray
        Integer array corresponding to the id of the node. Input
        will be cast to an integer (i.e. 2.0 -> 2, 1.9 -> 1)
    coordinate : ndarray
        Positions of the nodes in space
    color : ndarray
        color of the nodes as integers
    def_cs : ndarray
        Coordinate system where the node's position is defined
    disp_cs : ndarray
        Coordinate system where the node's displacements are defined
    structured_array : ndarray (structured)
        Alternatively to the individual attributes, a single numpy structured
        array can be passed, which should have the same name as the inputs to
        the function listed above.


    Returns
    -------
    node_array : NodeArray
    """
    if structured_array is not None:
        try:
            id = structured_array['id']
            coordinate = structured_array['coordinate']
            color = structured_array['color']
            def_cs = structured_array['def_cs']
            disp_cs = structured_array['disp_cs']
        except (ValueError, TypeError):
            raise ValueError(
                'structured_array must be numpy.ndarray with dtype names "id", "coordinate", "color", "def_cs" and "disp_cs"')
    else:
        id = np.array(id)
        coordinate = np.array(coordinate)
        color = np.array(color)
        def_cs = np.array(def_cs)
        disp_cs = np.array(disp_cs)

    # Don't check shapes because we want them to be broadcastable, but id have to
    # be unique so we will use that for the shape
    narray = NodeArray(id.shape)
    narray.id = id
    narray.coordinate = coordinate
    narray.color = color
    narray.def_cs = def_cs
    narray.disp_cs = disp_cs

    return narray


class TracelineArray(SdynpyArray):
    """Traceline information array

    Use the traceline_array helper function to create the array.
        """
    data_dtype = [('id', 'uint64'), ('color', 'uint16'), ('description', '<U40'),
                  ('connectivity', 'object')]

    def __new__(subtype, shape, buffer=None, offset=0,
                strides=None, order=None):
        # Create the ndarray instance of our type, given the usual
        # ndarray input arguments.  This will call the standard
        # ndarray constructor, but return an object of our type.
        # It also triggers a call to InfoArray.__array_finalize__

        obj = super(TracelineArray, subtype).__new__(subtype, shape,
                                                     TracelineArray.data_dtype, buffer, offset, strides, order)
        # Finally, we must return the newly created object:
        return obj

    def __repr__(self):
        string_out = '{:>8s}, {:>6s}, {:>20s}, {:>5s}, {:>7s}\n'.format(
            'Index', 'ID', 'Description', 'Color', '# Nodes')
        if self.size == 0:
            string_out += '----------- Empty -------------\n'
        for i, (key, val) in enumerate(self.ndenumerate()):
            if i >= MAX_NUMBER_REPR:
                string_out += '  .\n  .\n  .\n'
                break
            string_out += '{:>8s}, {:>6d}, {:>20s}, {:>5d}, {:>7d}\n'.format(
                str(key), val.id, val.description, val.color, len(val.connectivity))
        return string_out

    def __call__(self, index_by_id):
        """
        Select nodes by id rather than by index

        Parameters
        ----------
        index_by_id : int or np.ndarray
            ID number(s) to use to select nodes.

        Raises
        ------
        ValueError
            If specified ID(s) not found in array.

        Returns
        -------
        output : TracelineArray
            Subset of TracelineArray with the specified IDs.

        """
        index_dict = {node.id[()]: index for index, node in self.ndenumerate()}
        ids = np.array(index_by_id)
        output = np.empty(ids.shape, dtype=self.dtype).view(self.__class__)
        for key, val in np.ndenumerate(ids):
            try:
                index = index_dict[val]
            except KeyError:
                raise ValueError('ID {:} not found in array'.format(val))
            output[key] = self[index]
        return output

    def reduce(self, node_list):
        """
        Keep only tracelines fully contain nodes in node_list

        Parameters
        ----------
        node_list : iterable
            Iterable containing nodes to keep.

        Returns
        -------
        TracelineArray
            TracelineArray containing lines that contain only nodes in node_list.

        """
        output_list = []
        for key, traceline in self.ndenumerate():
            if np.all(np.in1d(traceline.connectivity, node_list) | (traceline.connectivity == 0)):
                output_list.append(traceline)
        return np.array(output_list, self.dtype).view(TracelineArray)

    def remove(self, traceline_list):
        """
        Removes tracelines with id numbers in traceline_list

        Parameters
        ----------
        traceline_list : interable
            Iterable containing tracelines to discard.

        Returns
        -------
        TracelineArray
            TracelineArray containing tracelines with id numbers not it
            traceline_list.

        """
        output_list = []
        for key,traceline in self.ndenumerate():
            if traceline.id not in traceline_list:
                output_list.append(traceline)
        return np.array(output_list, self.dtype).view(TracelineArray)
                

    @staticmethod
    def from_unv(unv_data_dict, combine=True):
        """
        Load TracelineArrays from universal file data from read_unv

        Parameters
        ----------
        unv_data_dict : dict
            Dictionary containing data from read_unv
        combine : bool, optional
            If True, return as a single TracelineArray object

        Returns
        -------
        output_arrays : TracelineArray
            Tracelines read from unv

        """
        try:
            datasets = unv_data_dict[82]
        except KeyError:
            if combine:
                return TracelineArray(0)
            else:
                return []
        output_arrays = []
        for dataset in datasets:
            output_arrays.append(traceline_array(dataset.traceline_number,
                                                 dataset.identification,
                                                 dataset.color,
                                                 dataset.nodes)[np.newaxis])
        if combine:
            output_arrays = np.concatenate(output_arrays)
        return output_arrays


def traceline_array(id=1, description='', color=1,
                    connectivity=None, structured_array=None):
    """
    Creates an array that specifies tracelines in the geometry

    Creates an array of tracelines that specify connectivity of sensors
    in a test or analysis.  Traceline arrays can be created using a
    numpy structured array or individual arrays for each attribute.
    Multidimensional arrays can be used.

    Parameters
    ----------
    id : ndarray
        Integer array corresponding to the id of the node. Input
        will be cast to an integer (i.e. 2.0 -> 2, 1.9 -> 1)
    Description : ndarray
        Traceline name
    color : ndarray
        color of the elements as integers
    connectivity : ndarray
        An object array of iterables defining the connectivity for each traceline.
    structured_array : ndarray (structured)
        Alternatively to the individual attributes, a single numpy structured
        array can be passed, which should have the same name as the inputs to
        the function listed above.


    Returns
    -------
    traceline_array : TracelineArray

    """
    if structured_array is not None:
        try:
            id = structured_array['id']
            description = structured_array['description']
            color = structured_array['color']
            connectivity = structured_array['connectivity']
        except (ValueError, TypeError):
            raise ValueError(
                'structured_array must be numpy.ndarray with dtype names "id", "color", "description", and "connectivity"')
    else:
        id = np.array(id)
        description = np.array(description)
        color = np.array(color)
        # Don't do this, need to keep as an object dtype
        # connectivity = np.array(connectivity)

    # Don't check shapes because we want them to be broadcastable, but id have to
    # be unique so we will use that for the shape
    tlarray = TracelineArray(id.shape)
    tlarray.id = id
    tlarray.description = description
    tlarray.color = color
    if tlarray.ndim == 0:
        np.ndarray.__getitem__(tlarray, 'connectivity')[()] = np.array(connectivity, dtype='uint32')
    else:
        connectivity = np.array(connectivity, dtype=object)
        for key, val in np.ndenumerate(id):
            tlarray.connectivity[key] = np.array(connectivity[key], dtype='uint32')

    return tlarray


from ..fem.sdynpy_exodus import Exodus, ExodusInMemory  # noqa: E402
from .sdynpy_shape import shape_array  # noqa: E402


class Geometry:
    """Container for nodes, coordinate systems, tracelines, and elements

    Geometry is the class that is most useful for working with the positioning
    and spatial visualization of test or analysis data.  It contains functions
    to plot and animate 3D geometry"""

    def __init__(self, node=NodeArray((0,)), coordinate_system=CoordinateSystemArray((0,)),
                 traceline=TracelineArray((0,)), element=ElementArray((0,))):
        """
        Initialize a geometry object with nodes, coordinate systems, tracelines,
        and elements.

        All input arguments will be flattened when passed to the Geometry, as
        the geometry does not support multi-dimensional object arrays.

        Parameters
        ----------
        node : NodeArray, optional
            The set of nodes in the geometry. The default is NodeArray((0,)),
            an empty NodeArray.
        coordinate_system : CoordinateSystemArray, optional
            The set of coordinate systems in the geometry. The default is
            CoordinateSystemArray((0,)), an empty CoordinateSystemArray
        traceline : TracelineArray, optional
            The set of tracelines defined in the geometry. The default is
            TracelineArray((0,)), an empty TracelineArray
        element : ElementArray, optional
            The set of elements defined in the geometry. The default is
            ElementArray((0,)), an empty element array.

        Returns
        -------
        None.

        """
        self.node = node.flatten()
        self.coordinate_system = coordinate_system.flatten()
        self.traceline = traceline.flatten()
        self.element = element.flatten()

    def __repr__(self):
        string_out = '\n'.join(val.capitalize(
        ) + '\n' + repr(self.__dict__[val]) for val in ['node', 'coordinate_system', 'traceline', 'element'])
        return string_out

    def reduce(self, node_list):
        """
        Reduce the geometry to only contain nodes in `node_list`

        Elements and tracelines will only be kept if all nodes in each element
        or traceline are in `node_list`.  Coordinate systems will only be kept
        if they are required by a node in `node_list`

        Parameters
        ----------
        node_list : iterable
            An iterable of integer node id numbers.

        Returns
        -------
        Geometry
            A geometry only containing the nodes in `node_list`

        """
        node = self.node.reduce(node_list).copy()
        traceline = self.traceline.reduce(node_list).copy()
        element = self.element.reduce(node_list).copy()
        css = np.unique(np.concatenate((np.unique(node.def_cs), node.disp_cs)))
        coordinate_system = self.coordinate_system(css).copy()
        return Geometry(node, coordinate_system, traceline, element)

    def add_traceline(self, connectivity, id=None, description='', color=1):
        """
        Adds a traceline to the geometry

        Parameters
        ----------
        connectivity : iterable
            Iterable containing the node_ids to connect with the traceline
        id : TYPE, optional
            The id number of the new traceline. The default is None, which
            results in the new traceline having an id of one greater than the
            current maximum traceline it.
        description : str, optional
            A string description of the new traceline. The default is ''.
        color : int, optional
            An integer corresponding to the color of the new traceline.
            The default is 1.

        Returns
        -------
        None.  Modifications are made in-place to the current geometry.

        """
        if id is None:
            if self.traceline.size > 0:
                id = np.max(self.traceline.id) + 1
            else:
                id = 1
        new_traceline = traceline_array((id,), description, color,
                                        [np.array(connectivity)])
        self.traceline = np.concatenate((self.traceline, new_traceline), axis=0)

    def add_element(self, elem_type, connectivity, id=None, color=1):
        if id is None:
            if self.element.size > 0:
                id = np.max(self.element.id) + 1
            else:
                id = 1
        new_element = element_array((id,), elem_type, color, [np.array(connectivity)])
        self.element = np.concatenate((self.element, new_element))

    @classmethod
    def from_unv(cls, unv_dict):
        """
        Create a geometry from universal file format data from readunv

        Parameters
        ----------
        unv_dict : dict
            Dictionary containing data from read_unv

        Returns
        -------
        Geometry
            Geometry object read from the unv data

        """
        try:
            coordinate_systems = CoordinateSystemArray.from_unv(unv_dict, combine=True)
        except KeyError:
            coordinate_systems = CoordinateSystemArray((0,))
        # Nodes
        try:
            nodes = NodeArray.from_unv(unv_dict, combine=True)
            cs_indices = coordinate_systems.id
            for node_index, node in nodes.ndenumerate():
                cs_index = np.where(cs_indices == node.def_cs)
                matching_cs = coordinate_systems[cs_index]
                if matching_cs.size == 0:
                    print('Warning: Coordinate System {:} not found in universal file structure.  Adding Global Cartesian with ID {:}'.format(
                        node.def_cs, node.def_cs))
                    coordinate_systems = np.concatenate(
                        (coordinate_systems, coordinate_system_array((node.def_cs,))))
                    cs_indices = coordinate_systems.id
            #         cs_index = np.where(cs_indices == node.def_cs)
            #     node.coordinate = local_coord(coordinate_systems[cs_index][0],node.coordinate)
        except KeyError:
            nodes = NodeArray((0,))
        try:
            elements = ElementArray.from_unv(unv_dict, combine=True)
        except KeyError:
            elements = ElementArray((0,))
        try:
            tracelines = TracelineArray.from_unv(unv_dict, combine=True)
        except KeyError:
            tracelines = TracelineArray((0,))

        return cls(nodes, coordinate_systems, tracelines, elements)

    from_uff = from_unv

    @staticmethod
    def write_excel_template(path_to_xlsx):
        """
        Writes an Excel File Template for Creating Geometry

        Parameters
        ----------
        path_to_xlsx : string
            Path to write xlsx Excel file

        Returns
        -------
        Nothing

        Notes
        -----
        See documentation for from_excel_template for instructions on filling
        out the template to create a geometry.
        """
        
        # Create a new workbook
        workbook = openpyxl.Workbook()
        
        # Add multiple sheets to the workbook
        sheet_names = ['Coordinate Systems', 'Nodes', 'Elements', 'Trace Lines']
        for sheet_name in sheet_names:
            workbook.create_sheet(sheet_name)
        
        # Remove the default sheet created at the start
        del workbook['Sheet']
        
        # Define column names for each sheet
        columns = {
            'Coordinate Systems': ['ID', 'Name', 'Color', 'Type', 'X Location', 'Y Location', 'Z Location', 'Rotation 1 Angle', 'Rotation 1 Axis', 'Rotation 2 Angle', 'Rotation 2 Axis', 'Rotation 3 Angle', 'Rotation 3 Axis'],
            'Nodes': ['ID', 'Color', 'X Location', 'Y Location', 'Z Location', 'Displacement CS', 'Definition CS'],
            'Elements': ['ID', 'Color', 'Type', 'Node 1', 'Node 2', 'Node 3', 'Node 4', 'Node 5'],
            'Trace Lines': ['ID', 'Description', 'Color', 'Node 1', 'Node 2', 'Node 3', 'Node 4', 'Node 5']
        }
        
        # Create data validations
        data_validation_ID = openpyxl.worksheet.datavalidation.DataValidation(type="whole", operator='greaterThan', formula1='-1',allow_blank=True,error='Must be interger greater than -1',errorStyle='stop',showErrorMessage=True)
        data_validation_Color = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"Black,Blue,Gray Blue,Light Blue,Cyan,Dark Olive,Dark Green,Green,Yellow,Golden Orange,Orange,Red,Magenta,Light Magenta,Pink,White"',allow_blank=True,error='Must be value from list',errorStyle='stop',showErrorMessage=True)
        data_validation_CSType = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"Cartesian,Cylindrical,Sypherical"',allow_blank=True,error='Must be value from list',errorStyle='stop',showErrorMessage=True)
        data_validation_Axis = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"X,Y,Z"',allow_blank=True,error='Must be value from list',errorStyle='stop',showErrorMessage=True)
        data_validation_ElemType = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"beam,triangle,quadrilateral,tetrahedral"',allow_blank=True,error='Must be value from list',errorStyle='stop',showErrorMessage=True)
        data_validation_Node = openpyxl.worksheet.datavalidation.DataValidation(type="whole", operator='greaterThan', formula1='-1',allow_blank=True,error='Must be value from list',errorStyle='stop',showErrorMessage=True)
        data_validation_Decimal = openpyxl.worksheet.datavalidation.DataValidation(type="decimal",allow_blank=True,error='Must be decimal',errorStyle='stop',showErrorMessage=True)
        
        # Specify Columns for data validation
        table_height = 10
        data_validation_ID_A = copy.deepcopy(data_validation_ID)
        data_validation_ID_A.add('A2:A'+str(table_height+1))
        data_validation_Color_B = copy.deepcopy(data_validation_Color)
        data_validation_Color_B.add('B2:B'+str(table_height+1))
        data_validation_Color_C = copy.deepcopy(data_validation_Color)
        data_validation_Color_C.add('C2:C'+str(table_height+1))
        data_validation_CSType.add('D2:D'+str(table_height+1))
        data_validation_Axis.add('I2:I'+str(table_height+1))
        data_validation_Axis.add('K2:K'+str(table_height+1))
        data_validation_Axis.add('M2:M'+str(table_height+1))
        data_validation_Decimal_Coordinate_System = data_validation_Decimal
        data_validation_Decimal_Coordinate_System.add('E2:H'+str(table_height+1))
        data_validation_Decimal_Coordinate_System.add('J2:J'+str(table_height+1))
        data_validation_Decimal_Coordinate_System.add('L2:L'+str(table_height+1))
        data_validation_Decimal_Nodes = copy.deepcopy(data_validation_Decimal)
        data_validation_Decimal_Nodes.add('C2:E'+str(table_height+1))
        data_validation_ID_Nodes = copy.deepcopy(data_validation_ID)
        data_validation_ID_Nodes.add('F2:G'+str(table_height+1))
        data_validation_ElemType.add('C2:C'+str(table_height+1))
        data_validation_Node.add('D2:H'+str(table_height+1))
        
        # Add data and tables to each sheet
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            
            # Define the column headers
            headers = columns[sheet_name]
            
            # Write the headers into the first row
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)
        
            table_range = f'A1:{chr(65 + len(headers) - 1)}{table_height + 1}'
        
            table = openpyxl.worksheet.table.Table(displayName=sheet_name.replace(' ','_'), ref=table_range)
            
            # Apply table style
            style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            # Add table to the worksheet
            sheet.add_table(table)

        # Apply Data Validations for "Coordinate Systems" Sheet
        sheet = workbook['Coordinate Systems']
        sheet.add_data_validation(data_validation_ID_A)
        sheet.add_data_validation(data_validation_Color_C)
        sheet.add_data_validation(data_validation_CSType)
        sheet.add_data_validation(data_validation_Decimal_Coordinate_System)
        sheet.add_data_validation(data_validation_Axis)

        # Apply Data Validations for "Nodes" Sheet
        sheet = workbook['Nodes']
        sheet.add_data_validation(data_validation_ID_A)
        sheet.add_data_validation(data_validation_ID_Nodes)
        sheet.add_data_validation(data_validation_Color_B)
        sheet.add_data_validation(data_validation_Decimal_Nodes)
        
        # Apply Data Validations for "Elements" Sheet
        sheet = workbook['Elements']
        sheet.add_data_validation(data_validation_ID_A)
        sheet.add_data_validation(data_validation_Color_B)
        sheet.add_data_validation(data_validation_ElemType)
        sheet.add_data_validation(data_validation_Node)
        
        # Apply Data Validations for "Trace Lines" Sheet
        sheet = workbook['Trace Lines']
        sheet.add_data_validation(data_validation_ID_A)
        sheet.add_data_validation(data_validation_Color_C)
        sheet.add_data_validation(data_validation_Node)

        # Save the workbook
        workbook.save(path_to_xlsx)

    @classmethod
    def from_excel_template(cls, path_to_xlsx):
        """
        Create a geometry from Excel file template

        Parameters
        ----------
        path_to_xlsx : string
            Path to xlsx Excel file containing geometry information

        Returns
        -------
        Geometry
            Geometry object created from the Excel file

        Notes
        -----
        To use this function, first save out an excel template file using the
        `write_excel_template` function.  This will construct an excel
        workbook with four worksheets on which the different portions of the
        Geometry are defined.

        On the Coordinate Systems tab, users will define the various global and
        local coordinate systems in their geometry.  Each geometry requires
        an ID number.   A Name can optionally be given.  The Color should be
        selectred from the dropdown list of colors in the Excel template.  The
        type of the coordinate system should be selected from the dropdown list of
        coordinate system types.
        
        The origin of the coordinate system can be specified with the X Location,
        Y Location, Z Location columns.  Then rotations of the coordinate system
        can be specified using rotations about axes.  Up to three axes and angles
        can be specified to create arbitrary compound rotations.  The rotation
        axes should be X, Y, or Z, and the rotation angles are in degrees.

        On the Nodes tab, all tabs must be filled out.  ID numbers must be unique.
        Colors should be selected from the dropdown list in the Excel template.  The
        position of the node is specified using the X Location, Y Location, Z
        Location columns.  Each node has a displacement coordinate system in which
        its position is defined, and a a definition coordinate system in which
        its displacements are defined.  These columns should consist of integers
        corresponding to ID numbers from the Coordinate Systems tab.

        On the Elements tab, elements connecting nodes are defined.  The ID
        column must consist of unique integer identifiers.  The Color tab should
        be selected from the dropdown list.  The type can be selected from the dropdown list.  
        If the type column
        is empty, an element type based on the number of connections given will
        be used. Defult element types for connection length of 2 is "Type 21 - Linear Beam",
        for a connection length of 3 is "Type 41 - Plane Stress Linear Triangle",
        and for a connection length of 4 is "Type 44 - Plane Stress Linear Quadrilateral".
        The columns Node 1 through Node 4 contain the nodes in each element.  Only the required
        number of nodes must be filled out (e.g. a tri element would only contain
        3 nodes, so only columns Node 1, Node 2, and Node 3 would be filled).

        On the Trace Lines tab, lines connecting the nodes are defined.  The
        ID column must consist of unique integer identifiers.  The Description
        column contains a string description of the line.  The Color should be 
        selected from the dropdown list.  The
        Node 1 through Node 5 columns should contain the nodes for each line.
        Only the number of nodes in the line must be filled out, so if a line
        only connects 5 nodes, only Node 1 though Node 5 must be filled. More columns can
        can be added for additional Nodes if longer tracelines are needed.


        """

        #% Read Geometry Information from Excel File
        df_coordinate_systems = pd.read_excel(io = path_to_xlsx,sheet_name='Coordinate Systems')
        df_nodes = pd.read_excel(io = path_to_xlsx,sheet_name='Nodes')
        df_elements = pd.read_excel(io = path_to_xlsx,sheet_name='Elements')
        df_trace_lines = pd.read_excel(io = path_to_xlsx,sheet_name='Trace Lines')

        color_number_dict = {
            'Black':1,
            'Blue':2,
            'Gray Blue':3,
            'Light Blue':4,
            'Cyan':5,
            'Dark Olive':6,
            'Dark Green':7,
            'Green':8,
            'Yellow':9,
            'Golden Orange':10,
            'Orange':11,
            'Red':12,
            'Magenta':13,
            'Light Magenta':14,
            'Pink':15,
            'White':16,
            }

        cs_type_dict = {
            'Cartesian':0,
            'Cylindrical':1,
            'Sypherical':2,
            }

        df_coordinate_systems = df_coordinate_systems.replace({'Color':color_number_dict})
        df_coordinate_systems = df_coordinate_systems.replace({'Type':cs_type_dict})
        df_nodes = df_nodes.replace({'Color':color_number_dict})
        df_elements = df_elements.replace({'Color':color_number_dict})
        df_trace_lines = df_trace_lines.replace({'Color':color_number_dict})

        #% Add Coordinate Systems
        # Start with an empty coordinate system
        coordinate_systems = coordinate_system_array([])
        for row_index, df_coordinate_system in df_coordinate_systems.iterrows():
            # Create Transformation Matrix
            T = np.eye(3)
            for rotation in range(1,(len(df_coordinate_system)-7)//2+1):
                rotation_angle = df_coordinate_system['Rotation ' + str(rotation) + ' Angle']
                rotation_axis = df_coordinate_system['Rotation ' + str(rotation) + ' Axis']
                rotation_axis_dict = {'X':0,'Y':1,'Z':2}
                if pd.notnull(rotation_angle) and pd.notnull(rotation_axis):
                    T = R(rotation_axis_dict[rotation_axis],rotation_angle,degrees=True).T@T

            # Add Coordinate System to Geometry
            origin = np.array((df_coordinate_system['X Location'],df_coordinate_system['Y Location'],df_coordinate_system['Z Location']))[np.newaxis,:]
            coordinate_systems = np.concatenate((coordinate_systems,coordinate_system_array(id=df_coordinate_system['ID'], name=df_coordinate_system['Name'], color=df_coordinate_system['Color']-1, cs_type=df_coordinate_system['Type'], matrix = np.concatenate((T,origin),axis=0))),axis=None)

        #% Add Nodes
        # print(df_nodes['Color'])
        nodes = node_array(
            id=np.array(df_nodes['ID']),
            coordinate=np.concatenate(
                [np.array(df_nodes['X Location'])[:,np.newaxis],np.array(df_nodes['Y Location'])[:,np.newaxis],np.array(df_nodes['Z Location'])[:,np.newaxis]],axis=-1),
            color=np.array(df_nodes['Color'])-1,
            def_cs=np.array(df_nodes['Definition CS']),
            disp_cs=np.array(df_nodes['Displacement CS']))

        #% Add Elements
        elements = element_array([])
        element_connection_length_dict = {
                                        2:21, # 21 : 'Linear Beam'
                                        3:41, # 41 : 'Plane Stress Linear Triangle'
                                        4:44, # 44 : 'Plane Stress Linear Quadrilateral'
                                            }
        for row_index, df_element in df_elements.iterrows():
            connectivity = df_element[3:].values
            connectivity = connectivity[pd.notnull(connectivity)]
            if isinstance(df_element['Type'],str):
                element_type = _exodus_elem_type_map[df_element['Type'].lower()]
            elif isinstance(df_element['Type'],int):
                element_type = df_element['Type']
            elif pd.isnull(df_element['Type']):
                element_type = element_connection_length_dict[len(connectivity)]
            new_element = element_array(id=df_element['ID'], type=element_type, color=df_element['Color']-1,connectivity=connectivity)
            elements = np.concatenate((elements,new_element),axis=None)

        #% Add Tracelines
        tracelines = traceline_array([])
        for row_index, df_trace_line in df_trace_lines.iterrows():
            connectivity=df_trace_line[3:].values
            connectivity = connectivity[pd.notnull(connectivity)]
            new_traceline = traceline_array(connectivity=connectivity,id=df_trace_line['ID'],description=df_trace_line['Description'],color=df_trace_line['Color']-1)
            tracelines = np.concatenate((tracelines,new_traceline),axis=None)

        return cls(nodes, coordinate_systems, tracelines, elements)

    @classmethod
    def from_exodus(cls, exo: Exodus, blocks=None, local=False,
                    preferred_local_orientation=np.array((0.0, 0.0, 1.0)),
                    secondary_preferred_local_orientation=np.array((1.0, 0.0, 0.0)),
                    local_nodes=None):
        """
        Generate a geometry from exodus file data

        Parameters
        ----------
        exo : Exodus or ExodusInMemory
            The exodus data from which geometry will be created.
        blocks : iterable, optional
            Iterable containing a set of block ids to import when creating the
            geometry.  The default is None, which uses all blocks in the model.
        local : bool, optional
            Flag to specify whether or not to create local coordinate systems
            for each node in the model.  This can be useful when creating
            instrumnetation positions from a finite element model, where the
            sensor will be oriented perpendicular to the surface it is mounted
            on.  The default is False, which returns all data in a single
            global coordinate system.
        preferred_local_orientation : np.ndarray, optional
            A preferred direction for the local coordinate system.  The first
            constraint is that the local Z+ axis is perpendicular to the
            surface.  The coordinate system will then try to align itself as
            much as it can with this direction.  The default is
            np.array((0.0,0.0,1.0)), which points the local coordinate systems
            along the global Z+ axis.
        secondary_preferred_local_orientation : np.ndarray, optional
            A secondary preferred direction is only used if the surface normal
            direction is parallel to the primary preferred_local_orientation.
            The default is np.array((1.0,0.0,0.0)), which points the local
            coordinate system along the local Z+ axis.
        local_nodes : np.ndarray, optional
            If specified, only create local coordinate systems at the specified
            nodes.

        Returns
        -------
        Geometry
            A geometry consisting of the finite element nodes and element
            connectivity.

        """
        if isinstance(exo, Exodus):
            exo = exo.load_into_memory(close=False, variables=[], timesteps=[], blocks=blocks)
        # Get nodes
        coordinates = exo.nodes.coordinates
        node_map = np.arange(
            coordinates.shape[0]) + 1 if exo.nodes.node_num_map is None else exo.nodes.node_num_map
        nodes = node_array(node_map, coordinates)
        tl_connectivity = []
        elem_connectivity = []
        tl_color = []
        elem_color = []
        elem_types = []
        if blocks is None:
            blocks = exo.blocks
        else:
            blocks = [block for block in exo.blocks if block.id in blocks]
        for i, block in enumerate(blocks):
            color_index = (i % 14) + 1
            try:
                connectivity = node_map[block.connectivity]
            except AttributeError:
                continue
            elem_type = block.elem_type
            if connectivity.shape[-1] <= 1:
                # Skip conmasses
                continue
            for element in connectivity:
                elem_connectivity.append(element)
                elem_color.append(color_index)
                elem_types.append(_exodus_elem_type_map[elem_type.lower()])
        tls = traceline_array(np.arange(len(tl_connectivity)) + 1,
                              color=tl_color,
                              connectivity=tl_connectivity)
        elems = element_array(np.arange(len(elem_connectivity)) + 1,
                              elem_types, elem_color, elem_connectivity)
        if local:
            if local_nodes is not None:
                node_map_dict = {val: index for index, val in enumerate(node_map)}
                local_node_indices = np.array([node_map_dict[node] for node in local_nodes])
                local_node_index_map = {node_index: index for index, node_index in enumerate(local_node_indices)}
            preferred_local_orientation /= np.linalg.norm(preferred_local_orientation)
            secondary_preferred_local_orientation /= np.linalg.norm(
                secondary_preferred_local_orientation)
            # Go through each node, find the block that it's in, get its elements
            normal_sum = np.zeros(
                (coordinates.shape[0] if local_nodes is None else len(local_nodes),
                 3))
            for block in blocks:
                if block.connectivity.shape[-1] <= 1:
                    # Skip conmasses
                    continue
                full_block_node_positions = coordinates[block.connectivity, :]
                normal_vectors = np.cross(full_block_node_positions[:, 2, :] - full_block_node_positions[:, 0, :],
                                          full_block_node_positions[:, 1, :] - full_block_node_positions[:, 0, :])
                normal_vectors /= np.linalg.norm(normal_vectors, axis=-1, keepdims=True)
                block_nodes = np.unique(block.connectivity)
                if local_nodes is not None:
                    block_nodes = local_node_indices[np.in1d(local_node_indices, block_nodes)]
                for index in block_nodes:
                    node_elements = np.any(block.connectivity == index, axis=-1)
                    node_normals = normal_vectors[node_elements]
                    if local_nodes is None:
                        normal_sum[index] += np.sum(node_normals, axis=0)
                    else:
                        normal_sum[local_node_index_map[index]] += np.sum(node_normals, axis=0)
            # Put small nodes to reasonable values to avoid divide by zero
            normal_sum[np.linalg.norm(normal_sum, axis=-1) < 1e-14] = preferred_local_orientation
            node_rotations = np.zeros((3,) + normal_sum.shape)
            node_rotations[2] = -normal_sum / np.linalg.norm(normal_sum, axis=-1, keepdims=True)
            # Find the nodes that are basically in the direction of the preferred direction
            nodes_in_preferred_direction = np.abs(
                np.einsum('ij,j->i', node_rotations[2], preferred_local_orientation)) > 0.99
            # Compute cross products if we are not parallel with the preferred orientation
            node_rotations[1, ~nodes_in_preferred_direction] = np.cross(
                node_rotations[2, ~nodes_in_preferred_direction], preferred_local_orientation)
            node_rotations[1, nodes_in_preferred_direction] = np.cross(
                node_rotations[2, nodes_in_preferred_direction], secondary_preferred_local_orientation)
            node_rotations[1] /= np.linalg.norm(node_rotations[1], axis=-1, keepdims=True)
            node_rotations[0] = np.cross(node_rotations[1], node_rotations[2])
            node_rotations[0] /= np.linalg.norm(node_rotations[0], axis=-1, keepdims=True)
            node_rotations = np.moveaxis(node_rotations, 0, 1)
            # Create the coordinate system matrix
            cs_matrix = np.concatenate((node_rotations, np.zeros(
                (node_rotations.shape[0], 1, node_rotations.shape[-1]))), axis=1)
            # Put a global coordinate system at the end
            cs_matrix = np.concatenate((cs_matrix, np.eye(4, 3)[np.newaxis]), axis=0)
            if local_nodes is None:
                cs_ids = np.concatenate((node_map, (node_map.max() + 1,)), axis=0)
                cs_descriptions = ['Node {:} Disp'.format(id) for id in node_map] + ['Global']
                css = coordinate_system_array(cs_ids, cs_descriptions, 1, 0, cs_matrix)
                nodes.def_cs = cs_ids[-1]
                nodes.disp_cs = cs_ids[:-1]
            else:
                cs_ids = np.concatenate((local_nodes, (node_map.max() + 1,)), axis=0)
                cs_descriptions = ['Node {:} Disp'.format(id) for id in local_nodes] + ['Global']
                css = coordinate_system_array(cs_ids, cs_descriptions, 1, 0, cs_matrix)
                nodes.def_cs = cs_ids[-1]
                nodes.disp_cs = cs_ids[-1]
                nodes.disp_cs[local_node_indices] = cs_ids[:-1]
        else:
            css = coordinate_system_array((1,))
        return Geometry(nodes, css, tls, elems)


    @classmethod
    def camera_visualization(cls, K, RT, image_size, size=1, colors=1):
        """
        Create a geometry used to visualize a camera specified with K and RT

        Parameters
        ----------
        K : ndarray
            A (...,3,3) matrix containing the intrinsic parameters of the cameras
        RT : ndarray
            A (...,3,4) matrix containing the extrinsic parameters of the cameras
        image_size : ndarray
            A (...,2) matrix containing the width,height of each camera
        size : float
            The distance that the rays will project from the pinhole of the camera
        colors : int or ndarray
            The colors assigned to each camera

        Returns
        -------
        geometry : cls
            A geoemtry containing the cameras that can be used for visualization

        """
        if K.ndim == 2:
            K = K[np.newaxis, ...]
        if K.ndim > 3:
            K = K.reshape(-1, 3, 3)
        if RT.ndim == 2:
            RT = RT[np.newaxis, ...]
        if RT.ndim > 3:
            RT = RT.reshape(-1, 3, 4)
        if image_size.ndim == 1:
            image_size = image_size[np.newaxis, ...]
        if image_size.ndim > 2:
            image_size = image_size.reshape(-1, 2)
        if np.array(colors).ndim == 0:
            colors = colors * np.ones(K.shape[0])
        if np.array(colors).ndim > 1:
            colors = np.array(colors).flatten()

        corners = np.array([(0, 0),
                            (1, 0),
                            (1, 1),
                            (0, 1)])

        geometry = cls(coordinate_system=coordinate_system_array((1,)))

        for i, (thisK, thisRT, this_size, this_color) in enumerate(zip(K, RT, image_size, colors)):
            pixel_corners = corners * this_size

            # Compute the positions
            image_corners = point_on_pixel(thisK, thisRT, pixel_corners, size)
            pinhole = -thisRT[:3, :3].T @ thisRT[..., :3, 3:]
            # Create a geometry
            pinhole_node = node_array(((i + 1) * 10,), pinhole.T, disp_cs=i + 2, color=this_color)
            pixel_nodes = node_array((i + 1) * 10 + np.arange(4) + 1,
                                     image_corners, disp_cs=i + 2, color=this_color)
            cs = coordinate_system_array((i + 2,), color=this_color)
            cs.matrix[..., :3, :3] = thisRT[:3, :3]
            geometry.node = np.concatenate((geometry.node, pinhole_node, pixel_nodes))
            geometry.coordinate_system = np.concatenate((geometry.coordinate_system, cs))
            geometry.add_traceline(np.array([0, 1, 2, 3, 4, 1]) + 10 * (i + 1), color=this_color)
            geometry.add_traceline(np.array([0, 2]) + 10 * (i + 1), color=this_color)
            geometry.add_traceline(np.array([0, 3]) + 10 * (i + 1), color=this_color)
            geometry.add_traceline(np.array([0, 4]) + 10 * (i + 1), color=this_color)
        return geometry

    def compress_ids(self, compress_nodes=True, compress_elements=True,
                     compress_tracelines=True, compress_cs=True,
                     return_maps=False):
        """
        Compresses ID numbers to make node, element, etc. contiguous

        Parameters
        ----------
        compress_nodes : bool, optional
            If True, compress the node ids.  The default is True.
        compress_elements : bool, optional
            If True, compress the element ids. The default is True.
        compress_tracelines : bool, optional
            If True, compress the traceline ids. The default is True.
        compress_cs : bool, optional
            If True, compress the coordinate system ids. The default is True.
        return_maps : bool, optional
            If True, return id maps for nodes, elements, tracelines, and
            coordinate systems.  Maps will be equal to `None` if no compression
            was done on that field. The default is False.

        Returns
        -------
        mapped_geometry : Geometry
            Geometry with contiguous id numbers for the selected fields.
        node_map : id_map or `None`
            Mapping from the old set of nodes to the new set of nodes, only
            returned if `return_maps` is True.  If `compress_nodes` is False,
            this will be `None`.
        traceline_map : id_map or `None`
            Mapping from the old set of tracelines to the new set of tracelines, only
            returned if `return_maps` is True.  If `compress_tracelines` is False,
            this will be `None`.
        element_map : id_map or `None`
            Mapping from the old set of elements to the new set of elements, only
            returned if `return_maps` is True.  If `compress_elements` is False,
            this will be `None`.
        cs_map : id_map or `None`
            Mapping from the old set of coordinate systems to the new set of
            coordinate systems, only
            returned if `return_maps` is True.  If `compress_cs` is False,
            this will be `None`.

        """
        if compress_nodes:
            to_nodes = np.arange(self.node.size)+1
            from_nodes = self.node.id.flatten()
            node_map = id_map(from_nodes, to_nodes)
        else:
            node_map = None
        if compress_cs:
            to_cs = np.arange(self.coordinate_system.size)+1
            from_cs = self.coordinate_system.id.flatten()
            cs_map = id_map(from_cs, to_cs)
        else:
            cs_map = None
        if compress_elements:
            to_elements = np.arange(self.element.size)+1
            from_elements = self.coordinate_system.id.flatten()
            element_map = id_map(from_elements, to_elements)
        else:
            element_map = None
        if compress_tracelines:
            to_tls = np.arange(self.traceline.size)+1
            from_tls = self.coordinate_system.id.flatten()
            traceline_map = id_map(from_tls, to_tls)
        else:
            traceline_map = None
        mapped_geometry = self.map_ids(node_map, traceline_map, element_map, cs_map)
        if return_maps:
            return mapped_geometry, node_map, traceline_map, element_map, cs_map
        else:
            return mapped_geometry

    @classmethod
    def from_imat_struct(cls, imat_fem_struct):
        """
        Constructs a Geometry from an imat_fem class saved to a Matlab structure

        In IMAT, a structure can be created from an `imat_fem` by using the get()
        function.  This can then be saved to a .mat file and loaded using
        `scipy.io.loadmat`.  The output from loadmat can be passed into this function

        Parameters
        ----------
        imat_fem_struct : np.ndarray
            structure from loadmat containing data from an imat_fem

        Returns
        -------
        Geometry
            Geometry constructed from the data in the imat structure

        """
        node_struct = imat_fem_struct['node'][0, 0]
        cs_array = node_struct['cs'][0, 0].reshape(-1, 3)
        nodes = node_array(id=node_struct['id'][0, 0].flatten(),
                           coordinate=node_struct['coord'][0, 0].reshape(-1, 3),
                           color=node_struct['color'][0, 0].flatten(),
                           def_cs=cs_array[:, 0], disp_cs=cs_array[:, 1])
        cs_struct = imat_fem_struct['cs'][0, 0]
        cs_ids = cs_struct['id'][0, 0].flatten()
        if cs_ids.size > 0:
            cs_names = np.concatenate(cs_struct['name'][0, 0].flatten()).flatten()
        else:
            cs_names = []
        css = coordinate_system_array(id=cs_ids,
                                      name=cs_names,
                                      color=cs_struct['color'][0, 0].flatten(),
                                      cs_type=cs_struct['type'][0, 0].flatten(),
                                      matrix=cs_struct['matrix'][0, 0].reshape(4, 3, -1).transpose(2, 0, 1))
        elem_struct = imat_fem_struct['elem'][0, 0]
        elems = element_array(id=elem_struct['id'][0, 0].flatten(),
                              type=elem_struct['type'][0, 0].flatten(),
                              color=elem_struct['color'][0, 0].flatten(),
                              connectivity=[val.flatten() for val in elem_struct['conn'][0, 0].flatten()])
        tl_struct = imat_fem_struct['tl'][0, 0]
        tl_ids = tl_struct['id'][0, 0].flatten()
        if tl_ids.size > 0:
            tl_names = np.concatenate(tl_struct['desc'][0, 0].flatten()).flatten()
        else:
            tl_names = []
        tls = traceline_array(id=tl_ids,
                              description=tl_names,
                              color=tl_struct['color'][0, 0].flatten(),
                              connectivity=[val.flatten() for val in tl_struct['conn'][0, 0].flatten()])
        return cls(nodes, css, tls, elems)

    def global_node_coordinate(self, node_ids=None):
        """
        Position of the Geometry's nodes in the global coordinate system

        Parameters
        ----------
        node_ids : np.ndarray
            An array of node id numbers to keep in the output coordinate.  If
            not specified, all nodes will be returned in the order they are in
            the model.


        Returns
        -------
        np.ndarray
            numpy array with shape (n,3) where n is the number of nodes in the
            geometry.  This array contains the 3d position of each node in the
            global coordinate system.

        """
        if node_ids is None:
            node = self.node
        else:
            node = self.node(node_ids)
        cs_array = self.coordinate_system(node.def_cs)
        return node.global_coordinate(cs_array)

    def node_by_global_position(self, global_position_array):
        """
        Select node by closest position

        Parameters
        ----------
        global_position_array : np.ndarray
            A (...,3) shape array containing positions of nodes to keep

        Returns
        -------
        NodeArray
            NodArray containing nodes that were closest to the positions in
            position_array.

        """
        node_differences = np.linalg.norm(
            self.global_node_coordinate() - global_position_array[(Ellipsis,) + (np.newaxis,) * self.node.ndim + (slice(None),)], axis=-1)
        node_differences = node_differences.reshape(*global_position_array.shape[:-1], -1)
        node_indices = np.argmin(
            node_differences,
            axis=-1)
        return self.node.flatten()[node_indices]

    def node_by_global_grid(self,grid_spacing, x_min = None, y_min = None,
                            z_min = None, x_max = None, y_max = None,
                            z_max = None):
        """
        Selects nodes in a grid

        Parameters
        ----------
        grid_spacing : float
            Approximate grid spacing between selected nodes
        x_min : float, optional
            Minimum X-value of the grid.
            The default is the minimum node x-coordinate.
        y_min : float, optional
            Minimun Y-value of the grid.
            The default is the minimum node y-coordinate.
        z_min : float, optional
            Minimum Z-value of the grid.
            The default is the minimum node z-coordinate.
        x_max : float, optional
            Maximum X-value of the grid.
            The default is the maximum node x-coordinate.
        y_max : float, optional
            Maximum Y-value of the grid.
            The default is the maximum node y-coordinate.
        z_max : float, optional
            Maximum Z-value of the grid.
            The default is the maximum node z-coordinate.

        Returns
        -------
        NodeArray
            NodeArray containing the closest nodes to the grid points
        """
        coordinate = self.global_node_coordinate()
        if x_min is None:
            x_min = coordinate[:, 0].min()
        if y_min is None:
            y_min = coordinate[:, 1].min()
        if z_min is None:
            z_min = coordinate[:, 2].min()
        if x_max is None:
            x_max = coordinate[:, 0].max()
        if y_max is None:
            y_max = coordinate[:, 1].max()
        if z_max is None:
            z_max = coordinate[:, 2].max()
        min_coords = np.array((x_min, y_min, z_min))
        max_coords = np.array((x_max, y_max, z_max))
        range_coords = max_coords - min_coords
        num_grids = np.ceil(range_coords / grid_spacing).astype(int)
        # Now we will create the grid in each dimension using linspace
        grid_arrays = [np.linspace(min_coord, max_coord, num_coord)
                       for min_coord, max_coord, num_coord
                       in zip(min_coords, max_coords, num_grids)]
        # We will then use meshgrid to assemble the array grid.  We will flatten the
        # point dimension and transpose so the array has shape (n_points x 3)
        grid_points = np.array(np.meshgrid(*grid_arrays, indexing='ij')).reshape(3, -1).T
        # Now we will select nodes that are closest to the points in the grid
        candidate_nodes = self.node_by_global_position(grid_points)
        # Reduce to only nodes that are within one grid spacing of their requested point
        # this reduces the nodes that are selected by points far from the model
        candidate_nodes = candidate_nodes[
            np.linalg.norm(self.global_node_coordinate(candidate_nodes.id) - grid_points, axis=-1) < grid_spacing]
        # Remove any duplicates
        candidate_node_ids = np.unique(candidate_nodes.id)
        return candidate_nodes(candidate_node_ids)

    def global_deflection(self, coordinate_array):
        """
        Direction of local deflection in the global coordinate system

        Parameters
        ----------
        coordinate_array : CoordinateArray
            A list of coordinates for which the global deformation direction
            will be computed

        Returns
        -------
        global_deflections : np.ndarray
            numpy array with shape (n,3) where n is the number of coordinates
            in the specified `coordinate_array`.  This array contains the 3d
            direction of motion for each coordinate in the global coordinate
            system.

        """
        local_deformations = coordinate_array.local_direction()
        ordered_nodes = self.node(coordinate_array.node)
        coordinate_systems = self.coordinate_system(ordered_nodes.disp_cs)
        points = global_coord(self.coordinate_system(
            ordered_nodes.def_cs), ordered_nodes.coordinate)
        global_deflections = global_deflection(coordinate_systems, local_deformations, points)
        return global_deflections

    @staticmethod
    def overlay_geometries(geometries, color_override=None, return_node_id_offset=False):
        """
        Combines several geometries, offsetting the id numbers to avoid conflicts

        Parameters
        ----------
        geometries : iterable
            An iterable of geometry objects that will be combined into a single
            geometry
        color_override : iterable, optional
            An iterble of integers specifying colors, which will override the
            existing geometry colors.  This should have the same length as the
            `geometries` input.  The default is None, which keeps the original
            geometry colors.
        return_node_id_offset : bool, optional
            Specifies whether or not the applied node offset should be returned,
            which is useful if the new node numbers are to be mapped to the old
            node numbers.  The default is False, which only returns the combined
            geometry.

        Returns
        -------
        Geometry
            A geometry consisting of a combination of the specified geometries
        node_offset
            An integer specifying the node id offset applied to avoid conflicts

        """
        node_ids = np.concatenate([geometry.node.id for geometry in geometries])
        cs_ids = np.concatenate([geometry.coordinate_system.id for geometry in geometries])
        tl_ids = np.concatenate([geometry.traceline.id for geometry in geometries])
        elem_ids = np.concatenate([geometry.element.id for geometry in geometries])
        try:
            max_node_id = np.max(node_ids)
        except ValueError:
            max_node_id = 1
        try:
            max_cs_id = np.max(cs_ids)
        except ValueError:
            max_cs_id = 1
        try:
            max_tl_id = np.max(tl_ids)
        except ValueError:
            max_tl_id = 1
        try:
            max_elem_id = np.max(elem_ids)
        except ValueError:
            max_elem_id = 1
        node_length = int(np.floor(np.log10(max_node_id)) + 1)
        tl_length = int(np.floor(np.log10(max_tl_id)) + 1)
        elem_length = int(np.floor(np.log10(max_elem_id)) + 1)
        cs_length = int(np.floor(np.log10(max_cs_id)) + 1)
        new_geometries = [geometry.modify_ids(
            node_change=10**(node_length) * (i + 1),
            traceline_change=10**(tl_length) * (i + 1),
            element_change=10**(elem_length) * (i + 1),
            coordinate_system_change=10**(cs_length) * (i + 1))
            for i, geometry in enumerate(geometries)]
        if color_override is not None:
            for i, color in enumerate(color_override):
                new_geometries[i].node.color = color
                new_geometries[i].traceline.color = color
                new_geometries[i].element.color = color
                new_geometries[i].coordinate_system.color = color

        new_geometry = new_geometries[0]
        for i in range(1, len(new_geometries)):
            new_geometry = new_geometry + new_geometries[i]

        if return_node_id_offset:
            return new_geometry, 10**(node_length)
        else:
            return new_geometry
    
    def map_coordinate(self,dof_list,to_geometry,node_map=None,
                       plot_maps = False):
        """
        Map degrees of freedom from one geometry to the closest aligned on a
        second geometry.

        Parameters
        ----------
        dof_list : CoordinateArray
            An array of degrees of freedom on the geometry to map.
        to_geometry : Geometry
            The geometry to which the degrees of freedom will be mapped.
        node_map : None, str, or id_map, optional
            If specified, it will map the node ids from the present geometry to
            the mapping geometry.  If the string 'position' is specified, then
            the closest node by position on the mapping geometry to the present
            geometry will be chosen.  To explicitly map between nodes, pass a 
            sdpy.id_map.  If not specified, it is assumed that the node numbers
            in the present geometry map to the nodes in the mapping geometry.
        plot_maps : bool
            If True, it will plot the CoordinateArray on the geometries to
            visualize the two mapped degrees of freedom.

        Raises
        ------
        ValueError
            If an invalid mapping is specified.

        Returns
        -------
        output_dof_list : CoordinateArray
            A CoordinateArray of the same size and shape of dof_list with each
            entry in dof_list corresponding to the same position in output_dof_list.

        """
        dof_list = dof_list.copy()
        if node_map == 'position':
            from_nodes = np.unique(dof_list.node)
            to_nodes = to_geometry.node_by_global_position(self.global_node_coordinate(from_nodes)).id
            to_geometry = to_geometry.reduce(to_nodes)
            node_map = id_map(from_nodes,to_nodes)
            self = self.reduce(from_nodes).map_ids(node_id_map = node_map)
            dof_list.node = node_map(dof_list.node)
        elif isinstance(node_map,id_map):
            self = self.reduce(node_map.from_ids).map_ids(node_id_map = node_map)
            to_geometry = to_geometry.reduce(node_map.to_nodes)
            dof_list.node = node_map(dof_list.node)
        elif node_map is None:
            self = self.reduce(np.unique(dof_list.node))
            to_geometry = to_geometry.reduce(np.unique(dof_list.node))
        else:
            raise ValueError('Invalid node_map.  Must be an id_map, "position", or None.')

        output_dof_list = dof_list.copy()
        global_directions = self.global_deflection(dof_list)
        to_directions = {id_num:to_geometry.global_deflection(coordinate_array(id_num,[1,2,3]))
                         for id_num in to_geometry.node.id}
        for index,value in dof_list.ndenumerate():
            dirs = to_directions[value.node]
            dots = np.einsum('ij,j',dirs,global_directions[index])
            direction = np.argmax(abs(dots))
            sign = np.sign(dots[direction])
            output_dof_list[index].direction = sign*(direction+1)
            if plot_maps:
                plotter = self.plot_coordinate(dof_list[index])
                to_geometry.plot_coordinate(output_dof_list[index],plot_kwargs={'plotter':plotter})
        
        return output_dof_list
    
    def convert_elements_to_tracelines(self,keep_ids = False, start_id = None,
                                       in_place = False):
        """
        Transforms all elements in the geometry into tracelines

        Parameters
        ----------
        keep_ids : bool, optional
            If True, the element ids will be used as the new traceline ids.
            The default is False.
        start_id : int, optional
            The starting id number for the tracelines converted from elements.
            Only used if keep_ids is False.  If not specified, each subsequent
            element will use the current highest traceline id number + 1.
        in_place : bool, optional
            If True, modify the geometry in place.  If False, a copy will be
            created and returned. The default is False.

        Raises
        ------
        ValueError
            If conflicts are expected to occur between the newly created
            traceline IDs and the current existing traceline IDs.
        NotImplementedError
            If solid element types are converted.

        Returns
        -------
        geometry : Geometry
            Only returned if in_place is False.

        """
        if in_place:
            geometry = self
        else:
            geometry = self.copy()
        if keep_ids:
            if np.any(np.isin(geometry.element.id,geometry.traceline.id)):
                raise ValueError('Keeping element IDs would result in a conflict with existing traceline IDs')
        if start_id is not None:
            if np.any(np.isin(np.arange(geometry.element.size)+start_id,geometry.traceline.id)):
                raise ValueError('Setting new traceline ID starting at {:} would result in a conflict with existing traceline IDs'.format(start_id))
        if not keep_ids and start_id is None:
            if len(geometry.traceline) > 0:
                start_id = np.max(geometry.traceline.id) + 1
            else:
                start_id = 1
        for idx, element in enumerate(geometry.element.flatten()):
            if keep_ids:
                new_id = element.id
            else:
                new_id = start_id + idx
            if element.type <= 96:
                conn = element.connectivity
                conn = np.concatenate((conn,[conn[0]]))
            else:
                raise NotImplementedError('Solid Elements are not implemented yet!')
            geometry.add_traceline(conn,new_id,color=element.color)
        geometry.element = ElementArray((0,))
        if not in_place:
            return geometry
        
    def split_tracelines_into_segments(self, in_place = False):
        """
        Splits long tracelines into many length-2 tracelines

        Parameters
        ----------
        in_place : bool, optional
            If True, the current geometry will be modified in place.  The
            default is False.

        Returns
        -------
        geometry : Geometry
            Only returned if in_place is False.

        """
        temp_geometry = Geometry()
        for traceline in self.traceline:
            connectivity = np.array((traceline.connectivity[:-1],traceline.connectivity[1:])).T
            connectivity = connectivity[np.all(connectivity != 0,axis=-1)]
            for conn in connectivity:
                temp_geometry.add_traceline(conn, traceline.id, traceline.description, traceline.color)
        if not in_place:
            geometry = self.copy()
            geometry.traceline = temp_geometry.traceline
            return geometry
        else:
            self.traceline = temp_geometry.traceline
            
    def remove_duplicate_tracelines(self, in_place = False):
        """
        Removes tracelines that are effectively equivalent to each other.
        
        If two tracelines have the same connectivity (or reversed connectivity),
        then the first will be kept and the second will be discarded.  This will
        not take into account other fields such as description or color.
        
        Parameters
        ----------
        in_place : bool, optional
            If True, the current geometry will be modified in place.  The
            default is False.

        Returns
        -------
        geometry : Geometry
            Only returned if in_place is False.

        """
        conn_set = set([])
        keep_array = []
        for traceline in self.traceline:
            conn = tuple(traceline.connectivity)
            if conn in conn_set or conn[::-1] in conn_set:
                keep_array.append(False)
            else:
                keep_array.append(True)
                conn_set.add(conn)
        if not in_place:
            geometry = self.copy()
            geometry.traceline = geometry.traceline[keep_array]
            return geometry
        else:
            self.traceline = self.traceline[keep_array]

    def plot(self, node_size: int = 5, line_width: int = 1, opacity=1.0, view_up=None, view_from=None, plotter=None,
             show_edges=False, label_nodes = False, label_tracelines = False,
             label_elements = False, label_font_size = 16, 
             plot_individual_items = False):
        """
        Plots the geometry in an interactive 3D window.

        Parameters
        ----------
        node_size : int, optional
            Size to display the nodes in pixels.  Set to 0 to not display nodes.
            The default is 5.
        line_width : int, optional
            Width to display tracelines and element edges in pixels.  Set to 0
            to not show tracelines or edges.  The default is 1.
        opacity : float, optional
            A float between 0 and 1 to specify the transparency of the geometry.
            Set to 1 for completely opaque, and 0 for completely transparent
            (invisible).  The default is 1.0, no transparency.
        view_up : np.ndarray, optional
            Set the "up" direction in the plot by passing in a size-3 numpy
            array.  The default is None.
        view_from : np.ndarray, optional
            Specify the direction from which the geometry is viewed.  The
            default is None.
        plotter : BackgroundPlotter, optional
            A plotter can be specified to plot the geometry in an existing
            plot.  The default is None, which creates a new window and plot.
        show_edges : bool, optional
            Specify whether or not to draw edges on elements. The default is False.
        label_nodes : bool or iteratble, optional
            Specify whether or not to label the nodes in the geometry.  If True,
            all nodes will be plotted.  If label_nodes is an array or list, it
            should contain the node numbers to label.  The default is False,
            which does not label nodes.
        label_tracelines : bool or iteratble, optional
            Specify whether or not to label the tracelines in the geometry.  If True,
            all tracelines will be plotted.  If label_tracelines is an array or list, it
            should contain the traceline numbers to label.  The default is False,
            which does not label tracelines.
        label_elements : bool or iteratble, optional
            Specify whether or not to label the elements in the geometry.  If True,
            all elements will be plotted.  If label_elements is an array or list, it
            should contain the element numbers to label.  The default is False,
            which does not label elements.
        label_font_size : int, optional
            Specify the font size of the node labels.  The default is 16.
        plot_individual_items : bool, optional
            If True, then each item will be plotted one at a time.  This will
            make the plotting much slower, but it is required for certain
            features (like exporting to 3D PDF content) where having all data
            in one mesh is not feasible.  Note that volume elements will NOT
            be plotted in this case.

        Raises
        ------
        KeyError
            If referenced id numbers are not found in the corresponding object,
            for example if a traceline references node 11 but there is no node
            11 in the NodeArray
        ValueError
            If an invalid or unknown element type is used

        Returns
        -------
        plotter : BackgroundPlotter
            A reference to the plotter object that the geometry was plotted in
        face_mesh : TYPE
            A reference to the mesh used to plot surface elements
        point_mesh : TYPE
            A reference to the mesh used to plot nodes and tracelines
        solid_mesh : TYPE
            A reference to the mesh used to plot volume elements

        """
        if IGNORE_PLOTS:
            return None
        # Get part information
        nodes = self.node.flatten()
        css = self.coordinate_system.flatten()
        elems = self.element.flatten()
        tls = self.traceline.flatten()
        
        
        num_labels = ((1 if (label_nodes is True or (label_nodes is not False and len(label_nodes) > 0)) else 0)
                      + (1 if (label_tracelines is True or (label_tracelines is not False and len(label_tracelines) > 0)) else 0)
                      + (1 if (label_elements is True or (label_elements is not False and len(label_elements) > 0)) else 0))

        coordinate_system = CoordinateSystemArray(nodes.shape)
        cs_map = {cs['id']:idx for idx,cs in np.ndenumerate(self.coordinate_system)}
        for key, node in nodes.ndenumerate():
            coordinate_system[key] = css[cs_map[node.def_cs]]
        global_node_positions = global_coord(coordinate_system, nodes.coordinate)
        node_index_dict = {node.id[()]: index[0] for index, node in nodes.ndenumerate()}
        node_index_map = np.vectorize(node_index_dict.__getitem__)
        tl_index_dict = {tl.id[()]: index[0] for index, tl in tls.ndenumerate()}
        tl_index_map = np.vectorize(tl_index_dict.__getitem__)
        elem_index_dict = {elem.id[()]: index[0] for index, elem in elems.ndenumerate()}
        elem_index_map = np.vectorize(elem_index_dict.__getitem__)

        # Now go through and get the element and line connectivity
        if not plot_individual_items:
            face_element_connectivity = []
            face_element_colors = []
            solid_element_connectivity = []
            solid_element_colors = []
            solid_element_types = []
            node_colors = []
            line_connectivity = []
            line_colors = []
            for index, node in nodes.ndenumerate():
                # element_connectivity.append(1)
                # element_connectivity.append(index[0])
                # element_colors.append(node.color)
                node_colors.append(node.color)
            for index, element in elems.ndenumerate():
                # Check which type of element it is
                if element.type in _beam_elem_types:  # Beamlike element, use a line
                    line_connectivity.append(len(element.connectivity))
                    try:
                        line_connectivity.extend(node_index_map(element.connectivity))
                    except KeyError:
                        raise KeyError(
                            'Element {:} contains a node id not found in the node array'.format(element.id))
                    line_colors.append(element.color)
                elif element.type in _face_element_types:
                    face_element_connectivity.append(len(element.connectivity))
                    try:
                        face_element_connectivity.extend(node_index_map(element.connectivity))
                    except KeyError:
                        raise KeyError(
                            'Element {:} contains a node id not found in the node array'.format(element.id))
                    face_element_colors.append(element.color)
                elif element.type in _solid_element_types:
                    try:
                        solid_element_types.append(_vtk_element_map[element.type])
                    except KeyError:
                        raise ValueError('Do not know equivalent VTK element type for {:}: {:}'.format(
                            element.type, _element_types[element.type]))
                    solid_element_connectivity.append(len(element.connectivity))
                    try:
                        if solid_element_types[-1] in _vtk_connectivity_reorder:
                            solid_element_connectivity.extend(node_index_map(
                                element.connectivity[..., _vtk_connectivity_reorder[solid_element_types[-1]]]))
                        else:
                            solid_element_connectivity.extend(node_index_map(element.connectivity))
                    except KeyError:
                        raise KeyError(
                            'Element {:} contains a node id not found in the node array'.format(element.id))
                    solid_element_colors.append(element.color)
                else:
                    raise ValueError('Unknown element type {:}'.format(element.type))
    
            for index, tl in tls.ndenumerate():
                for conn_group in split_list(tl.connectivity, 0):
                    if len(conn_group) == 0:
                        continue
                    line_connectivity.append(len(conn_group))
                    try:
                        line_connectivity.extend(node_index_map(conn_group))
                    except KeyError:
                        raise KeyError(
                            'Traceline {:} contains a node id not found in the node array'.format(tl.id))
                    line_colors.append(tl.color)
        else:
            face_element_connectivity = []
            face_element_colors = []
            node_colors = []
            line_connectivity = []
            line_colors = []
            for index, node in nodes.ndenumerate():
                # element_connectivity.append(1)
                # element_connectivity.append(index[0])
                # element_colors.append(node.color)
                node_colors.append(node.color)
            for index, element in elems.ndenumerate():
                # Check which type of element it is
                if element.type in _beam_elem_types:  # Beamlike element, use a line
                    try:
                        line_connectivity.append(node_index_map(element.connectivity))
                    except KeyError:
                        raise KeyError(
                            'Element {:} contains a node id not found in the node array'.format(element.id))
                    line_colors.append(element.color)
                elif element.type in _face_element_types:
                    try:
                        if len(element.connectivity) == 3:
                            face_element_connectivity.append(node_index_map(element.connectivity))
                            face_element_colors.append(element.color)
                        else:
                            face_element_connectivity.append(node_index_map(element.connectivity[[0,1,3]]))
                            face_element_colors.append(element.color)
                            face_element_connectivity.append(node_index_map(element.connectivity[[1,2,3]]))
                            face_element_colors.append(element.color)
                    except KeyError:
                        raise KeyError(
                            'Element {:} contains a node id not found in the node array'.format(element.id))
                elif element.type in _solid_element_types:
                    warnings.warn('Solid Elements are currently not supported and will be skipped!')
                else:
                    raise ValueError('Unknown element type {:}'.format(element.type))

            for index, tl in tls.ndenumerate():
                for conn_group in split_list(tl.connectivity, 0):
                    if len(conn_group) == 0:
                        continue
                    try:
                        mapped_conn_group = node_index_map(conn_group)
                    except KeyError:
                        raise KeyError(
                            'Traceline {:} contains a node id not found in the node array'.format(tl.id))
                    for indices in zip(mapped_conn_group[:-1],mapped_conn_group[1:]):
                        line_connectivity.append(indices)
                        line_colors.append(tl.color)

        # Now we start to plot
        if plotter is None:
            plotter = GeometryPlotter(editor=False)

        if not plot_individual_items:
            # Need to split up between point mesh, face/line mesh, and solid mesh
            if len(face_element_connectivity) == 0 and len(line_connectivity) == 0:
                face_mesh = None
            else:
                face_mesh = pv.PolyData(global_node_positions,
                                        faces=None if len(
                                            face_element_connectivity) == 0 else face_element_connectivity,
                                        lines=None if len(line_connectivity) == 0 else line_connectivity)
                face_mesh.cell_data['color'] = line_colors + face_element_colors
    
                plotter.add_mesh(face_mesh, scalars='color', cmap=colormap, clim=[0, 15],
                                 show_edges=show_edges,  # True if line_width > 0 else False,
                                 show_scalar_bar=False, line_width=line_width,
                                 opacity=opacity)
            if len(solid_element_connectivity) == 0:
                solid_mesh = None
            else:
                solid_mesh = pv.UnstructuredGrid(np.array(solid_element_connectivity),
                                                 np.array(solid_element_types, dtype='uint8'),
                                                 np.array(global_node_positions))
                solid_mesh.cell_data['color'] = solid_element_colors
    
                plotter.add_mesh(solid_mesh, scalars='color', cmap=colormap, clim=[0, 15],
                                 show_edges=show_edges,  # True if line_width > 0 else False,
                                 show_scalar_bar=False, line_width=line_width,
                                 opacity=opacity)
    
            if node_size > 0:
                point_mesh = pv.PolyData(global_node_positions)
                point_mesh.cell_data['color'] = node_colors
                plotter.add_mesh(point_mesh, scalars=node_colors, cmap=colormap, clim=[0, 15],
                                 show_edges=show_edges,  # True if line_width > 0 else False,
                                 show_scalar_bar=False, point_size=node_size,
                                 opacity=opacity)
            else:
                point_mesh = None
            
            if (label_nodes is True or (label_nodes is not False and len(label_nodes) > 0)):
                try:
                    label_node_ids = [id_num for id_num in label_nodes]
                    label_node_indices = node_index_map(label_node_ids)
                except TypeError:
                    label_node_ids = [id_num for id_num in self.node.id]
                    label_node_indices = node_index_map(label_node_ids)
                
                    
                node_label_mesh = pv.PolyData(global_node_positions[label_node_indices])
                node_label_mesh['NODE_ID'] = [
                    ('N' if num_labels > 1 else '') + str(val) for val in label_node_ids]
                plotter.add_point_labels(node_label_mesh, 'NODE_ID', tolerance=0.0, shape=None,
                                         show_points=False, always_visible=True, font_size=label_font_size)
                
            if (label_tracelines is True or (label_tracelines is not False and len(label_tracelines) > 0)):
                try:
                    label_traceline_ids = [id_num for id_num in label_tracelines]
                    label_traceline_indices = tl_index_map(label_traceline_ids)
                except TypeError:
                    label_traceline_ids = [id_num for id_num in self.traceline.id]
                    label_traceline_indices = tl_index_map(label_traceline_ids)
                
                # Need to get the point to plot for each traceline
                traceline_locations = []
                for idx in label_traceline_indices:
                    tl = tls[idx]
                    connectivity = tl.connectivity[tl.connectivity != 0]
                    # Find the middle of the traceline
                    center_nodes = [connectivity[len(connectivity)//2-1],connectivity[len(connectivity)//2]]
                    position = np.mean(global_node_positions[node_index_map(center_nodes)],axis=0)
                    traceline_locations.append(position)
                    
                tl_label_mesh = pv.PolyData(np.array(traceline_locations))
                tl_label_mesh['TL_ID'] = [
                    ('L' if num_labels > 1 else '') + str(val) for val in label_traceline_ids]
                plotter.add_point_labels(tl_label_mesh, 'TL_ID', tolerance=0.0, shape=None,
                                         show_points=False, always_visible=True, font_size=label_font_size)
                
            if (label_elements is True or (label_elements is not False and len(label_elements) > 0)):
                try:
                    label_element_ids = [id_num for id_num in label_elements]
                    label_element_indices = elem_index_map(label_element_ids)
                except TypeError:
                    label_element_ids = [id_num for id_num in self.element.id]
                    label_element_indices = elem_index_map(label_element_ids)
                
                # Need to get the point to plot for each traceline
                element_locations = []
                for idx in label_element_indices:
                    elem = elems[idx]
                    connectivity = elem.connectivity
                    # Find the middle of the traceline
                    position = np.mean(global_node_positions[node_index_map(connectivity)],axis=0)
                    element_locations.append(position)
                    
                elem_label_mesh = pv.PolyData(np.array(element_locations))
                elem_label_mesh['ELEM_ID'] = [
                    ('E' if num_labels > 1 else '') + str(val) for val in label_element_ids]
                plotter.add_point_labels(elem_label_mesh, 'ELEM_ID', tolerance=0.0, shape=None,
                                         show_points=False, always_visible=True, font_size=label_font_size)
                
        else:
            face_map = []
            face_mesh = []
            for conn,color in zip(face_element_connectivity,face_element_colors):
                node_indices,inverse = np.unique(conn,return_inverse=True)
                node_positions = global_node_positions[node_indices][inverse]
                node_connectivity = np.arange(node_positions.shape[0])
                mesh = pv.PolyData(node_positions,faces=[len(node_connectivity)]+[c for c in node_connectivity])
                mesh.cell_data['color'] = color
                mesh.SetObjectName('Elem: {:} {:} {:}'.format(*self.node.id[node_indices[inverse]]))
                plotter.add_mesh(mesh,scalars='color',cmap=colormap, clim = [0,15],
                                 show_edges=show_edges, show_scalar_bar = False,
                                 line_width=line_width,opacity=opacity,label='Elem: {:} {:} {:}'.format(*self.node.id[node_indices[inverse]]))
                face_map.append(self.node.id[node_indices[inverse]])
                face_mesh.append(mesh)

            line_map = []
            line_mesh = []
            for conn,color in zip(line_connectivity,line_colors):
                node_indices,inverse = np.unique(conn,return_inverse=True)
                node_positions = global_node_positions[node_indices][inverse]
                node_connectivity = np.arange(node_positions.shape[0])
                mesh = pv.PolyData(node_positions,lines=[len(node_connectivity)]+[c for c in node_connectivity])
                mesh.cell_data['color'] = color
                mesh.SetObjectName('Line: {:} {:}'.format(*self.node.id[node_indices[inverse]]))
                plotter.add_mesh(mesh,scalars='color',cmap=colormap, clim = [0,15],
                                 show_edges=show_edges, show_scalar_bar = False,
                                 line_width=line_width,opacity=opacity,label='Line: {:} {:}'.format(*self.node.id[node_indices[inverse]]))
                line_map.append(self.node.id[node_indices[inverse]])
                line_mesh.append(mesh)
            
            point_map = []
            point_mesh = []
            for node,position,color in zip(self.node.id,global_node_positions,self.node.color):
                if node_size > 0:
                    mesh = pv.PolyData(position)
                    mesh.cell_data['color'] = color
                    plotter.add_mesh(mesh, scalars = color, cmap=colormap, clim=[0,15],
                                     show_edges=show_edges, show_scalar_bar=False, point_size=node_size,
                                     opacity=opacity)
                    point_map.append(node)
                    point_mesh.append(mesh)
            
            if label_nodes:
                try:
                    label_node_ids = [id_num for id_num in label_nodes]
                    label_node_indices = node_index_map(label_node_ids)
                except TypeError:
                    label_node_ids = [id_num for id_num in self.node.id]
                    label_node_indices = node_index_map(label_node_ids)
                
                for index,label in zip(label_node_indices,label_node_ids):
                    node_label_mesh = pv.PolyData(global_node_positions[index])
                    node_label_mesh['NODE_ID'] = [str(label)]
                    plotter.add_point_labels(node_label_mesh, 'NODE_ID', tolerance=0.0, shape=None,
                                             show_points=False, always_visible=True, font_size=label_font_size)
            
        if view_from is not None:
            focus = plotter.camera.focal_point
            distance = plotter.camera.distance
            plotter.camera.position = np.array(
                focus) + distance * np.array(view_from) / np.linalg.norm(view_from)
            plotter.camera.focal_point = focus
        if view_up is not None:
            plotter.camera.up = view_up
        plotter.show()
        plotter.render()
        QApplication.processEvents()
        if plot_individual_items:
            return plotter, (face_mesh,face_map), (line_mesh,line_map), (point_mesh,point_map)
        else:
            return plotter, face_mesh, point_mesh, solid_mesh

    def global_geometry_table(self,coordinate_array_or_nodelist = None):
        if coordinate_array_or_nodelist is None:
            coord = coordinate_array(self.node.id,[1,2,3],force_broadcast=True)
        elif isinstance(coordinate_array_or_nodelist,CoordinateArray):
            coord = coordinate_array_or_nodelist.flatten()
        else:
            coord = coordinate_array(coordinate_array_or_nodelist,[1,2,3],force_broadcast=True)
        global_positions = self.global_node_coordinate(coord.node)
        global_directions = self.global_deflection(coord)
        coord_string = coord.string_array()
        df_dict = {}
        df_dict['Degree of Freedom'] = coord_string
        for label,column in zip('XYZ',global_positions.T):
            df_dict[f'Coordinate {label}'] = column
        for label,column in zip('XYZ',global_directions.T):
            df_dict[f'Direction {label}'] = column
        df = pd.DataFrame(df_dict)
        return df
            
    def plot_coordinate(self, coordinates: CoordinateArray = None,
                        arrow_scale=0.1,
                        arrow_scale_type='bbox', label_dofs=False,
                        label_font_size=16, opacity=1.0,
                        arrow_ends_on_node=False, plot_kwargs={},
                        plot_individual_items = False):
        """
        Plots coordinate arrows on the geometry

        Parameters
        ----------
        coordinates : CoordinateArray, optional
            Coordinates to draw on the geometry.  If no coordinates are specified,
            all translation degrees of freedom at each node will be plotted.
        arrow_scale : float, optional
            Size of the arrows in proportion to the length of the diagonal of
            the bounding box of the Geometry if `arrow_scale_type` is 'bbox',
            otherwise the raw length of the arrow.  The default is 0.1.
        arrow_scale_type : str, optional
            Specifies how to compute the size of the arrows.  If 'bbox', then
            the arrow is scaled based on the size of the geometry.  Otherwise,
            the arrow size is the specified length.  The default is 'bbox'.
        label_dofs : bool, optional
            Specify whether or not to label the coordinates with strings.
            The default is False.
        label_font_size : int, optional
            Specifies the font size for the node labels.
            Default is 16.
        opacity : float, optional
            A float between 0 and 1 to specify the transparency of the geometry.
            Set to 1 for completely opaque, and 0 for completely transparent
            (invisible).  The default is 1.0, no transparency.
        arrow_ends_on_node : bool, optional
            If True, arrow tip ends at the node, otherwise the arrow begins at
            node.
            Defualt is False
        plot_kwargs : dict, optional
            Any additional keywords that should be passed to the Geometry.plot
            function. The default is {}.
        plot_individual_items : bool, optional
            If True, then each item will be plotted one at a time.  This will
            make the plotting much slower, but it is required for certain
            features (like exporting to 3D PDF content) where having all data
            in one mesh is not feasible.  Note that volume elements will NOT
            be plotted in this case.

        Returns
        -------
        plotter : BackgroundPlotter
            A reference to the window the geometry was plotted in.

        """
        if IGNORE_PLOTS:
            return None
        plot_kwargs = plot_kwargs.copy()
        plot_kwargs['plot_individual_items'] = plot_individual_items
        plotter, face_mesh, point_mesh, solid_mesh = self.plot(opacity=opacity, **plot_kwargs)
        # Add deflection directions to each node
        # Get local deformation for each coordinate direction

        if coordinates is None:
            coordinates = from_nodelist(self.node.id)

        def build_coord_mesh(coordinates, geom):
            nodes = coordinates.node
            indices = np.in1d(coordinates.node, geom.node.id)
            coordinates = coordinates[indices]
            nodes = nodes[indices]
            local_deformations = coordinates.local_direction()
            coordinate_systems = geom.coordinate_system(geom.node(nodes).disp_cs)
            points = global_coord(geom.coordinate_system(
                geom.node(nodes).def_cs), geom.node(nodes).coordinate)
            global_deflections = global_deflection(coordinate_systems, local_deformations, points)
            # Now add the point array to the mesh
            coord_mesh = pv.PolyData(points)
            coord_mesh.point_data['Coordinates'] = global_deflections
            coord_mesh.point_data['Direction'] = abs(coordinates.direction)
            coord_mesh.point_data['Node ID'] = nodes
            return coord_mesh, points, global_deflections

        coordinates = coordinates.flatten()

        if arrow_scale_type == 'bbox':
            node_coords = self.global_node_coordinate()
            bbox_diagonal = np.linalg.norm(
                np.max(node_coords, axis=0) - np.min(node_coords, axis=0))
            arrow_factor = bbox_diagonal * arrow_scale
        else:
            arrow_factor = arrow_scale

        # Create Straight Arrows and Labels
        coordinates_straight = coordinates[np.abs(coordinates.direction) < 4]

        if coordinates_straight.size:
            if plot_individual_items:
                for coordinate_straight in coordinates_straight:
                    [coord_mesh_straight, points_straight, global_deflections_straight] = build_coord_mesh(
                        coordinate_straight[np.newaxis], self)
        
                    if arrow_ends_on_node:
                        arrow_start = (-1.0, 0.0, 0.0)
                    else:
                        arrow_start = (0.0, 0.0, 0.0)
                    coord_arrows_stright = coord_mesh_straight.glyph(
                        orient='Coordinates', scale=False, factor=arrow_factor, geom=pv.Arrow(start=arrow_start))
                    color = ((1.0,0.0,0.0) if abs(coordinate_straight.direction) == 1 else
                             (0.0,1.0,0.0) if abs(coordinate_straight.direction) == 2 else
                             (0.0,0.0,1.0))
                    plotter.add_mesh(coord_arrows_stright, color = color, show_scalar_bar=False)
            else:
                [coord_mesh_straight, points_straight, global_deflections_straight] = build_coord_mesh(
                    coordinates_straight, self)
    
                if arrow_ends_on_node:
                    arrow_start = (-1.0, 0.0, 0.0)
                else:
                    arrow_start = (0.0, 0.0, 0.0)
                coord_arrows_stright = coord_mesh_straight.glyph(
                    orient='Coordinates', scale=False, factor=arrow_factor, geom=pv.Arrow(start=arrow_start))
                plotter.add_mesh(coord_arrows_stright, scalars='Direction',
                                 cmap=coord_colormap, clim=[1, 6], show_scalar_bar=False)

            if label_dofs:
                if arrow_ends_on_node:
                    dof_label_mesh = pv.PolyData(
                        points_straight - global_deflections_straight * arrow_factor)
                else:
                    dof_label_mesh = pv.PolyData(
                        points_straight + global_deflections_straight * arrow_factor)
                dof_label_mesh['DOF'] = [val for val in coordinates_straight.string_array()]
                plotter.add_point_labels(dof_label_mesh, 'DOF', tolerance=0.0, shape=None,
                                         show_points=False, always_visible=True, font_size=label_font_size)

        # Create Rotational Arrows and Labels
        coordinates_rotation = coordinates[np.abs(coordinates.direction) > 3]
        if coordinates_rotation.size:
            if plot_individual_items:
                for coordinate_rotation in coordinates_rotation:
                    [coord_mesh_rotation, points_rotation, global_deflections_rotation] = build_coord_mesh(
                        coordinate_rotation[np.newaxis], self)
        
                    arrow_index = np.arange(4)
                    head_location_angles = 1 / 8 * np.pi + 1 / 2 * np.pi * arrow_index
                    tail_location_angles = 3 / 8 * np.pi + 1 / 2 * np.pi * arrow_index
        
                    r = 1
                    arrow_head_locations = np.array([np.zeros(
                        head_location_angles.size), r * np.sin(head_location_angles), r * np.cos(head_location_angles)]).T
                    arrow_tail_locations = np.array([np.zeros(
                        tail_location_angles.size), r * np.sin(tail_location_angles), r * np.cos(tail_location_angles)]).T
                    cone_vectors = np.array([np.zeros(head_location_angles.size), np.sin(
                        head_location_angles - np.pi / 2), np.cos(head_location_angles - np.pi / 2)]).T
        
                    arc = pv.merge([pv.CircularArc(pointa=start, pointb=stop, center=(0, 0, 0)).tube(
                        radius=0.05) for start, stop in zip(arrow_head_locations, arrow_tail_locations)])
                    cone = pv.merge([pv.Cone(center=cone_center, direction=cone_vector, height=0.3, radius=0.1, resolution=20)
                                    for cone_center, cone_vector in zip(arrow_head_locations, cone_vectors)])
        
                    curved_arrow = pv.merge([arc, cone])
                    coord_arrows_rotation = coord_mesh_rotation.glyph(
                        orient='Coordinates', scale=False, factor=arrow_factor, geom=curved_arrow)
                    color = ((1.0,0.0,0.0) if abs(coordinate_straight.direction) == 1 else
                             (0.0,1.0,0.0) if abs(coordinate_straight.direction) == 2 else
                             (0.0,0.0,1.0))
                    plotter.add_mesh(coord_arrows_rotation, color=color, show_scalar_bar=False)
            else:
                [coord_mesh_rotation, points_rotation, global_deflections_rotation] = build_coord_mesh(
                    coordinates_rotation, self)
    
                arrow_index = np.arange(4)
                head_location_angles = 1 / 8 * np.pi + 1 / 2 * np.pi * arrow_index
                tail_location_angles = 3 / 8 * np.pi + 1 / 2 * np.pi * arrow_index
    
                r = 1
                arrow_head_locations = np.array([np.zeros(
                    head_location_angles.size), r * np.sin(head_location_angles), r * np.cos(head_location_angles)]).T
                arrow_tail_locations = np.array([np.zeros(
                    tail_location_angles.size), r * np.sin(tail_location_angles), r * np.cos(tail_location_angles)]).T
                cone_vectors = np.array([np.zeros(head_location_angles.size), np.sin(
                    head_location_angles - np.pi / 2), np.cos(head_location_angles - np.pi / 2)]).T
    
                arc = pv.merge([pv.CircularArc(pointa=start, pointb=stop, center=(0, 0, 0)).tube(
                    radius=0.05) for start, stop in zip(arrow_head_locations, arrow_tail_locations)])
                cone = pv.merge([pv.Cone(center=cone_center, direction=cone_vector, height=0.3, radius=0.1, resolution=20)
                                for cone_center, cone_vector in zip(arrow_head_locations, cone_vectors)])
    
                curved_arrow = pv.merge([arc, cone])
                coord_arrows_rotation = coord_mesh_rotation.glyph(
                    orient='Coordinates', scale=False, factor=arrow_factor, geom=curved_arrow)
                # Make Colors of Rotations same as Straight
                coord_arrows_rotation['Direction'] = coord_arrows_rotation['Direction'] - 3
    
                plotter.add_mesh(coord_arrows_rotation, scalars='Direction',
                                 cmap=coord_colormap, clim=[1, 6], show_scalar_bar=False)

            if label_dofs:
                nodes = coordinates_rotation.node
                indices = np.in1d(coordinates_rotation.node, self.node.id)
                nodes = nodes[indices]
                local_deformations = coordinates_rotation.local_direction()
                local_deformation_transformation = np.array(([arrow_head_locations[0, 0], arrow_head_locations[0, 1], arrow_head_locations[0, 2]],
                                                             [arrow_head_locations[0, 2], arrow_head_locations[0,
                                                                                                               0], arrow_head_locations[0, 1]],
                                                             [arrow_head_locations[0, 1], arrow_head_locations[0, 2], arrow_head_locations[0, 0]]))
                local_deformations_new = np.transpose(
                    local_deformation_transformation.T @ local_deformations.T)
                coordinate_systems = self.coordinate_system(self.node(nodes).disp_cs)
                points = global_coord(self.coordinate_system(
                    self.node(nodes).def_cs), self.node(nodes).coordinate)
                global_deflections_rotation_new = global_deflection(
                    coordinate_systems, local_deformations_new, points)

                dof_label_mesh = pv.PolyData(
                    points_rotation + global_deflections_rotation_new * arrow_factor)

                dof_label_mesh['DOF'] = [val for val in coordinates_rotation.string_array()]
                plotter.add_point_labels(dof_label_mesh, 'DOF', tolerance=0.0, shape=None,
                                         show_points=False, always_visible=True, font_size=label_font_size)
        return plotter

    def plot_shape(self, shape, plot_kwargs={}, background_plotter_kwargs={'editor': False}, undeformed_opacity=0.25, deformed_opacity=1.0, starting_scale=1.0):
        """
        Plot mode shapes on the geometry

        Parameters
        ----------
        shape : ShapeArray
            The set of shapes to plot
        plot_kwargs : dict, optional
            Any additional keywords that should be passed to the Geometry.plot
            function. The default is {}.
        background_plotter_kwargs : dict, optional
            Any additional arguments that should be passed to the
            BackgroundPlotter initializer.  The default is {'editor':False}.
        undeformed_opacity : float, optional
            A float between 0 and 1 to specify the transparency of the undeformed
            geometry.  Set to 1 for completely opaque, and 0 for completely
            transparent (invisible).  The default is 0.25.
        deformed_opacity : float, optional
            A float between 0 and 1 to specify the transparency of the deformed
            geometry.  Set to 1 for completely opaque, and 0 for completely
            transparent (invisible). The default is 1.0.
        starting_scale : float, optional
            The starting scale factor of the animation. The default is 1.0.

        Returns
        -------
        ShapePlotter
            A reference to the ShapePlotter class that is created to plot the
            animated shapes

        """
        if IGNORE_PLOTS:
            return None
        return ShapePlotter(self, shape, plot_kwargs, background_plotter_kwargs,
                            undeformed_opacity, deformed_opacity=deformed_opacity, starting_scale=starting_scale)

    def plot_deflection_shape(self, deflection_shape_data, plot_kwargs={},
                              background_plotter_kwargs={'editor': False},
                              undeformed_opacity=0.25, deformed_opacity=1.0,
                              starting_scale=1.0):
        """
        Plot deflection shapes shapes on the geometry

        Parameters
        ----------
        deflection_shape_data : NDDataArray
            Data array containing the deflection shapes to plot
        plot_kwargs : dict, optional
            Any additional keywords that should be passed to the Geometry.plot
            function. The default is {}.
        background_plotter_kwargs : dict, optional
            Any additional arguments that should be passed to the
            BackgroundPlotter initializer.  The default is {'editor':False}.
        undeformed_opacity : float, optional
            A float between 0 and 1 to specify the transparency of the undeformed
            geometry.  Set to 1 for completely opaque, and 0 for completely
            transparent (invisible).  The default is 0.25.
        deformed_opacity : float, optional
            A float between 0 and 1 to specify the transparency of the deformed
            geometry.  Set to 1 for completely opaque, and 0 for completely
            transparent (invisible). The default is 1.0.
        starting_scale : float, optional
            The starting scale factor of the animation. The default is 1.0.

        Returns
        -------
        deflection_shape_data
            A reference to the deflection_shape_data object that is created to
            plot the animated shapes

        """
        if IGNORE_PLOTS:
            return None
        return DeflectionShapePlotter(self, deflection_shape_data, plot_kwargs, background_plotter_kwargs,
                                      undeformed_opacity, deformed_opacity=deformed_opacity, starting_scale=starting_scale)

    def plot_transient(self, displacement_data, displacement_scale=1.0,
                       frames_per_second=20,
                       undeformed_opacity=0.0, deformed_opacity=1.0,
                       plot_kwargs={},
                       transformation_shapes=None,
                       num_curves=50,
                       show: bool = True,
                       app=None,
                       window_size=None,
                       off_screen=None,
                       allow_quit_keypress=True,
                       toolbar=True,
                       menu_bar=True,
                       editor=False,
                       update_app_icon=None,
                       **kwargs):
        """
        Create a TransientPlotter object to plot displacements over time

        Parameters
        ----------
        displacement_data : TimeHistoryArray
            Transient displacement data that will be applied
        displacement_scale : float, optional
            Scale factor applied to displacements. The default is 1.0.
        frames_per_second : float, optional
            Number of time steps to plot per second while the displacement is
            animating.  Default is 20.
        undeformed_opacity : float, optional
            Opacity of the undeformed geometry. The default is 0.0, or
            completely transparent.
        deformed_opacity : float, optional
            Opacity of the deformed geometry. The default is 1.0, or completely
            opaque.
        plot_kwargs : dict, optional
            Keyword arguments passed to the Geometry.plot function
        transformation_shapes : ShapeArray
            Shape matrix that will be used to expand the data.  Must be the
            same size as the `displacement_data`
        num_curves : int, optional
            Maximum number of curves to plot on the time selector.  Default is
            50.
        show : bool, optional
            Show the plotting window.  If ``False``, show this window by
            running ``show()``. The default is True.
        app : QApplication, optional
            Creates a `QApplication` if left as `None`.  The default is None.
        window_size : list of int, optional
            Window size in pixels.  Defaults to ``[1024, 768]``
        off_screen : TYPE, optional
            Renders off screen when True.  Useful for automated
            screenshots or debug testing. The default is None.
        allow_quit_keypress : bool, optional
            Allow user to exit by pressing ``"q"``. The default is True.
        toolbar : bool, optional
            If True, display the default camera toolbar. Defaults to True.
        menu_bar : bool, optional
            If True, display the default main menu. Defaults to True.
        editor : TYPE, optional
            If True, display the VTK object editor. Defaults to False.
        update_app_icon : bool, optional
            If True, update_app_icon will be called automatically to update the
            Qt app icon based on the current rendering output. If None, the
            logo of PyVista will be used. If False, no icon will be set.
            Defaults to None. The default is None.
        title : str, optional
            Title of plotting window.
        multi_samples : int, optional
            The number of multi-samples used to mitigate aliasing. 4 is a
            good default but 8 will have better results with a potential
            impact on performance.
        line_smoothing : bool, optional
            If True, enable line smothing
        point_smoothing : bool, optional
            If True, enable point smothing
        polygon_smoothing : bool, optional
            If True, enable polygon smothing
        auto_update : float, bool, optional
            Automatic update rate in seconds.  Useful for automatically
            updating the render window when actors are change without
            being automatically ``Modified``.  If set to ``True``, update
            rate will be 1 second.

        Returns
        -------
        TransientPlotter
        """
        if IGNORE_PLOTS:
            return None
        return TransientPlotter(self, displacement_data, displacement_scale,
                                frames_per_second,
                                undeformed_opacity, deformed_opacity,
                                plot_kwargs,
                                transformation_shapes,
                                num_curves,
                                show,
                                app,
                                window_size,
                                off_screen,
                                allow_quit_keypress,
                                toolbar,
                                menu_bar,
                                editor,
                                update_app_icon,
                                **kwargs)

    def __add__(self, geometry):
        if not isinstance(self, geometry.__class__):
            return NotImplemented
        final_fields = {}
        for field in ['node', 'traceline', 'element', 'coordinate_system']:
            # Look through and see which are in common
            common_ids = np.intersect1d(getattr(self, field).id, getattr(geometry, field).id)
            if common_ids.size > 0:
                equal_ids = getattr(self, field)(common_ids) == getattr(geometry, field)(common_ids)
                if not all(equal_ids):
                    raise ValueError('Both geometries contain {:} with ID {:} but they are not equivalent'.format(
                        field, common_ids[~equal_ids]))
            self_ids = np.concatenate(
                (np.setdiff1d(getattr(self, field).id, getattr(geometry, field).id), common_ids))
            geometry_ids = np.setdiff1d(getattr(geometry, field).id, getattr(self, field).id)
            final_fields[field] = np.concatenate(
                (getattr(self, field)(self_ids), getattr(geometry, field)(geometry_ids)))
        return Geometry(**final_fields)

    def modify_ids(self, node_change=0, traceline_change=0, element_change=0, coordinate_system_change=0):
        """
        Shifts the id numbers in the geometry

        Parameters
        ----------
        node_change : int, optional
            The amount to shift the node ids. The default is 0.
        traceline_change : int, optional
            The amount to shift the traceline ids.. The default is 0.
        element_change : int, optional
            The amount to shift the element ids. The default is 0.
        coordinate_system_change : int, optional
            The amount to shift the coordinate system ids. The default is 0.

        Returns
        -------
        geom_out : Geometry
            A copy of the original geometry with id numbers modified

        """
        geom_out = self.copy()
        geom_out.node.id += node_change
        geom_out.traceline.id += traceline_change
        geom_out.element.id += element_change
        geom_out.coordinate_system.id += coordinate_system_change
        geom_out.node.def_cs += coordinate_system_change
        geom_out.node.disp_cs += coordinate_system_change
        geom_out.traceline.connectivity += node_change
        geom_out.element.connectivity += node_change
        # Set values of tracelines back to zero if they just equal node_change,
        # because that means they were zero and should represent picking up the
        # pen
        for connectivity_array in geom_out.traceline.connectivity:
            connectivity_array[connectivity_array == node_change] = 0
        return geom_out

    def map_ids(self, node_id_map=None, traceline_id_map=None, element_id_map=None, coordinate_system_id_map=None):
        """
        Maps id numbers from an original set of ids to a new set of ids.

        This function accepts id_map classes defining "from" and "to" ids
        Existing ids found in the "from" set are transformed to the corresponding
        id in the "to" set.

        Parameters
        ----------
        node_id_map : id_map, optional
            An id_map defining the mapping applied to node ids.
            The default is None, which results in the ids being unchanged
        traceline_id_map : id_map, optional
            An id_map defining the mapping applied to traceline ids.
            The default is None, which results in the ids being unchanged
        element_id_map : id_map, optional
            An id_map defining the mapping applied to element ids.
            The default is None, which results in the ids being unchanged
        coordinate_system_id_map : id_map, optional
            An id_map defining the mapping applied to coordinate system ids.
            The default is None, which results in the ids being unchanged

        Returns
        -------
        geom_out : Geometry
            A copy of the original geometry with id numbers modified

        """
        geom_out = self.copy()
        if node_id_map is not None:
            geom_out.node.id = node_id_map(self.node.id)
            for key, val in self.traceline.ndenumerate():
                geom_out.traceline.connectivity[key] = node_id_map(self.traceline.connectivity[key])
            for key, val in self.element.ndenumerate():
                geom_out.element.connectivity[key] = node_id_map(self.element.connectivity[key])
        if coordinate_system_id_map is not None:
            geom_out.coordinate_system.id = coordinate_system_id_map(self.coordinate_system.id)
            geom_out.node.disp_cs = coordinate_system_id_map(self.node.disp_cs)
            geom_out.node.def_cs = coordinate_system_id_map(self.node.def_cs)
        if traceline_id_map is not None:
            geom_out.traceline.id = traceline_id_map(self.traceline.id)
        if element_id_map is not None:
            geom_out.element.id = element_id_map(self.element.id)
        return geom_out

    def response_kinematic_transformation(self, response_coordinate, 
                                          virtual_point_node_number,
                                          virtual_point_location=[0,0,0]):
        """
        Creates a kinematic response transformation from the rigid body body 
        shapes for the associated geometry and response coordinates. 

        Parameters
        ----------
        response_coordinate : CoordinateArray
            The "physical" response coordinates for the kinematic transformation.
        virtual_point_node_number : int
            The numeric ID for the virtual point in the kinematic transformation.
        virtual_point_location : list or ndarray
            The [X, Y, Z] coordinates that defined the location of the virtual 
            point. The default is is [0, 0, 0].  

        Returns
        -------
        transformation : Matrix
            The kinematic transformation as a matrix object. It is organized with
            the virtual point CoordinateArray on the rows and the physical 
            response CoordinateArray on the columns. 

        Notes 
        -----
        The transformation automatically handles polarity differences in the geometry 
        and response_coordinate.

        References
        ----------
        .. [1] M. Van der Seijs, D. van den Bosch, D. Rixen, and D. Klerk, "An improved 
               methodology for the virtual point transformation of measured frequency 
               response functions in dynamic substructuring," in Proceedings of the 4th 
               International Conference on Computational Methods in Structural Dynamics 
               and Earthquake Engineering, Kos Island, 2013, pp. 4334-4347, 
               doi: 10.7712/120113.4816.C1539. 
        """
        shapes = self.rigid_body_shapes(response_coordinate, cg = virtual_point_location)
        transformation_array = np.linalg.pinv(shapes.shape_matrix.T)

        virtual_point_coordinate = coordinate_array(node=virtual_point_node_number, direction=[1,2,3,4,5,6])

        return matrix(transformation_array, virtual_point_coordinate, response_coordinate)
    
    def force_kinematic_transformation(self, force_coordinate, 
                                       virtual_point_node_number,
                                       virtual_point_location=[0,0,0]):
        """
        Creates a kinematic force transformation from the rigid body body 
        shapes for the associated geometry and force coordinates. 

        Parameters
        ----------
        force_coordinate : CoordinateArray
            The "physical" force coordinates for the kinematic transformation.
        virtual_point_node_number : int
            The numeric ID for the virtual point in the kinematic transformation.
        virtual_point_location : list or ndarray
            The [X, Y, Z] coordinates that defined the location of the virtual 
            point. The default is is [0, 0, 0].  

        Returns
        -------
        transformation : Matrix
            The kinematic transformation as a matrix object. It is organized with
            the virtual point CoordinateArray on the rows and the physical 
            force CoordinateArray on the columns. 

        Notes 
        -----
        The transformation automatically handles polarity differences in the geometry 
        and force_coordinate.

        References
        ----------
        .. [1] M. Van der Seijs, D. van den Bosch, D. Rixen, and D. Klerk, "An improved 
               methodology for the virtual point transformation of measured frequency 
               response functions in dynamic substructuring," in Proceedings of the 4th 
               International Conference on Computational Methods in Structural Dynamics 
               and Earthquake Engineering, Kos Island, 2013, pp. 4334-4347, 
               doi: 10.7712/120113.4816.C1539. 
        """
        shapes = self.rigid_body_shapes(force_coordinate, cg = virtual_point_location)
        virtual_point_coordinate = coordinate_array(node=virtual_point_node_number, direction=[1,2,3,4,5,6])

        return matrix(shapes.shape_matrix, virtual_point_coordinate, force_coordinate)

    def copy(self):
        """
        Return's a copy of the current Geometry

        Changes to the copy will not also be applied to the original geometry

        Returns
        -------
        Geometry
            A copy of the current Geometry

        """
        return Geometry(self.node.copy(), self.coordinate_system.copy(), self.traceline.copy(), self.element.copy())

    def save(self, filename):
        """Saves the geometry to a numpy .npz file

        The .npz file will have fields 'node', 'coordinate_system', 'element',
        and 'traceline', with each field storing the respective portion of the
        geometry

        Parameters
        ----------
        filename : str
            Filename to save the geometry to.  If the filename doesn't end with
            .npz, it will be appended.

        Returns
        -------
        None.

        """
        np.savez(filename,
                 node=self.node.view(np.ndarray),
                 coordinate_system=self.coordinate_system.view(np.ndarray),
                 element=self.element.view(np.ndarray),
                 traceline=self.traceline.view(np.ndarray))

    def rigid_body_shapes(self, coordinates, mass=1, inertia=np.eye(3), cg=np.zeros(3), principal_axes=False):
        """
        Creates a set of shapes corresponding to the rigid body motions

        Rigid body translation and rotation shapes are computed analytically
        from the Geoemtry

        Parameters
        ----------
        coordinates : CoordinateArray
            coordinates at which to compute deformations
        mass : float, optional
            The mass of the geometry used to scale the rigid body translation
            shapes.  The default is 1.
        inertia : np.ndarray, optional
            A 3x3 array consisting of the mass moments of inertia, used to
            scale the rotation shapes.  The default is np.eye(3).
        cg : np.ndarray, optional
            The center of gravity of the Geometry about which the rotations
            occur.  The default is np.zeros(3).
        principal_axes : bool, optional
            If True, compute the principal axes of the test article and
            perform rotations about those axes.  The default is False.

        Returns
        -------
        output_shape : ShapeArray
            A set of rigid body shapes for the current geometry

        """
        if principal_axes:
            rotation_scale, rotation_directions = np.linalg.eig(inertia)
            rotation_directions /= np.linalg.norm(rotation_directions, axis=0, keepdims=True)
        else:
            rotation_scale = np.diag(inertia)
            rotation_directions = np.eye(3)
        M = np.diag(np.concatenate((mass * np.ones(3), rotation_scale)))
        phi = np.sqrt(np.linalg.inv(M))
        # Construct the full coordinate array for all degrees of freedom for each node
        full_coordinates = coordinate_array(np.unique(coordinates.node)[:, np.newaxis],
                                            ['X+', 'Y+', 'Z+']).flatten()
        # Assume local coordinates for now, transform later
        translation_shape_matrix = full_coordinates.local_direction().T
        rotation_shape_matrix = np.zeros(translation_shape_matrix.shape)
        for i in range(3):
            for j in range(len(full_coordinates)):
                direction = full_coordinates[j].local_direction()
                coord = self.global_node_coordinate(full_coordinates[j].node)
                rotation_shape_matrix[i, j] = phi[i + 3, i + 3] * \
                    np.dot(direction, np.cross(rotation_directions[:, i], coord - cg))
        # Concatenate shapes together
        shape_matrix = np.concatenate((translation_shape_matrix, rotation_shape_matrix), axis=0)
        full_shapes = shape_array(full_coordinates, shape_matrix)
        # Now we need to transform to the local coordinate systems
        # First set up a geometry with just the global coordinate system in it
        global_geometry = Geometry(node=self.node.copy(),
                                   coordinate_system=coordinate_system_array((1,)))
        global_geometry.node.disp_cs = 1
        global_geometry.node.def_cs = 1
        transformed_shape = full_shapes.transform_coordinate_system(global_geometry, self)
        # Now turn it into our requested coordinates
        reduced_shape_matrix = transformed_shape[coordinates]
        output_shape = shape_array(coordinates, reduced_shape_matrix)
        return output_shape

    @classmethod
    def load(cls, filename):
        """
        Loads a geometry from a numpy .npz or .unv file

        The .npz file must have fields 'node', 'coordinate_system', 'element',
        and 'traceline', with each field storing the respective portion of the
        geometry

        The .unv file will need to have the proper datasets specified to define
        a geometry

        Parameters
        ----------
        filename : str
            Filename to load the geometry from.

        Raises
        ------
        AttributeError
            Raised if the calling class does not have a `from_unv` method
            defined

        Returns
        -------
        Geometry
            Geometry constructed from the data in the loaded file

        """
        if filename[-4:].lower() in ['.unv', '.uff']:
            try:
                from ..fileio.sdynpy_uff import readunv
                unv_dict = readunv(filename)
                return cls.from_unv(unv_dict)
            except AttributeError:
                raise AttributeError('Class {:} has no from_unv attribute defined'.format(cls))
        else:
            with np.load(filename, allow_pickle=True) as data:
                return Geometry(node=data['node'].view(NodeArray),
                                coordinate_system=data['coordinate_system'].view(
                                    CoordinateSystemArray),
                                traceline=data['traceline'].view(TracelineArray),
                                element=data['element'].view(ElementArray))

    def write_to_unv(self, filename, write_nodes=True, write_coordinate_systems=True,
                     write_tracelines=True, write_elements=True,
                     dataset_2412_kwargs={},
                     dataset_2420_kwargs={}):
        """
        Write the geometry to a unversal file format file

        Parameters
        ----------
        filename : str
            Filename to which the geometry will be written.  If None, a
            unv data dictionary will be returned instead, similar to that
            obtained from the readunv function in sdynpy
        write_nodes : bool, optional
            If True, write the geometry's nodes to dataset 2411 in the output
            file. The default is True.
        write_coordinate_systems : True, optional
            If True, write the geometry's coordinate systems to dataset 2420
            in the output file. The default is True.
        write_tracelines : bool, optional
            If True, write the geometry's tracelines to dataset 82 in the
            output file. The default is True.
        write_elements : TYPE, optional
            If True, write the geometry's elements to dataset 2412 in the
            output file. The default is True.
        dataset_2412_kwargs : dict, optional
            Allows users to specify additional element parameters not stored
            by the Geometry.  The default is {}.
        dataset_2420_kwargs : dict, optional
            Allows users to specify additional coordinate system parameters not
            stored by the Geometry.  The default is {}.

        Returns
        -------
        unv_dict : dict
            Dictionary containing unv information, similar to that obtained from
            readunv.  Only returned if filename is None.

        """
        from ..fileio.sdynpy_uff import dataset_82, dataset_2411, dataset_2412, dataset_2420
        # Write Coordinate systems
        if write_coordinate_systems:
            if self.coordinate_system.size > 0:
                cs_unv = dataset_2420.Sdynpy_UFF_Dataset_2420(
                    cs_labels=self.coordinate_system.id.flatten(),
                    cs_types=self.coordinate_system.cs_type.flatten(),
                    cs_colors=self.coordinate_system.color.flatten(),
                    cs_names=[cs.name for index, cs in self.coordinate_system.ndenumerate()],
                    cs_matrices=self.coordinate_system.matrix.reshape(-1, 4, 3),
                    **dataset_2420_kwargs)
            else:
                cs_unv = None
        if write_nodes:
            if self.node.size > 0:
                node_unv = dataset_2411.Sdynpy_UFF_Dataset_2411(
                    self.node.id.flatten(),
                    self.node.def_cs.flatten(),
                    self.node.disp_cs.flatten(),
                    self.node.color.flatten(),
                    self.node.coordinate.reshape(-1, 3))
            else:
                node_unv = None
        if write_elements:
            if self.element.size > 0:
                if 'physical_property_table_numbers' not in dataset_2412_kwargs:
                    dataset_2412_kwargs['physical_property_table_numbers'] = [1] * self.element.size
                if 'material_property_table_numbers' not in dataset_2412_kwargs:
                    dataset_2412_kwargs['material_property_table_numbers'] = [1] * self.element.size
                if 'beam_orientations' not in dataset_2412_kwargs:
                    dataset_2412_kwargs['beam_aft_cross_section_numbers'] = [
                        None] * self.element.size
                if 'beam_aft_cross_section_numbers' not in dataset_2412_kwargs:
                    dataset_2412_kwargs['beam_orientations'] = [None] * self.element.size
                if 'beam_fore_cross_section_numbers' not in dataset_2412_kwargs:
                    dataset_2412_kwargs['beam_fore_cross_section_numbers'] = [
                        None] * self.element.size
                elem_unv = dataset_2412.Sdynpy_UFF_Dataset_2412(
                    element_labels=self.element.id.flatten(),
                    fe_descriptor_ids=self.element.type.flatten(),
                    colors=self.element.color.flatten(),
                    connectivities=[elem.connectivity for index,
                                    elem in self.element.ndenumerate()],
                    **dataset_2412_kwargs)
            else:
                elem_unv = None
        if write_tracelines:
            tl_unv = []
            for index, tl in self.traceline.ndenumerate():
                tl_unv.append(
                    dataset_82.Sdynpy_UFF_Dataset_82(
                        tl.id, tl.color, tl.description, tl.connectivity,
                    ))
        if filename is None:
            unv_dict = {}
            unv_dict[2420] = [cs_unv]
            unv_dict[2411] = [node_unv]
            unv_dict[2412] = [elem_unv]
            unv_dict[82] = tl_unv
            return unv_dict
        else:
            with open(filename, 'w') as f:
                if cs_unv is not None:
                    f.write('    -1\n')
                    f.write('  2420\n')
                    f.write(cs_unv.write_string())
                    f.write('    -1\n')
                if node_unv is not None:
                    f.write('    -1\n')
                    f.write('  2411\n')
                    f.write(node_unv.write_string())
                    f.write('    -1\n')
                if elem_unv is not None:
                    f.write('    -1\n')
                    f.write('  2412\n')
                    f.write(elem_unv.write_string())
                    f.write('    -1\n')
                for dataset in tl_unv:
                    f.write('    -1\n')
                    f.write('    82\n')
                    f.write(dataset.write_string())
                    f.write('    -1\n')


from_excel_template = Geometry.from_excel_template
write_excel_template = Geometry.write_excel_template
from_imat_struct = Geometry.from_imat_struct
from_exodus = Geometry.from_exodus
from_unv = Geometry.from_unv
from_uff = Geometry.from_uff
load = Geometry.load


class id_map:
    """Class defining mapping between two sets of id numbers"""

    def __init__(self, from_ids, to_ids):
        """
        Initializes the id map

        Parameters
        ----------
        from_ids : np.ndarray
            Id numbers to map from
        to_ids : np.ndarray
            Id numbers to map to

        Returns
        -------
        None.

        """
        self.from_ids = from_ids
        self.to_ids = to_ids
        map_dict = {f: t for f, t in zip(from_ids, to_ids)}
        map_dict[0] = 0
        self.mapper = np.vectorize(map_dict.__getitem__, otypes=['uint64'])

    def __call__(self, val):
        """
        Map id numbers

        Parameters
        ----------
        val : np.ndarray
            A set of id numbers to map.  These must all be contained in the
            id_map's from_ids array.

        Returns
        -------
        np.ndarray
            The set of id numbers from the id_map's to_ids array corresponding
            to those passed in to the function from the from_ids array.

        """
        return self.mapper(val)

    def inverse(self):
        """
        Returns an inverse map, swapping the from and to ids.
        
        If duplicate id entries exist, they will be overwritten.

        Returns
        -------
        id_map
            An id_map object with swapped from and to ids.

        """
        return id_map(self.to_ids,self.from_ids)
"""
ADVfile_manager
Author: Avi Twil

Unified file abstractions for Python with safe I/O patterns, in-memory cache,
timestamped backups, context manager safety (auto-backup & restore-on-error),
and exit-time cleanup of ephemeral backups — under a consistent API for
Text, JSON, CSV, YAML, INI, TOML, XML, and Excel files.

Highlights
----------
- Consistent CRUD: read / write / append across formats
- Unified search(): one method signature for all file types
- Safe writes: temp file + os.replace() where applicable
- Backups & restore: timestamped .bak files + optional retention by user
- Context manager: backup on enter (optional), restore on exception, clear cache on exit
- Ephemeral backups: auto-removed via an atexit hook when keep_backup=False
- Async facade: aread / awrite / aappend / asearch and async with

Optional Dependencies
---------------------
YAML  : pip install pyyaml
TOML  : Python 3.11+ (tomllib) or pip install tomli for reading; pip install tomli-w for writing
Excel : pip install openpyxl

Python Version
--------------
3.8+ recommended
"""

from __future__ import annotations

import os
import csv
import json
import glob
import atexit
import shutil
import weakref
import datetime
import itertools
from typing import (
    Any,
    Dict,
    List,
    Iterable,
    Optional,
    Generator,
    Tuple,
    Iterator,
    Sequence,
    Union,
)

# ---------- Optional deps ----------
try:
    import yaml  # pip install pyyaml
except Exception:
    yaml = None

# TOML readers/writers
try:
    import tomllib as _toml_reader  # Python 3.11+
except Exception:
    try:
        import tomli as _toml_reader  # pip install tomli
    except Exception:
        _toml_reader = None

try:
    import tomli_w as _toml_writer  # pip install tomli-w
except Exception:
    _toml_writer = None

# INI
import configparser

# XML
import xml.etree.ElementTree as ET

# Excel
try:
    import openpyxl  # pip install openpyxl
    from openpyxl import Workbook, load_workbook
except Exception:
    openpyxl = None

# Search helpers / async
import re
import asyncio


# ============================================================
# Exit-cleanup machinery (for ephemeral backups / keep_backup)
# ============================================================

_exit_cleanup_enabled: bool = True
_files_to_clear_on_exit: "weakref.WeakSet[File]" = weakref.WeakSet()  # type: ignore[name-defined]


def set_exit_cleanup(enabled: bool) -> None:
    """
    Enable or disable automatic cleanup of ephemeral backups at interpreter exit.

    Purpose
    -------
    When a file instance is configured as ephemeral (`keep_backup=False`), its
    backups are registered to be removed automatically on interpreter shutdown.
    This function toggles that global behavior.

    Parameters
    ----------
    enabled : bool
        True to enable automatic cleanup (default behavior).
        False to disable automatic cleanup.

    Returns
    -------
    None

    Raises
    ------
    None

    Examples
    --------
    set_exit_cleanup(False)  # disable auto cleanup globally
    set_exit_cleanup(True)   # enable it again
    """
    global _exit_cleanup_enabled
    _exit_cleanup_enabled = enabled


def cleanup_backups_for_all() -> int:
    """
    Immediately remove backups for all **ephemeral** file instances.

    Purpose
    -------
    Performs a manual, synchronous cleanup of all backups recorded by
    ephemeral instances (those created or used with `keep_backup=False`).

    Returns
    -------
    int
        The total number of backup files removed across all tracked instances.

    Raises
    ------
    None
        Any internal removal exceptions are swallowed to avoid breaking the caller.

    Examples
    --------
    removed = cleanup_backups_for_all()
    print("Removed", removed, "ephemeral backups")
    """
    total = 0
    for f in list(_files_to_clear_on_exit):
        try:
            total += f.clear_backups()
        except Exception:
            # Swallow exceptions to avoid breaking caller flow.
            pass
    return total


def _clear_backups_on_exit() -> None:
    """Internal atexit hook: clears ephemeral backups if exit cleanup is enabled."""
    if not _exit_cleanup_enabled:
        return
    cleanup_backups_for_all()


# Register exit hook silently
atexit.register(_clear_backups_on_exit)


# ===================
# Base class: File
# ===================

class File:
    """
    Abstract base class for file handling across multiple formats.

    Core Features
    -------------
    - Path handling & existence status
    - In-memory cache (self.content)
    - Backups (.bak) and restore
    - Context manager: optional auto-backup on enter, restore on exception,
      clear in-memory cache on exit
    - Exit-time cleanup of backups for ephemeral instances (keep_backup=False)
    - Unified search() signature (implemented per concrete subclass)

    Parameters
    ----------
    file_name : str
        The file name (with extension).
    file_path : Optional[str], default=None
        Directory path. If None, uses the current working directory.
    keep_backup : bool, default=True
        When False, this instance is considered ephemeral. Its backups will be
        automatically removed at interpreter exit (and on context exit).

    Attributes
    ----------
    name : str
        File name (as provided).
    path : str
        Directory path of the file.
    full_path : str
        Absolute/combined path to the file (path + name).
    status : bool
        True if the file currently exists on disk.
    content : Any
        In-memory cache of the file content, format-specific.

    Examples
    --------
    # Basic usage with context manager (auto-backup + restore-on-error):
    with TextFile("notes.txt", "data") as f:
        f.write("safe content")
        # if an exception happens here, the last backup is restored

    # Marking an instance as ephemeral (auto cleanup of backups on exit):
    tmp = TextFile("temp.txt", "data", keep_backup=False)
    tmp.write("short-lived content")
    # backups for 'tmp' will be auto-removed on interpreter shutdown
    """

    # -------- lifecycle --------

    def __init__(self, file_name: str, file_path: Optional[str] = None, keep_backup: bool = True):
        """
        Initialize a file abstraction.

        Parameters
        ----------
        file_name : str
            File name, including extension (e.g., "data.json").
        file_path : Optional[str], default=None
            Directory where the file resides. Created if missing.
            Defaults to the current working directory when None.
        keep_backup : bool, default=True
            If False, backups for this instance are treated as ephemeral:
            they are auto-removed on context exit and interpreter shutdown.

        Raises
        ------
        OSError
            If the target directory cannot be created due to permissions or other OS errors.
        """
        base = file_path if file_path else os.getcwd()
        self.name = file_name
        self.path = base
        self.full_path = os.path.join(base, file_name)
        if not os.path.exists(base):
            os.makedirs(base)
        self.status = os.path.exists(self.full_path)
        self.content: Any = None
        self.__keep_backup: bool = keep_backup
        if not keep_backup:
            _files_to_clear_on_exit.add(self)

    def __call__(self, keep_backup: bool = True):
        """
        Configure `keep_backup` flag fluently, especially for context use.

        Purpose
        -------
        Allows:
            with TextFile("draft.txt","data")(keep_backup=False) as f:
                f.write("temp edit")

        Parameters
        ----------
        keep_backup : bool
            True to retain backups; False to make them ephemeral.

        Returns
        -------
        File
            The same instance (enables chaining).

        Raises
        ------
        None
        """
        self.__keep_backup = keep_backup
        if not keep_backup:
            _files_to_clear_on_exit.add(self)
        return self

    def __enter__(self):
        """
        Enter the context manager.

        Behavior
        --------
        - If keep_backup=True and the file exists, a backup is taken automatically.
        - No error is raised if the file does not yet exist.

        Returns
        -------
        File
            The file instance itself.

        Raises
        ------
        None
        """
        if self.__keep_backup:
            try:
                self.backup()
            except FileNotFoundError:
                # No source yet, ignore.
                pass
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Exit the context manager.

        Behavior
        --------
        - When an exception occurs and keep_backup=True, restores the last backup.
        - When keep_backup=False, clears backups for this instance on exit.
        - Always clears in-memory cache on exit.

        Parameters
        ----------
        exc_type, exc_val, exc_tb
            Standard context manager exception triple.

        Returns
        -------
        None

        Raises
        ------
        None
        """
        if exc_type is not None and self.__keep_backup:
            try:
                self.restore()
            except FileNotFoundError:
                pass
        if not self.__keep_backup:
            self.clear_backups()
        self.clear_cache()

    # -------- abstract I/O API --------

    def read(self) -> Any:
        """
        Read file content from disk into memory (and cache it).

        Returns
        -------
        Any
            Format-specific Python object representing file content.

        Raises
        ------
        NotImplementedError
            Subclasses must implement this method.
        """
        raise NotImplementedError

    def write(self, data: Any):
        """
        Write content to disk, replacing previous data.

        Parameters
        ----------
        data : Any
            Format-specific content to persist.

        Returns
        -------
        None

        Raises
        ------
        NotImplementedError
            Subclasses must implement this method.
        """
        raise NotImplementedError

    def append(self, data: Any):
        """
        Append data to the file in a format-appropriate way.

        Parameters
        ----------
        data : Any
            Format-specific content to append.

        Returns
        -------
        None

        Raises
        ------
        NotImplementedError
            Subclasses must implement this method.
        """
        raise NotImplementedError

    # -------- cache --------

    def clear_cache(self) -> None:
        """
        Clear the in-memory cache of the file content.

        Notes
        -----
        Does not affect the file on disk. The next call to `read()` will
        reload the content from disk.

        Returns
        -------
        None

        Raises
        ------
        None
        """
        self.content = None

    # -------- backups --------

    def backup(self) -> str:
        """
        Create a timestamped backup copy under `<path>/backups`.

        Returns
        -------
        str
            Full path to the created backup file.

        Raises
        ------
        FileNotFoundError
            If the source file does not exist yet.
        OSError
            If the backup directory cannot be created or copy fails.

        Examples
        --------
        t = TextFile("file.txt","data")
        t.write("v1")
        b = t.backup()
        print("Backup stored at:", b)
        """
        if not os.path.exists(self.full_path):
            raise FileNotFoundError(f"File does not exist: {self.full_path}")

        backup_dir = os.path.join(self.path, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_name = f"{self.name}.{ts}.bak"
        backup_path = os.path.join(backup_dir, backup_name)
        shutil.copy2(self.full_path, backup_path)
        return backup_path

    def list_backups(self) -> List[str]:
        """
        List all backups for this file, sorted oldest → newest.

        Returns
        -------
        List[str]
            Absolute paths to backup files.

        Raises
        ------
        None

        Examples
        --------
        backups = t.list_backups()
        for p in backups:
            print(p)
        """
        backup_dir = os.path.join(self.path, "backups")
        pattern = os.path.join(backup_dir, f"{self.name}.*.bak")
        return sorted(glob.glob(pattern))

    def restore(self, backup_path: Optional[str] = None) -> str:
        """
        Restore the file from a backup.

        Parameters
        ----------
        backup_path : Optional[str], default=None
            Specific backup path to restore from. If None, the latest backup
            is used.

        Returns
        -------
        str
            Full path to the restored file.

        Raises
        ------
        FileNotFoundError
            If no backups are found (only when called explicitly).
        OSError
            If copying fails due to I/O errors.

        Examples
        --------
        # Restore latest:
        txt.restore()

        # Restore specific:
        backups = txt.list_backups()
        txt.restore(backups[-2])
        """
        if backup_path is None:
            backups = self.list_backups()
            if not backups:
                raise FileNotFoundError(f"No backups found for {self.name}")
            backup_path = backups[-1]
        shutil.copy2(backup_path, self.full_path)
        self.status = True
        self.clear_cache()
        return self.full_path

    def clear_backups(self) -> int:
        """
        Delete all backups created for this file.

        Returns
        -------
        int
            Number of backup files removed.

        Raises
        ------
        None
            Any per-file removal `FileNotFoundError` is ignored.

        Examples
        --------
        removed = txt.clear_backups()
        print("Removed", removed, "backup files")
        """
        count = 0
        for b in self.list_backups():
            try:
                os.remove(b)
                count += 1
            except FileNotFoundError:
                pass
        return count

    # -------- size --------

    def get_size(self) -> int:
        """
        Get the on-disk file size in bytes.

        Returns
        -------
        int
            Size in bytes; returns 0 if the file does not exist.

        Raises
        ------
        None
        """
        if os.path.exists(self.full_path):
            return os.path.getsize(self.full_path)
        return 0

    def get_size_human(self) -> str:
        """
        Get a human-readable file size string (e.g., '12.3 KB').

        Returns
        -------
        str
            Human-friendly size string.

        Raises
        ------
        None
        """
        size = self.get_size()
        units = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        s = float(size)
        while s >= 1024 and i < len(units) - 1:
            s /= 1024.0
            i += 1
        return f"{s:.1f} {units[i]}"

    # -------- search (unified signature) --------

    def search(
        self,
        pattern: Optional[str] = None,
        *,
        regex: bool = False,
        case: bool = False,
        key: Optional[str] = None,
        value: Any = None,
        columns: Optional[Sequence[str]] = None,
        tag: Optional[str] = None,
        attr: Optional[Dict[str, str]] = None,
        sheet: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Iterator[Dict[str, Any]]:
        """
        Unified search entry point. Concrete subclasses implement this.

        Parameters
        ----------
        pattern : Optional[str]
            Substring or regex pattern to search for.
        regex : bool, default=False
            Interpret `pattern` as a regular expression.
        case : bool, default=False
            Case-sensitive search when True.
        key : Optional[str]
            Key filter for mapping-like formats (JSON/YAML/INI/TOML).
        value : Any
            Value filter for mapping/list-like formats (exact match).
        columns : Optional[Sequence[str]]
            Limit CSV/Excel search to specific header names.
        tag : Optional[str]
            XML element tag filter.
        attr : Optional[Dict[str, str]]
            XML attribute equality filter (all must match).
        sheet : Optional[str]
            Excel sheet name filter.
        limit : Optional[int]
            Maximum number of hits to yield.

        Yields
        ------
        dict
            A standardized hit object (subset keys may be None):
            {'path','value','line','row','col','sheet','context', ...}

        Raises
        ------
        NotImplementedError
            Subclasses must implement their specific matching logic.
        """
        raise NotImplementedError


# =====================================
# Mixins: search helpers and async API
# =====================================

class _SearchMixin:
    """
    Internal helper mixin used by all concrete classes to implement search().

    Not part of the public API; users should not depend on this class directly.
    """

    def _match_text(self, text: str, pattern: str, *, regex: bool = False, case: bool = False) -> bool:
        """
        Check whether `text` matches `pattern` according to flags.

        Parameters
        ----------
        text : str
            Text to test (None-safe; None never matches).
        pattern : str
            Pattern or substring.
        regex : bool
            Use regular expression when True.
        case : bool
            Case-sensitive match when True.

        Returns
        -------
        bool

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.
        """
        if text is None:
            return False
        if regex:
            flags = 0 if case else re.IGNORECASE
            return re.search(pattern, text, flags) is not None
        return (pattern in text) if case else (pattern.lower() in str(text).lower())

    def _maybe_limit(self, it: Iterator[dict], limit: Optional[int]) -> Iterator[dict]:
        """
        Yield up to `limit` hits from iterator `it` (or unlimited if None).

        Parameters
        ----------
        it : Iterator[dict]
            Source iterator of hits.
        limit : Optional[int]
            Maximum number of yielded items; None for unlimited.

        Returns
        -------
        Iterator[dict]
            Generator honoring the limit.
        """
        if limit is None:
            yield from it
            return
        n = 0
        for x in it:
            yield x
            n += 1
            if n >= limit:
                break


class _AsyncMixin:
    """
    Async facade wrapping blocking operations with asyncio.to_thread().

    Methods
    -------
    aread(*args, **kwargs) -> Any
        Async variant of read().
    awrite(*args, **kwargs) -> Any
        Async variant of write().
    aappend(*args, **kwargs) -> Any
        Async variant of append().
    asearch(*args, **kwargs) -> List[dict]
        Async variant of search(); collects generator to a list.
    async with
        Async context manager mirrors sync __enter__/__exit__ semantics.

    Notes
    -----
    All methods schedule the corresponding sync method in a thread pool,
    preserving the semantics (including exceptions).
    """

    async def aread(self, *args, **kwargs):
        """Async variant of read(); see the sync `read()` docstring for semantics."""
        return await asyncio.to_thread(self.read, *args, **kwargs)

    async def awrite(self, *args, **kwargs):
        """Async variant of write(); see the sync `write()` docstring for semantics."""
        return await asyncio.to_thread(self.write, *args, **kwargs)

    async def aappend(self, *args, **kwargs):
        """Async variant of append(); see the sync `append()` docstring for semantics."""
        return await asyncio.to_thread(self.append, *args, **kwargs)

    async def asearch(self, *args, **kwargs):
        """
        Async variant of search(); collects the generator to a list.

        Returns
        -------
        list[dict]
            List of hits (eagerly collected).
        """
        def _collect():
            return list(self.search(*args, **kwargs))
        return await asyncio.to_thread(_collect)

    async def __aenter__(self):
        """Enter async context; mirrors sync __enter__()."""
        self.__enter__()
        return self

    async def __aexit__(self, exc_type, exc, tb):
        """Exit async context; mirrors sync __exit__()."""
        self.__exit__(exc_type, exc, tb)
        return False


# ================
# TextFile
# ================

class TextFile(_SearchMixin, File):
    """
    Plain text file handler (UTF-8).

    Extras
    ------
    - lines() -> generator of (line_number, line_text)
    - read_line(n) -> direct 1-based line access
    - search() -> line-based lookup by substring/regex, with case sensitivity

    Notes
    -----
    append() inserts a newline when the file is non-empty to preserve line semantics.
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Create a text file abstraction.

        Parameters
        ----------
        file_name : str
            Target text file name (e.g., "notes.txt").
        file_path : str
            Directory path; created if missing.

        Raises
        ------
        OSError
            If directory creation fails.
        """
        super().__init__(file_name, file_path)
        if self.status:
            self.read()

    def write(self, data: str) -> None:
        """
        Write text, replacing all existing content.

        Parameters
        ----------
        data : str
            Text to write.

        Returns
        -------
        None

        Raises
        ------
        OSError
            If the file cannot be written due to I/O or permission errors.

        Examples
        --------
        txt = TextFile("a.txt", "data")
        txt.write("Hello\\nWorld")
        """
        with open(self.full_path, "wt", encoding="utf-8") as f:
            f.write(data)
        self.content = data
        self.status = True

    def read(self) -> str:
        """
        Read full file content as a string (cached after first read).

        Returns
        -------
        str
            File content.

        Raises
        ------
        FileNotFoundError
            If the file does not exist.
        OSError
            On I/O read errors.

        Examples
        --------
        txt = TextFile("a.txt", "data")
        txt.write("Hello")
        content = txt.read()
        """
        if self.content is None:
            with open(self.full_path, "rt", encoding="utf-8") as f:
                self.content = f.read()
        return self.content

    def append(self, data: str) -> None:
        """
        Append text to the end of file.

        Behavior
        --------
        A leading newline is added if the file already has content to keep
        line boundaries intact.

        Parameters
        ----------
        data : str
            Text to append.

        Returns
        -------
        None

        Raises
        ------
        OSError
            On write errors (permissions, disk full, etc.).

        Examples
        --------
        txt = TextFile("log.txt", "data")
        txt.write("first")
        txt.append("second")
        """
        with open(self.full_path, "at", encoding="utf-8") as f:
            if self.status and os.path.getsize(self.full_path) > 0:
                f.write("\n" + data)
            else:
                f.write(data)
        self.content = (self.content + "\n" + data) if self.content else data
        self.status = True

    def lines(self) -> Generator[Tuple[int, str], None, None]:
        """
        Iterate over file lines with line numbers.

        Yields
        ------
        (int, str)
            Tuple of (line_number, line_text), 1-based index.

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        OSError
            On I/O errors while reading.

        Examples
        --------
        for i, line in TextFile("a.txt", "data").lines():
            print(i, line)
        """
        with open(self.full_path, "rt", encoding="utf-8") as f:
            for i, line in enumerate(f, start=1):
                yield i, line.rstrip("\n")

    def read_line(self, line_number: int) -> str:
        """
        Read a specific line by its 1-based index.

        Parameters
        ----------
        line_number : int
            1-based line index.

        Returns
        -------
        str
            Line content without trailing newline.

        Raises
        ------
        IndexError
            If the requested line does not exist.
        FileNotFoundError
            If the file is missing.
        OSError
            On I/O errors.

        Examples
        --------
        txt = TextFile("a.txt", "data")
        txt.write("first\\nsecond\\nthird")
        second = txt.read_line(2)
        """
        with open(self.full_path, "rt", encoding="utf-8") as f:
            line = next(itertools.islice(f, line_number - 1, line_number), None)
            if line is None:
                raise IndexError(f"Line {line_number} does not exist in {self.full_path}")
            return line.rstrip("\n")

    def search(self, pattern: Optional[str] = None, *, regex: bool = False, case: bool = False,
               limit: Optional[int] = None, **_: Any) -> Iterator[Dict[str, Any]]:
        """
        Search lines by substring or regex.

        Parameters
        ----------
        pattern : Optional[str]
            Text/regex to find (required for text search).
        regex : bool, default=False
            Interpret `pattern` as regex.
        case : bool, default=False
            Case-sensitive matching when True.
        limit : Optional[int]
            Max number of results.

        Yields
        ------
        dict
            Standardized hit object per matching line:
            {'path','value','line','row','col','sheet','context'}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        txt = TextFile("grep.txt", "data")
        txt.write("Alpha\\nbeta\\nALPHA BETA\\nGamma")
        hits = list(txt.search(pattern="alpha"))
        """
        if pattern is None:
            return iter(())
        def _iter():
            for ln, line in self.lines():
                if self._match_text(line, pattern, regex=regex, case=case):
                    yield {
                        "path": f"{self.full_path}:line[{ln}]",
                        "value": line,
                        "line": ln,
                        "row": None,
                        "col": None,
                        "sheet": None,
                        "context": line,
                    }
        return self._maybe_limit(_iter(), limit)


# ================
# JsonFile
# ================

def _walk_jsonlike(obj, path=""):
    """
    Recursively yield (path, value) pairs across dict/list JSON-like structures.

    Paths
    -----
    Creates a dotted path with 1-based indices for lists:
    'users[1].name', 'config.flags.debug', etc.

    Parameters
    ----------
    obj : Any
        JSON-like structure (dict/list/scalars).
    path : str
        Current path prefix (used internally during recursion).

    Yields
    ------
    tuple[str, Any]
        (path, value) for each leaf and intermediate node.

    Raises
    ------
    None
    """
    if isinstance(obj, dict):
        for k, v in obj.items():
            p = f"{path}.{k}" if path else str(k)
            yield p, v
            yield from _walk_jsonlike(v, p)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            p = f"{path}[{i+1}]"
            yield p, v
            yield from _walk_jsonlike(v, p)


class JsonFile(_SearchMixin, File):
    """
    JSON handler (dict or list root).

    Extras
    ------
    - append(): list -> append/extend; dict -> shallow update
    - get_item(index_or_key): 1-based index for lists, key for dict
    - items(): iterate (index,item) for lists or (key,value) for dicts
    - search(): by key/value/pattern over the entire JSON tree
    """

    def __init__(self, file_name: str, file_path: str, *, indent: int = 2):
        """
        Create a JSON file abstraction.

        Parameters
        ----------
        file_name : str
            File name (e.g., "data.json").
        file_path : str
            Directory path; created if missing.
        indent : int, default=2
            Indentation level for JSON pretty printing in write().

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        self.indent = indent
        if self.status:
            self.read()

    def write(self, data: Any) -> None:
        """
        Write a Python object (dict/list) as JSON.

        Parameters
        ----------
        data : Any
            JSON-serializable object.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If the object is not JSON serializable.
        OSError
            If the file cannot be written.

        Examples
        --------
        j = JsonFile("d.json", "data")
        j.write({"users":[{"id":1}]})
        """
        with open(self.full_path, "wt", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=self.indent)
        self.content = data
        self.status = True

    def read(self) -> Any:
        """
        Read JSON content from disk (cached after first read).

        Returns
        -------
        Any
            Python object (dict or list).

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        json.JSONDecodeError
            If the file content is not valid JSON.
        OSError
            On read I/O errors.

        Examples
        --------
        obj = JsonFile("d.json","data").read()
        """
        if self.content is None:
            with open(self.full_path, "rt", encoding="utf-8") as f:
                self.content = json.load(f)
        return self.content

    def append(self, data: Any) -> None:
        """
        Append content to a JSON file.

        Behavior
        --------
        - If root is list: append element or extend with an iterable.
        - If root is dict: shallow update with provided dict keys.

        Parameters
        ----------
        data : Any
            Data to append (dict, list, or element).

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If root is dict and `data` is not dict.
            If root is neither list nor dict.
        OSError
            On write errors.

        Examples
        --------
        j = JsonFile("d.json", "data")
        j.write({"users":[{"id":1}]})
        j.append({"active": True})
        """
        if not self.status:
            self.write(data)
            return
        cur = self.read()
        if isinstance(cur, list):
            if isinstance(data, Iterable) and not isinstance(data, (str, bytes, dict)):
                cur.extend(list(data))
            else:
                cur.append(data)
        elif isinstance(cur, dict):
            if not isinstance(data, dict):
                raise TypeError("append on dict-backed JSON expects dict")
            cur.update(data)
        else:
            raise TypeError("JSON root must be list or dict to support append")
        self.write(cur)

    def get_item(self, index_or_key: Any) -> Any:
        """
        Access an element by 1-based list index or dict key.

        Parameters
        ----------
        index_or_key : int | str
            1-based index for list root; string key for dict root.

        Returns
        -------
        Any

        Raises
        ------
        TypeError
            If root type doesn't match the accessor (e.g., list root but key given).
        IndexError
            If list index is out of range.
        KeyError
            If dict key does not exist.
        OSError
            On read I/O errors.

        Examples
        --------
        jl = JsonFile("l.json","data")
        jl.write([{"id":10},{"id":20}])
        second = jl.get_item(2)

        jd = JsonFile("d.json","data")
        jd.write({"x":42})
        value = jd.get_item("x")
        """
        data = self.read()
        if isinstance(data, list):
            if not isinstance(index_or_key, int) or index_or_key < 1:
                raise TypeError("For list-backed JSON, provide a 1-based integer index")
            idx = index_or_key - 1
            if idx >= len(data):
                raise IndexError(f"Index {index_or_key} out of range")
            return data[idx]
        elif isinstance(data, dict):
            if index_or_key not in data:
                raise KeyError(f"Key {index_or_key!r} not found")
            return data[index_or_key]
        else:
            raise TypeError("JSON root must be list or dict")

    def items(self) -> Iterable:
        """
        Iterate over items in the root object.

        Yields
        ------
        tuple
            - list root: (index (1-based), item)
            - dict root: (key, value)

        Raises
        ------
        TypeError
            If JSON root is neither list nor dict.

        Examples
        --------
        for k, v in JsonFile("d.json","data").items():
            print(k, v)
        """
        data = self.read()
        if isinstance(data, list):
            for i, item in enumerate(data, start=1):
                yield i, item
        elif isinstance(data, dict):
            for k, v in data.items():
                yield k, v
        else:
            raise TypeError("JSON root must be list or dict")

    def search(
        self,
        pattern: str | None = None,
        *,
        key: str | None = None,
        value: Any = None,
        regex: bool = False,
        case: bool = False,
        limit: int | None = None,
        **_
    ) -> Iterator[dict]:
        """
        Search JSON content (keys, values, or both).

        Matching Rules
        --------------
        - If 'key' is provided: inspect only that key within dicts and apply
          `pattern`/`value` against its value.
        - If 'key' is None and 'pattern' is provided: match dict keys by name
          and/or any scalar value (stringified).
        - If 'value' is provided: does an exact equality check (post-parsing).

        Parameters
        ----------
        pattern : str | None
            Substring/regex to match (stringified) values or key names.
        key : str | None
            Dict key to focus on (case-insensitive unless `case=True`).
        value : Any
            Exact value to match.
        regex : bool, default=False
            Whether `pattern` is a regex.
        case : bool, default=False
            Case-sensitive comparisons when True.
        limit : int | None
            Stop after this many hits (None for unlimited).

        Yields
        ------
        dict
            JSON hit schema:
            {"path": full_path, "key": <key or None>, "value": <matched value>}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        j = JsonFile("u.json","data")
        j.write({"users":[{"id":1,"name":"Avi"},{"id":2,"name":"Dana"}], "active":True})
        hits = list(j.search(key="name", pattern="^A", regex=True))
        """
        data = self.read()
        count = 0

        def match_text(text: str) -> bool:
            if pattern is None:
                return True
            if regex:
                flags = 0 if case else re.IGNORECASE
                return re.search(pattern, text, flags) is not None
            else:
                hay = text if case else text.lower()
                needle = pattern if case else pattern.lower()
                return needle in hay

        def key_equal(k: str, target: str) -> bool:
            return k == target if case else k.lower() == target.lower()

        def walk(node):
            nonlocal count

            if isinstance(node, dict):
                # If key is provided, test only that key within this dict
                if key is not None:
                    for k, v in node.items():
                        if not key_equal(k, key):
                            continue
                        if value is not None and v != value:
                            continue
                        if pattern is not None and not match_text(str(v)):
                            continue
                        yield {"path": self.full_path, "key": k, "value": v}
                        count += 1
                        if limit and count >= limit:
                            return
                else:
                    # No key filter: keys and values both participate
                    for k, v in node.items():
                        if value is not None and v == value:
                            yield {"path": self.full_path, "key": k, "value": v}
                            count += 1
                            if limit and count >= limit:
                                return
                        if pattern is not None and match_text(str(k)):
                            yield {"path": self.full_path, "key": k, "value": v}
                            count += 1
                            if limit and count >= limit:
                                return
                        for hit in walk(v):
                            yield hit
                            if limit and count >= limit:
                                return

            elif isinstance(node, list):
                for _, v in enumerate(node):
                    for hit in walk(v):
                        yield hit
                        if limit and count >= limit:
                            return
            else:
                # scalar node
                if key is None:
                    if value is not None and node == value:
                        yield {"path": self.full_path, "value": node}
                        count += 1
                        return
                    if pattern is not None and match_text(str(node)):
                        yield {"path": self.full_path, "value": node}
                        count += 1

        for h in walk(data):
            yield h
            if limit and count >= limit:
                break


# ================
# CsvFile
# ================

class CsvFile(_SearchMixin, File):
    """
    CSV handler using DictReader/DictWriter.

    Extras
    ------
    - write(data, fieldnames=None): control header order; else inferred
    - read_row(n): 1-based row access
    - rows(): generator of (row_no, row_dict)
    - search(): match by pattern/value across columns (optionally restricted)
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Create a CSV file abstraction.

        Parameters
        ----------
        file_name : str
            CSV file name (e.g., "table.csv").
        file_path : str
            Directory path; created if missing.

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        self._fieldnames: Optional[List[str]] = None
        if self.status:
            self.read()

    def _infer_fieldnames(self, rows: Iterable[Dict[str, Any]]) -> List[str]:
        """
        Infer CSV header fieldnames from a list of dicts.

        Parameters
        ----------
        rows : Iterable[Dict[str, Any]]
            Data rows to infer from.

        Returns
        -------
        List[str]
            Fieldnames in first-seen order.

        Raises
        ------
        ValueError
            If rows are empty and no fieldnames can be inferred.
        """
        fields: List[str] = []
        seen = set()
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    fields.append(k)
        if not fields:
            raise ValueError("Cannot infer CSV header from empty data")
        return fields

    def write(self, data: Iterable[Dict[str, Any]], fieldnames: Optional[List[str]] = None) -> None:
        """
        Write rows to CSV, replacing existing content.

        Parameters
        ----------
        data : Iterable[Dict[str, Any]]
            Rows of {column: value}.
        fieldnames : Optional[List[str]]
            Header order. If None, it will be inferred from data.

        Returns
        -------
        None

        Raises
        ------
        ValueError
            If `data` is empty and `fieldnames` not provided.
        OSError
            On write I/O errors.

        Examples
        --------
        c = CsvFile("t.csv","data")
        c.write([{"name":"Avi","age":30},{"name":"Dana","age":25}])
        """
        rows = list(data)
        fieldnames = fieldnames or self._infer_fieldnames(rows)
        with open(self.full_path, "wt", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                writer.writerow({k: r.get(k, "") for k in fieldnames})
        self._fieldnames = fieldnames
        self.content = rows
        self.status = True

    def read(self) -> List[Dict[str, str]]:
        """
        Read CSV into a list of row dicts (values as strings).

        Returns
        -------
        List[Dict[str,str]]

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        OSError
            On read I/O errors.

        Examples
        --------
        rows = CsvFile("t.csv","data").read()
        """
        if self.content is None:
            with open(self.full_path, "rt", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                self._fieldnames = reader.fieldnames or []
                self.content = list(reader)
        return self.content

    def append(self, data: Any) -> None:
        """
        Append row(s) to CSV.

        Parameters
        ----------
        data : dict | Iterable[dict]
            Row(s) to append.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `data` is neither a dict nor an iterable of dicts.
        OSError
            On write I/O errors.

        Examples
        --------
        c = CsvFile("t.csv","data")
        c.write([{"name":"Avi","age":30}])
        c.append({"name":"Dana","age":25})
        c.append([{"name":"Noa","age":21},{"name":"Lior","age":28}])
        """
        if isinstance(data, dict):
            rows = [data]
        elif isinstance(data, Iterable) and not isinstance(data, (str, bytes)):
            rows = list(data)
            if rows and not isinstance(rows[0], dict):
                raise TypeError("append expects dict or iterable of dicts")
        else:
            raise TypeError("append expects dict or iterable of dicts")

        file_exists = self.status and os.path.exists(self.full_path) and os.path.getsize(self.full_path) > 0
        if file_exists and self._fieldnames:
            fieldnames = self._fieldnames
        else:
            fieldnames = self._infer_fieldnames(rows)

        with open(self.full_path, "at", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            for r in rows:
                writer.writerow({k: r.get(k, "") for k in fieldnames})

        if self.content is None:
            self.content = []
        self.content.extend([{k: str(r.get(k, "")) for k in fieldnames} for r in rows])
        self._fieldnames = fieldnames
        self.status = True

    def read_row(self, row_number: int) -> Dict[str, str]:
        """
        Read a specific CSV row by its 1-based index (excluding header).

        Parameters
        ----------
        row_number : int
            1-based row number.

        Returns
        -------
        Dict[str,str]
            Row content as {column: value}.

        Raises
        ------
        IndexError
            If the row does not exist.
        FileNotFoundError
            If the file is missing.
        OSError
            On read I/O errors.

        Examples
        --------
        row2 = CsvFile("t.csv","data").read_row(2)
        """
        with open(self.full_path, "rt", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, start=1):
                if i == row_number:
                    return row
        raise IndexError(f"Row {row_number} does not exist in {self.full_path}")

    def rows(self) -> Generator[Tuple[int, Dict[str, str]], None, None]:
        """
        Iterate over CSV rows lazily.

        Yields
        ------
        (int, Dict[str,str])
            Tuple of (row_number, row_dict).

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        OSError
            On read I/O errors.

        Examples
        --------
        for i, row in CsvFile("t.csv","data").rows():
            print(i, row)
        """
        with open(self.full_path, "rt", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, start=1):
                yield i, row

    def search(self, pattern: Optional[str] = None, *, regex: bool = False, case: bool = False,
               columns: Optional[Sequence[str]] = None, value: Any = None,
               limit: Optional[int] = None, **_: Any) -> Iterator[Dict[str, Any]]:
        """
        Search CSV cells by pattern or exact value.

        Parameters
        ----------
        pattern : Optional[str]
            Substring/regex to match in cell text.
        regex : bool, default=False
            Interpret `pattern` as regex.
        case : bool, default=False
            Case-sensitive match when True.
        columns : Optional[Sequence[str]]
            Restrict search to these headers; default: all.
        value : Any
            Exact value match (stringified for comparison).
        limit : Optional[int]
            Max number of hits.

        Yields
        ------
        dict
            Hit schema:
            {'path','value','line',None,'row','col','sheet',None,'context'}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        hits = list(CsvFile("t.csv","data").search(pattern="Avi", columns=["name"]))
        """
        def _iter():
            for r, row in self.rows():
                cols = columns or row.keys()
                for col in cols:
                    cell = row.get(col, "")
                    ok = True
                    if value is not None:
                        ok = ok and (cell == str(value))
                    if pattern is not None:
                        ok = ok and self._match_text(str(cell), pattern, regex=regex, case=case)
                    if ok:
                        yield {
                            "path": f"{self.full_path}:row[{r}].{col}",
                            "value": cell,
                            "line": None, "row": r, "col": col, "sheet": None,
                            "context": str(row),
                        }
        return self._maybe_limit(_iter(), limit)


# ================
# YamlFile
# ================

class YamlFile(_SearchMixin, File):
    """
    YAML handler (requires PyYAML).

    Extras
    ------
    - append(): list -> append/extend; dict -> shallow update
    - get_item(index_or_key), items()
    - search(): like JSON tree search
    """

    def __init__(self, file_name: str, file_path: str, *, sort_keys: bool = False):
        """
        Create a YAML file abstraction.

        Parameters
        ----------
        file_name : str
            YAML file name (e.g., "config.yaml").
        file_path : str
            Directory path; created if missing.
        sort_keys : bool, default=False
            Whether to sort keys when dumping.

        Raises
        ------
        ImportError
            If PyYAML is not installed.
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        if yaml is None:
            raise ImportError("PyYAML not installed. Install with: pip install pyyaml")
        self.sort_keys = sort_keys
        if self.status:
            self.read()

    def write(self, data: Any) -> None:
        """
        Write a Python object as YAML.

        Parameters
        ----------
        data : Any
            YAML-serializable object (dict/list/scalars).

        Returns
        -------
        None

        Raises
        ------
        ImportError
            If PyYAML is not installed.
        yaml.YAMLError
            If serialization fails.
        OSError
            On write errors.

        Examples
        --------
        y = YamlFile("c.yaml","data")
        y.write({"app":{"name":"demo"}, "features":["a"]})
        """
        with open(self.full_path, "wt", encoding="utf-8") as f:
            yaml.safe_dump(data, f, sort_keys=self.sort_keys, allow_unicode=True)
        self.content = data
        self.status = True

    def read(self) -> Any:
        """
        Read YAML from disk (cached after first read).

        Returns
        -------
        Any
            Python object (dict/list/...) — empty dict if file is empty.

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        yaml.YAMLError
            If YAML is invalid.
        OSError
            On read errors.

        Examples
        --------
        y = YamlFile("c.yaml","data")
        obj = y.read()
        """
        if self.content is None:
            with open(self.full_path, "rt", encoding="utf-8") as f:
                loaded = yaml.safe_load(f)
            self.content = loaded if loaded is not None else {}
        return self.content

    def append(self, data: Any) -> None:
        """
        Append content to a YAML file.

        Behavior
        --------
        - If root is list: append/extend.
        - If root is dict: shallow update.

        Parameters
        ----------
        data : Any
            Data to append.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If root is dict and data is not dict.
        OSError
            On write errors.

        Examples
        --------
        y = YamlFile("c.yaml","data")
        y.write({"features":["a"]})
        y.append({"features":["b"]})
        """
        if not self.status:
            self.write(data)
            return
        current = self.read()
        if isinstance(current, list):
            if isinstance(data, Iterable) and not isinstance(data, (str, bytes, dict)):
                current.extend(list(data))
            else:
                current.append(data)
        elif isinstance(current, dict):
            if not isinstance(data, dict):
                raise TypeError("append on dict-backed YAML expects dict")
            current.update(data)
        else:
            raise TypeError("YAML root must be list or dict to support append")
        self.write(current)

    def get_item(self, index_or_key: Any) -> Any:
        """
        Access a YAML element by list index or dict key.

        Parameters
        ----------
        index_or_key : int | str
            1-based index for list; key for dict.

        Returns
        -------
        Any

        Raises
        ------
        TypeError
            If wrong accessor type for the current root type.
        IndexError
            If list index out of range.
        KeyError
            If dict key missing.

        Examples
        --------
        y = YamlFile("c.yaml","data")
        y.write({"app":{"name":"demo"}})
        app = y.get_item("app")
        """
        data = self.read()
        if isinstance(data, list):
            if not isinstance(index_or_key, int) or index_or_key < 1:
                raise TypeError("For list-backed YAML, provide a 1-based integer index")
            idx = index_or_key - 1
            if idx >= len(data):
                raise IndexError(f"Index {index_or_key} out of range")
            return data[idx]
        elif isinstance(data, dict):
            if index_or_key not in data:
                raise KeyError(f"Key {index_or_key!r} not found")
            return data[index_or_key]
        else:
            raise TypeError("YAML root must be list or dict")

    def items(self) -> Iterable:
        """
        Iterate over top-level YAML items.

        Yields
        ------
        tuple
            (index, value) for list root; (key, value) for dict root.

        Raises
        ------
        TypeError
            If YAML root is neither list nor dict.

        Examples
        --------
        for k, v in YamlFile("c.yaml","data").items():
            print(k, v)
        """
        data = self.read()
        if isinstance(data, list):
            for i, item in enumerate(data, start=1):
                yield i, item
        elif isinstance(data, dict):
            for k, v in data.items():
                yield k, v
        else:
            raise TypeError("YAML root must be list or dict to iterate")

    def search(self, pattern: Optional[str] = None, *, regex: bool = False, case: bool = False,
               key: Optional[str] = None, value: Any = None, limit: Optional[int] = None, **_: Any
               ) -> Iterator[Dict[str, Any]]:
        """
        Search through YAML structure by key/value/text (JSON-like rules).

        Parameters
        ----------
        pattern : Optional[str]
            Substring/regex to match stringified values.
        regex : bool, default=False
            Whether `pattern` is regex.
        case : bool, default=False
            Case-sensitive matching when True.
        key : Optional[str]
            Dict key name to focus on (exact match).
        value : Any
            Exact value to match.
        limit : Optional[int]
            Maximum results.

        Yields
        ------
        dict
            {'path','value','line','row','col','sheet','context'}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        hits = list(YamlFile("c.yaml","data").search(pattern="demo"))
        """
        data = self.read()
        def _iter():
            for p, v in _walk_jsonlike(data):
                k = p.split(".")[-1]
                ok = True
                if key is not None:
                    ok = ok and (str(k) == key)
                if value is not None:
                    ok = ok and (v == value)
                if pattern is not None:
                    ok = ok and self._match_text(str(v), pattern, regex=regex, case=case)
                if ok:
                    yield {
                        "path": f"{self.full_path}:{p}",
                        "value": v,
                        "line": None, "row": None, "col": None, "sheet": None,
                        "context": str(v)[:200],
                    }
        return self._maybe_limit(_iter(), limit)


# ================
# IniFile
# ================

class IniFile(_SearchMixin, File):
    """
    INI handler using configparser.

    read()  -> dict (lowercased sections/options)
    write() -> expects dict-like: {section: {key: value}}
    append()-> merges sections/keys (shallow)
    search()-> match by key/value/pattern

    Notes
    -----
    configparser stores values as strings; this class keeps that behavior.
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Create an INI file abstraction.

        Parameters
        ----------
        file_name : str
            INI file name (e.g., "settings.ini").
        file_path : str
            Directory path; created if missing.

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        self._cfg = configparser.ConfigParser()
        if self.status:
            self.read()

    def read(self) -> dict:
        """
        Read INI file, returning a plain dict with lowercase section/option names.

        Returns
        -------
        dict
            Structure: {section_lc: {option_lc: value_str}}

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        OSError
            On read errors.
        configparser.Error
            For parse errors.

        Examples
        --------
        cfg = IniFile("settings.ini","data").read()
        host = cfg["server"]["host"]
        """
        if self.content is None:
            cfg = configparser.ConfigParser()
            cfg.read(self.full_path, encoding="utf-8")
            data = {}
            for sect in cfg.sections():
                sect_l = sect.lower()
                data[sect_l] = {}
                for k, v in cfg.items(sect):
                    data[sect_l][k.lower()] = v
            self.content = data
        return self.content

    def write(self, data: Dict[str, Dict[str, Any]]) -> None:
        """
        Write INI settings from a nested dict.

        Parameters
        ----------
        data : Dict[str, Dict[str, Any]]
            Mapping of sections to key/value mappings.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `data` is not dict-of-dicts.
        OSError
            On write errors.

        Examples
        --------
        ini = IniFile("settings.ini","data")
        ini.write({"server": {"host": "127.0.0.1", "port": "8000"}})
        """
        if not isinstance(data, dict):
            raise TypeError("INI write expects a dict of sections")
        cfg = configparser.ConfigParser()
        for sect, mapping in data.items():
            if not isinstance(mapping, dict):
                raise TypeError("INI write expects mapping per section")
            cfg[sect] = {str(k): str(v) for k, v in mapping.items()}
        with open(self.full_path, "w", encoding="utf-8") as f:
            cfg.write(f)
        self.content = None
        self.status = True

    def append(self, data: Dict[str, Dict[str, Any]]) -> None:
        """
        Merge additional sections/keys (shallow) into the INI content.

        Parameters
        ----------
        data : Dict[str, Dict[str, Any]]
            Mapping of {section: {key: value}}. Sections/keys are normalized
            to lowercase in-memory for consistent access.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If `data` is not dict-of-dicts.
        OSError
            On write errors.

        Examples
        --------
        ini = IniFile("settings.ini","data")
        ini.write({"server":{"host":"127.0.0.1","port":"8000"}})
        ini.append({"server":{"debug":"true"},"auth":{"enabled":"yes"}})
        """
        if not isinstance(data, dict):
            raise TypeError("INI append expects a dict of sections")
        current = self.read()  # dict with lowercase section/keys
        merged: Dict[str, Dict[str, str]] = {s: {k: str(v) for k, v in kv.items()} for s, kv in current.items()}

        for sec, kv in (data or {}).items():
            if not isinstance(kv, dict):
                raise TypeError("Each section must map to a dict of options")
            sec_l = sec.lower()
            if sec_l not in merged:
                merged[sec_l] = {}
            for k, v in (kv or {}).items():
                merged[sec_l][k.lower()] = str(v)

        # Persist via write() (dict -> ConfigParser -> file)
        self.write(merged)

    def search(
            self,
            pattern: str | None = None,
            *,
            regex: bool = False,
            case: bool = False,
            key: str | None = None,
            value: Any = None,
            columns: None = None,  # kept for unified signature
            tag: None = None,
            attr: None = None,
            sheet: None = None,
            limit: int | None = None,
    ) -> Iterator[dict]:
        """
        Search INI contents (case-normalized in-memory).

        Matching
        --------
        - `key` filters by option name (case-sensitive only if `case=True`).
        - `value` matches option value (string equality).
        - `pattern` applies to the value, and to the key name if `key` not provided.

        Parameters
        ----------
        pattern : str | None
            Substring/regex for values (and keys when key=None).
        regex : bool, default=False
            Treat pattern as regex.
        case : bool, default=False
            Case-sensitive match when True.
        key : str | None
            Option name filter.
        value : Any
            Exact value to match (stringified comparison).
        limit : int | None
            Max number of results.

        Yields
        ------
        dict
            {'path','section','key','value'}

        Raises
        ------
        re.error
            If `regex=True` and pattern invalid.

        Examples
        --------
        hits = list(IniFile("settings.ini","data").search(key="port"))
        """
        data = self.read()  # dict: {section_lc: {option_lc: value_str}}
        count = 0

        def _match_text(text: str) -> bool:
            if pattern is None:
                return True
            if regex:
                flags = 0 if case else re.IGNORECASE
                return re.search(pattern, text, flags) is not None
            else:
                hay = text if case else text.lower()
                needle = pattern if case else pattern.lower()
                return needle in hay

        def _key_equal(k: str, target: str) -> bool:
            return (k == target) if case else (k.lower() == target.lower())

        for sect, mapping in data.items():
            for k, v in mapping.items():
                if key is not None and not _key_equal(k, key):
                    continue
                if value is not None and str(v) != str(value):
                    continue
                if pattern is not None:
                    if not (_match_text(str(v)) or (key is None and _match_text(str(k)))):
                        continue
                yield {
                    "path": self.full_path,
                    "section": sect,
                    "key": k,
                    "value": v,
                }
                count += 1
                if limit is not None and count >= limit:
                    return


# ================
# TomlFile
# ================

def _deep_merge_dict(dst: dict, src: dict) -> dict:
    """
    Deep-merge `src` into `dst` in place and return `dst`.

    Rules
    -----
    For overlapping keys:
      - If both values are dicts -> recurse (deep merge).
      - Else -> `src` overwrites `dst`.

    Parameters
    ----------
    dst : dict
        Destination dict (modified in place).
    src : dict
        Source dict merged into destination.

    Returns
    -------
    dict
        The modified destination dictionary.

    Raises
    ------
    None
    """
    for k, v in src.items():
        if isinstance(v, dict) and isinstance(dst.get(k), dict):
            _deep_merge_dict(dst[k], v)
        else:
            dst[k] = v
    return dst

class TomlFile(_SearchMixin, File):
    """
    TOML handler.

    read()  -> dict (parsed)
    write() -> dict (requires tomli_w)
    append()-> deep merge of dicts
    search()-> like JSON/YAML
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Create a TOML file abstraction.

        Parameters
        ----------
        file_name : str
            TOML file name (e.g., "cfg.toml").
        file_path : str
            Directory path; created if missing.

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        if self.status:
            self.read()

    def read(self) -> Dict[str, Any]:
        """
        Read TOML into a Python dict.

        Returns
        -------
        Dict[str, Any]

        Raises
        ------
        ImportError
            If no TOML reader is available (needs Python 3.11+ `tomllib` or `tomli`).
        FileNotFoundError
            If file missing.
        OSError
            On read errors.
        tomli.TOMLDecodeError / tomllib.TOMLDecodeError
            If TOML is invalid.

        Examples
        --------
        cfg = TomlFile("cfg.toml","data").read()
        """
        if _toml_reader is None:
            raise ImportError("TOML reading requires Python 3.11+ (tomllib) or 'tomli'")
        with open(self.full_path, "rb") as f:
            data = _toml_reader.load(f)
        self.content = data
        return data

    def write(self, data: Dict[str, Any]) -> None:
        """
        Write a Python dict as TOML.

        Parameters
        ----------
        data : Dict[str, Any]

        Returns
        -------
        None

        Raises
        ------
        ImportError
            If 'tomli_w' (writer) is not installed.
        OSError
            On write errors.

        Examples
        --------
        TomlFile("cfg.toml","data").write({"app":{"name":"demo"}})
        """
        if _toml_writer is None:
            raise ImportError("TOML writing requires 'tomli_w'")
        tmp = self.full_path + ".tmp"
        with open(tmp, "wb") as f:
            f.write(_toml_writer.dumps(data).encode("utf-8"))
        os.replace(tmp, self.full_path)
        self.status = True
        self.content = data

    def append(self, data: Dict[str, Any]) -> None:
        """
        Deep-merge keys into TOML.

        Behavior
        --------
        Performs a deep merge: nested dicts are merged recursively.

        Parameters
        ----------
        data : Dict[str, Any]

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If either the root or `data` is not a dict.
        OSError
            On write errors.

        Examples
        --------
        t = TomlFile("cfg.toml","data")
        t.write({"flags":{"x":True}})
        t.append({"flags":{"y":False}})
        """
        if not self.status:
            self.write(data)
            return
        current = self.read()
        if not isinstance(current, dict) or not isinstance(data, dict):
            raise TypeError("TOML append expects dict for root and data")
        _deep_merge_dict(current, data)
        self.write(current)

    def search(self, pattern: Optional[str] = None, *, regex: bool = False, case: bool = False,
               key: Optional[str] = None, value: Any = None, limit: Optional[int] = None, **_: Any
               ) -> Iterator[Dict[str, Any]]:
        """
        Search TOML structure by key/value/pattern (JSON-like rules).

        Parameters
        ----------
        pattern, regex, case, key, value, limit
            See JsonFile.search for detailed semantics.

        Yields
        ------
        dict
            {'path','value','line','row','col','sheet','context'}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        hits = list(TomlFile("cfg.toml","data").search(pattern="demo"))
        """
        data = self.read()
        def _iter():
            for p, v in _walk_jsonlike(data):
                k = p.split(".")[-1]
                ok = True
                if key is not None:
                    ok = ok and (k == key)
                if value is not None:
                    ok = ok and (v == value)
                if pattern is not None:
                    ok = ok and self._match_text(str(v), pattern, regex=regex, case=case)
                if ok:
                    yield {
                        "path": f"{self.full_path}:{p}",
                        "value": v,
                        "line": None, "row": None, "col": None, "sheet": None,
                        "context": str(v)[:200],
                    }
        return self._maybe_limit(_iter(), limit)


# ================
# XmlFile
# ================

class XmlFile(_SearchMixin, File):
    """
    XML handler using xml.etree.ElementTree.

    read()  -> Element (root)
    write() -> accept Element or XML string
    append()-> attach child(ren) to root
    search()-> match by tag/attributes/pattern in textual content
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Create an XML file abstraction.

        Parameters
        ----------
        file_name : str
            XML file name (e.g., "data.xml").
        file_path : str
            Directory path; created if missing.

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        if self.status:
            self.read()

    def read(self) -> ET.Element:
        """
        Parse XML and return the root element.

        Returns
        -------
        xml.etree.ElementTree.Element
            Root element.

        Raises
        ------
        FileNotFoundError
            If the file is missing.
        xml.etree.ElementTree.ParseError
            If the XML is invalid.
        OSError
            On read errors.

        Examples
        --------
        root = XmlFile("data.xml","data").read()
        """
        tree = ET.parse(self.full_path)
        root = tree.getroot()
        self.content = root
        return root

    def write(self, data: Union[ET.Element, str]) -> None:
        """
        Write XML data to disk.

        Parameters
        ----------
        data : Element | str
            Either a pre-built Element (root) or an XML string.

        Returns
        -------
        None

        Raises
        ------
        xml.etree.ElementTree.ParseError
            If given an invalid XML string.
        OSError
            On write errors.

        Examples
        --------
        import xml.etree.ElementTree as ET
        root = ET.Element("books")
        XmlFile("books.xml","data").write(root)
        """
        if isinstance(data, ET.Element):
            tree = ET.ElementTree(data)
        else:
            root = ET.fromstring(data)
            tree = ET.ElementTree(root)
        tmp = self.full_path + ".tmp"
        tree.write(tmp, encoding="utf-8", xml_declaration=True)
        os.replace(tmp, self.full_path)
        self.status = True
        self.content = tree.getroot()

    def append(self, data: Union[ET.Element, Sequence[ET.Element]]) -> None:
        """
        Append child element(s) under the XML root.

        Parameters
        ----------
        data : Element | Sequence[Element]

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If data is not an Element or a sequence of Elements.
        OSError
            On write errors.

        Examples
        --------
        import xml.etree.ElementTree as ET
        x = XmlFile("books.xml","data")
        x.write(ET.Element("books"))
        x.append(ET.Element("book", attrib={"id":"1"}))
        """
        root = self.read()
        if isinstance(data, Sequence) and not isinstance(data, (bytes, str)):
            for el in data:
                if not isinstance(el, ET.Element):
                    raise TypeError("append sequence must contain Element nodes")
                root.append(el)
        elif isinstance(data, ET.Element):
            root.append(data)
        else:
            raise TypeError("append expects Element or sequence of Elements")
        self.write(root)

    def search(
        self,
        pattern: Optional[str] = None,
        *,
        regex: bool = False,
        case: bool = False,
        key: Optional[str] = None,          # not used; kept for unified signature
        value: Any = None,                   # match against element text
        columns: Optional[Sequence[str]] = None,  # not used
        tag: Optional[str] = None,           # tag filter
        attr: Optional[Dict[str, str]] = None,    # attributes equality filter
        sheet: Optional[str] = None,         # not used
        limit: Optional[int] = None,
    ) -> Iterator[Dict[str, Any]]:
        """
        Search XML elements by tag/attributes and/or element text.

        Matching
        --------
        - `tag`: exact tag match.
        - `attr`: all provided attributes must match exactly.
        - `value`: element text (stripped) must equal the given value.
        - `pattern`: substring or regex on element text (stripped).
        - `limit`: stop after N hits.

        Parameters
        ----------
        pattern : Optional[str]
            Substring/regex for element text.
        regex : bool, default=False
            Treat pattern as regex.
        case : bool, default=False
            Case-sensitive when True.
        tag : Optional[str]
            Filter by element tag.
        attr : Optional[Dict[str, str]]
            All key/value pairs must match element attributes.
        value : Any
            Exact text to match (compared as str).
        limit : Optional[int]
            Maximum number of results.

        Yields
        ------
        dict
            {'path','value','context'} where context is the matched Element.

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        x = XmlFile("chapter.xml","data")
        x.write('<chapter><title>Intro</title><title>Advanced</title></chapter>')
        hits = list(x.search(tag="title", pattern="adv"))
        """
        tree_or_root = self.read()
        if isinstance(tree_or_root, ET.Element):
            root = tree_or_root
        elif hasattr(tree_or_root, "getroot"):
            root = tree_or_root.getroot()
        else:
            raise TypeError("XML content must be Element or ElementTree")

        count = 0
        for elem in root.iter():
            if tag and elem.tag != tag:
                continue

            if attr:
                if not all(elem.get(k) == v for k, v in attr.items()):
                    continue

            text = (elem.text or "").strip()

            if value is not None and str(value) != text:
                continue

            if pattern:
                if regex:
                    flags = 0 if case else re.IGNORECASE
                    if re.search(pattern, text, flags) is None:
                        continue
                else:
                    hay = text if case else text.lower()
                    needle = pattern if case else pattern.lower()
                    if needle not in hay:
                        continue

            yield {
                "path": self.full_path,
                "value": text,
                "context": elem,
            }

            count += 1
            if limit is not None and count >= limit:
                break


# ================
# ExcelFile
# ================

class ExcelFile(_SearchMixin, File):
    """
    Excel handler using openpyxl.

    read(sheet=None)  -> List[Dict[str,Any]] (first row as header)
    write(List[Dict], sheet=None)
    append(Dict | Sequence[Dict], sheet=None)
    search(pattern/value, columns=[...], sheet=...)

    Notes
    -----
    - The first row is treated as the header.
    - Empty cells are coerced to empty strings during read().
    """

    def __init__(self, file_name: str, file_path: str, *, default_sheet: str = "Sheet1"):
        """
        Create an Excel (.xlsx) abstraction.

        Parameters
        ----------
        file_name : str
            Workbook file name (e.g., "report.xlsx").
        file_path : str
            Directory path; created if missing.
        default_sheet : str, default="Sheet1"
            Used when sheet is not specified.

        Raises
        ------
        OSError
            On directory creation errors.
        """
        super().__init__(file_name, file_path)
        self.default_sheet = default_sheet
        self._headers: dict[str, list[str]] = {}

    def _get_ws(self, wb, sheet: str | None):
        """
        Return a worksheet by name, creating it if missing.

        Parameters
        ----------
        wb : openpyxl.Workbook
        sheet : str | None

        Returns
        -------
        openpyxl.worksheet.worksheet.Worksheet
        """
        name = sheet or self.default_sheet
        if name in wb.sheetnames:
            return wb[name]
        # Create sheet if missing
        if wb.sheetnames:
            ws = wb.create_sheet(title=name)
        else:
            ws = wb.active
            ws.title = name
        return ws

    def _read_headers_from_ws(self, ws) -> list[str]:
        """
        Read header row (first row) from a worksheet.

        Returns
        -------
        list[str]
            Header names (coerced to strings); empty list if sheet empty.
        """
        if ws.max_row >= 1:
            headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
            return [str(h) for h in headers]
        return []

    def _clear_ws(self, ws):
        """
        Clear a worksheet by recreating it at the same index.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet

        Returns
        -------
        openpyxl.worksheet.worksheet.Worksheet
            A fresh, empty worksheet with the same title.

        Notes
        -----
        This approach avoids subtle artifacts left by iteratively deleting rows.
        """
        wb = ws.parent
        title = ws.title
        idx = wb.sheetnames.index(title)
        temp = wb.create_sheet(title=f"__tmp__{title}")
        wb._sheets.insert(idx, wb._sheets.pop(-1))
        del wb[title]
        temp.title = title
        return wb[title]

    def read(self, *, sheet: str | None = None) -> list[dict[str, Any]]:
        """
        Read a sheet into a list of dict rows.

        Parameters
        ----------
        sheet : str | None, default=None
            Target sheet; uses `default_sheet` if None.

        Returns
        -------
        list[dict[str, Any]]
            Rows with keys from header and values from subsequent rows.
            Missing cells become "".

        Raises
        ------
        FileNotFoundError
            If workbook does not exist.
        openpyxl.utils.exceptions.InvalidFileException
            If file is not a valid workbook.
        KeyError
            If the requested sheet cannot be resolved in unusual states.
        OSError
            On read errors.

        Examples
        --------
        data = ExcelFile("r.xlsx","data").read(sheet="S1")
        """
        wb = load_workbook(self.full_path)
        ws = self._get_ws(wb, sheet)

        headers = self._read_headers_from_ws(ws)
        if not headers:
            return []  # empty sheet

        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            values = list(row) if row is not None else []
            if len(values) < len(headers):
                values += [""] * (len(headers) - len(values))
            else:
                values = values[:len(headers)]
            out.append({h: ("" if v is None else v) for h, v in zip(headers, values)})

        self.content = None
        return out

    def write(self, rows: list[dict[str, Any]] | Iterable[dict[str, Any]], *, sheet: str | None = None):
        """
        Write rows to a sheet (replacing existing content).

        Parameters
        ----------
        rows : list[dict[str, Any]] | Iterable[dict[str, Any]]
            Dict rows to write. Header order is inferred from first occurrences.
        sheet : str | None, default=None
            Target sheet; uses `default_sheet` if None.

        Returns
        -------
        None

        Raises
        ------
        ImportError
            If openpyxl is not installed.
        TypeError
            If rows are not dict-like.
        OSError
            On write errors.

        Examples
        --------
        xls = ExcelFile("report.xlsx","data", default_sheet="S1")
        xls.write([{"name":"Avi","score":100},{"name":"Dana","score":90}], sheet="S1")
        """
        rows = list(rows)

        if os.path.exists(self.full_path) and os.path.getsize(self.full_path) > 0:
            wb = load_workbook(self.full_path)
        else:
            wb = Workbook()

        ws = self._get_ws(wb, sheet)

        ws = self._clear_ws(ws)

        headers: list[str] = []
        seen = set()
        for r in rows:
            if not isinstance(r, dict):
                raise TypeError("Excel write expects dict rows")
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    headers.append(k)
        if not headers:
            headers = ["value"]

        ws.append(headers)

        for r in rows:
            ws.append([r.get(h, "") for h in headers])

        tmp = self.full_path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, self.full_path)
        self._headers[ws.title] = headers
        self.status = True
        self.clear_cache()

    def append(self, data: dict[str, Any] | Iterable[dict[str, Any]], *, sheet: str | None = None):
        """
        Append row(s) to a sheet, preserving headers.

        Behavior
        --------
        - If sheet is empty, headers are inferred from first appended rows and
          written to the first row.
        - For subsequent rows, any missing headers are written as "" in that row.

        Parameters
        ----------
        data : dict[str, Any] | Iterable[dict[str, Any]]
            Row(s) to append.
        sheet : str | None, default=None
            Target sheet; uses `default_sheet` if None.

        Returns
        -------
        None

        Raises
        ------
        TypeError
            If data is not dict or iterable of dicts.
        OSError
            On write errors.

        Examples
        --------
        xls.append({"name":"Noa","score":95}, sheet="S1")
        xls.append([{"name":"Lior","score":88},{"name":"Omri"}], sheet="S1")
        """
        if isinstance(data, dict):
            rows = [data]
        else:
            rows = list(data)

        if os.path.exists(self.full_path) and os.path.getsize(self.full_path) > 0:
            wb = load_workbook(self.full_path)
        else:
            wb = Workbook()

        ws = self._get_ws(wb, sheet)

        # existing headers?
        if ws.title in self._headers:
            headers = self._headers[ws.title]
        else:
            headers = self._read_headers_from_ws(ws)
            if not headers:
                headers = []
                seen = set()
                for r in rows:
                    if not isinstance(r, dict):
                        raise TypeError("Excel append expects dict rows")
                    for k in r.keys():
                        if k not in seen:
                            seen.add(k)
                            headers.append(k)
                if not headers:
                    headers = ["value"]
                # write inferred headers if sheet truly empty
                if ws.max_row == 1 and all(c.value in (None, "") for c in ws[1]):
                    ws.append(headers)

        for r in rows:
            if not isinstance(r, dict):
                raise TypeError("Excel append expects dict rows")
            ws.append([r.get(h, "") for h in headers])

        tmp = self.full_path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, self.full_path)
        self._headers[ws.title] = headers
        self.status = True
        self.clear_cache()

    def search(self, pattern: Optional[str] = None, *, regex: bool = False, case: bool = False,
               columns: Optional[Sequence[str]] = None, value: Any = None,
               sheet: Optional[str] = None, limit: Optional[int] = None, **_: Any
               ) -> Iterator[Dict[str, Any]]:
        """
        Search Excel cells by substring/regex or exact value.

        Parameters
        ----------
        pattern : Optional[str]
            Substring/regex to match in cell text.
        regex : bool, default=False
            Interpret `pattern` as regex.
        case : bool, default=False
            Case-sensitive when True.
        columns : Optional[Sequence[str]]
            Restrict search to these headers; default: all columns.
        value : Any
            Exact value match (stringified for comparison).
        sheet : Optional[str]
            Sheet name to search; default: `default_sheet`.
        limit : Optional[int]
            Maximum number of hits.

        Yields
        ------
        dict
            {'path','value','row','col','sheet','context'}

        Raises
        ------
        re.error
            If `regex=True` and the pattern is invalid.

        Examples
        --------
        hits = list(xls.search(pattern="Noa", columns=["name"], sheet="S1"))
        """
        data = self.read(sheet=sheet)
        cols = columns or (list(data[0].keys()) if data else [])
        def _iter():
            for i, row in enumerate(data, start=1):
                for col in cols:
                    cell = str(row.get(col, ""))
                    ok = True
                    if value is not None:
                        ok = ok and (cell == str(value))
                    if pattern is not None:
                        ok = ok and self._match_text(cell, pattern, regex=regex, case=case)
                    if ok:
                        yield {
                            "path": f"{self.full_path}:row[{i}].{col}",
                            "value": cell,
                            "line": None, "row": i, "col": col,
                            "sheet": sheet or self.default_sheet,
                            "context": str(row),
                        }
        return self._maybe_limit(_iter(), limit)


# ==========================================
# Async wrappers (A* classes) for convenience
# ==========================================

class ATextFile(_AsyncMixin, TextFile):
    """
    Async variant of TextFile.

    Methods
    -------
    await aread()  -> str
    await awrite(text: str) -> None
    await aappend(text: str) -> None
    await asearch(...) -> list[dict]
    async with ATextFile(...) -> transactional context

    Notes
    -----
    See the corresponding sync methods for semantics and exceptions.
    """
    pass


class AJsonFile(_AsyncMixin, JsonFile):
    """
    Async variant of JsonFile.

    Notes
    -----
    `asearch` returns a list of hits rather than a generator.
    """
    async def asearch(self, *args, **kwargs):
        """Async variant of JsonFile.search(); returns a list of hits."""
        return list(self.search(*args, **kwargs))


class ACsvFile(_AsyncMixin, CsvFile):
    """
    Async variant of CsvFile.

    Methods mirror the sync API with `await` and `async with`.
    """
    pass


class AYamlFile(_AsyncMixin, YamlFile):
    """
    Async variant of YamlFile.

    Requires PyYAML installed.
    """
    pass


class AIniFile(_AsyncMixin, IniFile):
    """
    Async variant of IniFile.
    """
    pass


class ATomlFile(_AsyncMixin, TomlFile):
    """
    Async variant of TomlFile.

    Requires tomllib/tomli for reading and tomli-w for writing (for sync operations).
    """
    pass


class AXmlFile(_AsyncMixin, XmlFile):
    """
    Async variant of XmlFile.
    """
    pass


class AExcelFile(_AsyncMixin, ExcelFile):
    """
    Async variant of ExcelFile.

    Requires openpyxl installed in your environment.
    """
    pass

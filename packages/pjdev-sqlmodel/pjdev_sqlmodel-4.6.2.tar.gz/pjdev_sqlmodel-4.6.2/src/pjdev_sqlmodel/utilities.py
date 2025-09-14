import csv
import warnings
from pathlib import Path
from typing import Type, TypeVar, List, Optional, Tuple, Dict

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from pydantic import ValidationError, BaseModel
import pandas as pd
from openpyxl.reader.excel import load_workbook
from loguru import logger
from sqlalchemy import Engine
from sqlmodel import select

from pjdev_sqlmodel import session_context, sqlmodel_service
from pjdev_sqlmodel.models import ModelBase

T = TypeVar("T", bound=ModelBase)


def get_files_in_directory(directory: Path, file_name='*', force_presence=False) -> List[Path]:
    file_obj = [f for f in directory.glob(f"**/{file_name}.xlsx") if not f.name.startswith("~$")] + [
        f for f in directory.glob(f"**/{file_name}.csv")
    ]
    if force_presence and len(file_obj) == 0:
        logger.error(
            f'Failed to find files matching the {directory}/{file_name} pattern and presence is required. Exiting...')
        exit(1)
    return file_obj


def get_csv_columns(file_path):
    with open(file_path, "r", encoding="utf-8-sig") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)  # This will return the column names
        return header


def get_excel_columns(file_path, header_ndx: int = 1, col_range: Optional[str] = None, sheet_name: str | int = 0):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore")
        wb = load_workbook(filename=file_path, read_only=True)


    # TODO: need to ensure that sending an int will work
    sheet = wb[sheet_name] if sheet_name else wb.active

    if col_range is not None:
        return [c.value for c in [row for row in sheet[col_range]][0]]

    return [cell.value for cell in sheet[header_ndx]]


def load_csv_data(model_type: Type[T], data_files: List[Path]) -> None:
    fields = model_type.model_fields.keys()
    cols = [
        f
        if model_type.model_fields.get(f).alias is None
        else model_type.model_fields.get(f).alias
        for f in fields
        if f != "row_id"
    ]
    filtered_files = [
        f
        for f in data_files
        if f.name.endswith(".csv")
           and len(set(cols).difference(set(get_csv_columns(f)))) == 0
    ]

    data: List[model_type] = []
    for file in filtered_files:
        df = __read_csv(file, cols)
        data.extend(__convert_to_models(file.name, df, model_type))

    with session_context() as session:
        session.add_all(data)
        session.commit()
    logger.info("Loaded {} rows for {} table".format(len(data), model_type.__name__))


def load_excel_data(
        model_type: Type[T],
        data_files: List[Path],
        header_ndx: int = 0,
        sheet_name: str | int = 0,
        col_range: Optional[str] = None,
) -> None:
    fields = model_type.model_fields.keys()
    cols = [
        f
        if model_type.model_fields.get(f).alias is None
        else model_type.model_fields.get(f).alias
        for f in fields
        if f != "row_id"
    ]

    filtered_files = [
        f
        for f in data_files
        if f.name.endswith(".xlsx")
           and len(
            set(cols).difference(set(get_excel_columns(f, header_ndx + 1, col_range, sheet_name=sheet_name)))
        )
           == 0
    ]

    data: List[model_type] = []

    if len(filtered_files) == 0:
        raise Exception(
            f"No files found that matched the schema for {model_type.__name__}"
        )

    for file in filtered_files:
        df = __read_excel(file=file, cols=cols, header_ndx=header_ndx, sheet_name=sheet_name)
        data.extend(__convert_to_models(file.name, df, model_type))

    with session_context() as session:
        session.add_all(data)
        session.commit()
    logger.info("Loaded {} rows for {} table".format(len(data), model_type.__name__))


def __convert_to_models(
        filename: str, df: pd.DataFrame, model_type: Type[T]
) -> List[T]:
    data: List[model_type] = []
    try:
        for _, row in df.iterrows():
            d = model_type.model_validate(row.to_dict())
            data.append(d)
    except ValidationError as e:
        logger.error(f"Error when parsing {filename}: {e}")

    return data


def load_raw_csv_data(
        filename: Path,
        table_name: str,
        data_type_map: Optional[Dict[str, Type]] = None,
        engine: Optional[Engine] = None,
) -> None:
    df = __read_csv(filename, data_type_map=data_type_map)

    if df is None:
        return

    engine = sqlmodel_service.get_engine() if engine is None else engine
    df.to_sql(
        table_name,
        con=engine,
        if_exists="replace",
        index_label="row_id",
        dtype=data_type_map,
    )


def load_raw_excel_data(
        filename: Path,
        table_name: str,
        header_ndx: int = 0,
        sheet_name: str | int = 0,
        data_type_map: Optional[Dict[str, Type]] = None,
        engine: Optional[Engine] = None,
) -> None:
    df = __read_excel(filename, header_ndx=header_ndx, data_type_map=data_type_map, sheet_name=sheet_name)

    if df is None:
        return

    engine = sqlmodel_service.get_engine() if engine is None else engine
    df.to_sql(
        table_name,
        con=engine,
        if_exists="replace",
        index_label="row_id",
        dtype=data_type_map,
    )


def convert_table_to_df(InputTable: type[BaseModel]):
    with session_context() as session:
        statement = select(InputTable)
        results = session.exec(statement).all()

        headers = []
        field_names = []
        for name, annotation in InputTable.__annotations__.items():
            field = InputTable.__fields__[name]
            alias = field.alias if field.alias else name
            headers.append(alias)
            field_names.append(name)

        data = []
        for res in results:
            new_obj = {}
            for i in range(len(headers)):
                header = headers[i]
                field_name = field_names[i]
                new_obj[header] = getattr(res, field_name)
            data.append(new_obj)
        df = pd.DataFrame(data)
        return df


def convert_table_to_csv(InputTable: type[BaseModel], file: Path):
    df = convert_table_to_df(InputTable)
    df.to_csv(file, index=False)


def export_to_sheet(InputTable: type[BaseModel], wb: Workbook, sheet_name: str, table_name: str = None,
                    hide_sheet=False):
    df = convert_table_to_df(InputTable)
    ws = wb.create_sheet(title=sheet_name)
    if hide_sheet:
        ws.sheet_state = 'hidden'

    if table_name:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        cell = ws.cell(row=1, column=1)
        cell.value = table_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        start_row = 2
        ws.row_dimensions[1].height = 40.5
        ws.row_dimensions[2].height = 30.75
    else:
        start_row = 1
        ws.row_dimensions[1].height = 30.75

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width


def __read_csv(
        file: Path | str,
        cols: Optional[List[str]] = None,
        data_type_map: Optional[Dict[str, Type]] = None,
) -> pd.DataFrame:
    return pd.read_csv(
        file, engine="pyarrow", usecols=cols, na_filter=False, dtype=data_type_map
    )


def __read_excel(
        file: Path | str,
        sheet_name: str | int = 0,
        cols: Optional[List[str]] = None,
        header_ndx: int = 0,
        data_type_map: Optional[Dict[str, Type]] = None,
) -> pd.DataFrame:
    return pd.read_excel(
        io=file,
        usecols=cols,
        sheet_name=sheet_name,
        na_filter=False,
        header=header_ndx,
        engine="calamine",
        dtype=data_type_map,
    )


def convert_to_csv(
        data: List[BaseModel],
        col_mapping_tuple: Tuple[List[str], Dict[str, str], List[str]],
        filename: Path,
        index=False,
) -> None:
    include_set, col_mapping, cols = col_mapping_tuple
    dict_data = [
        d.model_dump(by_alias=True, include=dict.fromkeys(include_set)) for d in data
    ]

    # Create DataFrame from the list of dictionaries
    df = pd.DataFrame(dict_data)

    # Export DataFrame to CSV
    df.to_csv(filename, index=index, columns=cols)

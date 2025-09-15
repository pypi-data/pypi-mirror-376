import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle


def merge_xlsx_sheets(output_file='merged_output.xlsx', sheet_name='web高危指纹'):
    # Create a new workbook for the merged output if it doesn't exist
    if not os.path.exists(output_file):
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)  # Remove the default sheet
    else:
        merged_wb = load_workbook(output_file)

    # Create the target sheet in the merged workbook
    if sheet_name in merged_wb.sheetnames:
        merged_ws = merged_wb[sheet_name]
    else:
        merged_ws = merged_wb.create_sheet(sheet_name)

    # Collect all xlsx files in the current directory
    xlsx_files = [f for f in os.listdir('') if f.endswith('.xlsx')]

    # Create a set of the styles to preserve the styles
    styles_set = {}

    for xlsx_file in xlsx_files:
        wb = load_workbook(xlsx_file, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Create a named style for each unique style in the worksheet
            for cell in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for c in cell:
                    if c.style not in styles_set:
                        styles_set[c.style] = NamedStyle(name=c.style)
                        # styles_set[c.style].font = c.font
                        # styles_set[c.style].border = c.border
                        # styles_set[c.style].fill = c.fill
                        # styles_set[c.style].number_format = c.number_format
                        # styles_set[c.style].protection = c.protection
                        # styles_set[c.style].alignment = c.alignment
                        # merged_wb.add_named_style(styles_set[c.style])

            # Append data and styles to the merged worksheet
            for row in ws.iter_rows(values_only=False):
                merged_row = [c.value for c in row]
                merged_ws.append(merged_row)
                for c in row:
                    new_cell = merged_ws.cell(row=merged_ws.max_row, column=c.column)
                    new_cell.style = c.style

    # Save the merged workbook
    merged_wb.save(output_file)


merge_xlsx_sheets()

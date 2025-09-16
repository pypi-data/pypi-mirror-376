import os
import re
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from prefect import flow, get_run_logger, task
from nemo_library.core import NemoLibrary
import pandas as pd



TEMPLATE_PATH = Path(__file__).parent / "information" / "Lieferantenbewertung_template.xlsx"
OUTPUTPATH = Path(__file__).parent / "data"
PLACEHOLDER_PREFIX = "{{"
PLACEHOLDER_SUFFIX = "}}"

@flow(name="DIG Template Generator", log_prints=True)
@task(name="generate Template")
def generate_templates() -> None:
    
    logger = get_run_logger()
    logger.info("Starting template generator")
    
    # Implementation for generating templates
    nl = NemoLibrary()
    supplierdf = nl.LoadReport(
        projectname="Business Processes", report_name="{OA} Supplier "
    )
        
    for index, row in supplierdf.iterrows():
        logger.info(f"Generating template for supplier: {row['SUPPLIER_I_D']}")

        mapping = _as_text_mapping(row)

        # Load and fill template
        wb = load_workbook(TEMPLATE_PATH)
        # Replace {{FIELD}} placeholders in all worksheets
        for ws in wb.worksheets:
            _replace_placeholders_in_sheet(ws, mapping)

def _as_text_mapping(row: pd.Series) -> dict[str, str]:
    """Convert a pandas row to a dict[str, str] without NaNs."""
    return {k: ("" if pd.isna(v) else str(v)) for k, v in row.items()}

def _replace_placeholders_in_sheet(ws: Worksheet, mapping: Dict[str, str]) -> None:
    """
    Replace placeholders like {{KEY}} in all string cells of a worksheet.
    Only replaces inside str cells; formulas and non-str cells are left unchanged.
    """
    # Build regex that matches any {{KEY}} present in mapping (escape keys safely)
    # If mapping is large, use a generic replacer.
    pat = re.compile(r"\{\{([A-Za-z0-9_]+)\}\}")

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and PLACEHOLDER_PREFIX in val and PLACEHOLDER_SUFFIX in val:
                def _sub(m: re.Match) -> str:
                    key = m.group(1)
                    return mapping.get(key, m.group(0))  # keep placeholder if key not found
                cell.value = pat.sub(_sub, val)

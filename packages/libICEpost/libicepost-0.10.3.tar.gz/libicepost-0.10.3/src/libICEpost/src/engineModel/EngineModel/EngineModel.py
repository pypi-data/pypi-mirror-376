#####################################################################
#                                 DOC                               #
#####################################################################

"""
@author: F. Ramognino       <federico.ramognino@polimi.it>
Last update:        12/06/2023
"""

#####################################################################
#                               IMPORT                              #
#####################################################################

#Typing
from __future__ import annotations
from types import FunctionType
from typing import Iterable, Callable
from matplotlib.axes import Axes

#Other
from operator import attrgetter
import os
from tqdm import tqdm
import traceback

#Plotting
import numpy as np
import pandas as pd

import libICEpost as lice

#Base class
from libICEpost.src.base.BaseClass import BaseClass
from libICEpost.src.base.dataStructures.EngineData.EngineData import EngineData

#I/O and pre-processing
from libICEpost.src.base.dataStructures.Dictionary import Dictionary
from libICEpost.src.base.Filter.Filter import Filter

#Submodels
from libICEpost.src.engineModel.EngineTime.EngineTime import EngineTime
from libICEpost.src.engineModel.EngineGeometry.EngineGeometry import EngineGeometry
from libICEpost.src.thermophysicalModels.thermoModels.thermoMixture.ThermoMixture import ThermoMixture
from libICEpost.src.thermophysicalModels.thermoModels.ThermoModel import ThermoModel
from libICEpost.src.thermophysicalModels.thermoModels.CombustionModel.CombustionModel import CombustionModel
from libICEpost.src.thermophysicalModels.thermoModels.EgrModel.EgrModel import EgrModel
from libICEpost.src.engineModel.HeatTransferModel.HeatTransferModel import HeatTransferModel
from libICEpost.src.thermophysicalModels.thermoModels.CombustionModel.NoCombustion import NoCombustion
from libICEpost.src.engineModel.HeatTransferModel.Woschni import Woschni

#Database
from libICEpost.Database.chemistry.specie.Mixtures import Mixtures, Mixture

#Errors
from libICEpost.src.base.Functions.runtimeWarning import helpOnFail

#############################################################################
#                                  FUNCTIONS                                #
#############################################################################
def get_postfix(zone:str) -> str:
    """
    Return the postfix to append to a corresponding zone. By default, no postfix corresponds to 'cylinder' zone, while '_<zone>' to all the others.

    Args:
        zone (str): The zone to which compute the postfix to append to the data of that zone.

    Returns:
        str: The postfix
    """
    return "" if zone == "cylinder" else f"_{zone}"

#############################################################################
#                               MAIN CLASSES                                #
#############################################################################
# TODO:
#   Handle direct injection (injectionModel?)
#   Handle interaction with other zones (creviceModel? prechamberModel?)

# NOTE: to handle diesel combustion, need to compute the phi from the injected mass 
# (probably the main parameter for the combustion model)

# NOTE: This model handles a single-zone model of the cylinder. Sub-classes may be 
# defined to introduce additional zones, like in case of pre-chamber engines, crevice 
# modeling, or maybe gas-exchange analysis (ducts)

class EngineModel(BaseClass):
    """
    Base class for modeling of an engine and processing experimental/numerical data
    
    NOTE:
    For naming of variables:
        -> By default they refer to the "cylinder" zone
        -> Variables referred to a specific zone are allocated as "<variableName>_<zoneName>"
    
    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    TODO: documentation and user info
    
    """
    #Attibutes:
    _cylinder:ThermoModel
    """Cylinder thermodynamic region"""
    
    Types:dict[str:type] = \
        {
            "EngineGeometry":           EngineGeometry,
            "EngineTime":               EngineTime,
            "EgrModel":                 EgrModel,
            "CombustionModel":          CombustionModel,
            "HeatTransferModel":        HeatTransferModel,
        }
    """Types for each main model"""
    
    Submodels:dict[str:type] = \
        {
            "EgrModel":             EgrModel(), # 0% EGR
            "CombustionModel":      NoCombustion(reactants=Mixture.empty()), #Inhert (motoring)
            "HeatTransferModel":    Woschni(),  # Woschni model with default coeffs.
        }
    """The available sub-models and their default initializers"""
    
    Zones:list[str] = \
        [
            "cylinder"
        ]
    """The zones avaliable in the model"""
    
    thermophysicalProperties:Dictionary
    """Dictionary with thermophysical properties for mixtures"""
    
    combustionProperties:Dictionary
    """Dictionary with properties for combustion modeling and chemical composition of mixtures"""
    
    CombustionModel:CombustionModel
    """The combustion model"""
    
    EgrModel:EgrModel
    """The EGR model"""
    
    HeatTransferModel:HeatTransferModel
    """The wall heat transfer model"""
    
    _air:Mixture
    """Air mixture"""
    
    info:Dictionary
    """General information for pre-post processing"""
    
    #########################################################################
    # Properties
    @property
    def raw(self)-> EngineData:
        """
        The raw data

        Returns:
            EngineData
        """
        return self._raw
    
    #################################
    @property
    def data(self)-> EngineData:
        """
        The processed/filtered data

        Returns:
            EngineData
        """
        return self._data
    
    #########################################################################
    # Class methods
    @classmethod
    def fromDictionary(cls, dictionary:Dictionary) -> EngineModel:
        """
        Construct from dictionary like:
        {
            EngineTime:         str
                Name of the EngineTime model to use
            <EngineTime>Dict:   dict
                Dictionary containing the data specific of the selected 
                SngineTime model (e.g., if engineTime is 'SparkIgnitionTime',
                then this dictionary must be named 'SparkIgnitionTimeDict'). 
                See at the helper for function 'fromDictionary' of the specific 
                EngineTime model selected.
                
            EngineGeometry:         str
                Name of the EngineGeometry model to use
            <EngineGeometry>Dict:   dict
                Dictionary with data required from engineGeometry.
                See at the helper for function 'fromDictionary' of the specific 
                EngineGeometry model selected.
            
            thermoPhysicalProperties:   dict
                Dictionary with types and data for thermophysical modeling of mixtures
            {
                ThermoType: dict
                {
                    Thermo: str
                    EquationOfState:    str
                }
                <Thermo>Dict: dict
                <EquationOfState>Dict: dict
            }
            
            combustionProperties:   dict
                Dictionaries for data required for mixture preparation and combustion modeling.
            {
                injectionModels: dict
                {
                    TODO
                },
                
                air:    Mixture (default: database.chemistry.specie.Mixtures.dryAir)
                    The air mixture composition
                
                initialMixture: dict
                    Dictionary with data for initialization of the mixture 
                    in the thermodynamic zones
                {
                    <zoneName>:
                    {
                        [depends on specific engine model]
                    }
                },
                
                CombustionModel:         str
                    Name of the CombustionModel to use
                <CombustionModel>Dict:   dict
                    Dictionary with data required from CombustionModel
                    See at the helper for function 'fromDictionary' of the specific 
                    CombustionModel model selected.
            }
            
            dataDict (dictionary): Dictionary with info for loading data, pre-processing and setting initial conditions.
            {
                TODO
            }
            
            functionObjects (Iterable[tuple[str,dict]]): Function object to apply at every time-loop
            [
                tuple[str, dict]: FunctionObjectType, constructionDictionary
            ]
        }
        """
        cls.checkType(dictionary, [dict, Dictionary], "dictionary")
        if isinstance(dictionary, dict):
            dictionary = Dictionary(dictionary)
        
        print("Constructing engine model from dictionary\n")
        
        #Engine time:
        print("Construct EngineTime")
        etModel = dictionary.lookup("EngineTime")
        ET = EngineTime.selector(etModel, dictionary.lookup(etModel + "Dict"))
        print(ET,"\n")
        
        #EngineGeometry:
        print("Construct EngineGeometry")
        egModel = dictionary.lookup("EngineGeometry")
        EG = EngineGeometry.selector(egModel, dictionary.lookup(egModel + "Dict"))
        print(EG,"\n")

        #combustionProperties
        combustionProperties = dictionary.lookup("combustionProperties")
        
        #thermophysical properties
        thermophysicalProperties = dictionary.lookup("thermophysicalProperties")
        
        #Data for pre-processing
        dataDict = dictionary.lookup("dataDict")
        
        #Submodels
        subModels = {}
        smDict = dictionary.lookupOrDefault("submodels", Dictionary())
        for sm in cls.Submodels:
            if sm + "Type" in smDict:
                print(f"Constructing {sm} sub-model")
                smTypeName = smDict.lookup(sm + "Type")
                print(f"\tType: {smTypeName}")
                subModels[sm] = cls.Types[sm].selector(smTypeName, smDict.lookup(smTypeName + "Dict"))
        
        #Function objects
        from libICEpost.src.engineModel.functionObjects import FunctionObject
        functionObjects = []
        foList = dictionary.lookupOrDefault("functionObjects", default=[])
        for ii, fo in enumerate(foList):
            cls.checkType(fo, tuple, "dictionary[functionObjects][{ii}]")
            if not len(fo) == 2: raise ValueError(f"dictionary[functionObjects][{ii}] must be of length 2 with FunctionObjectType and corresponding construction dictionary")
            cls.checkType(fo[0], str, f"dictionary[functionObjects][{ii}][0]")
            cls.checkType(fo[1], dict, f"dictionary[functionObjects][{ii}][1]")
            functionObjects.append(FunctionObject.selector(*fo))
        
        out = cls(time=ET, geometry=EG, thermophysicalProperties=thermophysicalProperties, combustionProperties=combustionProperties, dataDict=dataDict, functionObjects=functionObjects, **subModels)
        return out
    
    #########################################################################
    #Constructor:
    def __init__(self, *,
                 time:EngineTime,
                 geometry:EngineGeometry,
                 thermophysicalProperties:dict|Dictionary,
                 combustionProperties:dict|Dictionary,
                 dataDict:dict=None,
                 functionObjects:Iterable[Callable]=None,
                 **submodels,
                 ):
        """
        Base class for engine model, used for type-checking and loading the sub-models.

        Args:
            time (EngineTime): The engine time
            geometry (EngineGeometry): The engine geometry
            thermophysicalProperties (dict|Dictionary): Dictionary with thermophysical properties of mixtures
            combustionProperties (dict|Dictionary): Dictionary with combustion data and chemical composition
            dataDict (dict, optional): Dictionary for loading data. If not given, data are not loaded 
                and thermodynamic regions not initialized. Defaults to None.
            functionObjects (Iterable[Callable], optional): A list of function objects to apply at 
                every time-step. They must be functions takining the EngineModel (self) as *ONLY* input and
                returning None. These can be used to further post-processing (example, save the mixture
                compositions of the thermodynamic models)
            **submodels (dict, optional): Optional sub-models to load. Defaults to {}.
        """
        #Main models
        self.checkType(geometry, self.Types["EngineGeometry"], "geometry")
        self.checkType(time, self.Types["EngineTime"], "engineTime")
        self.geometry = geometry
        self.time = time
        
        #Data structures
        self._raw = EngineData()     #Raw data
        self._data = EngineData()    #Filtered data
        
        #Submodels
        for model in self.Submodels:
            if model in submodels:
                #Get from input
                sm = submodels[model]
                self.checkType(sm, self.Types[model], f"{submodels}[{model}]")
            else:
                #Take default
                sm = self.Submodels[model].copy()
            #Set sub-model
            self.__setattr__(model, sm)
        
        #Thermos
        self.checkType(thermophysicalProperties, dict, "thermophysicalProperties")
        if isinstance(thermophysicalProperties, dict):
            thermophysicalProperties = Dictionary(**thermophysicalProperties)
        self.thermophysicalProperties = thermophysicalProperties.copy()
        
        #Combustion properties
        self.checkType(combustionProperties, dict, "combustionProperties")
        if isinstance(combustionProperties, dict):
            combustionProperties = Dictionary(**combustionProperties)
        self.combustionProperties = combustionProperties.copy()
        
        #Contruct the thermodynamic models
        self._constructThemodynamicModels(combustionProperties)
        
        #TODO: construct injection models
        
        #Construct the Egr model
        self._constructEgrModel(combustionProperties)
        
        #Construct the combustion model
        self._constructCombustionModel(combustionProperties)
        
        #Misc parameters
        self.info = Dictionary()
        self.info["path"] = None
        self.info["dataDict"] = None
        self.info["filter"] = None
        self.info["initialConditions"] = None
        
        #Function objects
        self.checkArray(functionObjects, Callable, "functionObjects")
        self.info["functionObjects"] = [fo for fo in functionObjects]
        
        #Pre-processing
        if not dataDict is None:
            self.preProcess(**dataDict)
    
    #########################################################################
    #Construction methods:
    def _constructThemodynamicModels(self, combustionProperties:dict|Dictionary) -> EngineModel:
        """
        Construct the thermodynamic models of the system, setting their initial 
        mixture composition. Here setting everything to air, might be overwritten
        in sub-classes to handle specific initializations (SI engine will use 
        the premixedFuel entry)
        
        Args:
            combustionProperties (dict|Dictionary): the combustion properties
        
        Returns:
            EngineModel: self
        """
        self.checkType(combustionProperties, dict, "combustionProperties")
        if not isinstance(combustionProperties, Dictionary):
            combustionProperties = Dictionary(**combustionProperties)
            
        #Air composition
        air = combustionProperties.lookupOrDefault("air", Mixtures.dryAir)
        self._air = air.copy()
        
        #Here set everything to air, in sub-classes update
        for zone in self.Zones:
            self.__setattr__("_" + zone, ThermoModel(ThermoMixture(self._air.copy(), **self.thermophysicalProperties)))
        return self
    
    ####################################
    def _constructEgrModel(self, combustionProperties:dict|Dictionary):
        """
        Construct the EGR model and apply it to the cylinder region.
        Might be overwritten in child for applying EGR also to sub-regions of cylinder.
        
        Args:
            combustionProperties (dict|Dictionary): the combustion properties
        
        Returns:
            EngineModel: self
        """
        print("Constructing EGR model")
        
        self.checkType(combustionProperties, dict, "combustionProperties")
        if not isinstance(combustionProperties, Dictionary):
            combustionProperties = Dictionary(**combustionProperties)
        
        if "EgrModel" in combustionProperties:
            #Construct egr model from combustion properties
            egrModelType:str = combustionProperties.lookup("EgrModel")
            egrModelDict = combustionProperties.lookupOrDefault(egrModelType + "Dict", Dictionary())
            egrModelDict.update(reactants=self._cylinder.mixture.mix) #Append to dictionary the cylinder properties
        else:
            #Use default
            self.EgrModel = self.Submodels["EgrModel"].copy().update(reactants=self._cylinder.mixture.mix)
        
        #NOTE: When introducing the injection models, need to compute EVO composition instead
        
        #Construct the EGR model
        self.EgrModel = EgrModel.selector(egrModelType, egrModelDict)
        
        print(f"\tType: {self.EgrModel.__class__.__name__}")
        
        #Apply EGR to reactants
        self._cylinder.mixture.mix.dilute(self.EgrModel.EgrMixture, self.EgrModel.egr)
    
    ####################################
    def _constructCombustionModel(self, combustionProperties:dict|Dictionary):
        """
        Construct the combustion model.
        Might be overwritten in child to set additional parameters to combustionModelDict 
        (fuel for PremixedCombustion model is updated in SparkIgnitionEngine)
        
        Args:
            combustionProperties (dict|Dictionary): the combustion properties
        
        Returns:
            EngineModel: self
        """
        print("Constructing combustion model")
        
        self.checkType(combustionProperties, dict, "combustionProperties")
        if not isinstance(combustionProperties, Dictionary):
            combustionProperties = Dictionary(**combustionProperties)
        
        combustionModelType = combustionProperties.lookupOrDefault("CombustionModel", None, fatal=False)
        if not combustionModelType is None:
            #Get dictionary
            combustionModelDict = combustionProperties.lookupOrDefault(combustionModelType + "Dict", Dictionary())
            combustionModelDict.update( #Append reactants to combustion model dictionary
                reactants=self._cylinder.mixture.mix    #Initial in-cylinder mixture
            )
            
            #Construct
            self.CombustionModel = CombustionModel.selector(combustionModelType, combustionModelDict)
        else:
            #Use default
            self.CombustionModel = self.Submodels["CombustionModel"].copy().update(reactants=self._cylinder.mixture.mix)
        
        print(f"\tType: {self.CombustionModel.__class__.__name__}")
        
        #Check consistency
        if not isinstance(self.CombustionModel, self.Types["CombustionModel"]):
            raise TypeError(f"Combustion model type {combustionModelType} in combustionProperties dictionaries not compatible with allowed type for engine model {self.__class__.__name__} ({self.Types['CombustionModel']})")
        
    #########################################################################
    #Updating methods:
    def _updateMixtures(self) -> None:
        """
        Update mixture compositions (might be overwritten in child classes)
        """
        #TODO: Update the in-cylinder mixture based on injection models (may have already injected some mass)
        
        #Update the combustion model at current time
        self._updateCombustionModel()
        
        #Update in-cylinder mixture based on combustion model current mixture
        self._cylinder.mixture.update(self.CombustionModel.mixture)
        
    ####################################
    def _updateCombustionModel(self):
        """
        Update combustion model
        """
        #TODO: Update the reactants mixture based on injection models (may have already injected some mass)
        
        #All data
        index = self.data.index[self.data.loc[:,'CA'] == self.time.time].tolist()[0]
        data = self.data.loc[index].to_dict()
        
        self.CombustionModel.update(**data) #NOTE: update also fuel when implementing injection models
    
    #########################################################################
    #Dunder methods:
    def __str__(self):
        STR = ""
        STR += "Engine model instance:\n"
        STR += "Engine time\n\t" + self.time.__str__().replace("\n", "\n\t")
        STR += "\n"
        STR += "Engine geometry:\n\t" + self.geometry.__str__().replace("\n", "\n\t")
        STR += "\n"
        STR += "EGR model:\n\t" + self.EgrModel.__str__().replace("\n", "\n\t")
        STR += "\n"
        STR += "Combustion model:\n\t" + self.CombustionModel.__str__().replace("\n", "\n\t")
        return STR
        
    #########################################################################
    #Pre-processing methods:
    def _loadFile(self,*args,**argv) -> EngineModel:
        """
        Loads a file with raw data to self.raw. See EngineData.loadFile 
        documentation for arguments:
        """
        self.raw.loadFile(*args,**argv)
        return self
    
    ####################################
    def _loadArray(self,*args,**argv):
        """
        Loads an array with raw data to self.raw. See EngineData.loadArray 
        documentation for arguments:
        """
        self._raw.loadArray(*args,**argv)
        return self
    
    ####################################
    def loadData(self, dataPath:str=None, *, data:dict|Dictionary) -> EngineModel:
        """
        Load raw data.
        
        TODO: Info
        
        Args:
            data (dict | Dictionary): Dictionary containing the data to load for each region.
            dataPath (str, optional): Global path where to load/write data. Defaults to None.

        Returns:
            EngineModel: self
        """
        print("Loading data")
        print(f"Data path: {dataPath}")
        
        #Seth path info
        self.info["dataPath"] = dataPath
        
        #Cast to Dictionary
        data = Dictionary(**data)
        self.info["data"] = data
        
        #Load data:
        for zone in self.Zones:
            zoneDict = data.lookup(zone)
            
            #Check that pressure is found (mandatory)
            if (not "p" in zoneDict) and (not "p" in self.raw.columns):
                raise ValueError(f"Mandatory entry 'p' in data dictionary for zone {zone} not found. Pressure trace must be loaded for each thermodynamic region.")
            
            #Loop over data to be loaded:
            for entry in zoneDict:
                dataDict = zoneDict.lookup(entry)
                self.checkType(dataDict, Dictionary, f"{zone}[{entry}]")
                
                #If the region is not cylinder, append its name to the field
                entryName:str = entry + get_postfix(zone)
                currData:Dictionary = dataDict.lookup("data")
                opts:Dictionary = currData.lookupOrDefault("opts", Dictionary())
                
                #Get format
                dataFormat:str = dataDict.lookup("format")
                if (dataFormat == "file"):
                    #File
                    fileName = currData.lookup("fileName")
                    
                    #relative to dataPath if given
                    fileName = (dataPath + os.path.sep if dataPath else "") + fileName
                    
                    #Load
                    self._loadFile(fileName, entryName, **opts)
                    
                elif (dataFormat == "array"):
                    #(CA,val) array
                    dataArray = currData.lookup("array")
                    self._loadArray(dataArray, entryName, **opts)
                
                elif (dataFormat == "function"):
                    #Function f(CA)
                    function:FunctionType = currData.lookup("function")
                    self.checkType(function, FunctionType, f"{zone}[{entry}][function]")
                    
                    CA = self.raw.loc[:,"CA"]
                    f = CA.apply(function)
                    
                    self._loadArray(np.array((CA,f)).T, entryName, **opts)
                    
                elif (dataFormat == "uniform"):
                    #Uniform value
                    value:float = currData.lookup("value")
                    self.checkType(value, float, f"{zone}[{entry}][value]")
                    self.raw.loc[:,entryName] = value
                
                elif (dataFormat == "calc"):
                    #Apply operation between alredy loaded data
                    function:FunctionType = currData.lookup("function")
                    self.checkType(function, FunctionType, f"{zone}[{entry}][function]")
                    
                    #Function arguments
                    argNames:list[str] = function.__code__.co_varnames[:function.__code__.co_argcount]
                    
                    #Check if are present:
                    for arg in argNames:
                        if not arg in self.raw.columns:
                            raise ValueError(f"Field '{arg}' was not loaded.")
                    
                    #Extract corresponding columns from data-frame:
                    cols = {c:self.raw.loc[:,c] for c in argNames}
                    
                    CA = self.raw.loc[:,"CA"]
                    f = function(**cols)
                    
                    self._loadArray(np.array((CA,f)).T, entryName, **opts)
                
                else:
                    raise ValueError(f"Unknown data format '{dataFormat}' for entry {zone}[{entry}]")
                    
        return self
    
    ####################################
    def filterData(self, filter:"Filter|FunctionType|None"=None) -> EngineModel:
        """
        filter: Filter|FunctionType|None (optional)
            Filter to apply to raw data (e.g. resampling, low-pass filter, etc.). Required
            a method __call__(xp, yp)->(x,y) that resamples the dataset (xp,yp) to the
            datapoints (x,y).
        
        Filter the data in self.raw. Save the corresponding 
        filtered data to self.data.
        If filter is None, data are cloned from self.raw
        """
        #Save filter
        self.info["filter"] = filter
        
        #Clone if no filter is given
        if filter is None:
            for field in self._raw.columns:
                self._data.loadArray(np.array((self._raw.loc[:,"CA"],self._raw.loc[:,field])).T, field)
            return self
        
        #Apply filter
        print(f"Applying filter {filter if isinstance(filter,Filter) else filter.__name__}")
        for var in self._raw.columns:
            #Filter data
            if var != "CA":
                self._data.loadArray(np.array(filter(self._raw.loc[:,"CA"], self._raw.loc[:,var])).T, var)
        
        return self
    
    ####################################
    def initializeThemodynamicModels(self, **initialConditions) -> EngineModel:
        """
        Set the initial conditions of all thermodynamic regions of the EngineModel.
        For region to be initialized, a dict is given for the inital conditions,
        according to the following convention:
        
        ->  If a float is given, the value is used
        ->  If a str is given, it refers to the name of the vabiables stored in the EngineModel,
            in which case the corresponding initial condition is sampled from the the corresponding
            data-set at self.time.startTime.
        ->  If string starting with @ is given, it applies that method with input (self.time.startTime)

        Ex:
        {
            "pressure": "p",            #This interpolates self.data.p at self.time.startTime
            "mass": 1.2e-3,             #Value
            "volume": "@geometry.V"     #Evaluates self.geometry.V(self.time.startTime)
        }
        
        Args:
            **initialConditions:  data initialization of each zone in the model.
        """
        #Update start-time of engine time so that it is bounded to first avaliable time-step
        
        if not "CA" in self.data.columns:
            raise ValueError("No data loaded yet.")
        
        #Set start-time
        self.time.updateStartTime(self.data.loc[:,"CA"])
        self.info["time"] = self.time.time
        
        #Update the mixtures at start-time (combustion models, injection models, etc.)
        self._updateMixtures()
        
        initialConditions = Dictionary(**initialConditions)
        #Store initial conditions
        self.info["initialConditions"] = initialConditions
        
        for zone in self.Zones:
            zoneDict = initialConditions.lookup(zone)
            self.checkType(zoneDict, dict, "zoneDict")
            
            attrgetter("_" + zone)(self).initializeState(**self._preprocessThermoModelInput(zoneDict, zone=zone))
        
        return self
    
    ####################################
    def _preprocessThermoModelInput(self, inputDict:dict, zone:str) -> dict:
        """
        Auxiliary function to pre-process input dictionary 
        to initialization of a thermodynamic region

        NOTE:
            Might be overwritten in child class to also initialize 
            the mixture composition, for example based on the combustion model.
            In such case, do so:
                def _preprocessThermoModelInput(self, inputDict:dict, zone:str) -> dict:
                    tempDict = super()._preprocessThermoModelInput(inputDict, zone)
                    
                    ... #Manipulate mixture based on output dictionary
                    
                    nameList = ["mass", "temperature", "pressure", "volume", "density"]
                    outputDict = {v:[tempDict[v]] for v in tempDict if v in nameList}   #Set state variables
                    outputDict["mixture"] = mix #Set mixture
                    
                    return outputDict
                    
        
        Args:
            inputDict (dict): dictionary for thermodynamic inputs
            zone (str): the zone name
        
        Returns:
            dict: processed dictionary
        """
        #TODO error handling
        
        outputDict = {}
        for key in inputDict:
            val = inputDict[key]
            
            if isinstance(val,float):
                #Float -> use value
                outputDict[key] = val
            elif isinstance(val,str):
                #str
                startTime:float = self.time.startTime
                
                if val.startswith("@"):
                    #str with @ -> apply method
                    code = f"outputDict[key] = self.{val[1:]}(startTime)"
                    exec(code)
                else:
                    #str -> interpolate
                    outputDict[key] = attrgetter(val + (f"_{zone}" if not (zone == "cylinder") else ""))(self.data)(startTime)
            else:
                #Error
                raise TypeError(f"Type '{val.__class__.__name__}' not supported ({key}).")
        
        return outputDict
    
    ####################################
    def preProcess(self, dataPath:str=None, *, data:dict|Dictionary, preProcessing:dict|Dictionary=None, initialConditions:dict|Dictionary, **junk) -> EngineModel:
        """
        Pre-processing:
            1) Loading data (from files or arrays)
            2) Pre-process the data (filtering, optional)
            3) Initialize thermodynamic regions
        
        NOTE:
        Naming of variables when loading data as follows:
            -> By default they refer to the "cylinder" zone
            -> Variables referred to a specific zone are allocated as "<variableName>_<zoneName>"

        TODO: example
        
        Args:
            data (dict | Dictionary): Dictionary with info for loading data.
            preProcessing (dict | Dictionary, optional): Dictionary with pre-processing information. Defaults to None.
            dataPath (str, optional): Master path of the tree. Defaults to None.
            initialConditions (dict | Dictionary): Dictionary with initial condition for thermodynamic models

        Returns:
            EngineModel: self
        """
        #NOTE: **junk used to have other miscellaneous during contruction from dictionary.
        
        print("Pre-processing")
        
        #Loading data:
        self.loadData(dataPath, data=data)
        
        # Filtering data
        self.info["preProcessing"] = preProcessing
        filter = None
        if not preProcessing is None:
            preProcessing = Dictionary(**preProcessing)
            filterType = preProcessing.lookupOrDefault("Filter", None, fatal=False)
            if isinstance(filterType, str):
                #Got type name for run-time construction
                filter:Filter = Filter.selector(filterType, preProcessing.lookup(f"{filterType}Dict"))
            elif isinstance(filterType, Filter):
                #Got filter item
                filter = filterType
            else:
                #No filtering
                pass
            
        self.filterData(filter)
        
        #Initial conditions for thermodinamic models:
        self.initializeThemodynamicModels(**initialConditions)
        
        return self
        
    #########################################################################
    #Processing methods:
    def process(self) -> EngineModel:
        """
        Process the data, main time-loop.
        
        This is split into two function calls, which may be overwritten in child classes to tailored processings:
        1) _process__pre__: Create the columns in self.data for the fields generted by post-processing
        2) _update: The state-updating procedure in the main time-loop
        3) _process__post__: Final post-processing (e.g., computation of wall heat fluxes and rohr)
        
        Returns:
            EngineModel: self
        """
        print("")
        print("Processing")
        print("startTime:",self.time.startTime)
        print("endTime:",self.time.endTime)
        
        #Create fields
        self._process__pre__()
        
        #Process cylinder data
        for t in tqdm(self.time(self.data.loc[:,"CA"]), "Progress: ", initial=0, total=(self.time.endTime-self.time.startTime), unit="CAD"):  #With progress bar :)
            self.info["time"] = t
            #Break the time loop if the updating fails:
            try:
                self._update()                                      #Update
                self._storeLatestTime()                             #Save results at this time
                [fo(self) for fo in self.info["functionObjects"]]   #Compute function objects
            except BaseException as err:
                self.runtimeError(f"Failed updating engine model at time {t}",stack=True)
                print(traceback.format_exc())
                break

        #Final updates (heat transfer, cumulatives, etc...)
        self._process__post__()
        
        return self
        
    ####################################
    def _process__pre__(self) -> None:
        """
        Creation of the post-processed fields.
        
        NOTE:
            When overwriting, first call this method:
            def _process__pre__(self) -> None:
                super()._process__pre__()
                ...
        """
        #Add fields to data:
        fields = {"dpdCA", "AHRR", "ROHR", "A"}
        for zone in self.Zones:
            fields |= {v + get_postfix(zone) for v in getattr(self, f"_{zone}").state.__dict__}
        for f in fields:
            if not f in self.data.columns:
                self.data.loc[:,f] = float("nan")
        
        #Loop over zones to set mixture compositions
        for zone in self.Zones:
            postfix = get_postfix(zone)
            Z:ThermoModel = getattr(self, f"_{zone}")
            
            # #Specie
            # for specie in Z.mixture.mix:
            #     self.data.loc[:,specie.specie.name + "_x" + postfix] = 0.0
            #     self.data.loc[:,specie.specie.name + "_y" + postfix] = 0.0
        
        #Store at startTime
        self._storeLatestTime()
    
    ####################################
    def _update(self) -> None:
        """
        Method for updating the state during the time-loop. Here updating 
        cylinder as single-zone model without interaction with other regions.
        Could be overwritten for more detailed models (e.g., two-zone SI model, TJI, etc.)
        
        NOTE:
            When overwriting, afterwards call this method:
            def _update(self) -> None:
                ...
                super()._update()
        """
        #TODO injection models for mass end energy souce terms
        #TODO heat transfer models for temperature (open systems only!)
        
        #Current time
        CA = self.time.time
        index = self.data.index[self.data.loc[:,'CA'] == CA].tolist()[0]
        
        #Update state
        p = self.data.loc[index,"p"]
        V = self.geometry.V(CA)
        dpdCA = (self.data.loc[index,"p"] - self.data.loc[index-1,"p"])/self.time.deltaT
        self._cylinder.update(pressure=p, volume=V)
        self._updateMixtures()
        
        #Apparent heat release rate [J/CA]
        #Generalization to allow other EoS
        T = self._cylinder.state.T
        m = self._cylinder.state.m
        TOld = self.data.loc[index-1,"T"]
        pOld = self.data.loc[index-1,"p"]
        mOld = self.data.loc[index-1,"m"]
        #Apporximating Us derivative backwards in time
        dUsdCA = (self._cylinder.mixture.us(p,T)*m - self._cylinder.mixture.us(pOld,TOld)*mOld)/self.time.deltaT
        ahrr = dUsdCA + p*self.geometry.dVdCA(CA) #- dmIndCA*mixtureIn.hs(p,T) + dmOutdCA*mixtureOut.hs(p,T)
        
        #Store main parameters
        self.data.loc[index, "dpdCA"] = dpdCA
        self.data.loc[index, "AHRR"] = ahrr
        
    ####################################
    def _storeLatestTime(self):
        """
        Store all properties at latest time.
        """
        #Current time
        CA = self.time.time
        index = self.data.index[self.data.loc[:,'CA'] == CA].tolist()
        
        #Loop over zones
        for zone in self.Zones:
            postfix = get_postfix(zone)
            Z:ThermoModel = getattr(self, f"_{zone}")
            
            #Properties
            self.data.loc[index, Z.state.__dict__.keys()] = [Z.state.__dict__[key] for key in Z.state.__dict__.keys()]
            
            #Specie
            # for specie in Z.mixture.mix:
            #     self.data.loc[index, specie.specie.name + "_x" + postfix] = specie.X
            #     self.data.loc[index, specie.specie.name + "_y" + postfix] = specie.Y
    
    ####################################
    def _process__post__(self) -> None:
        """
        Computing wall heat fluxes and rohr
        
        NOTE:
            When overwriting, first call this method:
            def _process__post__(self) -> None:
                ...
                super()._process__post__()
        """
        #WHF and ROHR
        self._computeWallHeatFlux()
        self.data.loc[:,"ROHR"] = self.data.loc[:,"AHRR"] + self.data.loc[:,"dQwalls"]
        
        #Cumulatives
        self.data.loc[:,"cumHR"] = self.cumulativeIntegral("ROHR")
        self.data.loc[:,"cumAHR"] = self.cumulativeIntegral("AHRR")
    
    ####################################
    def _computeWallHeatFlux(self) -> None:
        """
        Compute wall heat fluxes for each patch and global value in each region. Might be overloaded in child.
        """
        areas = self.geometry.areas(self.data.loc[:,"CA"])
        
        #Compute wall heat transfer coefficient:
        h = self.HeatTransferModel.h(engine=self, CA=self.data.loc[:,"CA"])
        self.data.loc[:,"heatTransferCoeff"] = h
        
        #Total whf
        self.data.loc[:,"dQwalls"] = 0.0
        self.data.loc[:,"Qwalls"] = 0.0
        self.data.loc[:,"wallsArea"] = 0.0
        
        for patch in [c for c in areas.columns if not (c == "CA")]:
            #Search temperature as "T<patchName>":
            if f"T{patch}" in self.data.columns:
                Twall = self.data.loc[:,f"T{patch}"]
            #Fallback to default "Twalls": 
            elif "Twalls" in self.data.columns:
                Twall = self.data.loc[:,"Twalls"]
            else:
                raise ValueError("Cannot compute wall heat flux. Either load patch temperatures in the form t<patchName> or default temperature Twalls to compute wall heat fluxes.")

            #Compute patch area:
            A = areas[patch]
            self.data.loc[:,"wallsArea"] += A
            
            name = patch + "Area"
            if not name in self.data.columns:
                self.data.loc[:,name] = A
            
            #Compute wall heat flux at patch [converted to J/CA]:
            self.data.loc[:,f"dQ{patch}"] = h * A * (self.data.loc[:,"T"] - Twall) / self.time.dCAdt
            
            #Compute cumulative
            self.data.loc[:,f"Q{patch}"] = self.cumulativeIntegral(f"dQ{patch}")
            
            #Add to total
            self.data.loc[:,"dQwalls"] += self.data.loc[:,f"dQ{patch}"]
            self.data.loc[:,"Qwalls"] += self.data.loc[:,f"Q{patch}"]
            
    ####################################
    def integrateVariable(self, y:str, *, x:str="CA", start:float=None, end:float=None) -> float:
        """
        Integrate a variable over another. 
        If inital or final CA are not given, are set to first/last in CA range.

        Args:
            y (str): name of y variable.
            x (str, optional): Name of x variable. Defaults to "CA".
            start (float, optional): Initial CA. Defaults to None.
            end (float, optional): Final CA. Defaults to None.
        
        Returns:
            float: Integrated value 
        """
        if not x in self.data.columns:
            raise ValueError(f"Variable '{x}' not present among data.")
        if not y in self.data.columns:
            raise ValueError(f"Variable '{y}' not present among data.")
        
        from scipy import integrate
        
        start = self.data.loc[0, "CA"] if start is None else start
        end = self.data.loc[len(self.data)-1, "CA"] if end is None else end
        
        self.checkType(start,float,"start")
        self.checkType(end,float,"end")
        
        index = self.data.index[np.array(self.data.loc[:,"CA"] >= start) & np.array(self.data.loc[:,"CA"] <= end)]
        data = self.data.iloc[index]
        
        #Filter out "nan"
        Yarray = data.loc[:,y].copy()
        Yarray[np.isnan(Yarray)] = 0.0
        
        return integrate.trapz(Yarray, x=data.loc[:,x])
    
    ####################################
    def cumulativeIntegral(self, y:str, *, x:str="CA", start:float=None) -> np.ndarray:
        """
        Compute the cumulative integral of a variable over another. 
        If start is not given, it is set to self.time.startOfCombustion.

        Args:
            y (str): name of y variable.
            x (str, optional): Name of x variable. Defaults to "CA".
            start (float, optional): Initial CA. Defaults to None.
        
        Returns:
            pd.DataFrame: Cumulative integral function
        """
        if not x in self.data.columns:
            raise ValueError(f"Variable '{x}' not present among data.")
        if not y in self.data.columns:
            raise ValueError(f"Variable '{y}' not present among data.")
        
        from scipy import integrate
        
        #Check for start
        start = self.time.startOfCombustion() if start is None else start
        #Check for motored
        start = self.data.loc[:,"CA"][0] if start is None else start
        #Check type
        self.checkType(start,float,"start")
        
        #Filter out "nan"
        Yarray = self.data.loc[:,y].copy()
        Yarray[np.isnan(Yarray)] = 0.0
        
        #Compute cumulative
        out = integrate.cumulative_trapezoid(Yarray, x=self.data.loc[:,x], initial=0.0)
        
        #Set zero at start
        valAtStart = np.interp(start, self.data.loc[:,"CA"], out)
        out -= valAtStart
        
        return out
    
    ####################################
    def IMEP(self, start:float=None, end:float=None) -> float:
        """
        Compute indicated mean effective pressure. 
        If inital or final CA are not given, are set to first/last in CA range.

        Args:
            start (float, optional): Initial CA. Defaults to None.
            end (float, optional): Final CA. Defaults to None.
        
        Returns:
            float: IMEP [Pa]
        """
        start = self.data.loc[0, "CA"] if start is None else start
        end = self.data.loc[len(self.data)-1, "CA"] if end is None else end
        
        self.checkType(start,float,"start")
        self.checkType(end,float,"end")
        
        index = self.data.index[np.array(self.data.loc[:,"CA"] >= start) & np.array(self.data.loc[:,"CA"] <= end)]
        data = self.data.iloc[index]
        data.loc[index,"V"] = self.geometry.V(data.loc[:,"CA"])
        
        return self.work(start=start, end=end)/(np.nanmax(data.loc[:,"V"]) - np.nanmin(data.loc[:,"V"]))
    
    ####################################
    def work(self, start:float=None, end:float=None) -> float:
        """
        Compute indicated work (positive outgoing). 
        If inital or final CA are not given, are set to first/last in CA range.

        Args:
            start (float, optional): Initial CA. Defaults to None.
            end (float, optional): Final CA. Defaults to None.
        
        Returns:
            float: Work [J]
        """
        from scipy import integrate
        
        start = self.data.loc[0, "CA"] if start is None else start
        end = self.data.loc[len(self.data)-1, "CA"] if end is None else end
        
        self.checkType(start,float,"start")
        self.checkType(end,float,"end")
        
        index = self.data.index[np.array(self.data.loc[:,"CA"] >= start) & np.array(self.data.loc[:,"CA"] <= end)]
        #Remove nan
        index = index[np.invert(np.isnan(self.data.loc[index, "p"]))]
        
        data = self.data.iloc[index]
        data.loc[index,"V"] = self.geometry.V(data.loc[:,"CA"])
        
        return integrate.trapz(data.loc[:,"p"], x=data.loc[:,"V"])
    
    ####################################
    def plotPV(self, /,*,start:float=None, end:float=None, loglog:bool=True, timingsParams:dict=dict(), showTimings:bool=True, ax:Axes=None, **kwargs):
        """
        Create the pressure-volume diagram of the thermodynamic cycle.

        Args:
            start (float, optional): The beginning of the plot (CA). Defaults to None.
            end (float, optional): The end of the plot (CA). Defaults to None.
            loglog (bool, optional): log-log scale. Defaults to True.
            timingsParams(dict, optional): The kwargs for the scatter for timings. Defaults to:
            {
                "edgecolor":"k",
                "zorder":2,
            }
            showTimings (bool, optional): Show the timings through markers? Defaults to True.
            ax (Axes, optional): The axes to use. Defaults to None (create new figure).
        """
        #Get start and end
        if start is None:
            start = self.data.iloc[0]["CA"]
        if end is None:
            end = self.data.iloc[len(self.data)-1]["CA"]
        
        #Set default timingsParams
        default = \
        {
            "edgecolor":"k",
            "zorder":2,
        }
        [timingsParams.update({p:default[p]}) for p in default if not p in timingsParams]
        
        #Check arguments
        self.checkType(start,float,"start")
        self.checkType(end,float,"end")
        self.checkType(loglog,bool,"loglog")
        
        #Check if data were already processed:
        if not "p" in self.data.columns:
            raise ValueError("Data were not yet loaded/pre-processed (field p not present in self.data).")
        
        #Compute volume
        self.data.loc[:,"V"] = self.geometry.V(self.data.loc[:,"CA"])
        
        #Compute pressure in bar
        self.data.loc[:,"pBar"] = self.data.loc[:,"p"]/1e5
        
        #Plot
        ax = self.data.plot(x="V", y="pBar", xlabel="V [$m^3$]", ylabel="p [bar]", loglog=loglog, ax=ax, **kwargs)
        
        #Timings
        if showTimings:
            timings = self.time.timings
            size = [timingsParams.pop("markersize")]*len(timings) if "markersize" in timingsParams else None #The size of the markers
            if not "facecolor" in timingsParams: #Use same color of plot if not specified
                timingsParams["facecolor"] = ax.lines[-1]. get_color()
            ax.scatter([self.geometry.V(timings[t]) for t in timings], [self.data.pBar(timings[t]) for t in timings], s=size, **timingsParams)
        
        #Return the Axes
        return ax
    
    ####################################
    def plot(self, *args, **kwargs):
        """
        Redirects to 'self.data.plot' (See helper of DataFrame.plot method)
        """
        return self.data.plot(*args, **kwargs)
        
    ####################################
    def plotP_ROHR(self, *, label="_noLabel", legend=False, apparent=False, pkwargs:dict=None, rohrkwargs:dict=None, axes:tuple[Axes,Axes]=None, adjustLims:bool=True, **kwargs):
        """
        Plot pressure (left axis) and ROHR (right axis).
        Automatically adjust limits of y axes to minimize overlapping between 
        left and right curves. X axis limits are set to [IVC,EVO].

        Args:
            label (str, optional): the label to use in the legend. Defaults to "_noLabel".
            legend (bool, optional): wether to make the legend. Defaults to False.
            apparent (bool, optional): plot apparent heat release instead of ROHR. Defaults to False.
            adjustLims (bool, optional): Automatically adjust x/y limits. Defaults to True.
            axes (tuple[matplotlib.axes.Axes,matplotlib.axes.Axes]): tuple with pressure and ROHR axes to use. Defaults to None.
            pkwargs(dict, optional): The kwargs for the pressure plot. Defaults to:\
                {\
                    "linestyle":"-"\
                }
            rohrkwargs(dict, optional): The kwargs for the pressure plot. Defaults to:\
                {\
                    "linestyle":"--"\
                }
            **kwargs: Keyword arguments common to both plots (for example, figsize)
        """
        if not rohrkwargs is None:
            self.checkType(rohrkwargs, dict, "rohrkwargs")
            rohrkwargs = {**rohrkwargs}
        else:
            rohrkwargs = {}
        
        if not pkwargs is None:
            self.checkType(pkwargs, dict, "pkwargs")
            pkwargs = {**pkwargs}
        else:
            pkwargs = {}
        
        #Check if data were already processed:
        if not all(var in self.data.columns for var in ["p", "ROHR"]):
            raise ValueError("Data were not yet processed (fields p and ROHR not present in self.data).")
        
        #Compute pressure in bar
        self.data.loc[:,"pBar"] = self.data.loc[:,"p"]/1e5
        
        #Check axes:
        pLim_old = [float("inf"), -float("inf")]
        rohrLim_old = [float("inf"), -float("inf")]
        if not axes is None:
            self.checkType(axes, tuple, "axes")
            if not len(axes) == 2:
                raise ValueError("Entry 'axes' must be of length 2.")
            
            from matplotlib import pyplot as plt
            if not axes[0] is None:
                self.checkType(axes[0], Axes, "axes[0]")
                pLim_old = axes[0].get_ylim()
                
            if not axes[1] is None:
                self.checkType(axes[1], Axes, "axes[1]")
                rohrLim_old = axes[1].get_ylim()

            axp = axes[0]
            axrohr = axes[1]
        else:
            axp = None
            axrohr = None
        
        
        #Set default kwargs
        default_p = \
        {
            "linestyle":"-",
        }
        [pkwargs.update({p:default_p[p]}) for p in default_p if not p in pkwargs]
        [pkwargs.update({p:kwargs[p]}) for p in kwargs]
        default_rohr = \
        {
            "linestyle":"--",
        }
        [rohrkwargs.update({p:default_rohr[p]}) for p in default_rohr if not p in rohrkwargs]
        [rohrkwargs.update({p:kwargs[p]}) for p in kwargs]
        
        #Plot pressure:
        axp = self.data.plot(x="CA", y="pBar", xlabel="CA [°]", ylabel="p [bar]", ax = axp, label=label, legend=legend, **pkwargs)
        
        #Plot rohr
        if axrohr is None:
            axrohr = axp.twinx()

        #Color
        if not any(var in rohrkwargs for var in ["c", "color"]): #Use same color of pressure plot if not specified
            rohrkwargs["color"] = axp.lines[-1].get_color()

        if apparent:
            self.data.plot(x="CA", y="AHRR", xlabel="CA [°]", ylabel="AHRR [J/CA]", legend=False, label="_no", ax=axrohr, **rohrkwargs)
        else:
            self.data.plot(x="CA", y="ROHR", xlabel="CA [°]", ylabel="ROHR [J/CA]", legend=False, label="_no", ax=axrohr, **rohrkwargs)
        
        #x limits
        xlim = kwargs["xlim"] if "xlim" in kwargs else [self.time.IVC if self.time.startOfCombustion() is None else self.time.startOfCombustion(), self.time.EVO]
        axp.set_xlim(xlim)
        where = np.array(self.data.loc[:,"CA"] >= xlim[0]) & np.array(self.data.loc[:,"CA"] <= xlim[-1])
        
        #Limits
        if adjustLims:
            a = 0.02
            b = 0.2
            c = 0.7
            
            #pressure limits
            if not "ylim" in pkwargs:
                data = self.data.loc[:,"pBar"][where]
                if len(data[np.invert(np.isnan(data))]) > 0:
                    lim = [np.nanmin(data), np.nanmax(data)]
                    lim = [lim[0] - (lim[1] - lim[0])*b, lim[1] + (lim[1] - lim[0])*a]
                    lim = [min(lim[0], pLim_old[0]), max(lim[1], pLim_old[1])]
                    axp.set_ylim(lim)
            
            #ROHR limits
            if not "ylim" in rohrkwargs:
                if apparent:
                    data = self.data.loc[:,"AHRR"][where]
                else:
                    data = self.data.loc[:,"ROHR"][where]
                if len(data[np.invert(np.isnan(data))]) > 0:
                    lim = [np.nanmin(data), np.nanmax(data)]
                    lim = [lim[0] - (lim[1] - lim[0])*a, lim[1] + (lim[1] - lim[0])*c]
                    lim = [min(lim[0], rohrLim_old[0]), max(lim[1], rohrLim_old[1])]
                    axrohr.set_ylim(lim)
        
        #Return the Axes
        return (axp, axrohr)
    
    ####################################
    def save(self, file:str, *, overwrite:bool=False) -> None:
        """Save the results to an excel file 

        Args:
            file (str): The name of the file where to save
        """
        self.checkType(file, str, "file")
        self.checkType(overwrite, bool, "overwrite")
        
        if os.path.isfile(file) and not overwrite:
            raise IOError(f"File {file} exists. Run with overwrite=True to overwrite.")

        file = file if file.endswith(".xlsx") else f"{file}.xlsx"
        
        #Write the data
        with pd.ExcelWriter(file, engine = 'xlsxwriter') as writer:
            self.data._data.to_excel(writer, sheet_name="Data")
        
        #Some info
        import openpyxl
        book = openpyxl.load_workbook(file)
        sh = book.create_sheet("Info")
        sh = book["Info"]
        sh.cell(row=1, column=1).value = f"File generated with {lice.__name__} package"
        for ii, row in enumerate(self.__str__().split("\n")):
            sh.cell(row=ii+2, column=1).value = row
        book.save(file)
        
    
#########################################################################
#Create selection table
EngineModel.createRuntimeSelectionTable()
    
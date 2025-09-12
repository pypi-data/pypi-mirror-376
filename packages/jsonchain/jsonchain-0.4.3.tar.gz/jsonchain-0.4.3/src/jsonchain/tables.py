from typing import Optional
import csv
import pathlib
from openpyxl import load_workbook

def load_csv(
    csv_file: str | pathlib.Path,
) -> list[list[str | float]]:
    """
    Reads the csv table at 'csv_file'

    str values that are numeric are converted to ints or floats,
    if possible.
    """
    with open(csv_file, 'r') as file:
        return to_numeric(list(csv.reader(file)))


def load_excel_sheet(
    wb_file: str | pathlib.Path,
    sheet_indexes: Optional[int | list[int]] = None,
) -> list[list[str | float]] | dict[str, list[list[str | float]]]:
    """
    Reads the 
    """
    wb = load_workbook(wb_file)
    if isinstance(sheet_indexes, int):
        ws = wb.worksheets[sheet_indexes]
        return table_cell_values(ws)
    else:
        wb_acc = {}
        for ws in wb.worksheets:
            wb_acc.update({ws.title: table_cell_values(ws)})
        return wb_acc


def table_cell_values(
    ws: "openpyxl.Worksheet",
) -> list[list[str | float]]:
    """
    Extract cell values from a worksheet by row
    """
    table_acc = [] # outer_acc
    for row in ws.iter_rows(): # Use this method to extract data row-wise
        row_acc = [] # inner_acc
        for cell in row:
            row_acc.append(cell.value)
        table_acc.append(row_acc)
    return table_acc


def to_numeric(table: list[list[str]]) -> list[list[str | float]]:
    """
    Converts the cells in 'table' to ints and floats, if possible.
    """
    table_acc = []
    for row in table:
        row_acc = []
        for cell in row:
            try:
                cell = int(cell)
            except:
                try:  
                    cell = float(cell)
                except:
                    cell = cell
            row_acc.append(cell)
        table_acc.append(row)
    return table_acc


def transpose(table: list[list[str | float]]) -> list[tuple[str | float]]:
    """
    Returns 'table' transposed (rows become columns, columns become rows)
    """
    return list(zip(*table))


def drop_rows(table: list[list[str | float]], rows_to_drop: int | list[int]) -> list[list[str | float]]:
    """
    Returns 'table' but with the row indexes in 'rows_to_drop' omitted.
    """
    if isinstance(rows_to_drop, int):
        rows_to_drop = [rows_to_drop]
    table_acc = []
    for idx, row in enumerate(table):
        if idx not in rows_to_drop:
            table_acc.append(row)
    return table_acc


def drop_columns(table: list[list[str | float]], cols_to_drop: int | list[int]) -> list[tuple[str | float]]:
    """
    Returns 'table' but with the col indexes in 'cols_to_drop' omitted.
    """
    table_cols = transpose(table)
    if isinstance(cols_to_drop, int):
        cols_to_drop = [cols_to_drop]
    table_acc = []
    for idx, col in enumerate(table_cols):
        if idx not in cols_to_drop:
            table_acc.append(col)
    return transpose(table_acc)


def filter_table(
    table: list[list[str | float]], 
    filter_col: int, 
    filter_rule: callable
):
    """
    Filters the based on the values in 'filter_col'.
    """
    filtered_table = []
    for row in table[1:]: # Skip the header row
        if filter_rule(row[filter_col]):
            filtered_table.append(row)
    # Put header row back on
    filtered_table = [table[0]] + filtered_table
    return filtered_table


def create_tree_table(
    table: list[list[str | float]],
    ordered_tree_indexes: list[int],
) -> dict:
    """
    Returns a nested dictionary that has a depth equal to the length of 'ordered_tree_indexes'.

    'ordered_tree_indexes' are column indexes in the table that are to be used to index the resulting
        tree. 

    Any column indexes that are not in 'ordered_tree_indexes' will become part of the "leaf dictionary"
    at the end of the nested dictionary path.
    """
    tree_acc = {}
    non_tree_indexes = [idx for idx in range(len(table[0])) if idx not in ordered_tree_indexes]
    header_row = table[0]
    for row in table[1:]:
        tree_branch = None
        for tree_index in ordered_tree_indexes:
            if tree_branch is None:
                tree_acc.setdefault(row[tree_index], {})
                tree_branch = tree_acc[row[tree_index]]
            else:
                tree_branch.setdefault(row[tree_index], {})
                tree_branch = tree_branch[row[tree_index]]
                
        tree_leaves = {}
        for idx in non_tree_indexes:
            tree_leaves.update({header_row[idx]: row[idx]})
        tree_branch.update(tree_leaves)
    return tree_acc 
    